
"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@dff.org.in>
"Arun Kelkar" <arunkelkar@dff.org.in>
"Ishan Masdekar " <imasdekar@dff.org.in>
"Navin Karkera" <navin@dff.org.in>
"Parabjyot Singh" <parabjyot1996@gmail.com>
"Rahul Chaurasiya" <crahul4133@gmail.com>
"Reshma Bhatawadekar" <reshma_b@riseup.net>
"""

from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from pyramid.response import Response
import cStringIO
import calendar
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import RED

'''
This function returns a spreadsheet form of Conventional Balance Sheet Report.
In the Conventional format, capital and liabilities are shown on left side and assets on the right side.
The spreadsheet in XLSX format is generated by the frontendend.
'''
@view_config(route_name="printconvbalsheetreport" , request_param="type=conventionalbalancesheet")
def printconvbalsheetreport(request):
    try:
        calculateto = request.params["calculateto"]
        calculatefrom = request.params["calculatefrom"]
        header={"gktoken":request.headers["gktoken"]}
        fystart = str(request.params["fystart"]);
        orgname = str(request.params["orgname"])
        fyend = str(request.params["fyend"]);
        orgtype = str(request.params["orgtype"])
        result = requests.get("http://127.0.0.1:6543/report?type=balancesheet&calculateto=%s&baltype=1&calculatefrom=%s"%(calculateto,calculatefrom), headers=header)
        calculateto = calculateto[8:10]+calculateto[4:8]+calculateto[0:4]
        sources = result.json()["gkresult"]["leftlist"]
        applications = result.json()["gkresult"]["rightlist"]
        conventionalwb = openpyxl.Workbook()
        sheet = conventionalwb.active
        sheet.title = "Conventional Balance Sheet"
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 12
        sheet.column_dimensions['H'].width = 12
        sheet.merge_cells('A1:H2')
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet.merge_cells('A3:H3')
        if orgtype == "Profit Making":
            sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
            sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
            sheet['A3'] = "Conventional Balance Sheet from "+calculatefrom+" to "+calculateto
            sheet['A4'] = "Capital and Liabilities"
            sheet['A4'].font = Font(name='Liberation Serif',size='12',bold=True)
        if orgtype == "Not For Profit":
            sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
            sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
            sheet['A3'] = "Conventional Statement of Affairs as on "+calculateto
            sheet['A4'] = "Corpus and Liabilities"
            sheet['A4'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['D4'] = "Amount"
        sheet['D4'].alignment = Alignment(horizontal='right')
        sheet['D4'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['E4'] = "Property and Assets"
        sheet['E4'].alignment = Alignment(horizontal='left')
        sheet['E4'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['H4'] = "Amount"
        sheet['H4'].alignment = Alignment(horizontal='right')
        sheet['H4'].font = Font(name='Liberation Serif',size='12',bold=True)
        """ 
        Looping each dictionaries in list sources to store data in cells and apply styles.
        If the advflag = 1 then 'amount' will be displayed in 'RED' color
        """
        row = 4
        for record in sources:
            if record["groupAccname"]!="":
                if record["groupAccname"]!="Sources:":
                    if record["groupAccname"]=="Total" or record["groupAccname"]=="Sources:" or record["groupAccname"]=="Difference" :
                        sheet['A'+str(row)] = record["groupAccname"].upper()
                        sheet['A'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True)
                    elif (record["groupAccflag"]=="" and record["subgroupof"]!=""):
                        sheet['A'+str(row)] = "            "+record["groupAccname"]
                        sheet['A'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
                    elif record["groupAccflag"]==1 :
                        sheet['A'+str(row)] = "                        "+record["groupAccname"]
                        sheet['A'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
                    elif record["groupAccflag"]==2:
                        sheet['A'+str(row)] = "                        "+record["groupAccname"]
                        sheet['A'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
                    else:
                        sheet['A'+str(row)] = record["groupAccname"].upper()
                        sheet['A'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
                if record["groupAccflag"]==2 or record["groupAccflag"]==1:
                    if record["advflag"]==1:
                        sheet['B'+str(row)] = float("%.2f"%float(record["amount"]))
                        sheet['B'+str(row)].number_format = '0.00'
                        sheet['B'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['B'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                    else:
                        sheet['B'+str(row)] = float("%.2f"%float(record["amount"]))
                        sheet['B'+str(row)].number_format = '0.00'
                        sheet['B'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['B'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
                elif (record["groupAccflag"]=="" and record["subgroupof"]!=""):
                    if record["advflag"]==1:
                        sheet['C'+str(row)] = float("%.2f"%float(record["amount"]))
                        sheet['C'+str(row)].number_format = '0.00'
                        sheet['C'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['C'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                    else:
                        sheet['C'+str(row)] = float("%.2f"%float(record["amount"]))
                        sheet['C'+str(row)].number_format = '0.00'
                        sheet['C'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['C'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
                else:
                    if record["advflag"]==1:
                        sheet['D'+str(row)] = float("%.2f"%float(record["amount"]))
                        sheet['D'+str(row)].number_format = '0.00'
                        sheet['D'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['D'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                    else:
                        if record["amount"]!= "":
                            sheet['D'+str(row)] = float("%.2f"%float(record["amount"]))
                            sheet['D'+str(row)].number_format = '0.00'
                            sheet['D'+str(row)].alignment = Alignment(horizontal = 'right')
                            sheet['D'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True)
                row += 1
        """ 
        Looping each dictionaries in list applications to store data in cells and apply styles.
        If the advflag = 1 then 'amount' will be displayed in 'RED' color
        """
        row = 4
        for record in applications:
            if record["groupAccname"]!="":
                if record["groupAccname"]!="Applications:":
                    if record["groupAccname"]=="Total" or record["groupAccname"]=="Sources:" or record["groupAccname"]=="Difference" :
                        sheet['E'+str(row)] = record["groupAccname"].upper()
                        sheet['E'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True)
                    elif (record["groupAccflag"]=="" and record["subgroupof"]!=""):
                        sheet['E'+str(row)] = "            "+record["groupAccname"]
                        sheet['E'+str(row)].font = Font(name='Liberation Serif',size='12')
                    elif record["groupAccflag"]==1 :
                        sheet['E'+str(row)] = "                        "+record["groupAccname"]
                        sheet['E'+str(row)].font = Font(name='Liberation Serif',size='12')
                    elif record["groupAccflag"]==2:
                        sheet['E'+str(row)] = "                        "+record["groupAccname"]
                        sheet['E'+str(row)].font = Font(name='Liberation Serif',size='12')
                    else:
                        sheet['E'+str(row)] = record["groupAccname"].upper()
                        sheet['E'+str(row)].font = Font(name='Liberation Serif',size='12')
                        
                if record["groupAccflag"]==2 or record["groupAccflag"]==1:
                    if record["advflag"]==1:
                        sheet['F'+str(row)] = float("%.2f"%float(record["amount"]))
                        sheet['F'+str(row)].number_format = '0.00'
                        sheet['F'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['F'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                    else:
                        sheet['F'+str(row)] = float("%.2f"%float(record["amount"]))
                        sheet['F'+str(row)].number_format = '0.00'
                        sheet['F'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['F'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
                elif (record["groupAccflag"]=="" and record["subgroupof"]!=""):
                    if record["advflag"]==1:
                        sheet['G'+str(row)] = float("%.2f"%float(record["amount"]))
                        sheet['G'+str(row)].number_format = '0.00'
                        sheet['G'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['G'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                    else:
                        sheet['G'+str(row)] = float("%.2f"%float(record["amount"]))
                        sheet['G'+str(row)].number_format = '0.00'
                        sheet['G'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['G'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
                else:
                    if record["advflag"]==1:
                        sheet['H'+str(row)] = float("%.2f"%float(record["amount"]))
                        sheet['H'+str(row)].number_format = '0.00'
                        sheet['H'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['H'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                    else:
                        if record["amount"]!= "":
                            sheet['H'+str(row)] = float("%.2f"%float(record["amount"]))
                            sheet['H'+str(row)].number_format = '0.00'
                            sheet['H'+str(row)].alignment = Alignment(horizontal = 'right')
                            sheet['H'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True)
                row += 1
        output = cStringIO.StringIO()
        conventionalwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx', 'Set-Cookie':'fileDownload=true; path=/'}
        return Response(contents, headerlist=headerList.items())
    except:
        print "File not Found"
        {"gkstatus":3}

'''
This function returns a spreadsheet form of Consolidated Balance Sheet Report.
The spreadsheet in XLSX format is generated by the frontendend.
'''
@view_config(route_name="printconvbalsheetreport" , request_param="type=consolidatedbalancesheet")
def printconsbalsheetreport(request):
    try:
        calculateto = request.params["calculateto"]
        header={"gktoken":request.headers["gktoken"]}
        fystart = str(request.params["fystart"]);
        orgname = str(request.params["orgname"])
        fyend = str(request.params["fyend"]);
        orgtype = str(request.params["orgtype"])
        sorgname = json.loads(request.params["sorgname"])
        listofselectedorg = json.loads(request.params["selectedorg"])
        result = requests.get("http://127.0.0.1:6543/report?type=consolidatedbalancesheet&calculateto=%s&financialStart=%s&orgtype=%s"%(calculateto,fystart,orgtype),data=json.dumps({"listoforg":listofselectedorg}), headers=header)
        fystart = fystart[8:10] + fystart[4:8] + fystart[0:4]
        fyend = fyend[8:10] + fyend[4:8] + fyend[0:4]
        calculateto = calculateto[8:10]+calculateto[4:8]+calculateto[0:4]
        sources = result.json()["gkresult"]["leftlist"]
        applications = result.json()["gkresult"]["rightlist"]
        consolidationwb = openpyxl.Workbook()
        sheet = consolidationwb.active
        sheet.title = "Consolidated Balance Sheet"
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 12
        sheet.column_dimensions['H'].width = 12
        sheet.merge_cells('A1:H2')
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet.merge_cells('A3:H3')
        if orgtype == "Profit Making":
            sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
            sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
            sheet['A3'] = "Consolidated Balance Sheet as on "+calculateto
            sheet['A4'] = "Capital and Liabilities"
            sheet['A4'].font = Font(name='Liberation Serif',size='12',bold=True)
        if orgtype == "Not For Profit":
            sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
            sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
            sheet['A3'] = "Consolidated Statement of Affairs as on "+calculateto
            sheet['A4'] = "Corpus and Liabilities"
            sheet['A4'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['D4'] = "Amount"
        sheet['D4'].alignment = Alignment(horizontal='right')
        sheet['D4'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['E4'] = "Property and Assets"
        sheet['E4'].alignment = Alignment(horizontal='left')
        sheet['E4'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['H4'] = "Amount"
        sheet['H4'].alignment = Alignment(horizontal='right')
        sheet['H4'].font = Font(name='Liberation Serif',size='12',bold=True)
        
        """ 
        Looping each dictionaries in list sources to store data in cells and apply styles.
        If the advflag = 1 then 'amount' will be displayed in 'RED' color
        """
        row = 4
        for record in sources:
            if record["groupAccname"]!="":
                if record["groupAccname"]!="Sources:":
                    if record["groupAccname"]=="Total" or record["groupAccname"]=="Sources:" or record["groupAccname"]=="Difference" :
                        sheet['A'+str(row)] = record["groupAccname"].upper()
                        sheet['A'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True)
                    elif (record["groupAccflag"]=="" and record["subgroupof"]!=""):
                        sheet['A'+str(row)]= "             "+record["groupAccname"]
                        sheet['A'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
                    elif record["groupAccflag"]==1 :
                        sheet['A'+str(row)].font = "                         "+record["groupAccname"]
                        sheet['A'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
                    elif record["groupAccflag"]==2:
                        sheet['A'+str(row)] = "                         "+record["groupAccname"]
                        sheet['A'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
                    else:
                        sheet['A'+str(row)] = record["groupAccname"].upper()
                        sheet['A'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)

                if record["groupAccflag"]==2 or record["groupAccflag"]==1:
                    if record["advflag"]==1:
                        sheet['B'+str(row)] = record["amount"]
                        sheet['B'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['B'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                    else:
                        sheet['B'+str(row)] = record["amount"]
                        sheet['B'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['B'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
                elif (record["groupAccflag"]=="" and record["subgroupof"]!=""):
                    if record["advflag"]==1:
                        sheet['C'+str(row)] = record["amount"]
                        sheet['C'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['C'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                    else:
                        sheet['C'+str(row)] = record["amount"]
                        sheet['C'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['C'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
                else:
                    if record["advflag"]==1:
                        sheet['D'+str(row)] = record["amount"]
                        sheet['D'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['D'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                    else:
                        sheet['D'+str(row)] = record["amount"]
                        sheet['D'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['D'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True)
                row += 1

        """ 
        Looping each dictionaries in list applications to store data in cells and apply styles.
        If the advflag = 1 then 'amount' will be displayed in 'RED' color
        """
        row = 4
        for record in applications:
            if record["groupAccname"]!="":
                if record["groupAccname"]!="Applications:":
                    if record["groupAccname"]=="Total" or record["groupAccname"]=="Sources:" or record["groupAccname"]=="Difference" :
                        sheet['E'+str(row)] = record["groupAccname"].upper()
                        sheet['E'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True)
                    elif (record["groupAccflag"]=="" and record["subgroupof"]!=""):
                        sheet['E'+str(row)] = "            "+record["groupAccname"]
                        sheet['E'+str(row)].font = Font(name='Liberation Serif',size='12')
                    elif record["groupAccflag"]==1 :
                        sheet['E'+str(row)] = "            "+record["groupAccname"]
                        sheet['E'+str(row)].font = Font(name='Liberation Serif',size='12')
                    elif record["groupAccflag"]==2:
                        sheet['E'+str(row)] = "            "+record["groupAccname"]
                        sheet['E'+str(row)].font = Font(name='Liberation Serif',size='12')
                    else:
                        sheet['E'+str(row)] = record["groupAccname"].upper()
                        sheet['E'+str(row)].font = Font(name='Liberation Serif',size='12')

                if record["groupAccflag"]==2 or record["groupAccflag"]==1:
                    if record["advflag"]==1:
                        sheet['F'+str(row)] = record["amount"]
                        sheet['F'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['F'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                    else:
                        sheet['F'+str(row)] = record["amount"]
                        sheet['F'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['F'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
                elif (record["groupAccflag"]=="" and record["subgroupof"]!=""):
                    if record["advflag"]==1:
                        sheet['G'+str(row)] = record["amount"]
                        sheet['G'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['G'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                    else:
                        sheet['G'+str(row)] = record["amount"]
                        sheet['G'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['G'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
                else:
                    if record["advflag"]==1:
                        sheet['H'+str(row)] = record["amount"]
                        sheet['H'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['H'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                    else:
                        sheet['H'+str(row)] = record["amount"]
                        sheet['H'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['H'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True)
                row += 1
        output = cStringIO.StringIO()
        consolidationwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx', 'Set-Cookie':'fileDownload=true; path=/'}
        return Response(contents, headerlist=headerList.items())
    except:
        print "File not Found"
        {"gkstatus":3}
            
'''
This function returns a spreadsheet form of Statement of Sources and Applications of Funds Report(Vertical format report).
In the Vertical format capital & liabilities are shown in the upper part and assets are shown in the lower part.
The spreadsheet in XLSX format is generated by the frontendend.
'''
@view_config(route_name="printsourcesandappfundreport")
def printsourcesandappfundreport(request):
    try:
        calculateto = request.params["calculateto"]
        calculatefrom = request.params["calculatefrom"]
        header={"gktoken":request.headers["gktoken"]}
        fystart = str(request.params["fystart"])
        orgname = str(request.params["orgname"])
        fyend = str(request.params["fyend"])
        result = requests.get("http://127.0.0.1:6543/report?type=balancesheet&calculateto=%s&baltype=2&calculatefrom=%s"%(calculateto,calculatefrom), headers=header)
        calculateto = calculateto[8:10]+calculateto[4:8]+calculateto[0:4]
        sources = result.json()["gkresult"]["leftlist"]
        applications = result.json()["gkresult"]["rightlist"]
        srcandappfundwb = openpyxl.Workbook()
        sheet = srcandappfundwb.active
        sheet.title = "Src & Appl of Funds"
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 14
        sheet.column_dimensions['C'].width = 14
        sheet.column_dimensions['D'].width = 14
        sheet.merge_cells('A1:D2')
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A1'] = orgname+" (FY: "+fystart+" to "+fyend+")"
        sheet.merge_cells('A3:D3')
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A3'] = "Statement of Sources and Applications of Funds from "+calculatefrom+" to "+calculateto
        row = 4
        # Looping each dictionaries in list sources to store data in cells and apply styles.
        #If the advflag = 1 then 'amount' will be displayed in 'RED' color
        for record in sources:
            if record["groupAccname"]=="Total" or record["groupAccname"]=="Sources:" or record["groupAccname"]=="Difference" :
                sheet['A'+str(row)] = record["groupAccname"].upper()
                sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=True)
            elif (record["groupAccflag"]=="" and record["subgroupof"]!=""):
                sheet['A'+str(row)] = "            "+record["groupAccname"]
                sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            elif record["groupAccflag"]==1 :
                sheet['A'+str(row)] = "                        "+record["groupAccname"]
                sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            elif record["groupAccflag"]==2:
                sheet['A'+str(row)] = "                        "+record["groupAccname"]
                sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            else:
                sheet['A'+str(row)] = record["groupAccname"].upper()
            if record["groupAccflag"]==2 or record["groupAccflag"]==1:
                if record["advflag"]==1:
                    sheet['B'+str(row)] = float("%.2f"%float(record["amount"]))
                    sheet['B'+str(row)].number_format = '0.00'
                    sheet['B'+str(row)].alignment = Alignment(horizontal = 'right')
                    sheet['B'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                    
                else:
                    sheet['B'+str(row)] = float("%.2f"%float(record["amount"]))
                    sheet['B'+str(row)].number_format = '0.00'
                    sheet['B'+str(row)].alignment = Alignment(horizontal = 'right')
                    sheet['B'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
            elif (record["groupAccflag"]=="" and record["subgroupof"]!=""):
                if record["advflag"]==1:
                    sheet['C'+str(row)] = float("%.2f"%float(record["amount"]))
                    sheet['C'+str(row)].number_format = '0.00'
                    sheet['C'+str(row)].alignment = Alignment(horizontal = 'right')
                    sheet['C'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                else:
                    sheet['C'+str(row)] = float("%.2f"%float(record["amount"]))
                    sheet['C'+str(row)].number_format = '0.00'
                    sheet['C'+str(row)].alignment = Alignment(horizontal = 'right')
                    sheet['C'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
            else:
                if record["advflag"]==1:
                    if record["amount"]!= "":
                        sheet['D'+str(row)] = float("%.2f"%float(record["amount"]))
                        sheet['D'+str(row)].number_format = '0.00'
                        sheet['D'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['D'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                else:
                    if record["amount"]!= "":
                        sheet['D'+str(row)] = float("%.2f"%float(record["amount"]))
                        sheet['D'+str(row)].number_format = '0.00'
                        sheet['D'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['D'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True)
            row += 1
        # Looping each dictionaries in list applications to store data in cells and apply styles.
        for record in applications:
            if record["groupAccname"]=="Total" or record["groupAccname"]=="Sources:" or record["groupAccname"]=="Difference" :
                sheet['A'+str(row)] = record["groupAccname"].upper()
                sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=True)
            elif (record["groupAccflag"]=="" and record["subgroupof"]!=""):
                sheet['A'+str(row)] = "            "+record["groupAccname"]
                sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            elif record["groupAccflag"]==1 :
                sheet['A'+str(row)] ="                         "+record["groupAccname"]
                sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            elif record["groupAccflag"]==2:
                sheet['A'+str(row)] = "                        "+record["groupAccname"]
                sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            else:
                sheet['A'+str(row)] = record["groupAccname"].upper()
                sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            if record["groupAccflag"]==2 or record["groupAccflag"]==1:
                if record["advflag"]==1:
                    sheet['B'+str(row)] = float("%.2f"%float(record["amount"]))
                    sheet['B'+str(row)].number_format = '0.00'
                    sheet['B'+str(row)].alignment = Alignment(horizontal = 'right')
                    sheet['B'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                else:
                    sheet['B'+str(row)] = float("%.2f"%float(record["amount"]))
                    sheet['B'+str(row)].number_format = '0.00'
                    sheet['B'+str(row)].alignment = Alignment(horizontal = 'right')
                    sheet['B'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
            elif (record["groupAccflag"]=="" and record["subgroupof"]!=""):
                if record["advflag"]==1:
                    sheet['C'+str(row)] = float("%.2f"%float(record["amount"]))
                    sheet['C'+str(row)].number_format = '0.00'
                    sheet['C'+str(row)].alignment = Alignment(horizontal = 'right')
                    sheet['C'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                else:
                    sheet['C'+str(row)] = float("%.2f"%float(record["amount"]))
                    sheet['C'+str(row)].number_format = '0.00'
                    sheet['C'+str(row)].alignment = Alignment(horizontal = 'right')
                    sheet['C'+str(row)].font = Font(name='Liberation Serif',size='12',bold=False)
            else:
                if record["advflag"]==1:
                    if record["amount"]!= "":
                        sheet['D'+str(row)] = float("%.2f"%float(record["amount"]))
                        sheet['D'+str(row)].number_format = '0.00'
                        sheet['D'+str(row)].alignment = Alignment(horizontal = 'right')
                        sheet['D'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True,color=RED)
                else:
                    if record["amount"]!= "":
                        sheet['D'+str(row)] = float("%.2f"%float(record["amount"]))
                        sheet['D'+str(row)].number_format = '0.00'
                    sheet['D'+str(row)].alignment = Alignment(horizontal = 'right')
                    sheet['D'+str(row)].font = Font(name='Liberation Serif',size='12',bold=True)
            row += 1

        output = cStringIO.StringIO()
        srcandappfundwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx', 'Set-Cookie':'fileDownload=true; path=/'}
        return Response(contents, headerlist=headerList.items())
    except:
        print "File note found"
        {"gkstatus":3}


@view_config(route_name="showbalancesheet", renderer="gkwebapp:templates/viewbalancesheet.jinja2")
def showbalancesheet(request):
    return {"gkstatus":0}

@view_config(route_name="showbalancesheetreport")
def showbalancesheetreport(request):
    calculateto = request.params["calculateto"]
    calculatefrom=request.params["calculatefrom"]
    balancesheettype = request.params["balancesheettype"]
    orgtype = request.params["orgtype"]
    header={"gktoken":request.headers["gktoken"]}
    flag = request.params["flag"]    
    if balancesheettype == "conventionalbalancesheet":
        result = requests.get("http://127.0.0.1:6543/report?type=balancesheet&calculateto=%s&baltype=1&calculatefrom=%s"%(calculateto,calculatefrom), headers=header)
        return render_to_response("gkwebapp:templates/conventionalbalancesheetreport.jinja2",{"records":result.json()["gkresult"],"balancesheettype":"verticalbalancesheet","to":calculateto,"from":calculatefrom,"orgtype":orgtype,"flag":flag},request=request)
    if balancesheettype == "verticalbalancesheet":
        result = requests.get("http://127.0.0.1:6543/report?type=balancesheet&calculateto=%s&baltype=2&calculatefrom=%s"%(calculateto,calculatefrom), headers=header)
        return render_to_response("gkwebapp:templates/sourcesandapplicationoffundsreport.jinja2",{"records":result.json()["gkresult"],"balancesheettype":"conventionalbalancesheet","to":calculateto,"from":calculatefrom,"orgtype":orgtype},request=request)


@view_config(route_name="allorgcode",request_param="type=orgcodelist", renderer="json")
def listoforg(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/organisations?type=orgcodelist", headers=header)
    orgdata=[]
    for record in result.json()["gkdata"]:
        gdata= {"orgname":str(record["orgname"]),"orgtype":str(record["orgtype"]),"orgcode":str(record["orgcode"]),"yearstart":str(record["yearstart"]),"yearend":str(record["yearend"])}
        orgdata.append(gdata)
    return {"gkresult":orgdata}

@view_config(route_name="listoforgselected",request_param="type=orgselected")
def selectedorg(request):
    header={"gktoken":request.headers["gktoken"]}
    selectedorg = request.params["selectedorg"]
    flag = request.params["flag"]
    selectedorg = json.loads(selectedorg)
    listofselectedorg = json.loads(selectedorg["ds"])
    calculateto = selectedorg["calculateto"]
    orgcode = selectedorg["orgcode"]
    financialStart = selectedorg["financialStart"]
    orgtype = selectedorg["orgtype"]
    horgname = request.params["horgname"]
    sorgname = json.loads(request.params["sorgname"])
    slen = len(sorgname)
    result = requests.get("http://127.0.0.1:6543/report?type=consolidatedbalancesheet&calculateto=%s&financialStart=%s&orgtype=%s"%(calculateto,financialStart,orgtype),data=json.dumps({"listoforg":listofselectedorg}), headers=header)
    return render_to_response("gkwebapp:templates/conventionalbalancesheetreport.jinja2",{"records":result.json()["gkresult"],"orglist":listofselectedorg,"to":calculateto,"orgtype":orgtype,"flag":flag,"horgname":horgname,"sorgname":sorgname,"slen":slen,"financialStart":financialStart,"orgcode":orgcode,"selectedorg":selectedorg},request=request)

@view_config(route_name="showconsolidationpopup", renderer="gkwebapp:templates/showconsolidationpopup.jinja2")
def showpopup(request):
    return {"gkresult":request.params}
