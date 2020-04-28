
"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020,2019 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@dff.org.in>
"Arun Kelkar" <arunkelkar@dff.org.in>
"Ishan Masdekar " <imasdekar@dff.org.in>
"Navin Karkera" <navin@dff.org.in>
"Bhavesh Bawadhane" <bbhavesh07@gmail.com>
'Prajkta Patkar' <prajkta@gnukhata.in>
"""

from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from pyramid.response import Response
import io
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment
import calendar

@view_config(route_name = "printprofitandloss", renderer = "")
def printprofitandloss(request):
    calculatefrom = request.params["calculatefrom"]
    calculateto = request.params["calculateto"]
    orgtype = request.params["orgtype"]
    header={"gktoken":request.headers["gktoken"]}
    fystart = str(request.params["fystart"])
    fyend = str(request.params["fyend"])
    orgname = str(request.params["orgname"])
    result = requests.get("http://127.0.0.1:6543/report?type=profitloss&calculatefrom=%s&calculateto=%s"%(calculatefrom,calculateto), headers=header)
    calculatefrom = calculatefrom[8:10]+calculatefrom[4:8]+calculatefrom[0:4]
    calculateto = calculateto[8:10]+calculateto[4:8]+calculateto[0:4]
    DirectIncome = result.json()["gkresult"]["Direct Income"]
    InDirectIncome = result.json()["gkresult"]["Indirect Income"]
    DirectExpense = result.json()["gkresult"]["Direct Expense"]
    InDirectExpense = result.json()["gkresult"]["Indirect Expense"]
    gross = {}
    try:
        gross["gspCF"] = result.json()["gkresult"]["grossprofitcf"]
    except:
        gross["gslCF"] = result.json()["gkresult"]["grosslosscf"]
    net = {}
    try:
        net["netprofit"] = result.json()["gkresult"]["netprofit"]
    except:
        net["netloss"] = result.json()["gkresult"]["netloss"]
    Total = result.json()["gkresult"]["Total"]
    # A workbook is opened.
    pandlwb = openpyxl.Workbook()
    # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
    sheet = pandlwb.active
    # Title of the sheet and width of columns are set.
    if orgtype == "Profit Making":
        sheet.title = 'Profit and Loss'
    else:
        sheet.title = 'Income and Expenditure'
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 16
    # Cells of first two rows are merged to display organisation details properly.
    sheet.merge_cells('A1:D2')
    # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
    sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
    sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
    # Organisation name and financial year are displayed.
    sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
    sheet.merge_cells('A3:D3')
    sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
    sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
    # Setting heading of spreadsheet
    if orgtype == "Profit Making":
        sheet['A3'] = 'Profit and Loss (' + calculatefrom + ' to ' + calculateto + ')'
    else:
        sheet['A3'] = 'Income and Expenditure (' + calculatefrom + ' to ' + calculateto + ')'
    # Setting title for columns.
    sheet['A4'] = 'Particulars'
    sheet['B4'] = 'Amount'
    sheet['B4'].alignment = Alignment(horizontal = "right")
    sheet['A4'].font = Font(name='Liberation Serif',size=12,bold=True)
    sheet['B4'].font = Font(name='Liberation Serif',size=12,bold=True)
    sheet['C4'] = 'Particulars'
    sheet['D4'] = 'Amount'
    sheet['D4'].alignment = Alignment(horizontal = "right")
    sheet['C4'].font = Font(name='Liberation Serif',size=12,bold=True)
    sheet['D4'].font = Font(name='Liberation Serif',size=12,bold=True)
    '''
    Conventions for presenting data:-
    Group - Bold, Capitals
    Subgroup - Normal, Capitals
    Account - Normal, Italics
    Spaces are used for indentation.
    '''
    #Loiading data for Direct Expense group
    sheet['A5'] = "DIRECT EXPENSE"
    sheet['B5'] = DirectExpense["direxpbal"]
    sheet["B5"]=float("%.2f"%float(DirectExpense["direxpbal"]))
    sheet["B5"].number_format="0.00"
    sheet["B5"].alignment = Alignment(horizontal = "right")
    sheet['A5'].font = Font(name='Liberation Serif',size=12,bold=True)
    sheet['B5'].font = Font(name='Liberation Serif',size=12,bold=True)
    row = 6
    # If Purchase accounts are there they are displayed on the top
    if "Purchase" in DirectExpense:
        sheet["A" + str(row)] = "        PURCHASE"
        sheet["A" + str(row)].font = Font(name='Liberation Serif',size=12,italic=False)
        sheet["B" + str(row)] = DirectExpense["Purchase"]["balance"]
        sheet["B" + str(row)]=float("%.2f"%float(DirectExpense["Purchase"]["balance"]))
        sheet["B" + str(row)].number_format="0.00"
        sheet["B" + str(row)].alignment = Alignment(horizontal = "right")
        row = row + 1
        # Purchase accounts
        for purchaseaccount in DirectExpense["Purchase"]:
            if purchaseaccount != "balance":
                sheet["A" + str(row)] = "                " + purchaseaccount
                sheet["A" + str(row)].font = Font(name='Liberation Serif',size=12,italic=True)
                sheet["B" + str(row)] = DirectExpense["Purchase"][purchaseaccount]
                sheet["B" + str(row)]=float("%.2f"%float(DirectExpense["Purchase"][purchaseaccount]))
                sheet["B" + str(row)].number_format="0.00"
                sheet["B" + str(row)].alignment = Alignment(horizontal = "right")
                row = row + 1
    #Loading other subgroups
    for subgroup in DirectExpense:
        if subgroup != "Purchase" and "balance" in DirectExpense[subgroup]:
            try:
                sheet["A" + str(row)] = "        " + subgroup.upper()
            except:
                sheet["A" + str(row)] = "        " + subgroup
            sheet["A" + str(row)].font = Font(name='Liberation Serif',size=12,italic=False)
            sheet["B" + str(row)] = DirectExpense[subgroup]["balance"]
            sheet["B" + str(row)]=float("%.2f"%float(DirectExpense[subgroup]["balance"]))
            sheet["B" + str(row)].number_format="0.00"
            sheet["B" + str(row)].alignment = Alignment(horizontal = "right")
            row = row + 1
            #Loading accounts of each subgroup
            for subgroupaccount in DirectExpense[subgroup]:
                if subgroupaccount != "balance":
                    sheet["A" + str(row)] = "                " + subgroupaccount
                    sheet["A" + str(row)].font = Font(name='Liberation Serif',size=12,italic=True)
                    sheet["B" + str(row)] = DirectExpense[subgroup][subgroupaccount]
                    sheet["B" + str(row)]=float("%.2f"%float(DirectExpense[subgroup][subgroupaccount]))
                    sheet["B" + str(row)].number_format="0.00"
                    sheet["B" + str(row)].alignment = Alignment(horizontal = "right")
                    row = row + 1
        # Loading accounts that are not part of any subgroup
        if subgroup != "Purchase" and subgroup != "direxpbal" and "balance" not in DirectExpense[subgroup]:
            sheet["A" + str(row)] = "        " + subgroup
            sheet["A" + str(row)].font = Font(name='Liberation Serif',size=12,italic=True)
            sheet["B" + str(row)] = DirectExpense[subgroup]
            sheet["B" + str(row)]=float("%.2f"%float(DirectExpense[subgroup]))
            sheet["B" + str(row)].number_format="0.00"
            sheet["B" + str(row)].alignment = Alignment(horizontal = "right")
            row = row + 1

    #If there is Gross Profit it is shown in Expense side
    if "gspCF" in gross:
        sheet["A" + str(row)] = "Gross Profit C/F"
        sheet["A" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet["B" + str(row)] = gross["gspCF"]
        sheet["B" + str(row)]=float("%.2f"%float(gross["gspCF"]))
        sheet["B" + str(row)].number_format="0.00"
        sheet["B" + str(row)].alignment = Alignment(horizontal = "right")
        sheet["B" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)
        row = row + 1

    #Loading data for Indirect Expense group.
    sheet["A" + str(row)] = "INDIRECT EXPENSE"
    sheet["B" + str(row)] = InDirectExpense["indirexpbal"]
    sheet["B" + str(row)]=float("%.2f"%float(InDirectExpense["indirexpbal"]))
    sheet["B" + str(row)].number_format="0.00"
    sheet["B" + str(row)].alignment = Alignment(horizontal = "right")
    sheet["A" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)
    sheet["B" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)
    row = row + 1
    
    for subgroup in InDirectExpense:
        if "balance" in InDirectExpense[subgroup]:
            try:
                sheet["A" + str(row)] = "        " + subgroup.upper()
            except:
                sheet["A" + str(row)] = "        " + subgroup
            sheet["A" + str(row)].font = Font(name='Liberation Serif',size=12,italic=False)
            sheet["B" + str(row)] = InDirectExpense[subgroup]["balance"]
            sheet["B" + str(row)]=float("%.2f"%float(InDirectExpense[subgroup]["balance"]))
            sheet["B" + str(row)].number_format="0.00"
            sheet["B" + str(row)].alignment = Alignment(horizontal = "right")
            row = row + 1
            for subgroupaccount in InDirectExpense[subgroup]:
                if subgroupaccount != "balance":
                    sheet["A" + str(row)] = "                " + subgroupaccount
                    sheet["A" + str(row)].font = Font(name='Liberation Serif',size=12,italic=True)
                    sheet["B" + str(row)] = InDirectExpense[subgroup][subgroupaccount]
                    sheet["B" + str(row)]=float("%.2f"%float(InDirectExpense[subgroup][subgroupaccount]))
                    sheet["B" + str(row)].number_format="0.00"
                    sheet["B" + str(row)].alignment = Alignment(horizontal = "right")
                    row = row + 1
        if subgroup != "indirexpbal" and "balance" not in InDirectExpense[subgroup]:
            sheet["A" + str(row)] = "        " + subgroup
            sheet["A" + str(row)].font = Font(name='Liberation Serif',size=12,italic=True)
            sheet["B" + str(row)] = InDirectExpense[subgroup]
            sheet["B" + str(row)]=float("%.2f"%float(InDirectExpense[subgroup]))
            sheet["B" + str(row)].number_format="0.00"
            sheet["B" + str(row)].alignment = Alignment(horizontal = "right")
            row = row + 1

    #If there is Net Profit it is shown in Expense side
    if "netprofit" in net:
        sheet["A" + str(row)] = "Net Profit"
        sheet["A" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet["B" + str(row)]=float("%.2f"%float(net["netprofit"]))
        sheet["B" + str(row)].number_format="0.00"
        sheet["B" + str(row)].alignment = Alignment(horizontal = "right")
        sheet["B" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)
    row = row + 1
    sheet["A" + str(row)] = "Total"
    sheet["A" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)
    sheet["B" + str(row)] = Total
    sheet["B" + str(row)]=float("%.2f"%float(Total))
    sheet["B" + str(row)].number_format="0.00"
    sheet["B" + str(row)].alignment = Alignment(horizontal = "right")
    sheet["B" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)

    #Loading data for Direct Income group
    sheet['C5'] = "DIRECT INCOME"
    sheet['D5'] = DirectIncome["dirincmbal"]
    sheet["D5"]=float("%.2f"%float(DirectIncome["dirincmbal"]))
    sheet["D5"].number_format="0.00"
    sheet['D5'].alignment = Alignment(horizontal = "right")
    sheet["C5"].font = Font(name='Liberation Serif',size=12,bold=True)
    sheet["D5"].font = Font(name='Liberation Serif',size=12,bold=True)
    row = 6
    # If Sales accounts are there they are displayed on the top
    if "Sales" in DirectIncome:
        sheet["C" + str(row)] = "        SALES"
        sheet["C" + str(row)].font = Font(name='Liberation Serif',size=12,italic=False)
        sheet["D" + str(row)] = DirectIncome["Sales"]["balance"]
        sheet["D" + str(row)]=float("%.2f"%float(DirectIncome["Sales"]["balance"]))
        sheet["D" + str(row)].number_format="0.00"
        sheet["D" + str(row)].alignment = Alignment(horizontal = "right")
        row = row + 1
    
        for salesaccount in DirectIncome["Sales"]:
            if salesaccount != "balance":
                sheet["C" + str(row)] = "                " + salesaccount
                sheet["C" + str(row)].font = Font(name='Liberation Serif',size=12,italic=True)
                sheet["D" + str(row)] = DirectIncome["Sales"][salesaccount]
                sheet["D" + str(row)]=float("%.2f"%float(DirectIncome["Sales"][salesaccount]))
                sheet["D" + str(row)].number_format="0.00"
                sheet["D" + str(row)].alignment = Alignment(horizontal = "right")
                row = row + 1

    for subgroup in DirectIncome:
        if subgroup != "Sales" and "balance" in DirectIncome[subgroup]:
            try:
                sheet["C" + str(row)] = "        " + subgroup.upper()
            except:
                sheet["C" + str(row)] = "        " + subgroup
            sheet["C" + str(row)].font = Font(name='Liberation Serif',size=12,italic=False)
            sheet["D" + str(row)] = DirectIncome[subgroup]["balance"]
            sheet["D" + str(row)]=float("%.2f"%float(DirectIncome[subgroup]["balance"]))
            sheet["D" + str(row)].number_format="0.00"
            sheet["D" + str(row)].alignment = Alignment(horizontal = "right")
            row = row + 1
            for subgroupaccount in DirectIncome[subgroup]:
                if subgroupaccount != "balance":
                    sheet["C" + str(row)] = "                " + subgroupaccount
                    sheet["C" + str(row)].font = Font(name='Liberation Serif',size=12,italic=True)
                    sheet["D" + str(row)] = DirectIncome[subgroup][subgroupaccount]
                    sheet["D" + str(row)]=float("%.2f"%float(DirectIncome[subgroup][subgroupaccount]))
                    sheet["D" + str(row)].number_format="0.00"
                    sheet["D" + str(row)].alignment = Alignment(horizontal = "right")
                    row = row + 1
        if subgroup != "Sales" and subgroup != "dirincmbal" and "balance" not in DirectIncome[subgroup]:
            sheet["C" + str(row)] = "        " + subgroup
            sheet["C" + str(row)].font = Font(name='Liberation Serif',size=12,italic=True)
            sheet["D" + str(row)] = DirectIncome[subgroup]
            sheet["D" + str(row)]=float("%.2f"%float(DirectIncome[subgroup]))
            sheet["D" + str(row)].number_format="0.00"
            sheet["D" + str(row)].alignment = Alignment(horizontal = "right")
            row = row + 1
    #If there is Gross Loss it is shown in Expense side
    if "gslCF" in gross:
        sheet["C" + str(row)] = "Gross Loss C/F"
        sheet["C" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet["D" + str(row)] = gross["gslCF"]
        sheet["D" + str(row)]=float("%.2f"%float(gross["gslCF"]))
        sheet["D" + str(row)].number_format="0.00"
        sheet["D" + str(row)].alignment = Alignment(horizontal = "right")
        sheet["D" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)
        row = row + 1
    sheet["C" + str(row)] = "Closing Stock"
    sheet["C" + str(row)].font = Font(name='Liberation Serif',size=12,italic=True)
    sheet["D" + str(row)]=float("%.2f"%float(result.json()["gkresult"]["Closing Stock"]))
    sheet["D" + str(row)].number_format="0.00"
    sheet["D" + str(row)].alignment = Alignment(horizontal = "right")
    row = row + 1
            
    #Loading data for Indirect Income group.
    sheet["C" + str(row)] = "INDIRECT INCOME"
    sheet["D" + str(row)] = InDirectIncome["indirincmbal"]
    sheet["D" + str(row)]=float("%.2f"%float(InDirectIncome["indirincmbal"]))
    sheet["D" + str(row)].number_format="0.00"
    sheet["D" + str(row)].alignment = Alignment(horizontal = "right")
    sheet["C" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)
    sheet["D" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)
    row = row + 1
    
    for subgroup in InDirectIncome:
        if "balance" in InDirectIncome[subgroup]:
            try:
                sheet["C" + str(row)] = "        " + subgroup.upper()
            except:
                sheet["C" + str(row)] = "        " + subgroup
            sheet["C" + str(row)].font = Font(name='Liberation Serif',size=12,italic=False)
            sheet["D" + str(row)] = InDirectIncome[subgroup]["balance"]
            sheet["D" + str(row)]=float("%.2f"%float(InDirectIncome[subgroup]["balance"]))
            sheet["D" + str(row)].number_format="0.00"
            sheet["D" + str(row)].alignment = Alignment(horizontal = "right")
            row = row + 1
            for subgroupaccount in InDirectIncome[subgroup]:
                if subgroupaccount != "balance":
                    sheet["C" + str(row)] = "                " + subgroupaccount
                    sheet["C" + str(row)].font = Font(name='Liberation Serif',size=12,italic=True)
                    sheet["D" + str(row)] = InDirectIncome[subgroup][subgroupaccount]
                    sheet["D" + str(row)]=float("%.2f"%float(InDirectIncome[subgroup][subgroupaccount]))
                    sheet["D" + str(row)].number_format="0.00"
                    sheet["D" + str(row)].alignment = Alignment(horizontal = "right")
                    row = row + 1
        if subgroup != "indirincmbal" and "balance" not in InDirectIncome[subgroup]:
            sheet["C" + str(row)] = "        " + subgroup
            sheet["C" + str(row)].font = Font(name='Liberation Serif',size=12,italic=True)
            sheet["D" + str(row)] = InDirectIncome[subgroup]
            sheet["D" + str(row)]=float("%.2f"%float(InDirectIncome[subgroup]))
            sheet["D" + str(row)].number_format="0.00"
            sheet["D" + str(row)].alignment = Alignment(horizontal = "right")
            row = row + 1

    #If there is Net Loss it is shown in Income side
    if "netloss" in net:
        sheet["C" + str(row)] = "Net Loss"
        sheet["C" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet["D" + str(row)]=float("%.2f"%float(net["netloss"]))
        sheet["D" + str(row)].number_format="0.00"
        sheet["D" + str(row)].alignment = Alignment(horizontal = "right")
        sheet["D" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)
    row = row + 1
    sheet["C" + str(row)] = "Total"
    sheet["C" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)
    sheet["D" + str(row)] = Total
    sheet["D" + str(row)]=float("%.2f"%float(Total))
    sheet["D" + str(row)].number_format="0.00"
    sheet["D" + str(row)].alignment = Alignment(horizontal = "right")
    sheet["D" + str(row)].font = Font(name='Liberation Serif',size=12,bold=True)
    output = io.BytesIO()
    pandlwb.save(output)
    contents = output.getvalue()
    output.close()
    headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx', 'X-Content-Type-Options':'nosniff','Set-Cookie':'fileDownload=true; path=/ [;HttpOnly]'}
    # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true; path=/'}

    return Response(contents, headerlist=list(headerList.items()))


@view_config(route_name="showprofitloss", renderer="gkwebapp:templates/viewprofitloss.jinja2")
def showprofitloss(request):
    orgtype = request.params["orgtype"]
    return {"gkstatus":0,"orgtype":orgtype}

@view_config(route_name="showprofitlossreport")
def showprofitlossreport(request):
    calculateto = request.params["calculateto"]
    calculatefrom = request.params["calculatefrom"]
    orgtype = request.params["orgtype"]
    header={"gktoken":request.headers["gktoken"]}

    result = requests.get("http://127.0.0.1:6543/report?type=profitloss&calculatefrom=%s&calculateto=%s"%(calculatefrom,calculateto), headers=header)
    DirectIncome = result.json()["gkresult"]["Direct Income"]
    InDirectIncome = result.json()["gkresult"]["Indirect Income"]
    DirectExpense = result.json()["gkresult"]["Direct Expense"]
    InDirectExpense = result.json()["gkresult"]["Indirect Expense"]
    gross = {}
    try:
        gross["gspCF"] = result.json()["gkresult"]["grossprofitcf"]
    except:
        gross["gslCF"] = result.json()["gkresult"]["grosslosscf"]
    
    net = {}
    try:
        net["netprofit"] = result.json()["gkresult"]["netprofit"]
    except:
        net["netloss"] = result.json()["gkresult"]["netloss"]
    Total = result.json()["gkresult"]["Total"]    
    return render_to_response("gkwebapp:templates/profitlossreport.jinja2",{"DirectIncome":DirectIncome,"ClosingStock":result.json()["gkresult"]["Closing Stock"],"InDirectIncome":InDirectIncome,"DirectExpense":DirectExpense,"InDirectExpense":InDirectExpense,"net":net,"gross":gross,"orgtype":orgtype,"from":datetime.strftime(datetime.strptime(str(calculatefrom),"%Y-%m-%d").date(),'%d-%m-%Y'),"to":datetime.strftime(datetime.strptime(str(calculateto),"%Y-%m-%d").date(),'%d-%m-%Y'),"Total":Total},request=request)

@view_config(route_name="showprofitlossreport", request_param="type=print", renderer="gkwebapp:templates/printprofitlossreport.jinja2")
def printprofitloss(request):
    calculateto = request.params["calculateto"]
    calculatefrom = request.params["calculatefrom"]
    orgtype = request.params["orgtype"]
    header={"gktoken":request.headers["gktoken"]}

    result = requests.get("http://127.0.0.1:6543/report?type=profitloss&calculatefrom=%s&calculateto=%s"%(calculatefrom,calculateto), headers=header)
    DirectIncome = result.json()["gkresult"]["Direct Income"]
    InDirectIncome = result.json()["gkresult"]["Indirect Income"]
    DirectExpense = result.json()["gkresult"]["Direct Expense"]
    InDirectExpense = result.json()["gkresult"]["Indirect Expense"]
    gross = {}
    try:
        gross["gspCF"] = result.json()["gkresult"]["grossprofitcf"]
    except:
        gross["gslCF"] = result.json()["gkresult"]["grosslosscf"]
    net = {}
    try:
        net["netprofit"] = result.json()["gkresult"]["netprofit"]
    except:
        net["netloss"] = result.json()["gkresult"]["netloss"]
    Total = result.json()["gkresult"]["Total"]
    return {"DirectIncome":DirectIncome,"InDirectIncome":InDirectIncome,"DirectExpense":DirectExpense,"ClosingStock":result.json()["gkresult"]["Closing Stock"],"InDirectExpense":InDirectExpense,"net":net,"gross":gross,"orgtype":orgtype,"from":datetime.strftime(datetime.strptime(str(calculatefrom),"%Y-%m-%d").date(),'%d-%m-%Y'),"to":datetime.strftime(datetime.strptime(str(calculateto),"%Y-%m-%d").date(),'%d-%m-%Y'),"Total":Total}
