
"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@dff.org.in>
"Arun Kelkar" <arunkelkar@dff.org.in>
"Ishan Masdekar " <imasdekar@dff.org.in>
"Navin Karkera" <navin@dff.org.in>
"Vaibhav Kurhe" <vaibspidy@openmailbox.org>
"""

from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from odslib import ODS
from pyramid.response import Response
import os
import calendar
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import RED

@view_config(route_name="showtrialbalance", renderer="gkwebapp:templates/viewtrialbalance.jinja2")
def showtrialbalance(request):
    return {"gkstatus":0}

@view_config(route_name="printtrialbalancereport")
def printtrialbalancereport(request):
    calculateto = request.params["calculateto"]
    financialstart = request.params["financialstart"]
    trialbalancetype = int(request.params["trialbalancetype"])

    header={"gktoken":request.headers["gktoken"]}
    if trialbalancetype == 1:
        result = requests.get("http://127.0.0.1:6543/report?type=nettrialbalance&calculateto=%s&financialstart=%s"%(calculateto,financialstart), headers=header)
        return render_to_response("gkwebapp:templates/printnettrialbalance.jinja2",{"records":result.json()["gkresult"],"trialbalancetype":1,"from":datetime.strftime(datetime.strptime(str(financialstart),"%Y-%m-%d").date(),'%d-%m-%Y'),"to":datetime.strftime(datetime.strptime(str(calculateto),"%Y-%m-%d").date(),'%d-%m-%Y')},request=request)
    elif trialbalancetype == 2:
        result = requests.get("http://127.0.0.1:6543/report?type=extendedtrialbalance&calculateto=%s&financialstart=%s"%(calculateto,financialstart), headers=header)
        return render_to_response("gkwebapp:templates/printgrosstrialbalance.jinja2",{"records":result.json()["gkresult"],"trialbalancetype":2,"from":datetime.strftime(datetime.strptime(str(financialstart),"%Y-%m-%d").date(),'%d-%m-%Y'),"to":datetime.strftime(datetime.strptime(str(calculateto),"%Y-%m-%d").date(),'%d-%m-%Y')},request=request)
    elif trialbalancetype == 3:
        result = requests.get("http://127.0.0.1:6543/report?type=extendedtrialbalance&calculateto=%s&financialstart=%s"%(calculateto,financialstart), headers=header)
        return render_to_response("gkwebapp:templates/printextendedtrialbalance.jinja2",{"records":result.json()["gkresult"],"trialbalancetype":3,"from":datetime.strftime(datetime.strptime(str(financialstart),"%Y-%m-%d").date(),'%d-%m-%Y'),"to":datetime.strftime(datetime.strptime(str(calculateto),"%Y-%m-%d").date(),'%d-%m-%Y')},request=request)


@view_config(route_name="showtrialbalancereport")
def showtrialbalancereport(request):
    calculateto = request.params["calculateto"]
    financialstart = request.params["financialstart"]
    trialbalancetype = int(request.params["trialbalancetype"])

    header={"gktoken":request.headers["gktoken"]}
    if trialbalancetype == 1:
        result = requests.get("http://127.0.0.1:6543/report?type=nettrialbalance&calculateto=%s&financialstart=%s"%(calculateto,financialstart), headers=header)
        return render_to_response("gkwebapp:templates/nettrialbalance.jinja2",{"records":result.json()["gkresult"],"trialbalancetype":1,"from":datetime.strftime(datetime.strptime(str(financialstart),"%Y-%m-%d").date(),'%d-%m-%Y'),"to":datetime.strftime(datetime.strptime(str(calculateto),"%Y-%m-%d").date(),'%d-%m-%Y')},request=request)
    elif trialbalancetype == 2:
        result = requests.get("http://127.0.0.1:6543/report?type=extendedtrialbalance&calculateto=%s&financialstart=%s"%(calculateto,financialstart), headers=header)
        return render_to_response("gkwebapp:templates/grosstrialbalance.jinja2",{"records":result.json()["gkresult"],"trialbalancetype":2,"from":datetime.strftime(datetime.strptime(str(financialstart),"%Y-%m-%d").date(),'%d-%m-%Y'),"to":datetime.strftime(datetime.strptime(str(calculateto),"%Y-%m-%d").date(),'%d-%m-%Y')},request=request)
    elif trialbalancetype == 3:
        result = requests.get("http://127.0.0.1:6543/report?type=extendedtrialbalance&calculateto=%s&financialstart=%s"%(calculateto,financialstart), headers=header)
        return render_to_response("gkwebapp:templates/extendedtrialbalance.jinja2",{"records":result.json()["gkresult"],"trialbalancetype":3,"from":datetime.strftime(datetime.strptime(str(financialstart),"%Y-%m-%d").date(),'%d-%m-%Y'),"to":datetime.strftime(datetime.strptime(str(calculateto),"%Y-%m-%d").date(),'%d-%m-%Y')},request=request)


'''
This function returns a spreadsheet form of Trial Balance Report.
The spreadsheet in XLSX format is generated by the frontendend.
'''
@view_config(route_name="printtrialbalance", renderer ="json")
def printtrialbalance(request):
    try :
        header = {"gktoken": request.headers["gktoken"]}
        orgname = request.params["orgname"]
        financialstart= request.params["financialstart"]
        fyend = str(request.params["fyend"])
        startdate =request.params["financialstart"]
        calculateto =request.params["calculateto"]
        trialbalancetype = int(request.params["trialbalancetype"])
        if trialbalancetype == 1:
            result = requests.get("http://127.0.0.1:6543/report?type=nettrialbalance&calculateto=%s&financialstart=%s"%(calculateto,financialstart), headers=header)
        elif trialbalancetype == 2:
            result = requests.get("http://127.0.0.1:6543/report?type=extendedtrialbalance&calculateto=%s&financialstart=%s"%(calculateto,financialstart), headers=header)
        elif trialbalancetype == 3:
            result = requests.get("http://127.0.0.1:6543/report?type=extendedtrialbalance&calculateto=%s&financialstart=%s"%(calculateto,financialstart), headers=header)
        records = result.json()["gkresult"]
        # A workbook is opened.
        trialbalancewb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = trialbalancewb.active
        # Title of the sheet self.assertNotIn(member, container)d width of columns are set.
        sheet.title = "Trial Balance of %s" %(str(orgname))
        # Condition for Net Trial Balance
        if trialbalancetype == 1:
            sheet.column_dimensions['A'].width = 8
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 14
            sheet.column_dimensions['D'].width = 16
            sheet.column_dimensions['E'].width = 22
            # Cells of first two rows are merged to display organisation details properly.
            sheet.merge_cells('A1:E2')
            # Name and Financial Year of organisation is fetched to be displayed on the first row.
            sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
            sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
            # Organisation name and financial year are displayed.
            sheet['A1'] = orgname + ' (FY: ' +  datetime.strptime(str(financialstart),"%Y-%m-%d").strftime("%d-%m-%Y") + ' to ' + fyend +')'
            sheet.merge_cells('A3:F3')
            sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
            sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
            sheet['A3'] = 'Net Trial Balance for the period from %s to %s' %( datetime.strptime(str(startdate),"%Y-%m-%d").strftime("%d-%m-%Y"), datetime.strptime(str(calculateto),"%Y-%m-%d").strftime("%d-%m-%Y"))
            sheet['A4'] = 'Sr. No.'
            sheet['B4'] = 'Account Name'
            sheet['C4'] = 'Debit'
            sheet['D4'] = 'Credit'
            sheet['E4'] = 'Group Name'
            titlerow = sheet.row_dimensions[4]
            titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['C4'].font= Font(name='Liberation Serif',size=12,bold=True)
            sheet['D4'].font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['E4'].font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['c4'].alignment = Alignment(horizontal='right')
            sheet['D4'].alignment = Alignment(horizontal='right')
            sheet['E4'].alignment = Alignment(horizontal='center')
            row =5
            for record in records:
                sheet['A'+str(row)] = record['srno']
                sheet['A'+str(row)].alignment = Alignment(horizontal='left')
                sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['B'+str(row)] = record['accountname']
                sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                if record["advflag"]==1:
                    sheet['C'+str(row)] = record['Dr']
                    sheet['C'+str(row)].alignment = Alignment(horizontal='right')
                    sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12',  bold=True, color=RED)
                    sheet['D'+str(row)] = record['Cr']
                    sheet['D'+str(row)].alignment = Alignment(horizontal='right')
                    sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12',  bold=True, color=RED)
                else:
                    sheet['C'+str(row)] = record['Dr']
                    sheet['C'+str(row)].alignment = Alignment(horizontal='right')
                    sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12',  bold=False)
                    sheet['D'+str(row)] = record['Cr']
                    sheet['D'+str(row)].alignment = Alignment(horizontal='right')
                    sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12',  bold=False)
                sheet['E'+str(row)] = record['groupname']
                sheet['E'+str(row)].alignment = Alignment(horizontal='center')
                sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12',  bold=False)
                row = row + 1
        # Condition for Gross Trial Balance
        elif trialbalancetype == 2:
            sheet.column_dimensions['A'].width = 8
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 18
            sheet.column_dimensions['D'].width = 18
            sheet.column_dimensions['E'].width = 20
            sheet.column_dimensions['F'].width = 20
            sheet.column_dimensions['G'].width = 20
            # Cells of first two rows are merged to display organisation details properly.
            sheet.merge_cells('A1:G2')
            # Name and Financial Year of organisation is fetched to be displayed on the first row.
            sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
            sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
            # Organisation name and financial year are displayed.
            sheet['A1'] = orgname + ' (FY: ' + datetime.strptime(str(financialstart),"%Y-%m-%d").strftime("%d-%m-%Y") + ' to ' + fyend +')'
            sheet.merge_cells('A3:G3')
            sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
            sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
            sheet['A3'] = 'Gross Trial Balance for the period from %s to %s' %(datetime.strptime(str(startdate),"%Y-%m-%d").strftime("%d-%m-%Y"), datetime.strptime(str(calculateto),"%Y-%m-%d").strftime("%d-%m-%Y"))
            sheet['A4'] = 'Sr. No. '
            sheet['B4'] = 'Account Name'
            sheet['C4'] = 'Debit'
            sheet['D4'] = 'Credit'
            sheet['E4'] = 'Dr Balance'
            sheet['F4'] = 'Cr Balance'
            sheet['G4'] = 'Group Name'
            titlerow = sheet.row_dimensions[4]
            titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['C4'].font= Font(name='Liberation Serif',size=12,bold=True)
            sheet['D4'].font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['E4'].font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['F4'].font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['G4'].font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['c4'].alignment = Alignment(horizontal='right')
            sheet['D4'].alignment = Alignment(horizontal='right')
            sheet['E4'].alignment = Alignment(horizontal='right')
            sheet['F4'].alignment = Alignment(horizontal='right')
            sheet['G4'].alignment = Alignment(horizontal='center')
            row =5
            for record in records:
                sheet['A'+str(row)] = record['srno']
                sheet['A'+str(row)].alignment = Alignment(horizontal='center')
                sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['B'+str(row)] = record['accountname']
                sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['C'+str(row)] = record['totaldr']
                sheet['C'+str(row)].alignment = Alignment(horizontal='right')
                sheet['C'+str(row)].font = Font(name='Liberation Serif', bold=False)
                sheet['D'+str(row)] = record['totalcr']
                sheet['D'+str(row)].alignment = Alignment(horizontal='right')
                sheet['D'+str(row)].font = Font(name='Liberation Serif', bold=False)
                if record["advflag"]==1:
                    sheet['E'+str(row)] = record['curbaldr']
                    sheet['E'+str(row)].alignment = Alignment(horizontal='right')
                    sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12',  bold=True, color=RED)
                    sheet['F'+str(row)] = record['curbalcr']
                    sheet['F'+str(row)].alignment = Alignment(horizontal='right')
                    sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12',  bold=True, color=RED)
                else:
                    sheet['E'+str(row)] = record['curbaldr']
                    sheet['E'+str(row)].alignment = Alignment(horizontal='right')
                    sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12',  bold=False)
                    sheet['F'+str(row)] = record['curbalcr']
                    sheet['F'+str(row)].alignment = Alignment(horizontal='right')
                    sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12',  bold=False)
                sheet['G'+str(row)] = record['groupname']
                sheet['G'+str(row)].alignment = Alignment(horizontal='center')
                sheet['G'+str(row)].font = Font(name='Liberation Serif', size='12',  bold=False)
                row = row + 1
        # Condition for Extended Trial Balance
        elif trialbalancetype == 3:
            sheet.column_dimensions['A'].width = 8
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 18
            sheet.column_dimensions['D'].width = 16
            sheet.column_dimensions['E'].width = 16
            sheet.column_dimensions['F'].width = 16
            sheet.column_dimensions['G'].width = 16
            sheet.column_dimensions['H'].width = 20
            # Cells of first two rows are merged to display organisation details properly.
            sheet.merge_cells('A1:H2')
            # Name and Financial Year of organisation is fetched to be displayed on the first row.
            sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
            sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
            # Organisation name and financial year are displayed.
            sheet['A1'] = orgname + ' (FY: ' + datetime.strptime(str(financialstart),"%Y-%m-%d").strftime("%d-%m-%Y") + ' to ' + fyend +')'
            sheet.merge_cells('A3:H3')
            sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
            sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
            sheet['A3'] = 'Extended Trial Balance for the period from %s to %s' %(datetime.strptime(str(startdate),"%Y-%m-%d").strftime("%d-%m-%Y"), datetime.strptime(str(calculateto),"%Y-%m-%d").strftime("%d-%m-%Y"))
            sheet['A4'] = 'Sr. No. '
            sheet['B4'] = 'Account Name'
            sheet['C4'] = 'Opening Balance'
            sheet['D4'] = 'Total Debit'
            sheet['E4'] = 'Total Credit'
            sheet['F4'] = 'Debit Balance'
            sheet['G4'] = 'Credit Balance'
            sheet['H4'] = 'Group Name'
            titlerow = sheet.row_dimensions[4]
            titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['C4'].font= Font(name='Liberation Serif',size=12,bold=True)
            sheet['D4'].font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['E4'].font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['F4'].font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['G4'].font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['H4'].font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['c4'].alignment = Alignment(horizontal='right')
            sheet['D4'].alignment = Alignment(horizontal='right')
            sheet['E4'].alignment = Alignment(horizontal='right')
            sheet['F4'].alignment = Alignment(horizontal='right')
            sheet['G4'].alignment = Alignment(horizontal='right')
            sheet['H4'].alignment = Alignment(horizontal='center')
            row =5
            for record in records:
                sheet['A'+str(row)] = record['srno']
                sheet['A'+str(row)].alignment = Alignment(horizontal='center')
                sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['B'+str(row)] = record['accountname']
                sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['C'+str(row)] = record['openingbalance']
                sheet['C'+str(row)].alignment = Alignment(horizontal='right')
                sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['D'+str(row)] = record['totaldr']
                sheet['D'+str(row)].alignment = Alignment(horizontal='right')
                sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['E'+str(row)] = record['totalcr']
                sheet['E'+str(row)].alignment = Alignment(horizontal='right')
                sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                if record["advflag"]==1:
                    sheet['F'+str(row)] = record['curbaldr']
                    sheet['F'+str(row)].alignment = Alignment(horizontal='right')
                    sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12',  bold=True, color=RED)
                    sheet['G'+str(row)] = record['curbalcr']
                    sheet['G'+str(row)].alignment = Alignment(horizontal='right')
                    sheet['G'+str(row)].font = Font(name='Liberation Serif', size='12',  bold=True, color=RED)
                else:
                    sheet['F'+str(row)] = record['curbaldr']
                    sheet['F'+str(row)].alignment = Alignment(horizontal='right')
                    sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12',  bold=False)
                    sheet['G'+str(row)] = record['curbalcr']
                    sheet['G'+str(row)].alignment = Alignment(horizontal='right')
                    sheet['G'+str(row)].font = Font(name='Liberation Serif', size='12',  bold=False)
                sheet['H'+str(row)] = record['groupname']
                sheet['H'+str(row)].alignment = Alignment(horizontal='center')
                sheet['H'+str(row)].font = Font(name='Liberation Serif', size='12',  bold=False)
                row = row + 1
        trialbalancewb.save('report.xlsx')
        xlsxfile = open("report.xlsx","r")
        reportxslx = xlsxfile.read()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(reportxslx),'Content-Disposition': 'attachment; filename=report.xlsx', 'Set-Cookie':'fileDownload=true; path=/'}
        xlsxfile.close()
        os.remove("report.xlsx")
        return Response(reportxslx, headerlist=headerList.items())
    except:
        print "file not found"
        return {"gkstatus":3}
