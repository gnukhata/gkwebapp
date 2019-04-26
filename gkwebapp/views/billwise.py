
"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@dff.org.in>
"Arun Kelkar" <arunkelkar@dff.org.in>
'Abhijith Balan'<abhijith@dff.org.in>
'Prajkta Patkar'<prajakta@dff.org.in>
"""

from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from pyramid.response import Response
import openpyxl
from openpyxl.styles import Font, Alignment
import cStringIO
'''
This function brings data of unpaid bills and unadjusted amounts.
This could be either called after a voucher is created or from the Unadjusted Amounts module.
If accountcode is received it must have been called after voucher entry. In this case we retrieve custid from backend by sending accountcode and then get the bill details by sending the custid.
In case of unadjusted amounts module custid is received instead of account code and bill details are subsequently fetched.
The details of bills, sum of invoice and pending amounts and custid are then returned.
'''

@view_config(route_name="billwise", request_param = "type=vchbillwise", renderer="gkwebapp:templates/billwiseaccounting.jinja2")
def getBillTable(request):
    header={"gktoken":request.headers["gktoken"]}
    accountcode = int(request.params["accountcode"])
    voucherdate = request.params["voucherdate"]
    voucherdate = datetime.strptime(voucherdate, "%d%m%Y")
    result = requests.get("http://127.0.0.1:6543/customersupplier?by=account&accountcode=%d"%accountcode, headers=header)
    if result.json()["gkstatus"] == 0:
        custid = result.json()["gkresult"]
        customer = requests.get("http://127.0.0.1:6543/customersupplier?qty=single&custid=%d"%custid, headers=header)
        customerdetails = customer.json()["gkresult"]
        billdetails = requests.get("http://127.0.0.1:6543/invoice?type=bwa&custid=%d"%custid, headers=header)
        unpaidbills = []
        sumofinvoiceamounts = 0.00
        sumofpendingamounts = 0.00
        if billdetails.json()["gkstatus"] == 0:
            for bill in billdetails.json()["gkresult"]["unpaidbills"]:
                invoicedate = bill["invoicedate"]
                invoicedate = datetime.strptime(invoicedate, "%d-%m-%Y")
                if invoicedate <= voucherdate:
                    sumofinvoiceamounts = sumofinvoiceamounts + float(bill["invoicetotal"])
                    sumofpendingamounts = sumofpendingamounts + float(bill["pendingamount"])
                    unpaidbills.append(bill)
            return {"gkstatus":result.json()["gkstatus"], "gkresult":unpaidbills, "sumofinvoiceamounts":sumofinvoiceamounts, "sumofpendingamounts":sumofpendingamounts, "custid":custid, "vouchercode":request.params["vouchercode"]}
    return {"gkresult":[]}
    """
    elif request.params.has_key("custid"):
        custid = int(request.params["custid"])
        customer = requests.get("http://127.0.0.1:6543/customersupplier?qty=single&custid=%d"%custid, headers=header)
        customerdetails = customer.json()["gkresult"]
        billdetails = requests.get("http://127.0.0.1:6543/invoice?type=bwa&custid=%d"%custid, headers=header)
        unpaidbills = []
        sumofinvoiceamounts = 0.00
        sumofpendingamounts = 0.00
        if billdetails.json()["gkstatus"] == 0:
            for bill in billdetails.json()["gkresult"]["unpaidbills"]:
                sumofinvoiceamounts = sumofinvoiceamounts + float(bill["invoicetotal"])
                sumofpendingamounts = sumofpendingamounts + float(bill["pendingamount"])
                unpaidbills.append(bill)
            return {"gkstatus":customer.json()["gkstatus"], "gkresult":unpaidbills, "sumofinvoiceamounts":sumofinvoiceamounts, "sumofpendingamounts":sumofpendingamounts, "custid":custid, "onaccount":"%.2f"%customerdetails["onaccamt"], "asadvance":"%.2f"%customerdetails["advamt"]}
        return {"gkresult":[]}
    else:
        return {"gkresult":[]}

"""
'''
The below function calls a function in API for invoice that updates the amount paid field in invoice table and advamt and onaccamt fields in customersupplier table.
It receives a list of dictionaries.
It contains a flag(payflag) to check for the type of payment. It could be settlement of a bill(payflag=2), advance payment(payflag=1) of amount or amount set as on account(payflag=15).
For advance and on account payments an additional flag(icflag) is also sent which tells the API whether to increment organisation decrement the advamt or onaccamt fields.
Each dictionary also has custid(id of customer or supplier) and amount.
'''
@view_config(route_name="billwise", request_param="action=updatepayment", renderer="json")
def updatepayment(request):
    header={"gktoken":request.headers["gktoken"]}
    payments = json.loads(request.params["billwisedata"])
    dataset = {"adjbills":payments}
    result = requests.post("http://127.0.0.1:6543/billwise",data=json.dumps(dataset),headers = header)
    return {"gkstatus":result.json()["gkstatus"]}

'''
The below function returns a list of customers and suppliers.
This is used in the Unudjusted Amounts module where a user is presented with this list to choose a customer/supplier.
'''
@view_config(route_name="billwise", request_param = "action=showcustomersupplierlist", renderer="gkwebapp:templates/customersupplierlist.jinja2")
def getCustomerSupplierList(request):
    header={"gktoken":request.headers["gktoken"]}
    customers = requests.get("http://127.0.0.1:6543/customersupplier?qty=custall", headers=header)
    suppliers = requests.get("http://127.0.0.1:6543/customersupplier?qty=supall", headers=header)
    return {"customers": customers.json()["gkresult"], "suppliers": suppliers.json()["gkresult"]}

'''
This function returns a list of unadjusted invoices and vouchers that can be used to adjust them.
'''
@view_config(route_name="billwise", request_param = "action=showunadjustedamounts", renderer="gkwebapp:templates/unadjustedamount.jinja2")
def getunadjustedamounts(request):
    header={"gktoken":request.headers["gktoken"]}
    unadjustedamounts = requests.get("http://127.0.0.1:6543/billwise?csid=%d&csflag=%d"%(int(request.params["csid"]), int(request.params["csflag"])), headers=header)
    invs = unadjustedamounts.json()["invoices"]
    totalinv = 0.00
    totalamtpending = 0.00
    for inv in invs:
        totalinv = totalinv + float(inv["invoiceamount"])
        totalamtpending = totalamtpending + float(inv["balanceamount"])
        
    return {"invoices":invs, "vouchers":unadjustedamounts.json()["vouchers"],"totalinv":"%.2f"%(totalinv),"totalamtpending":"%.2f"%(totalamtpending)}

'''
Function to render list of unpaid invoices report view page.
Number of invoices is also checked and returned.
'''
@view_config(route_name="billwise", request_param = "action=viewlistofunpaidinvoices", renderer="gkwebapp:templates/viewlistofunpaidinvoices.jinja2")
def viewListofInvoices(request):
    header={"gktoken":request.headers["gktoken"]}
    invoices = requests.get("http://127.0.0.1:6543/billwise?type=all",headers=header)
    return {"status":True,"numberofinvoices":len(invoices.json()["invoices"])}

'''
Function to render list of unpaid invoices report.
It takes order, type, period and inoutflag from javascript and sends a request to backend to fetch data for the report.
This is presented in a tabular form.
'''
@view_config(route_name="billwise", request_param="action=showlist", renderer="gkwebapp:templates/listofunpaidinvoices.jinja2")
def listofUnpaidInvoices(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/billwise?type=onlybillsforall&inoutflag=%d&orderflag=%d&typeflag=%d&startdate=%s&enddate=%s"%(int(request.params["inoutflag"]), int(request.params["orderflag"]),int(request.params["typeflag"]), request.params["fromdate"], request.params["todate"]), headers=header)
    return {"gkstatus":result.json()["gkstatus"], "gkresult": result.json()["invoices"], "type": result.json()["type"], "order": result.json()["order"], "inout": result.json()["inout"], "inoutflag":int(request.params["inoutflag"]), "orderflag": request.params["orderflag"], "typeflag": request.params["typeflag"], "fromdate": request.params["fromdate"], "todate": request.params["todate"]}

'''
This function returns a printable view of list of unpaid invoices report.
'''
@view_config(route_name="billwise", request_param="action=printlist", renderer="gkwebapp:templates/printlistofunpaidinvoices.jinja2")
def printListofUnpaidInvoices(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/billwise?type=onlybillsforall&inoutflag=%d&orderflag=%d&typeflag=%d&startdate=%s&enddate=%s"%(int(request.params["inoutflag"]), int(request.params["orderflag"]),int(request.params["typeflag"]), request.params["fromdate"], request.params["todate"]), headers=header)
    return {"gkstatus":result.json()["gkstatus"], "gkresult": result.json()["invoices"], "type": result.json()["type"], "order": result.json()["order"], "inout": result.json()["inout"], "inoutflag":int(request.params["inoutflag"]), "orderflag": request.params["orderflag"], "typeflag": request.params["typeflag"], "fromdate": request.params["fromdate"], "todate": request.params["todate"]}

'''
This function returns a spreadsheet form of List of Unpaid Invoices Report.
The spreadsheet in XLSX format is generated by the backend and sent in base64 encoded format.
It is decoded and returned along with mime information.
'''
@view_config(route_name="billwise", request_param="type=spreadsheet", renderer="")
def unpaidInvoicesSpreadsheet(request):
    try:
        header={"gktoken":request.headers["gktoken"]}
        result = requests.get("http://127.0.0.1:6543/billwise?type=onlybillsforall&inoutflag=%d&orderflag=%d&typeflag=%d&startdate=%s&enddate=%s"%(int(request.params["inoutflag"]), int(request.params["orderflag"]),int(request.params["typeflag"]), request.params["fromdate"], request.params["todate"]), headers=header)
        inoutflag = int(request.params["inoutflag"])
        orderflag = int(request.params["orderflag"])
        typeflag = int(request.params["typeflag"])
        inouts= {9:"Purchase", 15:"Sale"}
        orders = {1:"Ascending", 4:"Descending"}
        types = {1:"Amount Wise", 3:"Party Wise", 4:"Due Wise"}
        inout = inouts[inoutflag]
        order = orders[orderflag]
        reporttype = types[typeflag]
        startdate =datetime.strptime(str(request.params["fromdate"]),"%d-%m-%Y").strftime("%Y-%m-%d")
        enddate =datetime.strptime(str(request.params["todate"]),"%d-%m-%Y").strftime("%Y-%m-%d")
        fystart = str(request.params["fystart"]);
        fyend = str(request.params["fyend"]);
        orgname = str(request.params["orgname"]);
        # A workbook is opened.
        billwisewb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = billwisewb.active
        # Title of the sheet and width of columns are set.
        sheet.title = "List of Unpaid Invoices"
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 18
        sheet.column_dimensions['C'].width = 14
        sheet.column_dimensions['D'].width = 24
        sheet.column_dimensions['E'].width = 16
        sheet.column_dimensions['F'].width = 16
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells('A1:F2')
        # Name and Financial Year of organisation is fetched to be displayed on the first row.
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        # Organisation name and financial year are displayed.
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet.merge_cells('A3:F3')
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        invtype = 'Sale'
        if inoutflag == 9:
            invtype = 'Purchase'
        sheet['A3'] = '%s List of Outstanding %s Invoices in %s Order'%(str(reporttype), str(inout), str(order))
        sheet.merge_cells('A4:F4')
        sheet['A4'] = 'Period: ' + str(request.params["fromdate"]) + ' to ' + str(request.params["todate"])
        sheet['A4'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A4'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A5'] = 'Sr. No. '
        sheet['B5'] = 'Invoice No'
        sheet['C5'] = 'Invoice Date'
        custhead = "Customer Name"
        if inoutflag == 9:
            custhead = "Supplier Name"
        sheet['D5'] = custhead
        sheet['E5'] = 'Invoice Amount'
        sheet['F5'] = 'Amount Pending'
        titlerow = sheet.row_dimensions[5]
        titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['E5'].alignment = Alignment(horizontal='right')
        sheet['F5'].alignment = Alignment(horizontal='right')
        sheet['E5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['F5'].font = Font(name='Liberation Serif',size=12,bold=True)
        unAdjInvoices = result.json()["invoices"]
        row = 6
        # Looping each dictionaries in list unAdjInvoices to store data in cells and apply styles.
        srno = 1
        for uninv in unAdjInvoices:
            sheet['A'+str(row)] = srno
            sheet['A'+str(row)].alignment = Alignment(horizontal='left')
            sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['B'+str(row)] = uninv['invoiceno']
            sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['C'+str(row)] = uninv['invoicedate']
            sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['D'+str(row)] = uninv['custname']
            sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['E'+str(row)] = float("%.2f"%float(uninv['invoiceamount'])) 
            sheet['E'+str(row)].number_format = "0.00" 
            sheet['F'+str(row)] = float("%.2f"%float(uninv['balanceamount']))
            sheet['F'+str(row)].number_format = "0.00" 
            sheet['E'+str(row)].alignment = Alignment(horizontal='right')
            sheet['F'+str(row)].alignment = Alignment(horizontal='right')
            sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            row = row + 1
            srno = srno + 1
        output = cStringIO.StringIO()
        billwisewb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx', 'Set-Cookie':'fileDownload=true; path=/'}
        return Response(contents, headerlist=headerList.items()) 
    except:
        print "file not found"
        return {"gkstatus":3}
