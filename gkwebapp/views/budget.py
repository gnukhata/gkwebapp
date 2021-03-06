"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
   "Krishnakant Mane" <kk@gmail.com>
   "Karan Kamdar" <kamdar.karan@gmail.com>
   "Prajkta Patkar" <prajkta@riseup.com>
   "Abhijith Balan" <abhijith@dff.org.in>
   "rohan khairnar" <rohankhairnar@gmail.com>
"""

from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from pyramid.response import Response
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import os
from pyramid.i18n import default_locale_negotiator
s = requests.session()
from PIL import Image
import base64
import io
from io import BytesIO

@view_config(route_name="budget", renderer="gkwebapp:templates/budget.jinja2")
def budget(request):
    header={"gktoken":request.headers["gktoken"]}
    return {"status":True}

@view_config(route_name="budget",request_param="type=addtab", renderer="gkwebapp:templates/createbudget.jinja2")
def addbudgetpage(request):
    header={"gktoken":request.headers["gktoken"]}
    return {"status":True}

@view_config(route_name="budget",request_param="type=balance", renderer="gkwebapp:templates/createbudgetaccountstable.jinja2")
def balance(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/budget?type=addtab&financialstart=%s&uptodate=%s&btype=%d"%(request.params["financialstart"],request.params["uptodate"],int(request.params["budgettype"])), headers=header)
    
    return {"status":result.json()["gkstatus"],"gkresult":result.json()["gkresult"],"btype":int(request.params["budgettype"])}

@view_config(route_name="budget",request_param="type=edittab", renderer="gkwebapp:templates/editbudget.jinja2")
def editbudgetpage(request):
    header={"gktoken":request.headers["gktoken"]}
    return {"status":True}

@view_config(route_name="budget",request_param="type=viewbudgetreportpage", renderer="gkwebapp:templates/viewbudgetreport.jinja2")
def viewreportpage(request):
    header={"gktoken":request.headers["gktoken"]}
    return {"status":True,"menuflag":int(request.params["menuflag"])}

@view_config(route_name="budget",request_param="type=bdg_list", renderer="json")
def alllist(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/budget?bud=all&btype=%d"%int(request.params["btype"]), headers=header)
    buddata=[]
    for record in result.json()["gkresult"]:
        gdata= {"budname":str(record["budname"]),"budid":str(record["budid"]),"startdate": record["startdate"],"enddate": record["enddate"]}
        buddata.append(gdata)
    return {"gkresult":buddata,"numberofbudget":len(result.json()["gkresult"]),"status":True}
    
@view_config(route_name="budget",request_param="type=getbuddetails", renderer="json")
def getbuddetails(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/budget?bud=details&budid=%d"%int(request.params["budid"]), headers=header)
    
    if(result.json()["gkstatus"] == 0):
        record = result.json()["gkresult"]
        return {"gkstatus":0, "gkresult":record}
    return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="budget",request_param="type=add", renderer="json")
def addbudget(request):
    header={"gktoken":request.headers["gktoken"]}
    gkdata = {"budname":request.params["budname"],"startdate":request.params["startdate"],"enddate":request.params["enddate"], "contents":json.loads(request.params["contents"]), "budtype":int(request.params["btype"]), "gaflag":int(request.params["gaflag"])}

    result = requests.post("http://127.0.0.1:6543/budget", data =json.dumps(gkdata),headers=header)
    if result.json()["gkstatus"] == 0:
        gkdata = {"activity":request.params["budname"] + " budget created"}   
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
    return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="budget",request_param="type=edit", renderer="json")
def editbudget(request):
    header={"gktoken":request.headers["gktoken"]}
    gkdata = {"budid":request.params["budid"],"budname":request.params["budname"],"startdate":request.params["startdate"], "enddate":request.params["enddate"], "contents": json.loads(request.params["contents"]), "budtype":request.params["btype"],"gaflag":request.params["gaflag"]}
    result = requests.put("http://127.0.0.1:6543/budget", data =json.dumps(gkdata),headers=header)
    if result.json()["gkstatus"] == 0:
        gkdata = {"activity":request.params["budname"] + " budget edited"}   
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
    return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="budget",request_param="type=delete", renderer="json")
def deletebudget(request):
    header={"gktoken":request.headers["gktoken"]}
    budname = request.params["budname"]
    gkdata={"budid":request.params["budid"]}
    result = requests.delete("http://127.0.0.1:6543/budget",data =json.dumps(gkdata), headers=header)
    if result.json()["gkstatus"] == 0:
        gkdata = {"activity":budname + " budget deleted"}
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
    return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="budget",request_param="type=report", renderer="gkwebapp:templates/budgetreport.jinja2")
def budgetreport(request):
    header={"gktoken":request.headers["gktoken"]}
    financialstart = request.params["financialstart"]
    
    if (request.params["btype"] == '3'):
        result = requests.get("http://127.0.0.1:6543/budget?type=cashReport&budid=%d&financialstart=%s"%(int(request.params["budid"]),str(financialstart)), headers=header)
    if (request.params["btype"] == '16'):
        result = requests.get("http://127.0.0.1:6543/budget?type=profitlossReport&budid=%d&financialstart=%s"%(int(request.params["budid"]),str(financialstart)), headers=header)
    
    return {"gkstatus":result.json()["gkstatus"], "gkresult":result.json()["gkresult"], "budgetdetail":request.params["buddetails"], "btype":request.params["btype"],"budid":int(request.params["budid"]),"menuflag":request.params["menuflag"] }

@view_config(route_name="budget",request_param="type=printreport", renderer="gkwebapp:templates/printbudgetreport.jinja2")
def printbudgetreport(request):
    header={"gktoken":request.headers["gktoken"]}
    financialstart = request.params["financialstart"]
    if (request.params["btype"] == '3'):
        result = requests.get("http://127.0.0.1:6543/budget?type=cashReport&budid=%d&financialstart=%s"%(int(request.params["budid"]),str(financialstart)), headers=header)
    if (request.params["btype"] == '16'):
        result = requests.get("http://127.0.0.1:6543/budget?type=profitlossReport&budid=%d&financialstart=%s"%(int(request.params["budid"]),str(financialstart)), headers=header)

    return {"gkstatus":result.json()["gkstatus"], "gkresult":result.json()["gkresult"], "budgetdetails":request.params["buddetails"], "budid":int(request.params["budid"]), "financialstart":financialstart,"btype":request.params["btype"],"menuflag":request.params["menuflag"] }

@view_config(route_name="budget",request_param="type=cashspreadsheet", renderer="") 
def cashspreadsheet(request):
    try:
        header={"gktoken":request.headers["gktoken"]}
        financialstart = request.params["financialstart"]
        fystart = str(request.params["fystart"])
        fyend = str(request.params["fyend"])
        orgname = str(request.params["orgname"])
        budgetdetails = str(request.params["budgetdetails"])
        result = requests.get("http://127.0.0.1:6543/budget?type=cashReport&budid=%d&financialstart=%s"%(int(request.params["budid"]),str(financialstart)), headers=header)
        result = result.json()["gkresult"];
        budgetwb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = budgetwb.active
        # Title of the sheet and width of columns are set.
        sheet.title = "Cash Budget Report"

        sheet.column_dimensions['A'].width = 36
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells('A1:E2')
        # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        # Organisation name and financial year are displayed.
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A3'] = 'Cash Budget Report :'+ str(budgetdetails)
        sheet.merge_cells('A3:E3')

        sheet['A4'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['A4'].alignment = Alignment(horizontal = 'left', vertical='center')
        sheet['A4'] = ''
        sheet.merge_cells('A4:E4')

        sheet['A5'] = 'Particulars'
        sheet['B5'] = 'Budgeted'
        sheet['C5'] = 'Actuals'
        sheet['D5'] = 'Variance'
        sheet['E5'] = 'Variance (%)'
        titlerow = sheet.row_dimensions[5]
        titlerow.font = Font(name='Liberation Serif',size='12',bold=True)
        titlerow.alignment = Alignment(horizontal = 'center', vertical='center')

        titlecolumn = sheet.column_dimensions['A']
        titlecolumn.font = Font(name='Liberation Serif',size='12',bold=True)
        titlecolumn.alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A6'] = 'Opening'
        row = 7
        for ob in result["openingacc"]:
            accountsrow = sheet.row_dimensions[row]
            accountsrow.font = Font(name='Liberation Serif' )
            accountsrow.alignment = Alignment(horizontal = 'right', vertical='center')
            sheet['A'+str(row)] = ob["accountname"]
            sheet['A'+str(row)].font = Font(name='Liberation Serif',italic=True )
            sheet['A'+str(row)].alignment = Alignment(horizontal = 'right', vertical='center')
            sheet['B'+str(row)] = float(ob["balance"])
            sheet['B'+str(row)].number_format='0.00'
            sheet['C'+str(row)] = float(ob["balance"])
            sheet['C'+str(row)].number_format='0.00'
            sheet['D'+str(row)] = '-'
            sheet['E'+str(row)] = '-'
            row = row +1
        totalrow = sheet.row_dimensions[row]
        totalrow.font = Font(name='Liberation Serif' ,bold=True)
        totalrow.alignment = Alignment(horizontal = 'right', vertical='center')
        sheet['A'+str(row)] = 'Total'
        sheet['A'+str(row)].alignment = Alignment(horizontal = 'right', vertical='center')
        sheet['A'+str(row)].font = Font(name='Liberation Serif' ,bold=True)
        sheet['B'+str(row)] = float(result["opening"])
        sheet['B'+str(row)].number_format='0.00'
        sheet['C'+str(row)] = float(result["opening"])
        sheet['C'+str(row)].number_format='0.00'
        sheet['D'+str(row)] = '-'
        sheet['E'+str(row)] = '-'
        row = row+1
        sheet['A'+str(row)] = 'Inflow'
        row = row+1
        for ob in result["inflow"]:
            accountsrow = sheet.row_dimensions[row]
            accountsrow.font = Font(name='Liberation Serif' )
            accountsrow.alignment = Alignment(horizontal = 'right', vertical='center')
            sheet['A'+str(row)] = ob["accountname"]
            sheet['A'+str(row)].font = Font(name='Liberation Serif',italic=True )
            sheet['A'+str(row)].alignment = Alignment(horizontal = 'right', vertical='center')
            sheet['B'+str(row)] = float(ob["budget"])
            sheet['B'+str(row)].number_format='0.00'
            sheet['C'+str(row)] = float(ob["actual"])
            sheet['C'+str(row)].number_format='0.00'
            if ob["var"] != '-':
                sheet['D'+str(row)] = float(ob["var"])
                sheet['D'+str(row)].number_format='0.00'
            else:
                sheet['D'+str(row)] = ob["var"]
            if ob["varinpercent"] == '-':
                sheet['E'+str(row)] = ob["varinpercent"]
            else:
                sheet['E'+str(row)] = ob["varinpercent"] +" %"
            row = row +1
        totalrow = sheet.row_dimensions[row]
        totalrow.font = Font(name='Liberation Serif' ,bold=True)
        totalrow.alignment = Alignment(horizontal = 'right', vertical='center')
        sheet['A'+str(row)] = 'Total'
        sheet['A'+str(row)].alignment = Alignment(horizontal = 'right', vertical='center')
        sheet['A'+str(row)].font = Font(name='Liberation Serif' ,bold=True)
        sheet['B'+str(row)] = float(result["budgetin"])
        sheet['B'+str(row)].number_format='0.00'
        sheet['C'+str(row)] = float(result["actualin"])
        sheet['C'+str(row)].number_format='0.00'
        sheet['D'+str(row)] = float(result["varin"])
        sheet['D'+str(row)].number_format='0.00'
        sheet['E'+str(row)] = result["varpercentin"]+" %"
        row = row+1
        sheet['A'+str(row)] = 'Outflow'
        row = row+1
        for ob in result["outflow"]:
            accountsrow = sheet.row_dimensions[row]
            accountsrow.font = Font(name='Liberation Serif' )
            accountsrow.alignment = Alignment(horizontal = 'right', vertical='center')
            sheet['A'+str(row)] = ob["accountname"]
            sheet['A'+str(row)].font = Font(name='Liberation Serif',italic=True )
            sheet['A'+str(row)].alignment = Alignment(horizontal = 'right', vertical='center')
            sheet['B'+str(row)] = float(ob["budget"])
            sheet['B'+str(row)].number_format='0.00'
            sheet['C'+str(row)] = float(ob["actual"])
            sheet['C'+str(row)].number_format='0.00'
            if ob["var"] != '-':
                sheet['D'+str(row)] = float(ob["var"])
                sheet['D'+str(row)].number_format='0.00'
            else:
                sheet['D'+str(row)] = ob["var"]
            if ob["varinpercent"] == '-':
                sheet['E'+str(row)] = ob["varinpercent"]
            else:
                sheet['E'+str(row)] = ob["varinpercent"] +" %"
            row = row +1
        totalrow = sheet.row_dimensions[row]
        totalrow.font = Font(name='Liberation Serif' ,bold=True)
        totalrow.alignment = Alignment(horizontal = 'right', vertical='center')
        sheet['A'+str(row)] = 'Total'
        sheet['A'+str(row)].alignment = Alignment(horizontal = 'right', vertical='center')
        sheet['A'+str(row)].font = Font(name='Liberation Serif' ,bold=True)
        sheet['B'+str(row)] = float(result["budgetout"])
        sheet['B'+str(row)].number_format='0.00'
        sheet['C'+str(row)] = float(result["actualout"])
        sheet['C'+str(row)].number_format='0.00'
        sheet['D'+str(row)] = float(result["varout"])
        sheet['D'+str(row)].number_format='0.00'
        sheet['E'+str(row)] = result["varpercentout"]+" %"
        row = row+1
        sheet['A'+str(row)] = 'Closing'
        row = row+1
        for ob in result["closing"]:
            accountsrow = sheet.row_dimensions[row]
            accountsrow.font = Font(name='Liberation Serif' )
            accountsrow.alignment = Alignment(horizontal = 'right', vertical='center')
            sheet['A'+str(row)] = ob["accountname"]
            sheet['A'+str(row)].font = Font(name='Liberation Serif',italic=True )
            sheet['A'+str(row)].alignment = Alignment(horizontal = 'right', vertical='center')
            sheet['B'+str(row)] = float(ob["budget"])
            sheet['B'+str(row)].number_format='0.00'
            sheet['C'+str(row)] = float(ob["balance"])
            sheet['C'+str(row)].number_format='0.00'
            if ob["var"] != '-':
                sheet['D'+str(row)] = float(ob["var"])
                sheet['D'+str(row)].number_format='0.00'
            else:
                sheet['D'+str(row)] = ob["var"]
            if ob["varinpercent"] == '-':
                sheet['E'+str(row)] = ob["varinpercent"]
            else:
                sheet['E'+str(row)] = ob["varinpercent"] +" %"
            row = row +1

        output = io.BytesIO()
        budgetwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','X-Content-Type-Options':'nosniff' ,'Set-Cookie':'fileDownload=true ;path=/ [;HttpOnly]'}
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true ;path=/'}
        return Response(contents, headerlist=list(headerList.items()))       
    except:
        return {"gkstatus":3}

@view_config(route_name="budget",request_param="type=pnlspreadsheet", renderer="") 
def pnlpreadsheet(request):
    try:
        header={"gktoken":request.headers["gktoken"]}
        financialstart = request.params["financialstart"]
        fystart = str(request.params["fystart"])
        fyend = str(request.params["fyend"])
        Otype = str(request.params["orgtype"])
        orgname = str(request.params["orgname"])
        budgetdetails = str(request.params["budgetdetails"])
        result = requests.get("http://127.0.0.1:6543/budget?type=profitlossReport&budid=%d&financialstart=%s"%(int(request.params["budid"]),str(financialstart)), headers=header)
        result = result.json()["gkresult"]
        budgetwb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = budgetwb.active
        # Title of the sheet and width of columns are set.
        if Otype == "Not For Profit":
            sheet.title = "Income and Expenditure Budget Report"
        else:
            sheet.title = "Profit & Loss Budget Report"

        sheet.column_dimensions['A'].width = 36
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20        
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells('A1:E2')
        # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        # Organisation name and financial year are displayed.
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        if Otype == "Not For Profit":
            sheet['A3'] = 'Income and Expenditure Budget Report :'+ str(budgetdetails)
        else:
            sheet['A3'] = 'Profit & Loss Budget Report :'+ str(budgetdetails)
        
        sheet.merge_cells('A3:E3')
        
        sheet['A4'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['A4'].alignment = Alignment(horizontal = 'left', vertical='center')
        sheet['A4'] = ''
        sheet.merge_cells('A4:E4')
        sheet['A5'] = 'Particulars'
        sheet['B5'] = 'Budgeted'
        sheet['C5'] = 'Actuals'
        sheet['D5'] = 'Variance'
        sheet['E5'] = 'Variance (%)'
        titlerow = sheet.row_dimensions[5]
        titlerow.font = Font(name='Liberation Serif',size='12',bold=True)
        titlerow.alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A6'] = 'Incomes'
        sheet['A6'].font = Font(name='Liberation Serif' ,bold=True)
        sheet['A6'].alignment = Alignment(horizontal = 'left', vertical='center')
        row = 7
        for budget in result["incomeacc"]:
            accountsrow = sheet.row_dimensions[row]
            accountsrow.font = Font(name='Liberation Serif' )
            accountsrow.alignment = Alignment(horizontal = 'right', vertical='center')
            sheet['A'+str(row)] = budget["name"]
            sheet['A'+str(row)].font = Font(italic=True,name='Liberation Serif' )
            sheet['A'+str(row)].alignment = Alignment(horizontal = 'right', vertical='center')
            sheet['B'+str(row)] = float(budget["budget"])
            sheet['B'+str(row)].number_format='0.00'
            sheet['C'+str(row)] = float(budget["actual"])
            sheet['C'+str(row)].number_format='0.00'
            if budget["var"] !='-':
                sheet['D'+str(row)] = float(budget["var"])
                sheet['D'+str(row)].number_format='0.00'
            else:
                sheet['D'+str(row)] = budget["var"]
            if budget["varinpercent"] == '-':
                sheet['E'+str(row)] = budget["varinpercent"]
            else:
                sheet['E'+str(row)] = budget["varinpercent"] +" %"
            row=row+1
        totalrow = sheet.row_dimensions[row]
        totalrow.font = Font(name='Liberation Serif' ,bold=True)
        totalrow.alignment = Alignment(horizontal = 'right', vertical='center')
        sheet['A'+str(row)] = 'Total '
        sheet['A'+str(row)].font = Font(name='Liberation Serif' ,bold=True)
        sheet['A'+str(row)].alignment = Alignment(horizontal = 'right', vertical='center')
        sheet['B'+str(row)] = float(result["budgetincome"])
        sheet['B'+str(row)].number_format='0.00'
        sheet['C'+str(row)] = float(result["income"])
        sheet['C'+str(row)].number_format='0.00'
        sheet['D'+str(row)] = float(result["varincome"])
        sheet['D'+str(row)].number_format='0.00'
        sheet['E'+str(row)] = result["varinpercentincome"] + " %"
        row=row+1
        sheet['A'+str(row)] = 'Expenses'
        sheet['A'+str(row)].font = Font(name='Liberation Serif' ,bold=True)
        sheet['A'+str(row)].alignment = Alignment(horizontal = 'left', vertical='center')
        row = row+1
        for budget in result["expenseacc"]:
            accountsrow = sheet.row_dimensions[row]
            accountsrow.font = Font(name='Liberation Serif' )
            accountsrow.alignment = Alignment(horizontal = 'right', vertical='center')
            sheet['A'+str(row)] = budget["name"]
            sheet['A'+str(row)].font = Font(italic=True,name='Liberation Serif' )
            sheet['A'+str(row)].alignment = Alignment(horizontal = 'right', vertical='center')
            sheet['B'+str(row)] = float(budget["budget"])
            sheet['B'+str(row)].number_format='0.00'
            sheet['C'+str(row)] = float(budget["actual"])
            sheet['C'+str(row)].number_format='0.00'
            if budget["var"] != '-':
                sheet['D'+str(row)] = float(budget["var"])
                sheet['D'+str(row)].number_format='0.00'
            else:
                sheet['D'+str(row)] = budget["var"]
            if budget["varinpercent"] == '-':
                sheet['E'+str(row)] = budget["varinpercent"]
            else:
                sheet['E'+str(row)] = budget["varinpercent"] +" %"
            row=row+1
        totalrow = sheet.row_dimensions[row]
        totalrow.font = Font(name='Liberation Serif' ,bold=True)
        totalrow.alignment = Alignment(horizontal = 'right', vertical='center')
        sheet['A'+str(row)] = 'Total '
        sheet['A'+str(row)].font = Font(name='Liberation Serif' ,bold=True)
        sheet['A'+str(row)].alignment = Alignment(horizontal = 'right', vertical='center')
        sheet['B'+str(row)] = float(result["budgetexpense"])
        sheet['B'+str(row)].number_format='0.00'
        sheet['C'+str(row)] = float(result["expense"])
        sheet['C'+str(row)].number_format='0.00'
        sheet['D'+str(row)] = float(result["varexpense"])
        sheet['D'+str(row)].number_format='0.00'
        sheet['E'+str(row)] = result["varinpercentexp"] + " %"
        row = row+1
        totalrow = sheet.row_dimensions[row]
        totalrow.font = Font(name='Liberation Serif' ,bold=True)
        totalrow.alignment = Alignment(horizontal = 'right', vertical='center')
        sheet['A'+str(row)] = 'Net Profit'
        sheet['A'+str(row)].font = Font(name='Liberation Serif' ,bold=True)
        sheet['A'+str(row)].alignment = Alignment(horizontal = 'left', vertical='center')
        sheet['B'+str(row)] = float(result["budgetprofit"])
        sheet['B'+str(row)].number_format='0.00'
        sheet['C'+str(row)] = float(result["profit"])
        sheet['C'+str(row)].number_format='0.00'
        if result["varprofit"] != '-':
            sheet['D'+str(row)] = float(result["varprofit"])
            sheet['D'+str(row)].number_format='0.00'
        else:
            sheet['D'+str(row)] = result["varprofit"]
            if result["varinpercentprofit"]!='-':
                sheet['E'+str(row)] = result["varinpercentprofit"]+ " %"
            else:
                sheet['E'+str(row)]='-'

        output = io.BytesIO()
        budgetwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','X-Content-Type-Options':'nosniff', 'Set-Cookie':'fileDownload=true ;path=/ [;HttpOnly]'}
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true ;path=/'}
        return Response(contents, headerlist=list(headerList.items())) 
    except:
        return {"gkstatus":3}
