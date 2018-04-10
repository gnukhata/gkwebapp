
"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@dff.org.in>
"Arun Kelkar" <arunkelkar@dff.org.in>
"Ishan Masdekar " <imasdekar@dff.org.in>
"Navin Karkera" <navin@dff.org.in>
"Vaibhav Kurhe" <vaibspidy@openmailbox.org>
"""

from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from pyramid.response import Response
import openpyxl
from openpyxl.styles import Font, Alignment
import os
import calendar

@view_config(route_name="showcashflow", renderer="gkwebapp:templates/viewcashflow.jinja2")
def showcashflow(request):
	return {"gkstatus":0}

@view_config(route_name="cashflowreportprint")
def cashflowreportprint(request):
	calculateto = request.params["calculateto"]
	financialstart = request.params["financialstart"]
	calculatefrom = request.params["calculatefrom"]
	orgtype = request.params["orgtype"]
	header={"gktoken":request.headers["gktoken"]}
	result = requests.get("http://127.0.0.1:6543/report?type=cashflow&calculateto=%s&financialstart=%s&calculatefrom=%s"%(calculateto,financialstart,calculatefrom), headers=header)

	return render_to_response("gkwebapp:templates/printcashflowreport.jinja2",{"rcrecords":result.json()["rcgkresult"],"pyrecords":result.json()["pygkresult"],"orgtype":orgtype,"backflag":4,"from":datetime.strftime(datetime.strptime(str(calculatefrom),"%Y-%m-%d").date(),'%d-%m-%Y'),"to":datetime.strftime(datetime.strptime(str(calculateto),"%Y-%m-%d").date(),'%d-%m-%Y')},request=request)

@view_config(route_name="showcashflowreport")
def showcashflowreport(request):
	calculateto = request.params["calculateto"]
	financialstart = request.params["financialstart"]
	calculatefrom = request.params["calculatefrom"]
	orgtype = request.params["orgtype"]
	header={"gktoken":request.headers["gktoken"]}
	result = requests.get("http://127.0.0.1:6543/report?type=cashflow&calculateto=%s&financialstart=%s&calculatefrom=%s"%(calculateto,financialstart,calculatefrom), headers=header)

	return render_to_response("gkwebapp:templates/cashflowreport.jinja2",{"rcrecords":result.json()["rcgkresult"],"pyrecords":result.json()["pygkresult"],"orgtype":orgtype,"backflag":4,"from":datetime.strftime(datetime.strptime(str(calculatefrom),"%Y-%m-%d").date(),'%d-%m-%Y'),"to":datetime.strftime(datetime.strptime(str(calculateto),"%Y-%m-%d").date(),'%d-%m-%Y')},request=request)

'''
This function returns a spreadsheet form of Cash Flow Account Report.
The spreadsheet in XLSX format is generated by the backend and sent in base64 encoded format.
It is decoded and returned along with mime information.
'''
@view_config(route_name="printcashflowreport")
def printcashflowreport(request):
    try:
        calculateto = request.params["calculateto"]
        financialstart = request.params["fystart"]
        calculatefrom = request.params["calculatefrom"]
        orgname = request.params["orgname"]
        orgtype = request.params["orgtype"]
        fyend = str(request.params["fyend"])
        header={"gktoken":request.headers["gktoken"]}
        result = requests.get("http://127.0.0.1:6543/report?type=cashflow&calculateto=%s&financialstart=%s&calculatefrom=%s"%(calculateto,financialstart,calculatefrom), headers=header)
        receipt = result.json()["rcgkresult"]
        payment = result.json()["pygkresult"]
        fystart = datetime.strptime(request.params["fystart"],'%Y-%m-%d').strftime('%d-%m-%Y')
        orgname = str(request.params["orgname"])
        calculateto = calculateto[8:10]+calculateto[4:8]+calculateto[0:4]
        calculatefrom = calculatefrom[8:10]+calculatefrom[4:8]+calculatefrom[0:4]
        # A workbook is opened.
        cashwb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = cashwb.active
        # Title of the sheet and width of columns are set.
        #sheet.title = "List of Accounts"
        sheet.column_dimensions['A'].width = 4
        sheet.column_dimensions['B'].width = 32
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 4
        sheet.column_dimensions['F'].width = 32
        sheet.column_dimensions['G'].width = 12
        sheet.column_dimensions['H'].width = 12
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells('A1:H2')
        # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        # Organisation name and financial year are displayed.
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet.merge_cells('A3:H3')
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells('A1:H2')
        # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        # Organisation name and financial year are displayed.
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet.merge_cells('A3:H3')
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        if orgtype=="Profit Making":
            sheet.title = "Cash Flow Statement"
            sheet['A3'] ="Cash Flow Statement ("+calculatefrom+" to "+calculateto+")"
        if orgtype=="Not For Profit":
            sheet.title = "Receipt & Payment Account"
            sheet['A3']="Receipt & Payment Account ("+calculatefrom+" to "+calculateto+")"
        sheet.merge_cells('A3:H3')
        sheet['B4'] = 'Particulars'
        sheet['C4'] = 'Amount'
        sheet['D4'] = 'Amount'
        sheet['F4'] = 'Particulars'
        sheet['G4'] = 'Amount'
        sheet['H4'] = 'Amount'
        titlerow = sheet.row_dimensions[4]
        titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
        row = 5
        for account in receipt:
            sheet['A'+str(row)] = account["toby"]
            sheet['A'+str(row)].alignment = Alignment(horizontal='left')
            sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            if account["toby"]!="" and account["particulars"]!="Opening balance":
                sheet['B'+str(row)] = "			"+account["particulars"]
                sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['C'+str(row)] = account["amount"]
                sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['D'+str(row)] = ""
                sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            else:
                sheet['B'+str(row)] = account["particulars"]
                sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['C'+str(row)] = ""
                sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['D'+str(row)] = account["amount"]
                sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            row += 1
        row = 5
        for account in payment:
            sheet['E'+str(row)] = account["toby"]
            sheet['E'+str(row)].alignment = Alignment(horizontal='left')
            sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            if account["toby"]!="" and account["particulars"]!="Closing balance":
                sheet['F'+str(row)] = "			"+account["particulars"]
                sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['G'+str(row)] = account["amount"]
                sheet['G'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['H'+str(row)] = ""
                sheet['H'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            else:
                sheet['F'+str(row)] = account["particulars"]
                sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['G'+str(row)] = ""
                sheet['G'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['H'+str(row)] = account["amount"]
                sheet['H'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            row += 1
        cashwb.save('report.xlsx')
        xlsxfile = open("report.xlsx","r")
        reportxslx = xlsxfile.read()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(reportxslx),'Content-Disposition': 'attachment; filename=report.xlsx', 'Set-Cookie':'fileDownload=true; path=/'}
        xlsxfile.close()
        os.remove("report.xlsx")
        return Response(reportxslx, headerlist=headerList.items())
    except:
        return {"gkstatus":3}

        
