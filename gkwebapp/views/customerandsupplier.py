
"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@dff.org.in>
"Arun Kelkar" <arunkelkar@dff.org.in>
"Ishan Masdekar " <imasdekar@dff.org.in>
"Navin Karkera" <navin@dff.org.in>
"Sachin Patil" <sachpatil@openmailbox.org>
'Prajkta Patkar'<prajkta@dff.org.in>
"""

from pyramid.view import view_config
import requests, json
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import re

@view_config(route_name="customersuppliers",renderer="gkwebapp:templates/customersupplier.jinja2")
def showcustomersupplier(request):
    return {"status":True}

@view_config(route_name="customersuppliers",request_param="action=showadd",renderer="gkwebapp:templates/addcustomersupplier.jinja2")
def showaddcustomersupplier(request):
    header={"gktoken":request.headers["gktoken"]}
    customers = requests.get("http://127.0.0.1:6543/customersupplier?qty=custall", headers=header)
    suppliers = requests.get("http://127.0.0.1:6543/customersupplier?qty=supall", headers=header)
    groups = requests.get("http://127.0.0.1:6543/groupsubgroups?groupflatlist", headers=header)
    states = requests.get("http://127.0.0.1:6543/state", headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    debtgroupcode = groups.json()["gkresult"]["Sundry Debtors"]
    credgroupcode = groups.json()["gkresult"]["Sundry Creditors for Purchase"]
    return {"customers": customers.json()["gkresult"], "suppliers": suppliers.json()["gkresult"], "debtgroupcode":debtgroupcode, "credgroupcode":credgroupcode, "states":states.json()["gkresult"],"vatorgstflag":resultgstvat.json()["gkresult"]}

@view_config(route_name="customersuppliers",request_param="action=showaddpopup",renderer="gkwebapp:templates/createcustsuppopup.jinja2")
def showaddcustomersupplierpopup(request):
    header={"gktoken":request.headers["gktoken"]}

    customers = requests.get("http://127.0.0.1:6543/customersupplier?qty=custall", headers=header)
    suppliers = requests.get("http://127.0.0.1:6543/customersupplier?qty=supall", headers=header)
    groups = requests.get("http://127.0.0.1:6543/groupsubgroups?groupflatlist", headers=header)
    states = requests.get("http://127.0.0.1:6543/state", headers=header)
    debtgroupcode = groups.json()["gkresult"]["Sundry Debtors"]
    credgroupcode = groups.json()["gkresult"]["Sundry Creditors for Purchase"]
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    return {"gkstatus" : request.params["status"], "customers": customers.json()["gkresult"], "suppliers": suppliers.json()["gkresult"], "debtgroupcode":debtgroupcode, "credgroupcode":credgroupcode, "states":states.json()["gkresult"], "vatorgstflag":resultgstvat.json()["gkresult"]}

@view_config(route_name="customersuppliers",request_param="action=showedit",renderer="gkwebapp:templates/editcustomersupplier.jinja2")
def showeditcustomersupplier(request):
    header={"gktoken":request.headers["gktoken"]}
    customers = requests.get("http://127.0.0.1:6543/customersupplier?qty=custall", headers=header)
    suppliers = requests.get("http://127.0.0.1:6543/customersupplier?qty=supall", headers=header)
    states = requests.get("http://127.0.0.1:6543/state", headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    return {"customers": customers.json()["gkresult"], "suppliers": suppliers.json()["gkresult"],"vatorgstflag":resultgstvat.json()["gkresult"] ,"states":states.json()["gkresult"],"noofcustomer" :len(customers.json()["gkresult"]),"noofsupplier":len(suppliers.json()["gkresult"])}

@view_config(route_name="customersuppliers",request_param="action=get",renderer="json")
def getcustomersupplier(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/customersupplier?qty=single&custid=%d"%int(request.params["custid"]),headers=header)
    
    return {"gkstatus": result.json()["gkstatus"], "gkresult": result.json()["gkresult"]}

@view_config(route_name="customersuppliers",request_param="action=save",renderer="json")
def savecustomersupplier(request):
    header={"gktoken":request.headers["gktoken"]}
    dataset={"custname":request.params["custname"],"custaddr":request.params["custaddr"],"custphone":request.params["custphone"],"custemail":request.params["custemail"],"custfax":request.params["custfax"],"state":request.params["state"],"custpan":request.params["custpan"],"custtan":request.params["custtan"],"csflag":int(request.params["csflag"]),"pincode":request.params["pincode"],"subgroupcode":request.params["subgroupname"]}
    if request.params.has_key("bankdetails"): 
        dataset["bankdetails"]=json.loads(request.params["bankdetails"])
    if "gstin" in request.params:
        dataset["gstin"]=json.loads(request.params["gstin"])
    result=requests.post("http://127.0.0.1:6543/customersupplier",data=json.dumps(dataset),headers=header)
    if result.json()["gkstatus"] == 0:
        if dataset["csflag"] == 3:
            gkdata = {"activity":dataset["custname"] + " customer created"}
        else:
            gkdata = {"activity":dataset["custname"] + " supplier created"}
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
    return {"gkstatus": result.json()["gkstatus"]}


@view_config(route_name="customersuppliers",request_param="action=edit",renderer="json")
def editcustomersupplier(request):
    header={"gktoken":request.headers["gktoken"]}
    dataset={"custname":request.params["custname"],"custaddr":request.params["custaddr"],"custphone":request.params["custphone"],"custemail":request.params["custemail"],"custfax":request.params["custfax"],"custpan":request.params["custpan"],"state":request.params["state"],"custtan":request.params["custtan"],"custid":int(request.params["custid"]),"pincode":request.params["pincode"]}
    if request.params.has_key("bankdetails"):
        dataset["bankdetails"]=json.loads(request.params["bankdetails"])
    if request.params.has_key("gstin"):
        dataset["gstin"]=json.loads(request.params["gstin"])
    result=requests.put("http://127.0.0.1:6543/customersupplier",data=json.dumps(dataset),headers=header)
    if result.json()["gkstatus"] == 0:
        accs = requests.get("http://127.0.0.1:6543/accounts", headers=header)
        for acc in accs.json()["gkresult"]:
            if acc["accountname"] == request.params["oldcustname"]:
                updateacc = {"accountname":request.params["custname"],"openingbal":acc["openingbal"],"accountcode":acc["accountcode"]}
                resulteditacc = requests.put("http://127.0.0.1:6543/accounts", data =json.dumps(updateacc),headers=header)
                break
    return {"gkstatus": result.json()["gkstatus"],"custsup":request.params["custsup"]}

@view_config(route_name="customersuppliers", request_param="action=delete",renderer="json")
def deletecustomersupplier(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/customersupplier?qty=single&custid=%d"%(int(request.params["custid"])), headers=header)
    custname = result.json()["gkresult"]["custname"]
    custdetails = {"csflag": result.json()["gkresult"]["csflag"], "custname":result.json()["gkresult"]["custname"]}
    dataset={"custid":int(request.params["custid"])}
    result = requests.delete("http://127.0.0.1:6543/customersupplier",data =json.dumps(dataset), headers=header)
    if result.json()["gkstatus"] == 0:
        accs = requests.get("http://127.0.0.1:6543/accounts", headers=header)
        for acc in accs.json()["gkresult"]:
            if acc["accountname"] == custname:
                gkdata={"accountcode":acc["accountcode"]}
                result = requests.delete("http://127.0.0.1:6543/accounts",data =json.dumps(gkdata), headers=header)
                break
        gkdata = {"activity":custname + " account deleted"}
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
        if custdetails["csflag"] == 3:
                  gkdata = {"activity":custdetails["custname"] + " customer deleted"}
        else:
            gkdata = {"activity":custdetails["custname"] + " supplier deleted"}
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
    return {"gkstatus":result.json()["gkstatus"],"csflag":custdetails["csflag"]}

@view_config(route_name='customersuppliers', request_param='action=getallcusts',renderer ='json')
def getallcusts(request):
    header={"gktoken":request.headers["gktoken"]}
    customers = requests.get("http://127.0.0.1:6543/customersupplier?qty=custall",headers=header)
    return {"customers":customers.json()["gkresult"]}

@view_config(route_name='customersuppliers', request_param='action=getallsups',renderer ='json')
def getallsups(request):
    header={"gktoken":request.headers["gktoken"]}
    suppliers = requests.get("http://127.0.0.1:6543/customersupplier?qty=supall",headers=header)
    return {"customers":suppliers.json()["gkresult"]}

@view_config(route_name='import',request_param='action=cussupimport',renderer='json')
def cussupimport(request):
    # try:
        print "ok"
        header={"gktoken":request.headers["gktoken"]}
        xlsxfile = request.POST['xlsxfile'].file
        wb=load_workbook(xlsxfile)
        ws=wb.sheetnames
        if len(ws)!=2:
            return {"gkstatus":6}
        elif ws[0]!="Customers" or ws[1]!="Suppliers":
            return {"gkstatus":6}
        states = requests.get("http://127.0.0.1:6543/state",headers=header)
        stateList,statecodeList=[],[]
        for state in states.json()["gkresult"]:
            for cd in state.keys():
                statecodeList.append(str(cd))
            for st in state.values():
                stateList.append(st)
        # check errors in file contents
        errorTuples={"Customers":[],"Suppliers":[]}
        errorRows={"Customers":{},"Suppliers":{}}
        dupFlag=False
        for sheet_no in range(len(ws)):
            wb._active_sheet_index=sheet_no
            activeSheet=wb.active
            sheetList = tuple(activeSheet.rows)
            # empty sheet
            if len(sheetList)==1:
                continue
            sheetErrors=[]
            for sheetRow in sheetList[1:]:
                # if only GSTIN is present
                if sheet_no==0 and sheetRow[0].value==None and sheetRow[1].value==None and sheetRow[2].value==None and sheetRow[3].value==None and sheetRow[4].value==None and sheetRow[5].value==None and sheetRow[6].value==None and sheetRow[7].value==None and sheetRow[8].value!=None:
                        if int(str(sheetRow[8].value)[:2]) not in statecodeList and not re.match(r"^[a-zA-z]{5}\d{4}[a-zA-Z]{1}[0-9a-zA-Z]{1}[zZ]{1}[0-9]{1}$",str(sheetRow[8].value)[2:]):
                            sheetErrors.append(sheetRow[8].coordinate)
                elif sheet_no==1 and sheetRow[0].value==None and sheetRow[1].value==None and sheetRow[2].value==None and sheetRow[3].value==None and sheetRow[4].value==None and sheetRow[5].value==None and sheetRow[6].value==None and sheetRow[7].value==None and sheetRow[8].value!=None and sheetRow[9].value==None and sheetRow[10].value==None and sheetRow[11].value==None and sheetRow[12].value==None:
                        if int(str(sheetRow[8].value)[:2]) not in statecodeList and not re.match(r"^[a-zA-z]{5}\d{4}[a-zA-Z]{1}[0-9a-zA-Z]{1}[zZ]{1}[0-9]{1}$",str(sheetRow[8].value)[2:]):
                            sheetErrors.append(sheetRow[8].coordinate)
                else:
                    # name
                    if sheetRow[0].value==None:
                        sheetErrors.append(sheetRow[0].coordinate)
                    # email
                    if sheetRow[2].value!=None:
                        if not re.match(r"^\w+([\.-]?\w+)*@\w+([\.-]?\w+)*(\.\w{2,3})+$",str(sheetRow[2].value)):
                            sheetErrors.append(sheetRow[2].coordinate)
                    # address
                    if sheetRow[3].value==None:
                        sheetErrors.append(sheetRow[3].coordinate)
                    # pincode
                    if sheetRow[4].value==None:
                        sheetErrors.append(sheetRow[4].coordinate)
                    else:
                        if not re.match(r"^[0-9]{6}$",str(sheetRow[4].value)):
                            sheetErrors.append(sheetRow[4].coordinate)
                    # state
                    if sheetRow[5].value==None or sheetRow[5].value not in stateList:
                        sheetErrors.append(sheetRow[5].coordinate)
                    # PAN
                    if sheetRow[7].value!=None:
                        if not re.match(r"^[a-zA-z]{5}\d{4}[a-zA-Z]{1}$",str(sheetRow[7].value)):
                            sheetErrors.append(sheetRow[7].coordinate)
                    # GST
                    if sheetRow[8].value!=None:
                        if str(sheetRow[8].value)[:2] not in statecodeList or not re.match(r"^[a-zA-z]{5}\d{4}[a-zA-Z]{1}[0-9a-zA-Z]{1}[zZ]{1}[0-9]{1}$",str(sheetRow[8].value)[2:]):
                            sheetErrors.append(sheetRow[8].coordinate)
                    # for suppliers
                    if sheet_no==1:
                        # Bank Details
                        if sheetRow[9].value==None and sheetRow[10].value==None and sheetRow[11].value==None and sheetRow[12].value==None:
                            continue
                        elif sheetRow[9].value!=None and sheetRow[10].value!=None and sheetRow[11].value!=None and sheetRow[12].value!=None:
                            if not re.match(r"^[0-9]+$",str(sheetRow[9].value)):
                                sheetErrors.append(sheetRow[9].coordinate)
                        else:
                            if sheetRow[9].value==None:
                                sheetErrors.append(sheetRow[9].coordinate)
                            if sheetRow[10].value==None:
                                sheetErrors.append(sheetRow[10].coordinate)
                            if sheetRow[11].value==None:
                                sheetErrors.append(sheetRow[11].coordinate)
                            if sheetRow[12].value==None:
                                sheetErrors.append(sheetRow[12].coordinate)
            errorRowKey=str(ws[sheet_no])
            for errorCell in sheetErrors:
                cellTuple=openpyxl.utils.cell.coordinate_from_string(errorCell)
                errorRowNo=openpyxl.utils.cell.coordinate_from_string(errorCell)[1]
                errorTuples[errorRowKey].append(cellTuple)
                if errorRowNo not in errorRows[errorRowKey].keys():
                    errorRows[errorRowKey][errorRowNo]=[]
                    tempRow=sheetList[errorRowNo-1]
                    for tempCell in tempRow:
                        errorRows[errorRowKey][errorRowNo].append(tempCell.value)
        if len(errorTuples["Customers"])!=0 or len(errorTuples["Suppliers"])!=0:
            return {"gkstatus":6,"errorTuples":errorTuples,"errorRows":errorRows,"dupFlag":dupFlag}
        # import if no errors
        for sheet_no in range(len(ws)):
            wb._active_sheet_index=sheet_no
            activeSheet=wb.active
            sheetList = tuple(activeSheet.rows)
            # empty sheet
            if len(sheetList)==1:
                continue
            if ws[sheet_no]=="Customers":
                csflag=3
            if ws[sheet_no]=="Suppliers":
                csflag=19
            dupList=[]
            row_no=1
            while row_no < len(sheetList):
                sheetRow=sheetList[row_no]
                gstin_dict={}
                if sheetList[row_no][8].value!=None:
                    while(True):
                        gst_str=sheetList[row_no][8].value.encode('ascii','ignore')
                        state_str=gst_str[:2]
                        gstin_dict[state_str]=gst_str
                        row_no+=1
                        if row_no==len(sheetList):
                            break
                        if sheetList[row_no][0].value!=None:
                            break
                else:
                    row_no+=1
                cussupDict={}
                # customers
                if len(sheetRow)==9:
                    cussupDict= {"custname":sheetRow[0].value,"custphone":sheetRow[1].value, "custemail":sheetRow[2].value, "custaddr":sheetRow[3].value, "pincode":sheetRow[4].value, "state":sheetRow[5].value, "custfax":sheetRow[6].value, "custpan":sheetRow[7].value, "gstin":gstin_dict, "csflag":csflag}
                # suppliers without bank details
                elif sheetRow[9].value==None:
                    cussupDict= {"custname":sheetRow[0].value,"custphone":sheetRow[1].value, "custemail":sheetRow[2].value, "custaddr":sheetRow[3].value, "pincode":sheetRow[4].value, "state":sheetRow[5].value, "custfax":sheetRow[6].value, "custpan":sheetRow[7].value, "gstin":gstin_dict, "csflag":csflag}
                # suppliers with bank details
                else:
                    bank_dict={"ifsc":sheetRow[12].value,"bankname":sheetRow[10].value,"accountno":sheetRow[9].value,"branchname":sheetRow[11].value}
                    cussupDict= {"custname":sheetRow[0].value,"custphone":sheetRow[1].value, "custemail":sheetRow[2].value, "custaddr":sheetRow[3].value, "pincode":sheetRow[4].value, "state":sheetRow[5].value, "custfax":sheetRow[6].value, "custpan":sheetRow[7].value, "gstin":gstin_dict, "csflag":csflag, "bankdetails":bank_dict}
                result = requests.post("http://127.0.0.1:6543/customersupplier",data = json.dumps(cussupDict),headers=header)
                result=result.json()
                if result["gkstatus"]==1:
                    dupFlag=True
                    dupList.append(sheetRow[0].coordinate)
                elif result["gkstatus"]!=0:
                    return {"gkstatus":result["gkstatus"]}
            if dupFlag==True:
                errorRowKey=str(ws[sheet_no])
                for errorCell in dupList:
                    cellTuple=openpyxl.utils.cell.coordinate_from_string(errorCell)
                    errorRowNo=openpyxl.utils.cell.coordinate_from_string(errorCell)[1]
                    errorTuples[errorRowKey].append(cellTuple)
                    if errorRowNo not in errorRows[errorRowKey].keys():
                        errorRows[errorRowKey][errorRowNo]=[]
                        tempRow=sheetList[errorRowNo-1]
                        for tempCell in tempRow:
                            errorRows[errorRowKey][errorRowNo].append(tempCell.value)
        if dupFlag==True:
            return {"gkstatus":6,"errorTuples":errorTuples,"errorRows":errorRows,"dupFlag":dupFlag}
        return {"gkstatus":0}
    # except:
    #     return {"gkstatus":3}