
"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@dff.org.in>
"Arun Kelkar" <arunkelkar@dff.org.in>
"Ishan Masdekar " <imasdekar@dff.org.in>
"Navin Karkera" <navin@dff.org.in>
"Vanita Rajpurohit" <vanita.rajpurohit9819@gmail.com>
"Reshma Bhatawadekar" <reshma_b@riseup.net>
"Nitesh Chaughule <nitesh@disroot.org>" 
"Sanket Kolnoorkar" <sanketf123@gmail.com>
"""
from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from pyramid.response import Response
from PIL import Image
import base64
import cStringIO
import openpyxl
from openpyxl.styles import Font, Alignment
import os


@view_config(route_name="deliverychallan",renderer="gkwebapp:templates/deliverychallan.jinja2")
def showdeliverychallan(request):
    header={"gktoken":request.headers["gktoken"]}
    return {"status":True}
@view_config(route_name="deliverychallan",request_param="action=showadd",renderer="gkwebapp:templates/adddeliverychallan.jinja2")
def showadddeliverychallan(request):
    header={"gktoken":request.headers["gktoken"]}
    lastdelchaldata = {}
    if request.params["status"]=='in':
        podata = requests.get("http://127.0.0.1:6543/purchaseorder?psflag=16", headers=header)
        suppliers = requests.get("http://127.0.0.1:6543/customersupplier?qty=supall", headers=header)
        lastdelchaldata = requests.get("http://127.0.0.1:6543/delchal?delchal=last&status=9", headers=header)
    else:
        podata = requests.get("http://127.0.0.1:6543/purchaseorder?psflag=20", headers=header)
        suppliers = requests.get("http://127.0.0.1:6543/customersupplier?qty=custall", headers=header)
        lastdelchaldata = requests.get("http://127.0.0.1:6543/delchal?delchal=last&status=15", headers=header)
    productsnservices = requests.get("http://127.0.0.1:6543/products", headers=header)
    products = requests.get("http://127.0.0.1:6543/products?invdc=4", headers=header)
    states = requests.get("http://127.0.0.1:6543/state", headers=header)
    godowns = requests.get("http://127.0.0.1:6543/godown", headers=header)
    result = requests.get("http://127.0.0.1:6543/purchaseorder?psflag=16",headers=header)
    result1 = requests.get("http://127.0.0.1:6543/purchaseorder?psflag=20",headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    return {"gkstatus": request.params["status"],"suppliers": suppliers.json()["gkresult"],"products": products.json()["gkresult"],"productsnservices": productsnservices.json()["gkresult"],"godowns":godowns.json()["gkresult"],"purchaseorders":podata.json()["gkresult"], "lastdelchaldata":lastdelchaldata.json()["gkresult"], "numberofpurchaseorders":len(result.json()["gkresult"]),"numberofsalesorders":len(result1.json()["gkresult"]),"numberofgodowns":len(godowns.json()["gkresult"]),"states": states.json()["gkresult"], "resultgstvat":resultgstvat.json()["gkresult"]}
#delchalins and delchalouts are used to get the data of deliverych in i.e. 9 and deliverych out i.e. 15
@view_config(route_name="deliverychallan",request_param="action=showedit",renderer="gkwebapp:templates/editdeliverychallan.jinja2")
def showeditdeliverychallan(request):

    header={"gktoken":request.headers["gktoken"]}
    delchalins = requests.get("http://127.0.0.1:6543/delchal?delchal=all&inoutflag=9", headers=header)
    delchalouts = requests.get("http://127.0.0.1:6543/delchal?delchal=all&inoutflag=15", headers=header)
    suppliers = requests.get("http://127.0.0.1:6543/customersupplier?qty=supall", headers=header)
    customers = requests.get("http://127.0.0.1:6543/customersupplier?qty=custall", headers=header)
    godowns = requests.get("http://127.0.0.1:6543/godown", headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    return {"delchalins":delchalins.json()["gkresult"],"delchalouts":delchalouts.json()["gkresult"],"suppliers":suppliers.json()["gkresult"],"customers":customers.json()["gkresult"],"godowns":godowns.json()["gkresult"],"numberofgodowns":len(godowns.json()["gkresult"]),"resultgstvat":resultgstvat.json()["gkresult"],"numberofdeliveryinnotes":len(delchalins.json()["gkresult"]),"numberofdeliveryoutnotes":len(delchalouts.json()["gkresult"]),"status":True}
       
@view_config(route_name="deliverychallan", request_param="action=showeditpopup", renderer="gkwebapp:templates/editdeliverychallanpopup.jinja2")
def showeditpopupdeliverychallan(request):
    header={"gktoken":request.headers["gktoken"]}
    dcid = request.params["id"]
    delchal = requests.get("http://127.0.0.1:6543/delchal?delchal=single&dcid=%d"%(int(dcid)), headers=header)
    return {"gkstatus": delchal.json()["gkstatus"], "delchaldata": delchal.json()["gkresult"]["delchaldata"], "stockdata": delchal.json()["gkresult"]["stockdata"]}

@view_config(route_name="deliverychallan",request_param="action=getproducts",renderer="json")
def getproducts(request):
    header={"gktoken":request.headers["gktoken"]}
    products = requests.get("http://127.0.0.1:6543/products?invdc=4", headers=header)
    return {"gkstatus": products.json()["gkstatus"],"products": products.json()["gkresult"]}

#Get Applied Tax on Selected Product
@view_config(route_name="deliverychallan",request_param="action=getappliedtax",renderer="json")
def delchalgetappliedtax(request):
    header={"gktoken":request.headers["gktoken"]}
    try:
        taxdetails = requests.get("http://127.0.0.1:6543/products?type=pt&productcode=%d&source=%s&destination=%s&taxflag=%d"%(int(request.params["productcode"]),request.params["source"],request.params["destination"],int(request.params["taxflag"])), headers=header)
        data = taxdetails.json()["gkresult"]
        return{"gkstatus":taxdetails.json()["gkstatus"],"tax":data}
    except:
        return {"gkstatus":1}
    
#Get Product Details
@view_config(route_name="deliverychallan",request_param="action=getproduct",renderer="json")
def delchalgetproduct(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/products?qty=single&productcode=%d"%(int(request.params['productcode'])),headers=header)
    data = result.json()["gkresult"]
    if data.has_key("unitname"):
        return {"gkstatus": result.json()["gkstatus"],"unitname":data["unitname"],"gsflag":data["gsflag"],"gscode":data["gscode"]}
    else:
        return {"gkstatus": result.json()["gkstatus"],"gsflag":data["gsflag"],"gscode":data["gscode"]}


@view_config(route_name="deliverychallan",request_param="action=getpurchaseorder",renderer="json")
def getpurchaseorder(request):
    header={"gktoken":request.headers["gktoken"]}
    podata = requests.get("http://127.0.0.1:6543/purchaseorder?poso=single&orderid=%d"%(int(request.params["orderid"])), headers=header)
    return {"gkstatus": podata.json()["gkstatus"],"podata": podata.json()["gkresult"]}

@view_config(route_name="deliverychallan",request_param="action=getdelchal",renderer="json")
def getdelchal(request):
    header={"gktoken":request.headers["gktoken"]}
    delchaldata = requests.get("http://127.0.0.1:6543/delchal?delchal=single&dcid=%d"%(int(request.params["dcid"])), headers=header)
    delchalresult = {}
    delchalresult = delchaldata.json()["gkresult"]
    if int(delchalresult["delchaldata"]["inoutflag"])==9:
        inoutflag="in"
    else:
        inoutflag="out"
    return {"gkstatus": delchaldata.json()["gkstatus"],"delchaldata": delchaldata.json()["gkresult"],"inoutflag":inoutflag}


@view_config(route_name="deliverychallan",request_param="action=save",renderer="json")
def savedelchal(request):
    header={"gktoken":request.headers["gktoken"]}
    delchaldata = {"custid":int(request.params["custid"]),"dcno":request.params["dcno"],"dcdate":request.params["dcdate"],"dcflag":request.params["dcflag"], "noofpackages":request.params["noofpackages"], "modeoftransport":request.params["modeoftransport"],"taxstate":request.params["taxstate"],"tax":json.loads(request.params["tax"]),"cess":json.loads(request.params["cess"]), "delchaltotal":"%2f"%float(request.params["delchaltotal"]), "freeqty":json.loads(request.params["freeqty"]), "discount":json.loads(request.params["discount"]),"sourcestate":request.params["sourcestate"], "taxflag":request.params["taxflag"], "orgstategstin":request.params["orgstategstin"], "contents":json.loads(request.params["contents"]),"vehicleno":request.params["vehicleno"],"inoutflag":request.params["inoutflag"]}

    stockdata = {"inout":int(request.params["inout"])}
    if request.params.has_key("goid"):
        stockdata["goid"]=int(request.params["goid"])
    if request.params.has_key("issuername"):
        delchaldata["issuername"]=request.params["issuername"]
    if request.params.has_key("designation"):
        delchaldata["designation"]=request.params["designation"]
    if request.params.has_key("consignee"):
        delchaldata["consignee"]=json.loads(request.params["consignee"])
    if request.params["dateofsupply"] != "":
        delchaldata["dateofsupply"] = request.params["dateofsupply"]
    try:
        files = {}
        count = 0
        for i in request.POST.keys():
            if "file" not in i:
                continue
            else:
                img = request.POST[i].file
                image = Image.open(img)
                imgbuffer = cStringIO.StringIO()
                image.save(imgbuffer, format="JPEG")
                img_str = base64.b64encode(imgbuffer.getvalue())
                image.close()
                files[count] = img_str
                count += 1
        if len(files)>0:
            delchaldata["attachment"] = files
            delchaldata["attachmentcount"] = len(delchaldata["attachment"])
    except:
        print "no attachment found"
    delchalwholedata = {"delchaldata":delchaldata,"stockdata":stockdata}
    result=requests.post("http://127.0.0.1:6543/delchal",data=json.dumps(delchalwholedata),headers=header)
    if result.json()["gkstatus"]==0:
        return {"gkstatus":result.json()["gkstatus"],"gkresult":result.json()["gkresult"]}
    else:
        return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="deliverychallan",request_param="action=edit",renderer="json")
def editdelchal(request):
    header={"gktoken":request.headers["gktoken"]}
    delchaldata = {"custid":int(request.params["custid"]),"dcno":request.params["dcno"],"dcid":request.params["dcid"],"dcdate":request.params["dcdate"],"dcflag":request.params["dcflag"], "noofpackages":request.params["noofpackages"], "modeoftransport":request.params["modeoftransport"]}
    products = {}
    for  row in json.loads(request.params["products"]):
        products[row["productcode"]] = row["qty"]
    stockdata = {"inout":int(request.params["inout"]),"items":products}
    if request.params["goid"]!='':
        stockdata["goid"]=int(request.params["goid"])
    if request.params.has_key("issuername"):
        delchaldata["issuername"]=request.params["issuername"]
    if request.params.has_key("designation"):
        delchaldata["designation"]=request.params["designation"]
    delchalwholedata = {"delchaldata":delchaldata,"stockdata":stockdata}
    result=requests.put("http://127.0.0.1:6543/delchal",data=json.dumps(delchalwholedata),headers=header)
    return {"gkstatus":result.json()["gkstatus"]}

#Code for Export To spreadsheet
@view_config(route_name="deliverychallan",request_param="action=unbillspreadsheet", renderer="")
def unbillspreadsheet(request):
    try:
        header={"gktoken":request.headers["gktoken"]}
        fystart = str(request.params["fystart"]);
        fyend = str(request.params["fyend"]);
        orgname = str(request.params["orgname"])
        #orgname += " (FY: " + fystart+" to "+fyend +")"
        inputdate = request.params["inputdate"]
        del_unbilled_type = str(request.params["del_unbilled_type"]);
        if del_unbilled_type == "All":
            del_unbilled_type = "0"
            deltype = "All Types"
            merge = 6
        elif del_unbilled_type == "Approval":
            del_unbilled_type = "1"
            deltype = "Delivery Type : Approval"
            merge = 5
        elif del_unbilled_type == "Consignment":
            del_unbilled_type = "3"
            deltype = "Delivery Type : Consignment"
            merge = 5
        elif del_unbilled_type == "Sale":
            del_unbilled_type = "4"
            deltype = "Delivery Type : Sale"
            merge = 5
        elif del_unbilled_type == "Purchase":
            del_unbilled_type = "16"
            deltype = "Delivery Type : Purchase"
            merge = 5
        gkdata = {"inputdate": inputdate, "del_unbilled_type": del_unbilled_type}
        new_inputdate = datetime.strftime(datetime.strptime(str(inputdate),"%Y-%m-%d").date(),'%d-%m-%Y')
        inout = request.params["inout"]
        if inout == "9":
            result = requests.get("http://127.0.0.1:6543/report?type=del_unbilled&inout=i", data = json.dumps(gkdata), headers=header)
            headingtext="Inward Deliveries - Invoices Not Received | All Godowns | %s"%deltype
            title = "Supplier Name"
        elif inout == "15":
            result = requests.get("http://127.0.0.1:6543/report?type=del_unbilled&inout=o", data = json.dumps(gkdata), headers=header)
            headingtext = "Outward Deliveries - Invoices Not Received | All Godowns | %s"%deltype
            title = "Customer Name"
        result = result.json()["gkresult"]
        # A workbook is opened.
        unbilldelwb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = unbilldelwb.active
        # Title of the sheet and width of columns are set
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 18
        sheet.column_dimensions['C'].width = 14
        sheet.column_dimensions['D'].width = 24
        sheet.column_dimensions['E'].width = 16
        sheet.column_dimensions['F'].width = 16
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells('A1:F2')
        # Name and Financial Year of organisation is fetched to be displayed on the first row.
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        # Organisation name and financial year are displayed.
        sheet.merge_cells('A3:F3')
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet.merge_cells('A4:F4')
        sheet['A4'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A4'].alignment = Alignment(horizontal = 'center', vertical='center')
        if inout == "9":
               sheet.title = "Deliveries In"
        elif inout == "15":
            sheet.title = "Deliveries Out"
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet['A3'] = headingtext
        sheet['A4'] = "As on Date: "+new_inputdate
        if request.params["del_unbilled_type"] == "9":
            sheet.title = "Deliveries In"
            sheet['A4'] = "Deliveries In"
        elif request.params["del_unbilled_type"] == "15":
            sheet.title = "Deliveries Out"
            sheet['A4'] = "Deliveries Out"
        sheet['A5'] = "Sr. No."
        sheet['B5'] = "Deli. Note No."
        sheet['C5'] = "Deli. Note Date"
        sheet['D5'] = title
        sheet['E5'] = "Godown Name"
        if del_unbilled_type == "0":
            sheet['F5'] = "Delivery Type"
        titlerow = sheet.row_dimensions[5]
        titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
        titlerow.alignment = Alignment(horizontal = 'center', vertical='center')
        row = 6
        for deliverychallan in result:
            sheet['A'+str(row)] = deliverychallan["srno"]
            sheet['A'+str(row)].alignment = Alignment(horizontal='center')
            sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['B'+str(row)] = deliverychallan["dcno"]
            sheet['B'+str(row)].alignment = Alignment(horizontal='center')
            sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['C'+str(row)] = deliverychallan["dcdate"]
            sheet['C'+str(row)].alignment = Alignment(horizontal='center')
            sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['D'+str(row)] = deliverychallan["custname"]
            sheet['D'+str(row)].alignment = Alignment(horizontal='center')
            sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['E'+str(row)] = deliverychallan["goname"]
            sheet['E'+str(row)].alignment = Alignment(horizontal='center')
            sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            if del_unbilled_type == "0":
                sheet['F'+str(row)]  = deliverychallan["dcflag"]
                sheet['F'+str(row)].alignment = Alignment(horizontal='center')
                sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            row += 1
        unbilldelwb.save('report.xlsx')
        xlsxfile = open("report.xlsx","r")
        reportxslx = xlsxfile.read()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(reportxslx),'Content-Disposition': 'attachment; filename=report.xlsx', 'Set-Cookie':'fileDownload=true; path=/'}
        xlsxfile.close()
        os.remove("report.xlsx")
        return Response(reportxslx, headerlist=headerList.items())
    except:
        return {"gkstatus":3}

@view_config(route_name="deliverychallan",request_param="action=print",renderer="gkwebapp:templates/printdeliverychallan.jinja2")
def deliveryprint(request):
    header={"gktoken":request.headers["gktoken"]}
    delchaldata = requests.get("http://127.0.0.1:6543/delchal?delchal=single&dcid=%d"%(int(request.params["dcid"])), headers=header)
    if delchaldata.json()["gkresult"]["delchalflag"] ==15:
        org = requests.get("http://127.0.0.1:6543/organisation", headers=header)
    else:
        statecode = delchaldata.json()["gkresult"]["sourcestatecode"]
        org = requests.get("http://127.0.0.1:6543/organisations?billingdetails&statecode=%d"%(int(statecode)), headers=header)
    cust = requests.get("http://127.0.0.1:6543/customersupplier?qty=single&custid=%d"%(int(delchaldata.json()["gkresult"]["custSupDetails"]["custid"])), headers=header)
    if delchaldata.json()["gkresult"]["delchaldata"]["goid"]:
        godown = requests.get("http://127.0.0.1:6543/godown?qty=single&goid=%d"%(int(delchaldata.json()["gkresult"]["delchaldata"]["goid"])), headers=header)
        godowndata = godown.json()["gkresult"]
    else:
        godowndata = ''
    return {"gkstatus":org.json()["gkstatus"],"org":org.json()["gkdata"],"gkresult":delchaldata.json()["gkresult"],"cust":cust.json()["gkresult"],"godown":godowndata, "delchalflag":delchaldata.json()["gkresult"]["delchalflag"]}

@view_config(route_name="show_del_unbilled_report",renderer="gkwebapp:templates/unbilled_deliveries_report.jinja2")
def show_unbilled_deliveries_report(request):
    header={"gktoken":request.headers["gktoken"]}
    inputdate = request.params["inputdate"]
    inout = request.params["inout"]
    del_unbilled_type = request.params["del_unbilled_type"]
    if del_unbilled_type == "All":
        del_unbilled_type = "0"
    elif del_unbilled_type == "Approval":
        del_unbilled_type = "1"
    elif del_unbilled_type == "Consignment":
        del_unbilled_type = "3"
    elif del_unbilled_type == "Sale":
        del_unbilled_type = "4"
    elif del_unbilled_type == "Purchase":
        del_unbilled_type = "16"
    gkdata = {"inputdate": inputdate, "del_unbilled_type": del_unbilled_type}
    new_inputdate = datetime.strftime(datetime.strptime(str(inputdate),"%Y-%m-%d").date(),'%d-%m-%Y')
    if inout == "9":
        result = requests.get("http://127.0.0.1:6543/report?type=del_unbilled&inout=i", data = json.dumps(gkdata), headers=header)
    elif inout == "15":
        result = requests.get("http://127.0.0.1:6543/report?type=del_unbilled&inout=o", data = json.dumps(gkdata), headers=header)
    if del_unbilled_type == "0":
        del_unbilled_type = "All"
    elif del_unbilled_type == "1":
        del_unbilled_type = "Approval"
    elif del_unbilled_type == "3":
        del_unbilled_type = "Consignment"
    elif del_unbilled_type == "4":
        del_unbilled_type = "Sale"
    elif del_unbilled_type == "16":
        del_unbilled_type = "Purchase"

    return {"gkstatus":result.json()["gkstatus"], "gkresult": result.json()["gkresult"], "inputdate":inputdate, "new_inputdate":new_inputdate, "inout":inout, "del_unbilled_type": del_unbilled_type}

@view_config(route_name="del_unbilled", request_param="action=view", renderer="gkwebapp:templates/view_unbilled_deliveries.jinja2")
def view_unbilled_deliveries(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/godown", headers=header)
    goddata=[]
    for record in result.json()["gkresult"]:
        gdata= {"godownid": str(record["goid"]), "godownname" : str(record["goname"])}
        goddata.append(gdata)

    return {"gkstatus":result.json()["gkstatus"], "gkresult": goddata}

@view_config(route_name="print_unbilled_deliveries_report",renderer="gkwebapp:templates/print_unbilled_deliveries.jinja2")
def print_del_unbilled(request):
    header={"gktoken":request.headers["gktoken"]}
    inputdate = request.params["inputdate"];
    del_unbilled_type = request.params["del_unbilled_type"];
    if del_unbilled_type == "All":
        del_unbilled_type = "0"
    elif del_unbilled_type == "Approval":
        del_unbilled_type = "1"
    elif del_unbilled_type == "Consignment":
        del_unbilled_type = "3"
    elif del_unbilled_type == "Sale":
        del_unbilled_type = "4"
    elif del_unbilled_type == "Purchase":
        del_unbilled_type = "16"
    gkdata = {"inputdate": inputdate, "del_unbilled_type": del_unbilled_type}
    new_inputdate = datetime.strftime(datetime.strptime(str(inputdate),"%Y-%m-%d").date(),'%d-%m-%Y')
    inout = request.params["inout"]
    if inout == "9":
        result = requests.get("http://127.0.0.1:6543/report?type=del_unbilled&inout=i", data = json.dumps(gkdata), headers=header)
    elif inout == "15":
        result = requests.get("http://127.0.0.1:6543/report?type=del_unbilled&inout=o", data = json.dumps(gkdata), headers=header)
    if del_unbilled_type == "0":
        del_unbilled_type = "All"
    elif del_unbilled_type == "1":
        del_unbilled_type = "Approval"
    elif del_unbilled_type == "3":
        del_unbilled_type = "Consignment"
    elif del_unbilled_type == "4":
        del_unbilled_type = "Sale"
    elif del_unbilled_type == "16":
        del_unbilled_type = "Purchase"
    return {"gkstatus":result.json()["gkstatus"], "gkresult": result.json()["gkresult"], "inputdate":inputdate, "new_inputdate":new_inputdate, "inout":inout, "del_unbilled_type": del_unbilled_type}

@view_config(route_name="deliverychallan", request_param="action=getattachment", renderer="gkwebapp:templates/viewdelchalattachment.jinja2")
def getattachment(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/delchal?attach=image&dcid=%d"%(int(request.params["dcid"])), headers=header)
    return {"attachment":result.json()["gkresult"],"dcid":request.params["dcid"], "cancelflag":result.json()["cancelflag"],"userrole":result.json()["userrole"],"dcno":result.json()["dcno"]}
