"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Ishan Masdekar " <imasdekar@dff.org.in>
"Bhavesh Bawadhane" <bbhavesh07@gmail.com>
"Mohd. Talha Pawaty" <mtalha456@gmail.com>
"Aditya Shukla" <adityashukla9158.as@gmail.com>
"Reshma Bhatawadekar" <reshma_b@riseup.net>
"""

from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from pyramid.response import Response
from PIL import Image
import io
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import base64
from io import BytesIO

@view_config(route_name="invoice",renderer="gkwebapp:templates/invoice.jinja2")
def showinvoice(request):
    return {"status":True}

@view_config(route_name="invoice",request_param="action=invoiceviewlist",renderer="gkwebapp:templates/invoiceviewlist.jinja2")
def showinvoicelist(request):
    return {"status":True}

@view_config(route_name="invoice", request_param="action=showviewregister", renderer="gkwebapp:templates/viewregister.jinja2")
def showviewregister(request):
    return {"status":True}

@view_config(route_name="invoice", request_param="action=viewlist", renderer="gkwebapp:templates/viewlistofinvoices.jinja2")
def showlistofinv(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/invoice?inv=all", headers=header)
    return {"status":True, "numberofinvoices": len(result.json()["gkresult"]),"deleteflag":0}

@view_config(route_name="invoice", request_param="action=viewlistdeleted", renderer="gkwebapp:templates/viewlistofinvoices.jinja2")
def showlistofdeletedinv(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/invoice?inv=alldeleted", headers=header)
    return {"status":True, "numberofinvoices": len(result.json()["gkresult"]),"deleteflag":1}

@view_config(route_name="invoice",request_param="action=showadd",renderer="gkwebapp:templates/addinvoice.jinja2")
def showaddinvoice(request):
    header={"gktoken":request.headers["gktoken"]}
    inputdate = request.params["inputdate"]
    gkdata = {"inputdate": inputdate, "type": "invoice"}
    unbilled_delnotes = requests.get("http://127.0.0.1:6543/invoice?unbilled_delnotes", data=json.dumps(gkdata), headers=header)
    if request.params["status"]=='in':
        suppliers = requests.get("http://127.0.0.1:6543/customersupplier?qty=supall", headers=header)
    else:
        suppliers = requests.get("http://127.0.0.1:6543/customersupplier?qty=custall", headers=header)
    productsnservices = requests.get("http://127.0.0.1:6543/products", headers=header)
    products = requests.get("http://127.0.0.1:6543/products?invdc=4", headers=header)
    states = requests.get("http://127.0.0.1:6543/state", headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    resultAvflag = requests.get("http://127.0.0.1:6543/organisation?autovoucher",headers=header)
    return {"gkstatus": request.params["status"],"suppliers": suppliers.json()["gkresult"],"products": products.json()["gkresult"],"productsnservices": productsnservices.json()["gkresult"],"deliverynotes":unbilled_delnotes.json()["gkresult"],"states": states.json()["gkresult"], "resultgstvat":resultgstvat.json()["gkresult"]}

@view_config(route_name="invoice",request_param="action=getproducts",renderer="json")
def getproducts(request):
    header={"gktoken":request.headers["gktoken"]}
    if int(request.params["taxflag"]) == 7:
        products = requests.get("http://127.0.0.1:6543/products", headers=header)
    elif int(request.params["taxflag"]) == 22:
        products = requests.get("http://127.0.0.1:6543/products?invdc=4", headers=header)
    return {"gkstatus": products.json()["gkstatus"],"products": products.json()["gkresult"]}


@view_config(route_name="invoice",request_param="action=save",renderer="json")
def saveinvoice(request):
    header={"gktoken":request.headers["gktoken"]}
    invoicedata = {"invoiceno":request.params["invoiceno"],"ewaybillno":request.params["ewaybillno"], "taxstate":request.params["taxstate"],"invoicedate":request.params["invoicedate"],"tax":json.loads(request.params["tax"]), "cess":json.loads(request.params["cess"]),"custid":request.params["custid"],"invoicetotal":request.params["invtotal"],"invoicetotalword":request.params["invtotalword"], "contents":json.loads(request.params["contents"]),"issuername":request.params["issuername"],"designation":request.params["designation"],"freeqty":json.loads(request.params["freeqty"]), "discount":json.loads(request.params["discount"]), "consignee":json.loads(request.params["consignee"]), "taxflag":request.params["taxflag"],"sourcestate":request.params["sourcestate"],"transportationmode":request.params["transportationmode"], "reversecharge":request.params["reversecharge"], "vehicleno":request.params["vehicleno"],"orgstategstin":request.params["orgstategstin"], "paymentmode":request.params["paymentmode"],"inoutflag":request.params["inoutflag"],"av":json.loads(request.params["av"]) ,"roundoffflag":int(request.params["roundoff"]),"discflag":int(request.params["discflag"])}
    if ("invoice_narration" in request.params):
        s = str(request.params["invoice_narration"]).isspace()
        if (s == False):
            invoicedata["invnarration"] = request.params["invoice_narration"]
    if request.params["dateofsupply"] != "":
        invoicedata["dateofsupply"] = request.params["dateofsupply"]
    if "bankdetails" in request.params:
        invoicedata["bankdetails"]=json.loads(request.params["bankdetails"])
    if "address" in request.params:
        invoicedata["address"] = request.params["address"]
    if "pincode" in request.params:
        invoicedata["pincode"] = request.params["pincode"]
    if "pricedetails" in request.params:
        invoicedata["pricedetails"] = json.loads(request.params["pricedetails"])
    try:
        files = {}
        count = 0
        for i in list(request.POST.keys()):
            if "file" not in i:
                continue
            else:
                img = request.POST[i].file
                image = Image.open(img)
                imgbuffer = io.BytesIO()
                image.save(imgbuffer, format="JPEG")
                img_str = base64.b64encode(imgbuffer.getvalue())
                img_str = img_str.decode("ascii")
                image.close()
                files[count] = img_str
                count += 1
        if len(files)>0:
            invoicedata["attachment"] = files
            invoicedata["attachmentcount"] = len(invoicedata["attachment"])
    except:
        print("no attachment found")
    stock = json.loads(request.params["stock"])
    if request.params["dcid"]!="":
        invoicedata["dcid"] = request.params["dcid"]
    
    invoicewholedata = {"invoice":invoicedata,"stock":stock}
    result=requests.post("http://127.0.0.1:6543/invoice",data=json.dumps(invoicewholedata),headers=header)
    if result.json()["gkstatus"]==0:
        return {"gkstatus":result.json()["gkstatus"],"gkresult":result.json()["gkresult"],"gkvch":result.json()["vchData"]}
    else:
        return {"gkstatus":result.json()["gkstatus"]}

'''
This is a function to edit an invoice.
This function receives data regarding invoice, delivery note linked if any and stock details if any.
The data received is sent to core engine.
'''
@view_config(route_name="invoice",request_param="action=update",renderer="json")
def updateinvoice(request):
    header={"gktoken":request.headers["gktoken"]}

    invoicedata = {"ewaybillno":request.params["ewaybillno"],"invid":request.params["invid"],"invoiceno":request.params["invoiceno"],"taxstate":request.params["taxstate"],"invoicedate":request.params["invoicedate"],"tax":json.loads(request.params["tax"]), "cess":json.loads(request.params["cess"]),"custid":request.params["custid"],"invoicetotal":request.params["invtotal"],"invoicetotalword":request.params["invtotalword"], "contents":json.loads(request.params["contents"]),"issuername":request.params["issuername"],"designation":request.params["designation"],"freeqty":json.loads(request.params["freeqty"]), "discount":json.loads(request.params["discount"]), "consignee":json.loads(request.params["consignee"]), "taxflag":request.params["taxflag"],"sourcestate":request.params["sourcestate"],"transportationmode":request.params["transportationmode"], "reversecharge":request.params["reversecharge"], "vehicleno":request.params["vehicleno"],"orgstategstin":request.params["orgstategstin"], "paymentmode":request.params["paymentmode"],"inoutflag":request.params["inoutflag"],"roundoffflag":int(request.params["roundoff"])}
    if ("invoice_narration" in request.params):
        s = str(request.params["invoice_narration"]).isspace()
        if (s == False):
            invoicedata["invnarration"] = request.params["invoice_narration"]
        else:
            invoicedata["invnarration"] = None
    if request.params["dateofsupply"] != "":
        invoicedata["dateofsupply"] = request.params["dateofsupply"]
    if "bankdetails" in request.params:
        invoicedata["bankdetails"]=json.loads(request.params["bankdetails"])
    if "address" in request.params:
        invoicedata["address"] = request.params["address"]
    if "pricedetails" in request.params:
        invoicedata["pricedetails"] = json.loads(request.params["pricedetails"])
    try:
        files = {}
        count = 0
        for i in list(request.POST.keys()):
            if "file" not in i:
                continue
            else:
                img = request.POST[i].file
                image = Image.open(img)
                imgbuffer = io.BytesIO()
                image.save(imgbuffer, format="JPEG")
                img_str = base64.b64encode(imgbuffer.getvalue())
                img_str = img_str.decode("ascii")
                image.close()
                files[count] = img_str
                count += 1
        if len(files)>0:
            invoicedata["attachment"] = files
            invoicedata["attachmentcount"] = len(invoicedata["attachment"])
    except:
        print("no attachment found")
    stock = json.loads(request.params["stock"])
    if request.params["dcid"]!="":
        invoicedata["dcid"] = request.params["dcid"]
   # invoicedata["roundoff"] = int(request.params["roundoff"])
    invoicewholedata = {"invoice":invoicedata,"stock":stock,"av":json.loads(request.params["av"])}
    result=requests.put("http://127.0.0.1:6543/invoice",data=json.dumps(invoicewholedata),headers=header)
    if result.json()["gkstatus"]==0:
        return {"gkstatus":result.json()["gkstatus"],"gkvch":result.json()["vchData"]}
    else:
        return {"gkstatus":result.json()["gkstatus"]}
@view_config(route_name="invoice",request_param="action=getdeliverynote",renderer="json")
def getdeliverynote(request):
    header={"gktoken":request.headers["gktoken"]}
    delchal = requests.get("http://127.0.0.1:6543/delchal?delchal=single&dcid=%d"%(int(request.params["dcid"])), headers=header)
    return {"gkstatus": delchal.json()["gkstatus"],"delchal": delchal.json()["gkresult"]}

@view_config(route_name="invoice",request_param="action=gettax",renderer="json")
def getstatetax(request):
    header={"gktoken":request.headers["gktoken"]}
    if request.params["state"]=="":
        taxdata = requests.get("http://127.0.0.1:6543/tax?pscflag=i&productcode=%d"%(int(request.params["productcode"])), headers=header)
    else:
        taxdata = requests.get("http://127.0.0.1:6543/tax?pscflag=i&productcode=%d&state=%s"%(int(request.params["productcode"]),request.params["state"]), headers=header)
    result = requests.get("http://127.0.0.1:6543/products?qty=single&productcode=%d"%(int(request.params['productcode'])),headers=header)
    unit = result.json()["gkresult"]
    return {"gkstatus": taxdata.json()["gkstatus"],"taxdata": taxdata.json()["gkresult"],"unitname":unit["unitname"]}

@view_config(route_name="invoice",request_param="action=getappliedtax",renderer="json")
def getappliedtax(request):
    header={"gktoken":request.headers["gktoken"]}
    try:
        taxdetails = requests.get("http://127.0.0.1:6543/products?type=pt&productcode=%d&source=%s&destination=%s&taxflag=%d"%(int(request.params["productcode"]),request.params["source"],request.params["destination"],int(request.params["taxflag"])), headers=header)
        data = taxdetails.json()["gkresult"]
        return{"gkstatus":taxdetails.json()["gkstatus"],"tax":data}
    except:
        return {"gkstatus":1}


@view_config(route_name="invoice",request_param="action=getproduct",renderer="json")
def getproduct(request):
    header={"gktoken":request.headers["gktoken"]}
    pricedata = 0.00
    result = requests.get("http://127.0.0.1:6543/products?qty=single&productcode=%d"%(int(request.params['productcode'])),headers=header)
    data = result.json()["gkresult"]
    if "custid" in request.params and request.params['custid'] != "":
       lastprice = requests.get("http://127.0.0.1:6543/products?type=lastprice&productcode=%d&inoutflag=%d&custid=%d"%(int(request.params['productcode']),int(request.params['inoutflag']),int(request.params['custid'])),headers=header)
       if lastprice.json()["gkstatus"] == 0:
            pricedata = lastprice.json()["gkresult"] 
    if "unitname" in data:
        return {"gkstatus": result.json()["gkstatus"],"unitname":data["unitname"],"gsflag":data["gsflag"],"gscode":data["gscode"], "prodsp":data["prodsp"], "prodmrp":data["prodmrp"], "prodlp":"%.2f"%float(pricedata), "productdesc":data["productdesc"],"discountpercent":data["discountpercent"],"discountamount":data["discountamount"]}
    else:
        return {"gkstatus": result.json()["gkstatus"],"gsflag":data["gsflag"],"gscode":data["gscode"], "prodsp":data["prodsp"], "prodmrp":data["prodmrp"], "prodlp":"%.2f"%float(pricedata), "productdesc":data["productdesc"],"discountpercent":data["discountpercent"],"discountamount":data["discountamount"]}

@view_config(route_name="invoice",request_param="type=getstatess",renderer="json")
def showstate(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/organisation", headers=header)
    orgstateforinv = ""
    if ("orgstate" in result.json()["gkdata"] and result.json()["gkdata"]["orgstate"] != ""):
            orgstateforinv = result.json()["gkdata"]["orgstate"]
    else:
        orgstateforinv = 'null'
    return {"gkresult":orgstateforinv}

@view_config(route_name="invoice",request_param="type=getstatelist",renderer="json")
def showstatelist(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/state", headers=header)
    return {"gkresult": result.json()["gkresult"]}
    

@view_config(route_name="invoice",request_param="action=showeditinv",renderer="gkwebapp:templates/editinvoice.jinja2")
def showeditableinvoices(request):
    header={"gktoken":request.headers["gktoken"]}
    inputdate = request.params["inputdate"]
    gkdata = {"inputdate": inputdate, "type": "invoice"}
    if request.params["status"] == "out":
        if "invid" in request.params:
            invlist = [{"invid":request.params["invid"], "invoiceno":"", "invoicedata":"", "custname":""}]
        else:
            result = requests.get("http://127.0.0.1:6543/invoice?type=rectifyinvlist&invtype=15", headers=header)
            invlist = result.json()["invoices"]
        suppliers = requests.get("http://127.0.0.1:6543/customersupplier?qty=custall", headers=header)
    else:
        if "invid" in request.params:
            invlist = [{"invid":request.params["invid"], "invoiceno":"", "invoicedata":"", "custname":""}]
        else:
            result = requests.get("http://127.0.0.1:6543/invoice?type=rectifyinvlist&invtype=9", headers=header)
            invlist = result.json()["invoices"]
        suppliers = requests.get("http://127.0.0.1:6543/customersupplier?qty=supall", headers=header)
    unbilled_delnotes = requests.get("http://127.0.0.1:6543/invoice?unbilled_delnotes", data=json.dumps(gkdata), headers=header)
    productsnservices = requests.get("http://127.0.0.1:6543/products", headers=header)
    products = requests.get("http://127.0.0.1:6543/products?invdc=4", headers=header)
    states = requests.get("http://127.0.0.1:6543/state", headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    return {"gkresult": invlist, "products": products.json()["gkresult"],"productsnservices": productsnservices.json()["gkresult"],"deliverynotes":unbilled_delnotes.json()["gkresult"],"states": states.json()["gkresult"], "resultgstvat":resultgstvat.json()["gkresult"], "suppliers": suppliers.json()["gkresult"], "status":request.params["status"]}

@view_config(route_name="invoice",request_param="action=getinvdetails",renderer="json")
def getInvoiceDetails(request):
    header={"gktoken":request.headers["gktoken"]}
    invoicedata = requests.get("http://127.0.0.1:6543/invoice?inv=single&invid=%d"%(int(request.params["invid"])), headers=header)
    return {"gkstatus": invoicedata.json()["gkstatus"],"invoicedata": invoicedata.json()["gkresult"]}

@view_config(route_name="invoice", request_param="action=showlist", renderer="gkwebapp:templates/listofinvoices.jinja2")
def listofinv(request):
    header={"gktoken":request.headers["gktoken"]}
    if "orderflag" in request.params:
        result = requests.get("http://127.0.0.1:6543/invoice?type=list&flag=%s&fromdate=%s&todate=%s&orderflag=%d"%(request.params["flag"], request.params["fromdate"], request.params["todate"],int(request.params["orderflag"])), headers=header)
        orderflag = "1"
    else:
       result = requests.get("http://127.0.0.1:6543/invoice?type=list&flag=%s&fromdate=%s&todate=%s"%(request.params["flag"], request.params["fromdate"], request.params["todate"]), headers=header)
       orderflag = "4"
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    return {"gkstatus":result.json()["gkstatus"], "gkresult": result.json()["gkresult"], "flag": request.params["flag"], "fromdate": request.params["fromdate"], "todate": request.params["todate"], "displayfromdate": datetime.strptime(request.params["fromdate"],'%Y-%m-%d').strftime('%d-%m-%Y'), "displaytodate": datetime.strptime(request.params["todate"],'%Y-%m-%d').strftime('%d-%m-%Y'), "resultgstvat":resultgstvat.json()["gkresult"],"orderflag":orderflag,"deleteflag":0}

@view_config(route_name="invoice", request_param="action=showdeletedlist", renderer="gkwebapp:templates/listofinvoices.jinja2")
def listofdeletedinv(request):
    header={"gktoken":request.headers["gktoken"]}
    if "orderflag" in request.params:
        result = requests.get("http://127.0.0.1:6543/invoice?type=listdeleted&flag=%s&fromdate=%s&todate=%s&orderflag=%d"%(request.params["flag"], request.params["fromdate"], request.params["todate"],int(request.params["orderflag"])), headers=header)
        orderflag = "1"
    else:
       result = requests.get("http://127.0.0.1:6543/invoice?type=listdeleted&flag=%s&fromdate=%s&todate=%s"%(request.params["flag"], request.params["fromdate"], request.params["todate"]), headers=header)
       orderflag = "4"
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    return {"gkstatus":result.json()["gkstatus"], "gkresult": result.json()["gkresult"], "flag": request.params["flag"], "fromdate": request.params["fromdate"], "todate": request.params["todate"], "displayfromdate": datetime.strptime(request.params["fromdate"],'%Y-%m-%d').strftime('%d-%m-%Y'), "displaytodate": datetime.strptime(request.params["todate"],'%Y-%m-%d').strftime('%d-%m-%Y'), "resultgstvat":resultgstvat.json()["gkresult"],"orderflag":orderflag,"deleteflag":1}

@view_config(route_name="invoice", request_param="action=printlist", renderer="gkwebapp:templates/printlistofinvoices.jinja2")
def printlistofinv(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/invoice?type=list&flag=%s&fromdate=%s&todate=%s"%(request.params["flag"], request.params["fromdate"], request.params["todate"]), headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    return {"gkstatus":result.json()["gkstatus"], "gkresult": result.json()["gkresult"], "flag": request.params["flag"], "fromdate": request.params["fromdate"], "todate": request.params["todate"], "displayfromdate": datetime.strptime(request.params["fromdate"],'%Y-%m-%d').strftime('%d-%m-%Y'), "displaytodate": datetime.strptime(request.params["todate"],'%Y-%m-%d').strftime('%d-%m-%Y'), "resultgstvat":resultgstvat.json()["gkresult"],"deleteflag":0}

@view_config(route_name="invoice", request_param="action=printcanceled", renderer="gkwebapp:templates/printlistofinvoices.jinja2")
def printlistofcanceledinv(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/invoice?type=listdeleted&flag=%s&fromdate=%s&todate=%s"%(request.params["flag"], request.params["fromdate"], request.params["todate"]), headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    return {"gkstatus":result.json()["gkstatus"], "gkresult": result.json()["gkresult"], "flag": request.params["flag"], "fromdate": request.params["fromdate"], "todate": request.params["todate"], "displayfromdate": datetime.strptime(request.params["fromdate"],'%Y-%m-%d').strftime('%d-%m-%Y'), "displaytodate": datetime.strptime(request.params["todate"],'%Y-%m-%d').strftime('%d-%m-%Y'), "resultgstvat":resultgstvat.json()["gkresult"],"deleteflag":1}

@view_config(route_name="invoice",request_param="action=showinv",renderer="gkwebapp:templates/viewsingleinvoice.jinja2")
def showsingleinvoice(request):
    header={"gktoken":request.headers["gktoken"]}
    invoicedata = requests.get("http://127.0.0.1:6543/invoice?inv=single&invid=%d"%(int(request.params["invid"])), headers=header)
    return {"gkstatus": invoicedata.json()["gkstatus"],"gkresult": invoicedata.json()["gkresult"],"deleteflag":0}

@view_config(route_name="invoice",request_param="action=showdeletedinv",renderer="gkwebapp:templates/viewsingleinvoice.jinja2")
def showsingledeletedinvoice(request):
    header={"gktoken":request.headers["gktoken"]}
    invoicedata = requests.get("http://127.0.0.1:6543/invoice?inv=deletedsingle&invid=%d"%(int(request.params["invid"])), headers=header)
    return {"gkstatus": invoicedata.json()["gkstatus"],"gkresult": invoicedata.json()["gkresult"],"deleteflag":1}

@view_config(route_name="invoice",request_param="action=showvoucher",renderer="gkwebapp:templates/voucherininvoice.jinja2")
def showvoucherininvoice(request):
    header={"gktoken":request.headers["gktoken"]}  
    result = requests.get("http://127.0.0.1:6543/transaction?searchby=invoice&invid=%d"%(int(request.params["invid"])),headers=header)
    total=[]
    totaldr = 0.00
    totalcr = 0.00
    for vc in result.json()["gkresult"]:
        for dr in vc["drs"]:
            totaldr = float(totaldr) + float(vc["drs"][dr])
        vc["totaldr"] = "%.2f"%totaldr

        for cr in vc["crs"]:
            totalcr = float(totalcr) + float(vc["crs"][cr])
        vc["totalcr"] = "%.2f"%totalcr

        type = vc["vouchertype"]
        projects = requests.get("http://127.0.0.1:6543/projects", headers=header)
        if type=="contra" or type=="journal":
            result1 = requests.get("http://127.0.0.1:6543/accountsbyrule?type=%s"%(type), headers=header)
            if result1.json()["gkstatus"]==0:
                draccounts=result1.json()["gkresult"]
                craccounts=result1.json()["gkresult"]
        elif type=="creditnote" or type=="debitnote" or type=="salesreturn" or type=="purchasereturn":
            result1 = requests.get("http://127.0.0.1:6543/accountsbyrule?type=journal", headers=header)
            if result1.json()["gkstatus"]==0:
                draccounts=result1.json()["gkresult"]
                craccounts=result1.json()["gkresult"]

        else:
            drresult = requests.get("http://127.0.0.1:6543/accountsbyrule?type=%s&side=Dr"%(type), headers=header)
            crresult = requests.get("http://127.0.0.1:6543/accountsbyrule?type=%s&side=Cr"%(type), headers=header)
            if drresult.json()["gkstatus"]==0 and crresult.json()["gkstatus"]==0:
                draccounts=drresult.json()["gkresult"]
                craccounts=crresult.json()["gkresult"]
            else:
                return render_to_response("gkwebapp:templates/index.jinja2",{"status":"Please select an organisation and login again"},request=request)
        
        total.append({"projects":projects.json()["gkresult"],"vtype":type,"voucher":vc,"draccounts":draccounts,"craccounts":craccounts,"invoicedata":0})
        totaldr = 0.00
        totalcr = 0.00
    if result.json()["gkstatus"]==0:
    	return {"total":total}
    else:
    	return render_to_response("gkwebapp:templates/index.jinja2",{"status":"Please select an organisation and login again"},request=request)

@view_config(route_name="invoice",request_param="type=cancelinvoice",renderer="json")
def InvoiceCancel(request):
    header={"gktoken":request.headers["gktoken"]}
    invoicedata = requests.delete("http://127.0.0.1:6543/invoice?type=cancel",data =json.dumps({"invid":request.params["invid"]}), headers=header)
    return {"gkstatus": invoicedata.json()["gkstatus"]}


@view_config(route_name="invoice",request_param="action=print")
def Invoiceprint(request):
    header={"gktoken":request.headers["gktoken"]}
    invoicedata = requests.get("http://127.0.0.1:6543/invoice?inv=single&invid=%d"%(int(request.params["invid"])), headers=header)
    if  (invoicedata.json()["gkresult"]['custSupDetails']['custgstin']) == None:
        invoicedata.json()["gkresult"]['custSupDetails']['custgstin'] = ''
    statecode = invoicedata.json()["gkresult"]["sourcestatecode"]
    org = requests.get("http://127.0.0.1:6543/organisations?billingdetails&statecode=%d"%(int(statecode)), headers=header)
    if(request.params["pflag"] == '2'):
     
        return render_to_response("gkwebapp:templates/printinvoice2.jinja2",{"gkstatus":org.json()["gkstatus"],"org":org.json()["gkdata"],"gkresult":invoicedata.json()["gkresult"],"invid":(int(request.params["invid"]))},request=request)
    elif(request.params["pflag"] == '3'):
        return render_to_response("gkwebapp:templates/printinvoice3.jinja2",{"gkstatus":org.json()["gkstatus"],"org":org.json()["gkdata"],"gkresult":invoicedata.json()["gkresult"],"invid":(int(request.params["invid"]))},request=request)
    elif(request.params["pflag"] == '1'):
        return render_to_response("gkwebapp:templates/printinvoice1.jinja2",{"gkstatus":org.json()["gkstatus"],"org":org.json()["gkdata"],"gkresult":invoicedata.json()["gkresult"],"invid":(int(request.params["invid"]))},request=request)
    elif(request.params["pflag"] == '4'):
        return render_to_response("gkwebapp:templates/printinvoice4.jinja2",{"gkstatus":org.json()["gkstatus"],"org":org.json()["gkdata"],"gkresult":invoicedata.json()["gkresult"],"invid":(int(request.params["invid"]))},request=request)
    elif(request.params["pflag"] == '5'):
        return render_to_response("gkwebapp:templates/printinvoice5.jinja2",{"gkstatus":org.json()["gkstatus"],"org":org.json()["gkdata"],"gkresult":invoicedata.json()["gkresult"],"invid":(int(request.params["invid"]))},request=request)
    else:
        return render_to_response("gkwebapp:templates/printinvoice.jinja2",{"gkstatus":org.json()["gkstatus"],"org":org.json()["gkdata"],"gkresult":invoicedata.json()["gkresult"],"invid":(int(request.params["invid"]))},request=request)


@view_config(route_name="invoice", request_param="action=getattachment", renderer="gkwebapp:templates/viewinvoiceattachment.jinja2")
def getattachment(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/invoice?attach=image&invid=%d"%(int(request.params["invid"])), headers=header)
    return {"attachment":result.json()["gkresult"],"invid":request.params["invid"], "userrole":result.json()["userrole"],"invoiceno":result.json()["invoiceno"]}


@view_config(route_name="invoice", request_param="action=getdelinvprods", renderer="json")
def getdelinvproducts(request):
    header={"gktoken":request.headers["gktoken"]}
    dcid = request.params["dcid"]
    result = requests.get("http://127.0.0.1:6543/delchal?action=getdcinvprods&dcid=%d"%(int(dcid)), headers=header)
    return {"gkstatus": result.json()["gkstatus"], "items": result.json()["gkresult"]}

@view_config(route_name="invoice",request_param="action=showregisterreport",renderer="gkwebapp:templates/registerreport.jinja2")
def showregisterreport(request):
    header={"gktoken":request.headers["gktoken"]}
    if "orderflag" in request.params:
        result = requests.get("http://127.0.0.1:6543/report?type=register&flag=%d&calculatefrom=%s&calculateto=%s&orderflag=%d"%(int(request.params["flag"]), str(request.params["calculatefrom"]), str(request.params["calculateto"]),int(request.params["orderflag"])), headers=header)
        orderflag = "1"
    else:
        result = requests.get("http://127.0.0.1:6543/report?type=register&flag=%d&calculatefrom=%s&calculateto=%s"%(int(request.params["flag"]), str(request.params["calculatefrom"]), str(request.params["calculateto"])), headers=header)
        orderflag = "4"
        
    registerheader = {"flag": request.params["flag"], "calculatefrom": request.params["calculatefrom"], "calculateto": request.params["calculateto"]}
 
    return {"gkstatus":result.json()["gkstatus"], "gkresult": result.json()["gkresult"], "totalrow": result.json()["totalrow"], "taxcolumns":result.json()["taxcolumns"], "registerheader": registerheader,"orderflag":orderflag}

@view_config(route_name="invoice",request_param="action=getinvoiceno",renderer="json")
def getinvoiceno(request):
        header = {"gktoken": request.headers["gktoken"]}
        type=request.params["inoutflag"]
        result = requests.get("http://127.0.0.1:6543/invoice?getinvid&type=%s"%type, headers=header)
        if result.json()["gkstatus"]==0:
            return {"gkstatus":0,"invoiceno":result.json()["invoiceid"]}

'''
This function returns a spreadsheet form of List of Invoices Report.
The spreadsheet in XLSX format is generated by the frontend.
'''
@view_config(route_name="invoice",request_param="action=listofinvspreadsheet", renderer="")
def listofinvspreadsheet(request):
    try:
        header={"gktoken":request.headers["gktoken"]}
        result = requests.get("http://127.0.0.1:6543/invoice?type=list&flag=%s&fromdate=%s&todate=%s"%(request.params["flag"], request.params["fromdate"], request.params["todate"]), headers=header)
        result = result.json()["gkresult"]
        resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
        resultgstvat = resultgstvat.json()["gkresult"]
        fystart = str(request.params["fystart"]);
        fyend = str(request.params["fyend"]);
        orgname = str(request.params["orgname"])
        invflag = int(request.params["flag"])
        invoicewb = openpyxl.Workbook()
        sheet = invoicewb.active
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 16
        sheet.column_dimensions['E'].width = 16
        sheet.column_dimensions['F'].width = 16
        sheet.column_dimensions['G'].width = 16
        sheet.column_dimensions['H'].width = 10
        sheet.column_dimensions['I'].width = 16
        sheet.merge_cells('A1:K2')
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet.merge_cells('A3:K3')
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        if invflag == 0:
            sheet.title = "List of All Invoices"
            sheet['A3'] = "List of All Invoices"
        elif invflag == 1:
            sheet.title = "List of Sales Invoices"
            sheet['A3'] = "List of Sales Invoices"
        elif invflag == 2:
            sheet.title = "List of Purchase Invoices"
            sheet['A3'] = "List of Purchase Invoices"
        sheet.merge_cells('A4:K4')
        sheet['A4'] = 'Period: ' + datetime.strptime(request.params["fromdate"],'%Y-%m-%d').strftime('%d-%m-%Y') + ' to ' + datetime.strptime(request.params["todate"],'%Y-%m-%d').strftime('%d-%m-%Y')
        
        sheet['A4'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A4'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A5'] = 'Sr. No.'
        sheet['B5'] = 'INV No.'
        sheet['C5'] = 'INV Date'
        sheet['D5'] = 'Deli. Note '
        if invflag == 0:
            sheet['E5'] = 'Cust/Supp Name'
        elif invflag == 1:
            sheet['E5'] = 'Customer Name'
        elif invflag == 2:
            sheet['E5'] = 'Supplier Name'
        sheet['F5'] = 'Gross Amt'
        sheet['G5'] = 'Net Amt'
        sheet['H5'] = 'Tax Amt'
        sheet['I5'] = 'Godown'
        titlerow = sheet.row_dimensions[5]
        titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['A5'].alignment = Alignment(horizontal='center')
        sheet['B5'].alignment = Alignment(horizontal='center')
        sheet['C5'].alignment = Alignment(horizontal='center')
        sheet['D5'].alignment = Alignment(horizontal='center')
        sheet['E5'].alignment = Alignment(horizontal='center')
        sheet['F5'].alignment = Alignment(horizontal='right')
        sheet['G5'].alignment = Alignment(horizontal='right')
        sheet['H5'].alignment = Alignment(horizontal='right')
        sheet['I5'].alignment = Alignment(horizontal='center')
        sheet['A5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['B5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['C5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['D5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['E5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['F5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['G5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['H5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['I5'].font = Font(name='Liberation Serif',size=12,bold=True)
        row = 6
        # Looping each dictionaries in list result to store data in cells and apply styles.
        for invoice in result:
            sheet['A'+str(row)] = invoice['srno']
            sheet['A'+str(row)].alignment = Alignment(horizontal='center')
            sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['B'+str(row)] = invoice['invoiceno']
            sheet['B'+str(row)].alignment = Alignment(horizontal='center')
            sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['C'+str(row)] = invoice['invoicedate']
            sheet['C'+str(row)].alignment = Alignment(horizontal='center')
            sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            if invoice['dcno'] != "" and invoice['dcdate'] != "":
                sheet['D'+str(row)] = invoice['dcno'] +','+ invoice['dcdate']
                sheet['D'+str(row)].alignment = Alignment(horizontal='center')
                sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['E'+str(row)] = invoice['custname']
            sheet['E'+str(row)].alignment = Alignment(horizontal='center')
            sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['F'+str(row)] =float("%.2f"%float(invoice['grossamt']))
            sheet['F'+str(row)].number_format="0.00"
            sheet['F'+str(row)].alignment = Alignment(horizontal='right')
            sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['G'+str(row)] = float("%.2f"%float(invoice['netamt']))
            sheet['G'+str(row)].number_format="0.00"
            sheet['G'+str(row)].alignment = Alignment(horizontal='right')
            sheet['G'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['H'+str(row)] = float("%.2f"%float(invoice['taxamt']))
            sheet['H'+str(row)].number_format="0.00"
            sheet['H'+str(row)].alignment = Alignment(horizontal='right')
            sheet['H'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['I'+str(row)] = invoice['godown']
            sheet['I'+str(row)].alignment = Alignment(horizontal='center')
            sheet['I'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            row = row + 1
        output = io.BytesIO()
        invoicewb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','X-Content-Type-Options':'nosniff', 'Set-Cookie':'fileDownload=true ;path=/ [;HttpOnly]'}
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true ;path=/'}
        return Response(contents, headerlist=list(headerList.items()))
    except:
        print("file not found")
        return {"gkstatus":3}
    
'''
This function returns a spreadsheet form of List of Invoices Report.
The spreadsheet in XLSX format is generated by the frontend.
'''
@view_config(route_name="invoice",request_param="action=listofcanceledinvspreadsheet", renderer="")
def listofcanceledinvspreadsheet(request):
    try:
        header={"gktoken":request.headers["gktoken"]}
        result = requests.get("http://127.0.0.1:6543/invoice?type=listdeleted&flag=%s&fromdate=%s&todate=%s"%(request.params["flag"], request.params["fromdate"], request.params["todate"]), headers=header)
        result = result.json()["gkresult"]
        resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
        resultgstvat = resultgstvat.json()["gkresult"]
        fystart = str(request.params["fystart"]);
        fyend = str(request.params["fyend"]);
        orgname = str(request.params["orgname"])
        invflag = int(request.params["flag"])
        invoicewb = openpyxl.Workbook()
        sheet = invoicewb.active
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 16
        sheet.column_dimensions['E'].width = 16
        sheet.column_dimensions['F'].width = 16
        sheet.column_dimensions['G'].width = 16
        sheet.column_dimensions['H'].width = 10
        sheet.column_dimensions['I'].width = 16
        sheet.merge_cells('A1:K2')
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet.merge_cells('A3:K3')
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        if invflag == 0:
            sheet.title = "List of All Cancelled Invoices"
            sheet['A3'] = "List of All Cancelled Invoices"
        elif invflag == 1:
            sheet.title = "List of Cancelled Sales Invoices"
            sheet['A3'] = "List of Cancelled Sales Invoices"
        elif invflag == 2:
            sheet.title = "List of Cancelled Purchase Invoices"
            sheet['A3'] = "List of Cancelled Purchase Invoices"
        sheet.merge_cells('A4:K4')
        sheet['A4'] = 'Period: ' + datetime.strptime(request.params["fromdate"],'%Y-%m-%d').strftime('%d-%m-%Y') + ' to ' + datetime.strptime(request.params["todate"],'%Y-%m-%d').strftime('%d-%m-%Y')
        
        sheet['A4'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A4'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A5'] = 'Sr. No.'
        sheet['B5'] = 'INV No.'
        sheet['C5'] = 'INV Date'
        sheet['D5'] = 'Deli. Note '
        if invflag == 0:
            sheet['E5'] = 'Cust/Supp Name'
        elif invflag == 1:
            sheet['E5'] = 'Customer Name'
        elif invflag == 2:
            sheet['E5'] = 'Supplier Name'
        sheet['F5'] = 'Gross Amt'
        sheet['G5'] = 'Net Amt'
        sheet['H5'] = 'Tax Amt'
        sheet['I5'] = 'Godown'
        titlerow = sheet.row_dimensions[5]
        titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['A5'].alignment = Alignment(horizontal='center')
        sheet['B5'].alignment = Alignment(horizontal='center')
        sheet['C5'].alignment = Alignment(horizontal='center')
        sheet['D5'].alignment = Alignment(horizontal='center')
        sheet['E5'].alignment = Alignment(horizontal='center')
        sheet['F5'].alignment = Alignment(horizontal='right')
        sheet['G5'].alignment = Alignment(horizontal='right')
        sheet['H5'].alignment = Alignment(horizontal='right')
        sheet['I5'].alignment = Alignment(horizontal='center')
        sheet['A5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['B5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['C5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['D5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['E5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['F5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['G5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['H5'].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['I5'].font = Font(name='Liberation Serif',size=12,bold=True)
        row = 6
        # Looping each dictionaries in list result to store data in cells and apply styles.
        for invoice in result:
            sheet['A'+str(row)] = invoice['srno']
            sheet['A'+str(row)].alignment = Alignment(horizontal='center')
            sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['B'+str(row)] = invoice['invoiceno']
            sheet['B'+str(row)].alignment = Alignment(horizontal='center')
            sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['C'+str(row)] = invoice['invoicedate']
            sheet['C'+str(row)].alignment = Alignment(horizontal='center')
            sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            if invoice['dcno'] != "" and invoice['dcdate'] != "":
                sheet['D'+str(row)] = invoice['dcno'] + ',' + invoice['dcdate']
                sheet['D'+str(row)].alignment = Alignment(horizontal='center')
                sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['E'+str(row)] = invoice['custname']
            sheet['E'+str(row)].alignment = Alignment(horizontal='center')
            sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['F'+str(row)] =float("%.2f"%float(invoice['grossamt']))
            sheet['F'+str(row)].number_format="0.00"
            sheet['F'+str(row)].alignment = Alignment(horizontal='right')
            sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['G'+str(row)] = float("%.2f"%float(invoice['netamt']))
            sheet['G'+str(row)].number_format="0.00"
            sheet['G'+str(row)].alignment = Alignment(horizontal='right')
            sheet['G'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['H'+str(row)] = float("%.2f"%float(invoice['taxamt']))
            sheet['H'+str(row)].number_format="0.00"
            sheet['H'+str(row)].alignment = Alignment(horizontal='right')
            sheet['H'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['I'+str(row)] = invoice['godown']
            sheet['I'+str(row)].alignment = Alignment(horizontal='center')
            sheet['I'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            row = row + 1
        output = io.BytesIO()
        invoicewb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','X-Content-Type-Options':'nosniff', 'Set-Cookie':'fileDownload=true ;path=/ [;HttpOnly]'}
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true ;path=/'}
        return Response(contents, headerlist=list(headerList.items()))
    except:
        print("file not found")
        return {"gkstatus":3}
    

'''
This function returns a spreadsheet form of Purchase and Sale Report.
The spreadsheet in XLSX format is generated by the frontend.
'''
@view_config(route_name="invoice",request_param="type=spreadsheet", renderer="")
def registerspreadsheet(request):
    try:
        header={"gktoken":request.headers["gktoken"]}
        result = requests.get("http://127.0.0.1:6543/report?type=register&flag=%d&calculatefrom=%s&calculateto=%s"%(int(request.params["flag"]), str(request.params["calculatefrom"]), str(request.params["calculateto"])), headers=header)
        taxcolumns = result.json()["taxcolumns"]
        totalrow = result.json()["totalrow"]
        result = result.json()["gkresult"]
        fystart = str(request.params["fystart"]);
        fyend = str(request.params["fyend"]);
        orgname = str(request.params["orgname"])
        orgname += " (FY: " + fystart+" to "+fyend +")"
        registerwb = openpyxl.Workbook()
        sheet = registerwb.active
        if request.params["flag"] == "0":
            sheet.title= "Sales Register"
        else:
            sheet.title = "Purchase Register"
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 16
        sheet.column_dimensions['E'].width = 18
        sheet.column_dimensions['F'].width = 18
        sheet.column_dimensions['G'].width = 18
        sheet.column_dimensions['H'].width = 18
        sheet.merge_cells('A1:G2')
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A1'] = orgname
        sheet.merge_cells('A3:G3')
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        if request.params["flag"] == "0":
            sheet['A3'] = 'Sales Register'
        else:
            sheet['A3'] = 'Purchase Register'
        sheet.merge_cells('A4:G4')
        sheet['A4'] = 'Period: ' + request.params["calculatefrom"] + ' to ' + request.params["calculateto"]
        sheet['A4'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A4'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A5'] = 'Sr. No. '
        sheet['B5'] = 'Inv No.'
        sheet['C5'] = 'Inv Date'
        if request.params["flag"] == "0":
            sheet['D5'] = 'Cust. Name '
            sheet['E5'] = 'Cust. TIN'
            sheet['F5'] = 'Cust. GSTIN'
        else:
            sheet['D5'] = 'Suppl. Name '
            sheet['E5'] = 'Suppl. TIN'
            sheet['F5'] = 'Suppl. GSTIN'
        sheet['G5'] = 'Gross Amt.'
        sheet['H5'] = 'TAX Free'
        sheet['A5'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['B5'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['C5'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['D5'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['E5'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['F5'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['G5'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['H5'].font = Font(name='Liberation Serif',size='12',bold=True)
        sheet['A5'].alignment = Alignment(horizontal='center')
        sheet['B5'].alignment = Alignment(horizontal='center')
        sheet['C5'].alignment = Alignment(horizontal='center')
        sheet['D5'].alignment = Alignment(horizontal='center')
        sheet['E5'].alignment = Alignment(horizontal='center')
        sheet['F5'].alignment = Alignment(horizontal='center')
        sheet['G5'].alignment = Alignment(horizontal='right')
        sheet['H5'].alignment = Alignment(horizontal='right')
        #Looping each dictionaries in list taxcolumns to store data in cells and apply styles.
        colvar = 8
        for taxc in taxcolumns:
            sheet.column_dimensions[get_column_letter(colvar+1)].width = 18
            sheet.cell(column = colvar+1, row=5).value = "Net @" + taxc
            sheet.cell(column = colvar+1, row=5).font = Font(name='Liberation Serif',size='12',bold=True)
            sheet.cell(column = colvar+1, row=5).alignment = Alignment(horizontal='right')
            colvar +=1
            sheet.column_dimensions[get_column_letter(colvar+1)].width = 18
            sheet.cell(column = colvar+1, row=5).font = Font(name='Liberation Serif',size='12',bold=True)
            sheet.cell(column = colvar+1, row=5).alignment = Alignment(horizontal='right')
            sheet.cell(column = colvar+1, row=5).value = taxc
            colvar +=1
        #Looping each dictionaries in list result to store data in cells and apply styles.
        rowCount= 6
        for invoice in result:
            sheet['A'+str(rowCount)] = invoice["srno"]
            sheet['A'+str(rowCount)].font = Font(name='Liberation Serif',size='12',bold=False)
            sheet['A'+str(rowCount)].alignment = Alignment(horizontal='center')
            sheet['B'+str(rowCount)] = invoice["invoiceno"]
            sheet['B'+str(rowCount)].font = Font(name='Liberation Serif',size='12',bold=False)
            sheet['B'+str(rowCount)].alignment = Alignment(horizontal='center')
            sheet['C'+str(rowCount)] = invoice["invoicedate"]
            sheet['C'+str(rowCount)].font = Font(name='Liberation Serif',size='12',bold=False)
            sheet['C'+str(rowCount)].alignment = Alignment(horizontal='center')
            sheet['D'+str(rowCount)] = invoice["customername"]
            sheet['D'+str(rowCount)].font = Font(name='Liberation Serif',size='12',bold=False)
            sheet['D'+str(rowCount)].alignment = Alignment(horizontal='center')
            sheet['E'+str(rowCount)] = invoice["customertin"]
            sheet['E'+str(rowCount)].font = Font(name='Liberation Serif',size='12',bold=False)
            sheet['E'+str(rowCount)].alignment = Alignment(horizontal='center')
            if "custgstin" in invoice:
                sheet['F'+str(rowCount)] = invoice["custgstin"]
                sheet['F'+str(rowCount)].font = Font(name='Liberation Serif',size='12',bold=False)
                sheet['F'+str(rowCount)].alignment = Alignment(horizontal='center')
            else:
                sheet['F'+str(rowCount)] = " "
            sheet['G'+str(rowCount)] = float("%.2f"%float(invoice["grossamount"]))
            sheet['G'+str(rowCount)].number_format = '0.00'
            sheet['G'+str(rowCount)].font = Font(name='Liberation Serif',size='12',bold=False)
            sheet['G'+str(rowCount)].alignment = Alignment(horizontal='right')
            sheet['H'+str(rowCount)] = float("%.2f"%float(invoice["taxfree"]))
            sheet['H'+str(rowCount)].number_format = '0.00'
            sheet['H'+str(rowCount)].font = Font(name='Liberation Serif',size='12',bold=False)
            sheet['H'+str(rowCount)].alignment = Alignment(horizontal='right')
            #Looping each dictionaries in list taxcolumns to store data in cells and apply styles.
            colvar = 9
            for taxc in taxcolumns:
                if taxc in invoice["tax"]:
                    sheet.cell(column = colvar, row=rowCount).value = float("%.2f"%float(invoice["tax"][taxc]))
                    sheet.cell(column = colvar, row=rowCount).number_format = "0.00"
                    sheet.cell(column = colvar, row=rowCount).font = Font(name='Liberation Serif',size='12',bold=False)
                    sheet.cell(column = colvar, row=rowCount).alignment = Alignment(horizontal='right')
                else:
                    sheet.cell(column = colvar, row=rowCount).value = float("%.2f"%float("0.00"))
                    sheet.cell(column = colvar, row=rowCount).number_format = "0.00"
                    sheet.cell(column = colvar, row=rowCount).font = Font(name='Liberation Serif',size='12',bold=False)
                    sheet.cell(column = colvar, row=rowCount).alignment = Alignment(horizontal='right')
                colvar += 1
                if taxc in invoice["taxamount"]:
                    sheet.cell(column = colvar, row=rowCount).value = float("%.2f"%float(invoice["taxamount"][taxc]))
                    sheet.cell(column = colvar, row=rowCount).number_format = "0.00"
                    sheet.cell(column = colvar, row=rowCount).font = Font(name='Liberation Serif',size='12',bold=False)
                    sheet.cell(column = colvar, row=rowCount).alignment = Alignment(horizontal='right')
                else:
                    sheet.cell(column = colvar, row=rowCount).value = float("%.2f"%float("0.00"))
                    sheet.cell(column = colvar, row=rowCount).number_format = "0.00"
                    sheet.cell(column = colvar, row=rowCount).font = Font(name='Liberation Serif',size='12',bold=False)
                    sheet.cell(column = colvar, row=rowCount).alignment = Alignment(horizontal='right')
                colvar += 1
            rowCount += 1
        #Merging the columns for "Total"
        sheet.merge_cells(start_row= rowCount, start_column = 1, end_row = rowCount, end_column = 6)
        sheet['A'+str(rowCount)] = "Total"
        sheet['A'+str(rowCount)].alignment = Alignment(horizontal='right')
        sheet['A'+str(rowCount)].font = Font(name='Liberation Serif', size='12', bold=True)
        sheet['G'+str(rowCount)] = float("%.2f"%float(totalrow["grossamount"]))
        sheet['G'+str(rowCount)].number_format = "0.00"                         
        sheet['G'+str(rowCount)].alignment = Alignment(horizontal='right')
        sheet['G'+str(rowCount)].font = Font(name='Liberation Serif', size='12', bold=True)
        sheet['H'+str(rowCount)] = float("%.2f"%float(totalrow["taxfree"]))
        sheet['H'+str(rowCount)].number_format = "0.00"
        sheet['H'+str(rowCount)].alignment = Alignment(horizontal='right')
        sheet['H'+str(rowCount)].font = Font(name='Liberation Serif', size='12', bold=True)
        #Looping each dictionaries in list taxcolumns to store data in cells and apply styles.
        colvar = 9
        for taxc in taxcolumns:
            if taxc in totalrow["tax"]:
                sheet.cell(column = colvar, row=rowCount).value = float("%.2f"%float(totalrow["tax"][taxc]))
                sheet.cell(column = colvar, row=rowCount).number_format = "0.00"
                sheet.cell(column = colvar, row=rowCount).font = Font(name='Liberation Serif',size='12',bold=True)
                sheet.cell(column = colvar, row=rowCount).alignment = Alignment(horizontal='right')
            else:
                sheet.cell(column = colvar, row=rowCount).value = float("%.2f"%float("0.00"))
                sheet.cell(column = colvar, row=rowCount).number_format = "0.00"
                sheet.cell(column = colvar, row=rowCount).font = Font(name='Liberation Serif',size='12',bold=True)
                sheet.cell(column = colvar, row=rowCount).alignment = Alignment(horizontal='right')
            colvar += 1
            if taxc in totalrow["taxamount"]:
                sheet.cell(column = colvar, row=rowCount).value = float("%.2f"%float(totalrow["taxamount"][taxc]))
                sheet.cell(column = colvar, row=rowCount).number_format = "0.00"
                sheet.cell(column = colvar, row=rowCount).font = Font(name='Liberation Serif',size='12',bold=True)
                sheet.cell(column = colvar, row=rowCount).alignment = Alignment(horizontal='right')
            else:
                sheet.cell(column = colvar, row=rowCount).value = float("%.2f"%float("0.00"))
                sheet.cell(column = colvar, row=rowCount).number_format = "0.00"
                sheet.cell(column = colvar, row=rowCount).font = Font(name='Liberation Serif',size='12',bold=True)
                sheet.cell(column = colvar, row=rowCount).alignment = Alignment(horizontal='right')
            colvar += 1
            
        output = io.BytesIO()
        registerwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','X-Content-Type-Options':'nosniff', 'Set-Cookie':'fileDownload=true; path=/ [;HttpOnly]'}
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true; path=/ '}        
        return Response(contents, headerlist=list(headerList.items()))
    except:
        print("File not found")
        {"gkstatus":3}
