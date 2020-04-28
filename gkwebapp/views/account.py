
"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc.,51 Franklin Street, Fifth Floor, 
  Boston, MA 02110, United States


Contributors:
"Krishnakant Mane" <kk@dff.org.in>
"Arun Kelkar" <arunkelkar@dff.org.in>
"Ishan Masdekar " <imasdekar@dff.org.in>
"Navin Karkera" <navin@dff.org.in>
"Bhavesh Bawadhane" <bbhavesh07@gmail.com>
'Prajkta Patkar' <prajkta@riseup.net>
"""

from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from pyramid.response import Response
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
from io import BytesIO
'''
This function returns a spreadsheet form of List of Accounts Report.
The spreadsheet in XLSX format is generated by the backend and sent in base64 encoded format.
It is decoded and returned along with mime information.
'''
@view_config(route_name="printlistofaccounts", renderer="")
def spreadsheetofaccounts(request):
    try:
        header={"gktoken":request.headers["gktoken"]}
        result = requests.get("http://127.0.0.1:6543/accounts", headers=header)
        result = result.json()["gkresult"]
        fystart = str(request.params["fystart"]);
        fyend = str(request.params["fyend"]);
        orgname = str(request.params["orgname"])
        # A workbook is opened.
        accountwb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = accountwb.active
        # Title of the sheet and width of columns are set.
        sheet.title = "List of Accounts"
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 18
        sheet.column_dimensions['C'].width = 14
        sheet.column_dimensions['D'].width = 24
        sheet.column_dimensions['E'].width = 24

        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells('A1:G2')
        # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        # Organisation name and financial year are displayed.
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet.merge_cells('A3:G3')
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A3'] = 'List of Accounts'
        sheet.merge_cells('A3:G3')
        sheet['A4'] = 'Sr. No.'
        sheet['B4'] = 'Account Name'
        sheet['C4'] = 'Group Name'
        sheet['D4'] = 'Sub-group Name'
        sheet['E4'] = 'Default A/C for'
        titlerow = sheet.row_dimensions[4]
        titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
        row=5
        srno = 1
        for accrow in result:
            sheet['A'+str(row)] = srno
            sheet['A'+str(row)].alignment = Alignment(horizontal='left')
            sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['B'+str(row)] = accrow['accountname']
            sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['C'+str(row)] = accrow["groupname"]
            sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['D'+str(row)] = accrow["subgroupname"]
            sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['E'+str(row)] = accrow["defaultflag"]
            sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)            
            row = row + 1
            srno += 1
        output = io.BytesIO()
        accountwb.save(output)
        contents = (output.getvalue())
        output.close()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','X-Content-Type-Options':'nosniff','Accept-Charset':'UTF-8', 'Set-Cookie':'fileDownload=true; path=/[;HttpOnly]'}
        return Response(contents, headerlist=list(headerList.items()))
    except:
        return {"gkstatus":3}

@view_config(route_name="showaccount",renderer="gkwebapp:templates/createeditlistofacc.jinja2")
def showaddeditaccount(request):
    return {"status":True}
    
@view_config(route_name="showaccount",request_param="action=showadd", renderer="gkwebapp:templates/createaccount.jinja2")
def showaccount(request):

    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/groupsubgroups", headers=header)
    grpdata=[]
    for record in result.json()["gkresult"]:
        gdata= {"groupname":str(record["groupname"]),"groupcode":str(record["groupcode"])}
        grpdata.append(gdata)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    states = requests.get("http://127.0.0.1:6543/state", headers=header)
    return {"gkresult":grpdata,"baltbl":result.json()["baltbl"], "vatorgstflag":resultgstvat.json()["gkresult"],"states": states.json()["gkresult"]}

@view_config(route_name="listofaccountsprint", renderer="gkwebapp:templates/printlistofaccounts.jinja2")
def listofaccountprint(request):

    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/accounts", headers=header)

    return {"gkresult":result.json()["gkresult"],"gkstatus":result.json()["gkstatus"]}



@view_config(route_name="showlistofaccounts", renderer="gkwebapp:templates/listofaccounts.jinja2")
def showlistofaccount(request):

    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/accounts", headers=header)

    return {"gkresult":result.json()["gkresult"],"gkstatus":result.json()["gkstatus"]}


@view_config(route_name="accountexists", renderer="json")
def accountexists(request):

    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/accounts?find=exists&accountname=%s"%(request.params["accountname"]), headers=header)

    return {"gkstatus":result.json()["gkstatus"]}


@view_config(route_name="showmultiacc", renderer="gkwebapp:templates/multipleaccounts.jinja2")
def showmultiacc(request):

    return {"gkresult":request.params}

@view_config(route_name="showmultiacc", request_param="taxes",renderer="gkwebapp:templates/multipleaccounts_taxes.jinja2")
def showMultiAccTaxes(request):
    header={"gktoken":request.headers["gktoken"]}
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    states = requests.get("http://127.0.0.1:6543/state", headers=header)
    return {"gkresult":request.params, "vatorgstflag":resultgstvat.json()["gkresult"],"states": states.json()["gkresult"]}

@view_config(route_name="showeditaccount", renderer="gkwebapp:templates/editaccount.jinja2")
def showeditaccount(request):

    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/accounts?editaccount", headers=header)
    accdata=[]
    for record in result.json()["gkresult"]:
        adata= {"accountname":record["accountname"],"accountcode":record["accountcode"], "sysaccount":record["sysaccount"]}
        accdata.append(adata)
    result = requests.get("http://127.0.0.1:6543/groupsubgroups", headers=header)
    grpdata=[]
    for record in result.json()["gkresult"]:
        gdata= {"groupname":str(record["groupname"]),"groupcode":str(record["groupcode"])}
        grpdata.append(gdata)

    return {"gkresult":accdata, "baltbl":result.json()["baltbl"], "groupdata":grpdata}


@view_config(route_name="deleteaccount", renderer="json")
def deleteaccount(request):

    header={"gktoken":request.headers["gktoken"]}
    gkdata={"accountcode":request.params["accountcode"]}
    result = requests.get("http://127.0.0.1:6543/account/%s"%(request.params["accountcode"]), headers=header)
    accountname = result.json()["gkresult"]["accountname"]
    groupcode = result.json()["gkresult"]["groupcode"]
    result = requests.delete("http://127.0.0.1:6543/accounts",data =json.dumps(gkdata), headers=header)
    if result.json()["gkstatus"] == 0:
        groups = requests.get("http://127.0.0.1:6543/groupsubgroups?groupflatlist", headers=header)
        debtgroupcode = groups.json()["gkresult"]["Sundry Debtors"]
        credgroupcode = groups.json()["gkresult"]["Sundry Creditors for Purchase"]
        custid = -1
        if debtgroupcode == groupcode:
            resultcust = requests.get("http://127.0.0.1:6543/customersupplier?qty=custall", headers=header)
            for cust in resultcust.json()["gkresult"]:
                if cust["custname"] == accountname:
                    custid = cust["custid"]
                    break
            if custid != -1:
                gkdata = {"custid":custid}
                resultdelc = requests.delete("http://127.0.0.1:6543/customersupplier", data =json.dumps(gkdata),headers=header)
                gkdata = {"activity":accountname + " customer deleted"}
                resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
        elif credgroupcode == groupcode:
            resultsup = requests.get("http://127.0.0.1:6543/customersupplier?qty=supall", headers=header)
            for cust in resultsup.json()["gkresult"]:
                if cust["custname"] == accountname:
                    custid = cust["custid"]
                    break
            if custid != -1:
                gkdata = {"custid":custid}
                resultdels = requests.delete("http://127.0.0.1:6543/customersupplier", data =json.dumps(gkdata),headers=header)
                gkdata = {"activity":accountname + " supplier deleted"}
                resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
        gkdata = {"activity":accountname + " account deleted"}
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
    return {"gkstatus":result.json()["gkstatus"]}




@view_config(route_name="getaccdetails", renderer="json")
def getaccdetails(request):

    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/account/%s"%(request.params["accountcode"]), headers=header)
    record = result.json()["gkresult"]
    result = requests.get("http://127.0.0.1:6543/groupsubgroup/%s"%(record["groupcode"]), headers=header)
    grprecord = result.json()["gkresult"]
    
    if grprecord["groupcode"]==grprecord["subgroupcode"]:
        grprecord["subgroupname"] = "None"
        
    accdetails={"accountcode":record["accountcode"],"accountname":record["accountname"],"openingbal":record["openingbal"],"groupname":grprecord["groupname"],"subgroupname":grprecord["subgroupname"],"groupcode":grprecord["groupcode"],"subgroupcode":grprecord["subgroupcode"],"defaultflag":record["defaultflag"]}

    return {"gkresult":accdetails}

@view_config(route_name="getaccdetails", request_param="getAccCode", renderer="json")
def getAccCode(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/accounts?type=getAccCode&accountname=%s"%(request.params["accountname"]), headers=header)
    return {"accountcode":result.json()["accountcode"]}
    

@view_config(route_name="getsubgroup", renderer="json")
def getsubgroup(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/groupDetails/%s"%(request.params["groupcode"]), headers=header)
    subgrpdata=[]
    for record in result.json()["gkresult"]:
        sgdata= {"subgroupname":record["subgroupname"],"subgroupcode":record["groupcode"]}

        subgrpdata.append(sgdata)

    return {"gkresult":subgrpdata}

@view_config(route_name="addaccount", renderer="json")
def addaccount(request):
    header={"gktoken":request.headers["gktoken"]}
    newgkdata = {}
    gkdata = {"accountname":request.params["accountname"],"openingbal":request.params["openbal"]}
    if "moredata" in request.params and "custname" in request.params["moredata"]:
        newgkdata["moredata"] = json.loads(request.params["moredata"])
    else:
        newgkdata["moredata"] = {}
    if ("defaultflag" in request.params and int(request.params["defaultflag"]) != 0) :
        gkdata["defaultflag"] = int(request.params["defaultflag"])
    if request.params["subgroupname"]=="New":
        gkdata1={"groupname":request.params["newsubgroup"],"subgroupof":request.params["groupname"]}
        result = requests.post("http://127.0.0.1:6543/groupsubgroups", data =json.dumps(gkdata1),headers=header)
        if result.json()["gkstatus"]==0:
            gkdata["groupcode"] = result.json()["gkresult"]

        else:
            return {"gkstatus":False}

    elif request.params["subgroupname"]=="None":
        grpcode= request.params["groupname"]

        gkdata["groupcode"] = grpcode
    else:
        gkdata["groupcode"] = request.params["subgroupname"]
    newgkdata["gkdata"]=gkdata
    newgkdata["origin"]="createaccount"
    result = requests.post("http://127.0.0.1:6543/accounts", data =json.dumps(newgkdata),headers=header)
    if result.json()["gkstatus"] == 0:
        gkdata = {"activity":request.params["accountname"] + " account created"}
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
    return {"gkstatus":result.json()["gkstatus"]}



@view_config(route_name="multiacc", renderer="json")
def addmultiaccount(request):
    header={"gktoken":request.headers["gktoken"]}
    accdetails = json.loads(request.params["accdetails"])
    
    gkdata = {}
    if accdetails[0]["subgroupname"]=="New":
        gkdata1={"groupname":accdetails[0]["newsubgroup"],"subgroupof":accdetails[0]["groupname"]}
        result = requests.post("http://127.0.0.1:6543/groupsubgroups", data =json.dumps(gkdata1),headers=header)

        if result.json()["gkstatus"]==0:
            gkdata["groupcode"] = result.json()["gkresult"]

        else:
            return {"gkstatus":False}

    elif accdetails[0]["subgroupname"]=="None":
        gkdata["groupcode"] = accdetails[0]["groupname"]

    else:
        gkdata["groupcode"] = accdetails[0]["subgroupname"]

    for acc in accdetails:
        gkdata["accountname"]=acc["accountname"]
        gkdata["openingbal"]=acc["openbal"]
        if int(acc["defaultflag"]) != 0:
            gkdata["defaultflag"] = int(acc["defaultflag"])

        result = requests.post("http://127.0.0.1:6543/accounts", data =json.dumps(gkdata),headers=header)
        if result.json()["gkstatus"] == 0:
            gkdata2 = {"activity":acc["accountname"] + " account created"}
            resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata2),headers=header)
    return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="multiacc", request_param="type=GST", renderer="json")
def addGSTaccounts(request):
    header={"gktoken":request.headers["gktoken"]}
    accdetails = json.loads(request.params["accdetails"])
    gkdata = {}
    gstaccounts = []
    gstaccountslist = requests.get("http://127.0.0.1:6543/organisation?getgstaccounts", headers=header)
    if gstaccountslist.json()["gkstatus"] == 0:
        gstaccounts = gstaccountslist.json()["accounts"]
    if accdetails[0]["subgroupname"]=="New":
        gkdata1={"groupname":accdetails[0]["newsubgroup"],"subgroupof":accdetails[0]["groupname"]}
        addsubgroup = requests.post("http://127.0.0.1:6543/groupsubgroups", data =json.dumps(gkdata1),headers=header)

        if addsubgroup.json()["gkstatus"]==0:
            gkdata["groupcode"] = addsubgroup.json()["gkresult"]

        else:
            return {"gkstatus":3}

    elif accdetails[0]["subgroupname"]=="None":
        gkdata["groupcode"] = accdetails[0]["groupname"]

    else:
        gkdata["groupcode"] = accdetails[0]["subgroupname"]

    addedaccounts = []
    for acc in accdetails:
        if acc["accountname"] not in gstaccounts:
            gkdata["accountname"]=acc["accountname"]
            gkdata["openingbal"]=acc["openbal"]
            gkdata["sysaccount"]=1
            result = requests.post("http://127.0.0.1:6543/accounts", data =json.dumps(gkdata),headers=header)
            if result.json()["gkstatus"] == 0:
                addedaccounts.append(acc["accountname"])
                gkdata2 = {"activity":acc["accountname"] + " account created"}
                resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata2),headers=header)
    return {"accounts":addedaccounts}

#the functionality to edit customer after editing account and other such functionality should be done in core please make a note of this and change it later.
@view_config(route_name="editaccount", renderer="json")
def editaccount(request):
    header={"gktoken":request.headers["gktoken"]}
    newgkdata={}
    gkdata = {"accountname":request.params["accountname"],"openingbal":request.params["openingbal"],"accountcode":request.params["accountcode"]}
    newgkdata["oldcustname"] = request.params["oldcustname"]
    newgkdata["custsupflag"] = int(request.params["custsupflag"])
    if "custname" in request.params["moredata"]:
        newgkdata["moredata"] = json.loads(request.params["moredata"])
    if ("defaultflag" in request.params and int(request.params["defaultflag"]) != 0):
        gkdata["defaultflag"] = int(request.params["defaultflag"])

    '''New Sub-group created, then "New sub-group name" and "group code" under sub-group is created is store in "groupsubgroups" table. 
       return "groupcode" store with 'accountname' in 'accounts' table.
    ''' 
    if request.params["subgrpcode"]=="New":
        gkdatagrp={"groupname":request.params["newgrpname"],"subgroupof":request.params["groupcode"]}
        result = requests.post("http://127.0.0.1:6543/groupsubgroups", data =json.dumps(gkdatagrp),headers=header)

        if result.json()["gkstatus"]==0:
            gkdata["groupcode"] = result.json()["gkresult"]
        else:
            return {"gkstatus":False}
    elif request.params["subgrpname"]=="None":
        grpcode= request.params["groupcode"]

        gkdata["groupcode"] = grpcode
    else:
        gkdata["groupcode"] = request.params["subgrpcode"]
    newgkdata["gkdata"]=gkdata
    result = requests.get("http://127.0.0.1:6543/account/%s"%(request.params["accountcode"]), headers=header)
    accountname = result.json()["gkresult"]["accountname"]
    groupcode = result.json()["gkresult"]["groupcode"]
    
    result = requests.put("http://127.0.0.1:6543/accounts", data =json.dumps(newgkdata),headers=header)
    return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="addaccount", renderer="json", request_param="type=abbreviation")
def abbreviation(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/state?abbreviation&statecode=%d"%(int(request.params["statecode"])), headers=header)
    abbreviation = result.json()["abbreviation"]
    return {"gkstatus":result.json()["gkstatus"], "abbreviation":abbreviation}

@view_config(route_name="addaccount", request_param="type=subgroup", renderer="json")
def subgroup(request):
    header={"gktoken":request.headers["gktoken"]}
    data={"groupname":request.params["newsubgroup"],"subgroupof":request.params["groupname"]}
    result = requests.post("http://127.0.0.1:6543/groupsubgroups", data =json.dumps(data),headers=header)
    if result.json()["gkstatus"] == 0:
        return{"gkstatus":result.json()["gkstatus"], "subgroupcode":result.json()["gkresult"]}
    else:
        return{"gkstatus":result.json()["gkstatus"]}

@view_config(route_name='import',request_param='action =accountimport',renderer="json")
def accImpotrt(request):
    try:
        header={"gktoken":request.headers["gktoken"]}
        xlsxfile = request.POST['xlsxfile'].file
        wbTally = load_workbook(xlsxfile)
        wbTally._active_sheet_index = 0
        accountSheet = wbTally.active
        accountList = tuple(accountSheet.rows)
        gsResult = requests.get("http://127.0.0.1:6543/groupsubgroups?groupflatlist",headers=header)
        groups = gsResult.json()["gkresult"]
        curgrpid = None
        parentgroupid = None
        parentgroup = ""
        openingBl = 0.00
        for accRow in accountList:
            if accRow[0].value == None:
                continue
            if accRow[0].font.b:
                curgrpid = groups[accRow[0].value.strip()]
                parentgroupid = groups[accRow[0].value.strip()]
                parentgroup = accRow[0].value.strip()
                continue
            if accRow[0].font.b == False and accRow[0].font.i == False:
                if accRow[0].value in groups:
                    curgrpid = groups[accRow[0].value.strip()]
                else:
                    newsub = requests.post("http://127.0.0.1:6543/groupsubgroups",data = json.dumps({"groupname":accRow[0].value,"subgroupof":parentgroupid}),headers=header)                   
                    curgrpid = newsub.json()["gkresult"]
            if accRow[0].font.i:
                if len(accRow)>2:
                    if accRow[1].value==None and accRow[2].value==None:
                        newacc = requests.post("http://127.0.0.1:6543/accounts",data = json.dumps({"accountname":accRow[0].value,"groupcode":curgrpid,"openingbal":0.00}),headers=header)
                    # checking if opening Balance is not in Debit column. i.e. column no. 2 (B).
                    #It means value is in credit column 
                    if accRow[1].value==None and accRow[2].value!=None:
                        openingBl = accRow[2].value
                        #Check parent group so that opening balance type (cr/dr) can be determined.
                        if parentgroup == 'Current Assets' or parentgroup == 'Fixed Assets' or parentgroup == 'Investments' or parentgroup == 'Loans(Asset)' or parentgroup == 'Miscellaneous Expenses(Asset)':
                            openingBl = float(-openingBl)
                        newacc = requests.post("http://127.0.0.1:6543/accounts",data = json.dumps({"accountname":accRow[0].value,"groupcode":curgrpid,"openingbal":openingBl}),headers=header)
                        continue
                    #It means value is in debit column 
                    if accRow[2].value==None and accRow[1].value!=None:
                        openingBl = accRow[1].value
                        if parentgroup == 'Corpus' or parentgroup == 'Capital' or parentgroup == 'Current Liabilities' or parentgroup == 'Loans(Liabilities)' or parentgroup == "Reserves":
                            openingBl = float(-openingBl)
                        newacc = requests.post("http://127.0.0.1:6543/accounts",data = json.dumps({"accountname":accRow[0].value,"groupcode":curgrpid,"openingbal":openingBl}),headers=header)
                        continue
                if len(accRow)==2:
                    newsub = requests.post("http://127.0.0.1:6543/accounts",data = json.dumps({"accountname":accRow[0].value,"groupcode":curgrpid,"openingbal":accRow[1].value}),headers=header) 
        return {"gkstatus":0} 
    except:
        print("file not found")
        return {"gkstatus":3}       
