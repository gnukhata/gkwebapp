"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018 Digital Freedom Foundation & Accion Labs PVT LTD 
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@dff.org.in>
"Arun Kelkar" <arunkelkar@dff.org.in>
"Abhijith Balan" <abhijithb21@openmailbox.org>
"Prajkta Patkar" <prajakta@dff.org.in>
"""

from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from pyramid.response import Response
import base64
import os, shutil
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.graphic import GroupShape

@view_config(route_name='import',request_param='action=show',renderer='gkwebapp:templates/tallyimport.jinja2')
def showtallyImport(request):
    return {"gkstatus":0}

@view_config(route_name='import',request_param='action=import',renderer='json')
def tallyImport(request):
    """
    This function will take a spreadsheet containing data from tally or GNUKhata
    Then the code will read the file using parsing library (openpyxl).
    With a number of post calls to REST API, the data is added to GNUKhata.
    The data consists of :
    *new subgroups if they don't exist,
    *new accounts under existing or new subgroups
    * new accounts undr group as per data provided.
    The data   should be in the following format.
    * first sheet must contain the list of accounts
    * Structure should be groups with their optional subgroups
    * if accounts are to be under a group then they should come immediately below the group
    * if there are subgroups under the group they should imediately follow the group
    * groups are in bold
    *accounts are italics
    * subgroups are normal
    * list of groups should be exactly as per GNUKhata (13 at the most ).
    The the code will move to the next sheet which contains all the vouchers.
    This time the function will take list of rows.
    
    Following is the structure of voucher presentation.
    Column 1 = Date , Column 2 = Contains accounts / narration when Drs & Crs do not have value , Column 5 = Voucher Type, Column 6 = Voucher no (to be ignored) ,Column 7 = Dr Amount , Column 8 = Dr amount. In this pattern all the vouchers will be taken & send to the core engine for data entry by making successive POST request.
    
    Take all rows from sheet. Loop through all rows till you get 'Date' word in the first column, Then set the current row index to the next row and exit the loop.
   Start a new loop for list of vouchers and start checking following conditions :
    *If 1st, 2nd & 3rd columns don't have data then continue.
    * If 1st column & 5th column contains data then pick up the date and voucher type respectively also pick up account name and debit amount from 2nd & 6th columns respectively.
    * if 1st and (5th or 6th) column contains data then it is another debit  or credit entry. In this very condition have 2 nested conditions. 
    * If the 5th coulmn has data then data from the 2nd & 5th column gets added to DR dictionary.
    * If the 6th coulmn has data then data from the 2nd & 6th column gets added to CR dictionary.
    * Finally if 2nd column has data but 5th and 6th column does not pickup data from 2nd column as narration. 

   """
    #First we will get list of existing groups and subgroups for this organisation.
    #we will of course lead the workbook from the request.
    try:
        header={"gktoken":request.headers["gktoken"]}
        xlsxfile = request.POST['xlsxfile'].file
        wbTally = load_workbook(xlsxfile)
        wbTally._active_sheet_index = 0
        accountSheet = wbTally.active
        accountList = tuple(accountSheet.rows)
        gsResult = requests.get("http://127.0.0.1:6543/groupsubgroups?groupflatlist",headers=header)
        groups = gsResult.json()["gkresult"]
        curgrpid = None
        parentgroupid = None
        parentgroup = ""
        openingBl = 0.00
        for accRow in accountList:
            if accRow[0].value == None:
                continue
            if accRow[0].font.b:
                curgrpid = groups[accRow[0].value.strip()]
                parentgroupid = groups[accRow[0].value.strip()]
                parentgroup = accRow[0].value.strip()
                continue
            if accRow[0].font.b == False and accRow[0].font.i == False:
                if groups.has_key(accRow[0].value):
                    curgrpid = groups[accRow[0].value.strip()]
                else:
                    newsub = requests.post("http://127.0.0.1:6543/groupsubgroups",data = json.dumps({"groupname":accRow[0].value,"subgroupof":parentgroupid}),headers=header)
                    curgrpid = newsub.json()["gkresult"]
            if accRow[0].font.i:

                if len(accRow)>2:
                    if accRow[1].value==None and accRow[2].value==None:
                        newacc = requests.post("http://127.0.0.1:6543/accounts",data = json.dumps({"accountname":accRow[0].value,"groupcode":curgrpid,"openingbal":0.00}),headers=header)
                        continue
                    # checking if opening Balance is not in Debit column. i.e. column no. 2 (B).
                    #It means value is in credit column 
                    if accRow[1].value==None:
                        openingBl = accRow[2].value
                        #Check parent group so that opening balance type (cr/dr) can be determined.
                        if parentgroup == 'Current Assets' or parentgroup == 'Fixed Assets' or parentgroup == 'Investments' or parentgroup == 'Loans(Asset)' or parentgroup == 'Miscellaneous Expenses(Asset)':
                            openingBl = float(-openingBl)
                        newacc = requests.post("http://127.0.0.1:6543/accounts",data = json.dumps({"accountname":accRow[0].value,"groupcode":curgrpid,"openingbal":openingBl}),headers=header)
                        continue
                    # checking if opening Balance is not in Credit column. i.e. column no. 2 (A).
                    #It means value is in debit column 
                    if accRow[2].value==None:
                        openingBl = accRow[1].value
                        if parentgroup == 'Corpus' or parentgroup == 'Capital' or parentgroup == 'Current Liabilities' or parentgroup == 'Loans(Liabilities)' or parentgroup == "Reserves":
                            openingBl = float(-openingBl)
                        newacc = requests.post("http://127.0.0.1:6543/accounts",data = json.dumps({"accountname":accRow[0].value,"groupcode":curgrpid,"openingbal":openingBl}),headers=header)
                        continue

                if len(accRow)==2:
                    newsub = requests.post("http://127.0.0.1:6543/accounts",data = json.dumps({"accountname":accRow[0].value,"groupcode":curgrpid,"openingbal":accRow[1].value}),headers=header)

        #the dictionary thus returned will have
        #accountname as key and accountcode as value.
        acclist = requests.get("http://127.0.0.1:6543/accounts?acclist",headers=header)
        accounts = acclist.json()["gkresult"]
        Wsheets = wbTally.worksheets
        # When data is imported from GNUKhata exported file
        if Wsheets[1].title == "Vouchers List":
            gVchList =tuple(Wsheets[1].rows) 
            for gVch in gVchList:
                if gVch[1].value == None and gVch[2].value == None:
                    continue
                voucherno = gVch[0].value
                voucherdt = gVch[1].value
                vdates = voucherdt.split('-')
                voucherDt = vdates[2]+'/'+vdates[1]+'/'+vdates[0]
                vouchertype = gVch[2].value
                drs = {}
                crs = {}
                if (gVch[3].value) == "(as per details)":
                    try:
                        Vindex = gVchList.index(gVch) + 1
                        while gVchList[Vindex][3].value != None:
                            drs[accounts[gVchList[Vindex][3].value]] = gVchList[Vindex][4].value
                            Vindex = Vindex + 1 
                    except IndexError:
                        pass       
                else:
                    drs[accounts[gVch[3].value]] = gVch[4].value

                if (gVch[5].value) == "(as per details)":
                    Vindex = gVchList.index(gVch) + 1
                    while gVchList[Vindex][5].value != None: 
                        crs[accounts[gVchList[Vindex][5].value]] = gVchList[Vindex][6].value
                        Vindex = Vindex + 1
                else:
                    crs[accounts[gVch[5].value]] = gVch[6].value
                narration = gVch[7].value
                result = {"voucherdate":voucherDt,"vouchernumber":gVch[0].value,"vouchertype":gVch[2].value,"drs":drs,"crs":crs,"narration":gVch[7].value}
                gNewvch = requests.post("http://127.0.0.1:6543/transaction",data = json.dumps(result),headers=header)
            return {"gkstatus":0}   

        else:
            gVchList =tuple(Wsheets[1].rows)
            currentRowIndex = 0
            for gVch in gVchList:
                drs = {}
                crs = {}
                narrations = ""
                if gVch[0].value == None:
                    continue
                if gVch[0].value == "Date":
                    continue
                if gVch[0].value != None and  gVch[0].value != "Date":
                    voucherdt = str(gVch[0].value)
                    voucherDt = voucherdt[0:4]+'/'+voucherdt[5:7]+'/'+voucherdt[8:10]
                    voucherType = (gVch[4].value).lower().replace(" ","")
                    if gVch[6].value != None:
                            drs[accounts[gVch[1].value ]] =  str(gVch[6].value)
                    if gVch[7].value != None:
                            crs[accounts[gVch[1].value ]] =  str(gVch[7].value)
                    gVch = gVchList[gVchList.index(gVch)+1 ]
                    while gVch[0].value == None:
                        if gVch[6].value != None:
                            drs[accounts[gVch[1].value ]] =  str(gVch[6].value)
                        if gVch[7].value != None:
                            crs[accounts[gVch[1].value ]] =  str(gVch[7].value)
                        if gVch[6].value == None and gVch[7].value == None:
                            narrations  = gVch[1].value
                        try:
                            gVch = gVchList[gVchList.index(gVch)+1]
                        except IndexError:
                            break
                    result = {"voucherdate":voucherDt,"vouchertype":voucherType,"drs":drs,"crs":crs,"narration":narrations}
                    gNewvch = requests.post("http://127.0.0.1:6543/transaction",data = json.dumps(result),headers=header)
            return {"gkstatus":0}
    except:
        print "file not found"
        return {"gkstatus":3}
        
                            

@view_config(route_name="exportledger", renderer="")
def exportLedger(request):
    try:
        header={"gktoken":request.headers["gktoken"]}
        gkwb = Workbook()
        accountList = gkwb.active
        accountList.title = "Account List"
        accountList.column_dimensions["A"].width = 80
        accountList.column_dimensions["B"].width = 30
        gsResult = requests.get("http://127.0.0.1:6543/groupsubgroups",headers=header)
        groupList = gsResult.json()["gkresult"]
        ob = accountList.cell(row=1,column=2,value= "Opening Balance")
        cellCounter = 2
        for group in groupList:
            g = accountList.cell(row=cellCounter,column=1,value=group["groupname"])
            g.font = Font(name=g.font.name,bold=True)
            cellCounter = cellCounter + 1
            accResult = requests.get("http://127.0.0.1:6543/accounts?accbygrp&groupcode=%s"%group["groupcode"],headers=header)
            accList = accResult.json()["gkresult"]
            for acc in accList:
                a = accountList.cell(row=cellCounter,column=1,value=(acc["accountname"]))
                a.font = Font(name=g.font.name,italic=True)
                ob = accountList.cell(row=cellCounter,column=2,value=acc["openingbal"])
                ob2 = accountList.cell(row=cellCounter,column=3,value="")
                cellCounter = cellCounter + 1
            sgResult = requests.get("http://127.0.0.1:6543/groupDetails/%s"%(group["groupcode"]),headers=header)
            subgrpList = sgResult.json()["gkresult"]
            for sg in subgrpList:
                sbg = accountList.cell(row=cellCounter,column=1,value=sg["subgroupname"])
                cellCounter = cellCounter + 1
                accsgResult = requests.get("http://127.0.0.1:6543/accounts?accbygrp&groupcode=%s"%sg["groupcode"],headers=header)
                accListsg = accsgResult.json()["gkresult"]
                for accsg in accListsg:
                    a = accountList.cell(row=cellCounter,column=1,value=(accsg["accountname"]))
                    a.font = Font(name=g.font.name,italic=True)
                    ob = accountList.cell(row=cellCounter,column=2,value=accsg["openingbal"])
                    cellCounter = cellCounter + 1

                    
        Voucher = gkwb.create_sheet()
        Voucher.title = "List of all vouchers"
        yearStart = str(request.params["yearstart"])
        yearEnd = str(request.params["yearend"])
        vchResult = requests.get("http://127.0.0.1:6543/transaction?getdataby=date&from=%s&to=%s"%(yearStart,yearEnd),headers=header)
        vchList = vchResult.json()["gkresult"] 
        rowCounter = crRowCounter = counter = mCounter= 1
        
        #  1= Date, 2 = Particulars 5 = vchtype 6 = vchno 7 = debit 8 = credit.(column no. = Column title).
        # set size of column.
        Voucher.column_dimensions["A"].width = 15
        Voucher.column_dimensions["B"].width = 40
        Voucher.column_dimensions["E"].width = 10
        Voucher.column_dimensions["F"].width = 10
        Voucher.column_dimensions["G"].width = 10
        Voucher.column_dimensions["H"].width = 10
        # Columns title 
        Voucher.cell(row=rowCounter,column=1,value="Date")
        Voucher.cell(row=rowCounter,column=2,value="Particulars")
        Voucher.cell(row=rowCounter,column=5,value="Type")
        Voucher.cell(row=rowCounter,column=6,value= "Voucher No.")
        Voucher.cell(row=rowCounter,column=7,value="Debit")
        Voucher.cell(row=rowCounter,column=8,value= "Credit")
        # increase row counter so that data writing will start from next row
        rowCounter = rowCounter + 1
        # loop through all data(vouchers) 
        for v in vchList:
            # write data in their resepective cell, & which belong to same row.
            Voucher.cell(row=rowCounter,column=1,value=v["voucherdate"])
            Voucher.cell(row=rowCounter,column=5,value=v["vouchertype"])
            Voucher.cell(row=rowCounter,column=6,value=v["voucherno"])
            Crs = v["crs"]
            Drs = v["drs"]
            # loop through drs & crs dictionaries, there can be multiple crs / drs prsent . write account name in 2nd column & debit , credit in column 7 & 8 respectively.
            for k in Drs:
                Voucher.cell(row=rowCounter,column=2,value=k)
                Voucher.cell(row=rowCounter,column=7,value="%.2f"%float(Drs[k]))
                rowCounter = rowCounter + 1
            for k in Crs:
                Voucher.cell(row=rowCounter,column=2,value=k)
                Voucher.cell(row=rowCounter,column=8,value="%.2f"%float(Crs[k]))
                rowCounter = rowCounter + 1
                # write narration after all particular details is wrote , font style & size is set different for better presentation
            if v["narration"] !="":   
                a = Voucher.cell(row=rowCounter,column=2,value=v["narration"])
                a.font = Font(italic=True , size = 10)
                rowCounter = rowCounter + 1      
        gkwb.save(filename = "AllLedger.xlsx")
        AllLedgerfile = open("AllLedger.xlsx","r")
        bf = AllLedgerfile.read()
        AllLedgerfile.close()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(bf),'Content-Disposition': 'attachment; filename=AllLedger.xlsx', 'Set-Cookie':'fileDownload=true; path=/'}
        os.remove("AllLedger.xlsx")
        return Response(bf, headerlist=headerList.items())
    except:
        print "file not found"
        return {"gkstatus":3}
