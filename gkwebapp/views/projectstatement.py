
"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@dff.org.in>
"Arun Kelkar" <arunkelkar@dff.org.in>
"Ishan Masdekar " <imasdekar@dff.org.in>
"Navin Karkera" <navin@dff.org.in>
"Bhavesh Bawadhane" <bbhavesh07@gmail.com>
"""

from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from pyramid.response import Response
import os
import openpyxl
from openpyxl.styles import Font, Alignment
import calendar

@view_config(route_name="printprojectstatementreport", renderer = "")
def printprojectstatementreport(request):
    #try:
        calculateto = request.params["calculateto"]
        financialstart = request.params["fystart"]
        projectcode = int(request.params["projectcode"])
        projectname = request.params["projectname"]
        header={"gktoken":request.headers["gktoken"]}
        result = requests.get("http://127.0.0.1:6543/report?type=projectstatement&calculateto=%s&financialstart=%s&projectcode=%d"%(calculateto,financialstart,projectcode), headers=header)
        result = result.json()["gkresult"]
        fystart = str(request.params["fystart"]);
        fyend = str(request.params["fyend"]);
        fystart = fystart[8:10]+fystart[4:8]+fystart[0:4]
        fyend = fyend[8:10]+fyend[4:8]+fyend[0:4]
        calculateto = str(request.params["calculateto"])
        calculateto = calculateto[8:10]+calculateto[4:8]+calculateto[0:4]
        orgname = str(request.params["orgname"])
        projstmtwb = openpyxl.Workbook()
        sheet = projstmtwb.active
        sheet.title = "Project Statement ("+projectname+")"
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 18
        sheet.column_dimensions['C'].width = 18
        sheet.column_dimensions['D'].width = 16
        sheet.column_dimensions['E'].width = 16
        sheet.merge_cells('A1:E2')
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet.merge_cells('A3:E3')
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A3'] = 'Statement for :'+projectname+ " ("+fystart+" to "+calculateto+")"
        sheet['A4'] = 'Sr. No. '
        sheet['B4'] = 'Account Name'
        sheet['C4'] = 'Group Name'
        sheet['D4'] = 'Total Outgoing'
        sheet['E4'] = 'Total Incoming'
        row = 3
        for transaction in result:
            sheet['A'+str(row)] = transaction["srno"]
            sheet['A'+str(row)].alignment = Alignment(horizontal='left')
            sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)

        projstmtwb.save('report.xlsx')
        xlsxfile = open("report.xlsx","r")
        reportxslx = xlsxfile.read()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(reportxslx),'Content-Disposition': 'attachment; filename=report.xlsx', 'Set-Cookie':'fileDownload=true; path=/'}
        xlsxfile.close()
        os.remove("report.xlsx")
        return Response(reportxslx, headerlist=headerList.items())
    #except:
        print "File not found"
        return {"gkstatus":3}



@view_config(route_name="showviewprojectstatement", renderer="gkwebapp:templates/viewprojectstatement.jinja2")
def showviewprojectstatement(request):
    header={"gktoken":request.headers["gktoken"]}
    projects = requests.get("http://127.0.0.1:6543/projects", headers=header)
    return {"projects":projects.json()["gkresult"]}

@view_config(route_name="showprojectstatementreport")
def showprojectstatementreport(request):
    calculateto = request.params["calculateto"]
    financialstart = request.params["financialstart"]
    projectcode = int(request.params["projectcode"])
    projectname = request.params["projectname"]
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/report?type=projectstatement&calculateto=%s&financialstart=%s&projectcode=%d"%(calculateto,financialstart,projectcode), headers=header)
    return render_to_response("gkwebapp:templates/projectstatementreport.jinja2",{"records":result.json()["gkresult"],"projectcode":projectcode,"projectname":projectname,"from":datetime.strftime(datetime.strptime(str(financialstart),"%Y-%m-%d").date(),'%d-%m-%Y'),"to":datetime.strftime(datetime.strptime(str(calculateto),"%Y-%m-%d").date(),'%d-%m-%Y')},request=request)

@view_config(route_name="printprojectstatement")
def printprojectstatement(request):
    calculateto = request.params["calculateto"]
    financialstart = request.params["financialstart"]
    projectcode = int(request.params["projectcode"])
    projectname = request.params["projectname"]
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/report?type=projectstatement&calculateto=%s&financialstart=%s&projectcode=%d"%(calculateto,financialstart,projectcode), headers=header)
    return render_to_response("gkwebapp:templates/printprojectstatement.jinja2",{"records":result.json()["gkresult"],"projectcode":projectcode,"projectname":projectname,"from":datetime.strftime(datetime.strptime(str(financialstart),"%Y-%m-%d").date(),'%d-%m-%Y'),"to":datetime.strftime(datetime.strptime(str(calculateto),"%Y-%m-%d").date(),'%d-%m-%Y')},request=request)
