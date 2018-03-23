
"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@dff.org.in>
"Arun Kelkar" <arunkelkar@dff.org.in>
"Ishan Masdekar " <imasdekar@dff.org.in>
"Navin Karkera" <navin@dff.org.in>
"Vanita Rajpurohit" <vanita.rajpurohit9819@gmail.com>
"Ravishankar Purne" <ravismail96@gmail.com>
"Reshma Bhatawadekar" <reshma_b@riseup.net>
"""

from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from pyramid.response import Response
import openpyxl
from openpyxl.styles import Font, Alignment
import os


#This function calls User Tab 
@view_config(route_name="showuser", request_param="type=usertab", renderer="gkwebapp:templates/user.jinja2")
def newuser(request):
    header={"gktoken":request.headers["gktoken"]}
    result= requests.get("http://127.0.0.1:6543/users", headers=header)
    return {"gkstatus":result.json()["gkstatus"], "gkresult": result.json()["gkresult"]}    

#This function call New User Tab Inside User Module
@view_config(route_name="showuser",request_param="type=addtab", renderer="gkwebapp:templates/createuser.jinja2")
def createusers(request):
   header={"gktoken":request.headers["gktoken"]}
   result = requests.get("http://127.0.0.1:6543/users", headers=header)
   return {"gkresult":result.json()["gkstatus"],"gkresult":result.json()["gkresult"]}

#This function call Edituser Tab Inside User Module
@view_config(route_name="showuser",request_param="type=edittab", renderer="gkwebapp:templates/edituser.jinja2")
def editusers(request):
   header={"gktoken":request.headers["gktoken"]}
   result = requests.get("http://127.0.0.1:6543/users", headers=header)
   return {"gkresult":result.json()["gkstatus"],"gkresult":result.json()["gkresult"]}

#This function gives Userdetails of selected users 
@view_config(route_name="showuser",request_param="type=getuserdetails", renderer="json")
def getuserdetails(request):
    header={"gktoken":request.headers["gktoken"]}
    userid = int(request.params["userid"])
    result = requests.get("http://127.0.0.1:6543/user?userAllData&userid=%d"%(userid), headers=header)
    if(result.json()["gkstatus"] == 0):
        return {"gkstatus":0, "gkresult":result.json()["gkresult"]}
    return {"gkstatus":result.json()["gkstatus"]}

#This function gives status (i.e valid or not) of current password in edituser 
@view_config(route_name="showuser",request_param="type=userloginstatus", renderer="json")
def UserLoginstatus(request):
    header = {"gktoken":request.headers["gktoken"]}
    gkdata = {"username":request.params["username"], "userpassword":request.params["userpassword"]}
    result = requests.post("http://127.0.0.1:6543/users?userloginstatus", data =json.dumps(gkdata), headers=header)
    return {"gkstatus":result.json()["gkstatus"]}

#This function update user in database 
@view_config(route_name="showuser",request_param="type=updateuser", renderer="json")
def addedituser(request):
    header = {"gktoken":request.headers["gktoken"]}
    gkdata = {"userid":request.params["userid"],"username":request.params["username"],"userquestion":request.params["userquestion"],"useranswer":request.params["useranswer"]}
    if request.params.has_key("userpassword"):
        gkdata["userpassword"]=request.params["userpassword"]
    if request.params.has_key("userrole"):
        gkdata["userrole"]=int(request.params["userrole"])
    if request.params.has_key("godowns"):
        gkdata["golist"]=json.loads(request.params["godowns"])
    result = requests.put("http://127.0.0.1:6543/users?editdata", headers=header, data=json.dumps(gkdata))
    return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="showedituser", renderer="gkwebapp:templates/changepassword.jinja2")
def showedituser(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/users?user=single", headers=header)
    return {"gkstatus": result.json()["gkstatus"], "gkresult": result.json()["gkresult"]}

@view_config(route_name="edituser", renderer="json")
def edituser(request):
    header={"gktoken":request.headers["gktoken"]}
    gkdata={"userid":request.params["userid"], "username":request.params["username"], "userpassword":request.params["userpassword"]}
    result = requests.put("http://127.0.0.1:6543/users", headers=header, data=json.dumps(gkdata))
    return {"gkstatus": result.json()["gkstatus"]}

@view_config(route_name="createuser", renderer="json")
def createuser(request):
    headers={"gktoken":request.headers["gktoken"]}
    goflag = False
    if int(request.params["userrole"]) == 3: 
        gkdata = {"username":request.params["username"],"userpassword":request.params["userpassword"],"userrole":int(request.params["userrole"]),"userquestion":request.params["userquestion"],"useranswer":request.params["useranswer"],"golist":json.loads(request.params["godowns"])}
    else:
        gkdata = {"username":request.params["username"],"userpassword":request.params["userpassword"],"userrole":int(request.params["userrole"]),"userquestion":request.params["userquestion"],"useranswer":request.params["useranswer"]}
    result = requests.post("http://127.0.0.1:6543/users", data =json.dumps(gkdata), headers=headers)
    if result.json()["gkstatus"] == 0:
        if request.params["userrole"] == "-1":
            userrole = "Admin"
        elif request.params["userrole"] == "0":
            userrole = "Manager"
        elif request.params["userrole"] == "1":
            userrole = "Operator"
        elif request.params["userrole"] == "2":
            userrole = "Internal Auditor"
        else:
            userrole = "Godown In Charge"
            godnames = ""
            j = 1;
            godlist = json.loads(request.params["godowns"])
            for i in godlist:
                resultgodown = requests.get("http://127.0.0.1:6543/godown?qty=single&goid=%d"%(int(i)), headers=headers)
                godnames += resultgodown.json()["gkresult"]["goname"] + "(" + resultgodown.json()["gkresult"]["goaddr"] + ")"
                if j != len(godlist):
                    godnames += ", "
                j += 1
        if request.params["userrole"] == "3":
            gkdata = {"activity":gkdata["username"] + "(" + userrole + ")" + " user created for " + godnames + " godown"}
        else:
            gkdata = {"activity":gkdata["username"] + "(" + userrole + ")" + " user created"}
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=headers)
    return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="removeuser", renderer="gkwebapp:templates/removeUser.jinja2")
def removeuser(request):
    headers={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/users", headers=headers)
    return {"gkstatus": result.json()["gkstatus"], "Users": result.json()["gkresult"]}


@view_config(route_name="deleteuser", renderer="json")
def deleteuser(request):
    headers={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/user?userAllData&userid=%d"%(int(request.params["username"])), headers=headers)
    uname = result.json()["gkresult"]["username"]
    #urole in terms of integer
    urole = result.json()["gkresult"]["userrole"]
    #urole in terms of string
    userrole = result.json()["gkresult"]["userroleName"]
    if urole == 3:
        resultgodown = requests.get("http://127.0.0.1:6543/godown?type=byuser&userid=%d"%(int(request.params["username"])), headers=headers)
        resultgodown = resultgodown.json()["gkresult"]
    gkdata={"userid":request.params["username"]}
    result = requests.delete("http://127.0.0.1:6543/users", data=json.dumps(gkdata), headers=headers)
    if result.json()["gkstatus"] == 0:
        if urole == 3:
            godnames = ""
            j = 1
            for godown in resultgodown:
                godnames += godown["goname"] + "(" + godown["goaddr"] + ")"
                if j != len(resultgodown):
                    godnames += ", "
                j += 1
            gkdata = {"activity":uname + "(" + userrole + ")" + " user deleted from " + godnames + " godown"}
        else:
            gkdata = {"activity":uname + "(" + userrole + ")" + " user deleted"}
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=headers)
    return {"gkstatus":result.json()["gkstatus"]}


@view_config(route_name="forgotpassword", renderer="gkwebapp:templates/forgotpassword.jinja2")
def forgotpassword(request):
    code = request.params["orgcode"]
    return {"orgcode":code}

@view_config(route_name="securityquestion", renderer="json")
def securityquestion(request):
    result = requests.get("http://127.0.0.1:6543/forgotpassword?orgcode=%s&username=%s" % ((request.params["orgcode"]),(request.params["username"])))
    if result.json()["gkstatus"] == 0:
        userdata=[]
        userdata.append(result.json()["gkresult"])
        return {"gkresult":userdata, "gkstatus":result.json()["gkstatus"]}
    else:
        return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="securityanswer", renderer="json")
def securityanswer(request):
    result = requests.get("http://127.0.0.1:6543/forgotpassword?type=securityanswer&userid=%s&useranswer=%s" % ((request.params["userid"]),(request.params["useranswer"])))
    return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="newpassword", renderer="json")
def verifypassword(request):
    gkdata = {"userid":request.params["userid"],"userpassword":request.params["userpassword"],"useranswer":request.params["useranswer"]}
    result = requests.put("http://127.0.0.1:6543/forgotpassword", data =json.dumps(gkdata))
    return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="showuser",request_param="type=list", renderer="gkwebapp:templates/listofusers.jinja2")
def listofusers(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/users?type=list", headers=header)
    return {"gkresult":result.json()["gkresult"]}

@view_config(route_name="showuser",request_param="type=printable", renderer="gkwebapp:templates/printlistofusers.jinja2")
def printlistofusers(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/users?type=list", headers=header)
    return {"gkresult":result.json()["gkresult"]}

    '''
    This function returns a spreadsheet form of List of User Report.
    The spreadsheet in XLSX format is generated by the frontend.
    It is decoded and returned along with mime information.
    '''
@view_config(route_name="showuser",request_param="type=spreadsheet", renderer="")
def listofusersspreadsheet(request):
    try:
        header={"gktoken":request.headers["gktoken"]}
        result = requests.get("http://127.0.0.1:6543/users?type=list", headers=header)
        fystart = str(request.params["fystart"]);
        fyend = str(request.params["fyend"]);
        orgname = str(request.params["orgname"]);
        # A workbook is opened.
        userwb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = userwb.active
        # Title of the sheet and width of columns are set.
        sheet.title = "List of Users"
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 16
        sheet.column_dimensions['C'].width = 14
        sheet.column_dimensions['D'].width = 40
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells('A1:F2')
        # Name and Financial Year of organisation is fetched to be displayed on the first row.
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        # Organisation name and financial year are displayed.
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet.merge_cells('A3:F3')
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A3'] = 'List of Users'
        sheet['A4'] = 'Sr. No. '
        sheet['B4'] = 'User Name'
        sheet['C4'] = 'User Role'
        sheet['D4'] = 'Associated Godowns(s)'
        titlerow = sheet.row_dimensions[4]
        titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
        userList = result.json()["gkresult"]
        row=5
        #Looping to store the data in the cells and apply styles.
        srno = 1
        for user in userList:
            sheet['A'+str(row)] = srno
            sheet['A'+str(row)].alignment = Alignment(horizontal='left')
            sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['B'+str(row)] = user["username"]
            sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['C'+str(row)] = user["userrole"]
            sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            gostring = ""
            i = 1
            for godown in user["godowns"]:
                if i == user["noofgodowns"]:
                    gostring += godown
                else:
                    gostring = gostring + godown +","
                i += 1
            sheet['D'+str(row)] = gostring
            sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            row = row + 1
            srno = srno + 1
        userwb.save('report.xlsx')
        xlsxfile = open("report.xlsx","r")
        reportxslx = xlsxfile.read()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(reportxslx),'Content-Disposition': 'attachment; filename=report.xlsx', 'Set-Cookie':'fileDownload=true; path=/'}
        xlsxfile.close()
        os.remove("report.xlsx")
        return Response(reportxslx, headerlist=headerList.items())
    except:
        return {"gkstatus":3}
