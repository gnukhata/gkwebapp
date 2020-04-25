
"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@dff.org.in>
"Arun Kelkar" <arunkelkar@dff.org.in>
"Ishan Masdekar " <imasdekar@dff.org.in>
"Navin Karkera" <navin@dff.org.in>
"Bhavesh Bawadhane" <bbhavesh07@gmail.com>
"Abhijith Balan" <abhijithb21@openmailbox.org>
"Mohd. Talha Pawaty" <mtalha456@gmail.com>
"""

from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from pyramid.response import Response
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook 
from openpyxl.styles import Font, Alignment
import io
from io import BytesIO

@view_config(route_name="godown",renderer="gkwebapp:templates/godown.jinja2")
def godown(request):
    header={"gktoken":request.headers["gktoken"]}
    return {"status":True}

@view_config(route_name="godown",request_param="type=addtab", renderer="gkwebapp:templates/creategodown.jinja2")
def showgodown(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/godown?type=lastfivegodown", headers=header)
    goddata=[]
    states = requests.get("http://127.0.0.1:6543/state", headers=header)
    for record in result.json()["gkresult"]:
        gdata= {"godownname" : str(record["goname"]), "godownaddress": str(record["goaddr"]), "godownstate": str(record["state"])}
        
        goddata.append(gdata)

    return {"gkresult":goddata,"states":states.json()["gkresult"]}

@view_config(route_name="godown",request_param="type=addpopup", renderer="gkwebapp:templates/creategodownpopup.jinja2")
def showgodownpopup(request):
    return {"status":True}

@view_config(route_name="godown",request_param="type=multigodown", renderer="gkwebapp:templates/multiplegodowns.jinja2")
def showmultigodown(request):
    return {"status":True}

@view_config(route_name="godown",request_param="type=add", renderer="json")
def addgodown(request):
    header={"gktoken":request.headers["gktoken"]}
    gkdata = {"goname":request.params["godownname"], "goaddr":request.params["godownaddress"], "state":request.params["godownstate"], "gocontact":request.params["godowncontact"], "contactname":request.params["godowncontactname"]}
    result = requests.post("http://127.0.0.1:6543/godown", data =json.dumps(gkdata),headers=header)
    if result.json()["gkstatus"] == 0:
        gkdata = {"activity":request.params["godownname"] + " godown created"}   
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
    return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="godown",request_param="type=addmulti", renderer="json")
def addmultigodowns(request):
    header={"gktoken":request.headers["gktoken"]}
    goddetails = json.loads(request.params["goddetails"])
    gkdata = {}
    for godown in goddetails:
        gkdata["goname"] = godown["godownname"]
        gkdata["goaddr"] = godown["godownaddress"]
        gkdata["gocontact"] = godown["godowncontact"]
        result = requests.post("http://127.0.0.1:6543/godown", data =json.dumps(gkdata),headers=header)
        if result.json()["gkstatus"] == 0:
            gkdata = {"activity":godown["godownname"] + " godown created"}
            resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
    return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="godown",request_param="type=edittab", renderer="gkwebapp:templates/editgodown.jinja2")
def showeditgodown(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/godown", headers=header)
    goddata=[]
    for record in result.json()["gkresult"]:
        gdata= {"godownname":str(record["goname"]),"godownid":str(record["goid"]),"godownaddress": str(record["goaddr"])}
        goddata.append(gdata)

    return {"gkresult":goddata,"numberofgodown":len(result.json()["gkresult"]),"status":True}

@view_config(route_name="godown",request_param="type=getallgodowns", renderer="json")
def getallgodowns(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/godown", headers=header)
    goddata=[]
    for record in result.json()["gkresult"]:
        gdata= {"godownname":str(record["goname"]),"godownid":str(record["goid"]),"godownaddress": str(record["goaddr"])}
        goddata.append(gdata)
    return {"gkresult":goddata}

@view_config(route_name="godown",request_param="type=numOfProdInGodown", renderer="json")
def getProdInGodown(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/godown?type=goproduct&goid=%d"%int(request.params["goid"]), headers=header)
    return{"gkresult":result.json()["gkresult"],"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="godown",request_param="type=getgoddetails", renderer="json")
def getgoddetails(request):
    header={"gktoken":request.headers["gktoken"]}
    goid = int(request.params["goid"])
    result = requests.get("http://127.0.0.1:6543/godown?qty=single&goid=%d"%(goid), headers=header)
    if(result.json()["gkstatus"] == 0):
        record = result.json()["gkresult"]
        resp = {"godownid": str(record["goid"]), "godownname" : str(record["goname"]), "godownaddress": str(record["goaddr"]), "godownstate": str(record["state"]), "godowncontact": str(record["gocontact"]), "godowncontactname":str(record["contactname"]), "godowndesignation": str(record["designation"])}
        return {"gkstatus":0, "gkresult":resp}
    return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="godown",request_param="type=delete", renderer="json")
def deletegodown(request):
    header={"gktoken":request.headers["gktoken"]}
    goname = request.params["goname"]
    gkdata={"goid":request.params["goid"]}
    result = requests.delete("http://127.0.0.1:6543/godown",data =json.dumps(gkdata), headers=header)
    
    if result.json()["gkstatus"] == 0:
        gkdata = {"activity":goname + " godown deleted"}
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
    return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="godown",request_param="type=edit", renderer="json")
def editgodown(request):
        header={"gktoken":request.headers["gktoken"]}
        gkdata = {"goid":request.params["goid"],"goname":request.params["goname"],"goaddr":request.params["goaddr"], "state":request.params["gostate"], "gocontact": request.params["gocontact"], "contactname":request.params["gocontactname"]}
        result = requests.put("http://127.0.0.1:6543/godown", data =json.dumps(gkdata),headers=header)
        return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="godown",request_param="type=list", renderer="gkwebapp:templates/listofgodowns.jinja2")
def listofgodowns(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/godown", headers=header)
    goddata=[]

    for record in result.json()["gkresult"]:
        gdata= { "godownstatus":str(record["godownstatus"]),"srno":int(record["srno"]), "godownid": str(record["goid"]), "godownname" : str(record["goname"]), "godownaddress": str(record["goaddr"]), "godownstate": str(record["state"]), "godowncontact": str(record["gocontact"]), "godowncontactname":str(record["contactname"]), "godowndesignation": str(record["designation"])}
        goddata.append(gdata)

    return {"gkresult":goddata}

@view_config(route_name="godown",request_param="type=role_list", renderer="gkwebapp:templates/createusergodowntable.jinja2")
def godownsusers(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/godown", headers=header)
    goddata=[]
    for record in result.json()["gkresult"]:
        gdata= {"godownstatus":str(record["godownstatus"]), "srno":int(record["srno"]), "godownid": str(record["goid"]), "godownname" : str(record["goname"]), "godownaddress": str(record["goaddr"]), "godownstate": str(record["state"]), "godowncontact": str(record["gocontact"]), "godowncontactname":str(record["contactname"]), "godowndesignation": str(record["designation"])}
        goddata.append(gdata)
        
    return {"gkresult":goddata}

@view_config(route_name="godown",request_param="type=check", renderer="json")
def checkgodownsusers(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/godown", headers=header)
    goddata=[]
    for record in result.json()["gkresult"]:
        gdata= {"godownstatus":str(record["godownstatus"]), "srno":int(record["srno"]), "godownid": str(record["goid"]), "godownname" : str(record["goname"]), "godownaddress": str(record["goaddr"]), "godownstate": str(record["state"]), "godowncontact": str(record["gocontact"]), "godowncontactname":str(record["contactname"]), "godowndesignation": str(record["designation"])}
        goddata.append(gdata)
    return {"gkresult":goddata}

@view_config(route_name="godown",request_param="type=lists", renderer="gkwebapp:templates/listofgodownspopup.jinja2")
def listsofgodowns(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/godown", headers=header)
    goddata=[]
    for record in result.json()["gkresult"]:
        gdata= {"godownstatus":str(record["godownstatus"]), "srno":int(record["srno"]), "godownid": str(record["goid"]), "godownname" : str(record["goname"]), "godownaddress": str(record["goaddr"]), "godownstate": str(record["state"]), "godowncontact": str(record["gocontact"]), "godowncontactname":str(record["contactname"]), "godowndesignation": str(record["designation"])}
        goddata.append(gdata)
    return {"gkresult":goddata}


@view_config(route_name="godown",request_param="type=printable", renderer="gkwebapp:templates/printlistofgodowns.jinja2")
def printlistofgodowns(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/godown", headers=header)
    goddata=[]
    for record in result.json()["gkresult"]:
        gdata= {"godownstatus":str(record["godownstatus"]), "srno":int(record["srno"]), "godownid": str(record["goid"]), "godownname" : str(record["goname"]), "godownaddress": str(record["goaddr"]), "godownstate": str(record["state"]), "godowncontact": str(record["gocontact"]), "godowncontactname":str(record["contactname"])}
        goddata.append(gdata)
    return {"gkresult":goddata}

'''
This function returns a spreadsheet form of List of Godowns Report.
The spreadsheet in XLSX format is generated by the backend and sent in base64 encoded format.
It is decoded and returned along with mime information.
'''
@view_config(route_name="godown",request_param="type=spreadsheet", renderer="")
def listofgodownssspreadsheet(request):
    
    try:
        header={"gktoken":request.headers["gktoken"]}
        result = requests.get("http://127.0.0.1:6543/godown", headers=header)
        result = result.json()["gkresult"]
        fystart = str(request.params["fystart"])
        fyend = str(request.params["fyend"])
        orgname = str(request.params["orgname"])
        # A workbook is opened.
        godownwb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = godownwb.active
        # Title of the sheet and width of columns are set.
        sheet.title = "List of Godowns"
        
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 18
        sheet.column_dimensions['C'].width = 36
        sheet.column_dimensions['D'].width = 24
        sheet.column_dimensions['E'].width = 16
        sheet.column_dimensions['F'].width = 16
 
            
        # Cells of first two rows are merged to display organisation details properly.
        
        sheet.merge_cells('A1:F2')
        # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        # Organisation name and financial year are displayed.
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A3'] = 'List of Godowns'
        sheet.merge_cells('A3:F3')
        sheet['A4'] = 'Sr. No.'
        sheet['B4'] = 'Godown Name'
        sheet['C4'] = 'Address'
        sheet['D4'] = 'Contact Person'
        sheet['E4'] = 'Contact Number'
        sheet['F4'] = 'Status'
        titlerow = sheet.row_dimensions[4]
        titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
        row=5
        srno = 1
        for godown in result:
            sheet['A'+str(row)] = srno
            sheet['A'+str(row)].alignment = Alignment(horizontal='left')
            sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['B'+str(row)] = godown['goname']
            sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['C'+str(row)] = godown["goaddr"]+" , "+godown["state"]
            sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['D'+str(row)] = godown["contactname"]
            sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['E'+str(row)] = godown["gocontact"]
            sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['F'+str(row)] = godown["godownstatus"]
            sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)

            row = row + 1
            srno += 1
        output = io.BytesIO()
        godownwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','X-Content-Type-Options':'nosniff', 'Set-Cookie':'fileDownload=true; path=/ [;HttpOnly]'}
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true; path=/'}
        return Response(contents, headerlist=list(headerList.items()))
    except:
        return {"gkstatus":3}

@view_config(route_name='import',request_param='action=gdnimport',renderer='json')
def godownImport(request):
     try:
        header={"gktoken":request.headers["gktoken"]}
        xlsxfile = request.POST['xlsxfile'].file
        wb= load_workbook(xlsxfile)
        wb._active_sheet_index = 0
        godownSheet = wb.active
        godownList = tuple(godownSheet.rows)
        errorList=[]
        errorRows=[]
        states = requests.get("http://127.0.0.1:6543/state", headers=header)
        stateList=[]
        duplicateFlag=False
        for state in states.json()["gkresult"]:
            for st in list(state.values()):
                stateList.append(st)
        for godownrow in godownList[1 : ]:
            if godownrow[2].value not in stateList or godownrow[2].value ==None:
                errorList.append(godownrow[2].coordinate)
            if godownrow[1].value ==None:
                errorList.append(godownrow[1].coordinate)
            if godownrow[0].value ==None:
                errorList.append(godownrow[0].coordinate)
        if len(errorList)==0:
            for godownrow in godownList[1 : ]:
                try:
                    godownDict= {"goname":godownrow[0].value,"goaddr":godownrow[1].value, "state":godownrow[2].value, "contactname":godownrow[3].value, "gocontact":godownrow[4].value}
                    result = requests.post("http://127.0.0.1:6543/godown",data = json.dumps(godownDict),headers=header)
                    result=result.json()
                    if result["gkstatus"]==1:
                        duplicateFlag=True
                        errorList.append(godownrow[0].coordinate)
                        continue
                except Exception as e:
                    print (e)                
        if len(errorList)!=0:
            for error in errorList:
                row=openpyxl.utils.cell.coordinate_from_string(error)[1]-1
                errorRow=[row+1,godownList[row][0].value,godownList[row][1].value,godownList[row][2].value,godownList[row][3].value,godownList[row][4].value]
                if errorRow not in errorRows:
                    errorRows.append(errorRow)
            return{"gkstatus":6,"errorRows":errorRows, "errorList":errorList, "duplicateFlag":duplicateFlag} 
        else:
            return {"gkstatus":0}
     except Exception as e:
        print("file not found")
        return {"gkstatus":3}

