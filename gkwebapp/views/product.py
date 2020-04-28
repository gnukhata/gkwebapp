
"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018, 2019, 2020 Digital Freedom Foundation & Accion Labs Pvt. Ltd.
  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@dff.org.in>
"Ishan Masdekar " <imasdekar@dff.org.in>
"Navin Karkera" <navin@dff.org.in>
"Abhijith Balan" <abhijithb21@openmailbox.org>
"Prajkta Patkar" <prajkta.patkar007@gmail.com>
"Mohd. Talha Pawaty" <mtalha456@gmail.com>
"Ravishankar Purne" <ravismail96@gmail.com>
"Reshma Bhatawadekar" <bhatawadekar1reshma@gmail.com>
"Sakshi Agrawal" <agrawalsakshi1850@gmail.com>
"""

from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from pyramid.response import Response
import openpyxl
from openpyxl.styles import Font, Alignment
import io
from io import BytesIO
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.graphic import GroupShape

def is_number(stringwithhfloat):
    try:
        float(stringwithhfloat)
        return True
    except ValueError:
        return False

@view_config(route_name="product",request_param="type=tab", renderer="gkwebapp:templates/producttab.jinja2")
def showproducttab(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/products",headers=header)
    return{"numberofproducts":len(result.json()["gkresult"]),"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="product",request_param="type=stkmodal", renderer="gkwebapp:templates/addstockgstpopup.jinja2")
def addstkgstpopup(request):
    header={"gktoken":request.headers["gktoken"]}
    result2 = requests.get("http://127.0.0.1:6543/godown", headers=header)
    return{"godown":result2.json()["gkresult"], "tax":request.params["tax"]}
    
@view_config(route_name="product",request_param="type=prodlistdata", renderer="json")
def addprodstkpopup(request):
    header={"gktoken":request.headers["gktoken"]}
    productdetails=requests.get("http://127.0.0.1:6543/products?list=allprod&goid=%d"%(int(request.params["goid"])),headers=header)
    return{"products":productdetails.json()["gkresult"]}

@view_config(route_name="product",request_param="type=addtabvat", renderer="gkwebapp:templates/addproductvat.jinja2")
def addProductTabVat(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/categories", headers=header)
    result1 = requests.get("http://127.0.0.1:6543/unitofmeasurement?qty=all", headers=header)
    result2 = requests.get("http://127.0.0.1:6543/godown", headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    states = requests.get("http://127.0.0.1:6543/state", headers=header)
    userrole = requests.get("http://127.0.0.1:6543/user?type=role", headers=header)
    extrabuttons = True
    if "extrabuttons" in request.params:
        extrabuttons = False
    return{"gkresult":{"category":result.json()["gkresult"],"uom":result1.json()["gkresult"]},"godown":result2.json()["gkresult"],"gkstatus":result.json()["gkstatus"],"vatorgstflag":resultgstvat.json()["gkresult"], "states": states.json()["gkresult"], "userrole": userrole.json()["gkresult"], "extrabuttons":extrabuttons}

@view_config(route_name="product",request_param="type=addtab", renderer="gkwebapp:templates/addproduct.jinja2")
def addproducttab(request):
    header={"gktoken":request.headers["gktoken"]}
    uom = requests.get("http://127.0.0.1:6543/unitofmeasurement?qty=all", headers=header).json()["gkresult"]
    result = requests.get("http://127.0.0.1:6543/categories", headers=header)
    result1 = requests.get("http://127.0.0.1:6543/unitofmeasurement?qty=all", headers=header)
    result2 = requests.get("http://127.0.0.1:6543/godown", headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    states = requests.get("http://127.0.0.1:6543/state", headers=header)
    userrole = requests.get("http://127.0.0.1:6543/user?type=role", headers=header)
    extrabuttons = True
    if "extrabuttons" in request.params:
        extrabuttons = False
    return{"uom":uom,"gkresult":{"category":result.json()["gkresult"],"uom":result1.json()["gkresult"]},"godown":result2.json()["gkresult"],"gkstatus":result.json()["gkstatus"],"vatorgstflag":resultgstvat.json()["gkresult"], "states": states.json()["gkresult"], "userrole": userrole.json()["gkresult"], "extrabuttons":extrabuttons}

@view_config(route_name="product",request_param="type=hsnuom", renderer="json")
def gethsnuom(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/products?type=hsnuom&productcode=%d"%(int(request.params["productcode"])), headers=header)
    return{"gkresult":result.json()["gkresult"],"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="product",request_param="type=stockproduct", renderer="json")
def productstock(request):
    header = {"gktoken":request.headers["gktoken"]}
    gkdata = {"goid":request.params["goid"],"productdetails":json.loads(request.params["productdetails"])}
    result = requests.post("http://127.0.0.1:6543/products?type=addstock",data=json.dumps(gkdata), headers=header)
    return {"gkstatus": result.json()["gkstatus"]}

@view_config(route_name="product",request_param="type=specs", renderer="gkwebapp:templates/addproductspecs.jinja2")
def getcatspecs(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/categoryspecs?categorycode=%d"%(int(request.params["categorycode"])), headers=header)
    return{"gkresult":result.json()["gkresult"],"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="product",request_param="type=cattax", renderer="json")
def getcattax(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/tax?pscflag=c&categorycode=%d"%(int(request.params["categorycode"])), headers=header)
    return{"gkresult":result.json()["gkresult"],"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="product",request_param="by=category", renderer="json")
def getprodbycat(request):
    header={"gktoken":request.headers["gktoken"]}
    if request.params["categorycode"]=="":
        result = requests.get("http://127.0.0.1:6543/products?by=category&categorycode=", headers=header)
    else:
        result = requests.get("http://127.0.0.1:6543/products?by=category&categorycode=%d"%(int(request.params["categorycode"])), headers=header)

    if len(result.json()["gkresult"])==0:
        return{"gkresult":0,"gkstatus":result.json()["gkstatus"]}
    else:
        return{"gkresult":result.json()["gkresult"],"gkstatus":result.json()["gkstatus"]}

'''This method is used to get opening stock and godown id's of given productcode(i.e. opening stock of a product in different godowns).'''
@view_config(route_name="product",request_param="by=godown", renderer="json")
def getgodownproduct(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/products?by=godown&productcode=%d"%int(request.params["productcode"]), headers=header)
    return{"gkresult":result.json()["gkresult"],"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="product",request_param="type=prodtax", renderer="json")
def getprodtax(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/tax?pscflag=p&productcode=%d"%(int(request.params["productcode"])), headers=header)
    return{"gkresult":result.json()["gkresult"],"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="product",request_param="type=uom", renderer="json")
def produom(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.post("http://127.0.0.1:6543/unitofmeasurement", data=json.dumps({"unitname":request.params["unitname"]}),headers=header)
    if result.json()["gkstatus"] ==0:
        result = requests.get("http://127.0.0.1:6543/unitofmeasurement?qty=all", headers=header)
        return {"gkstatus": result.json()["gkstatus"], "gkresult": result.json()["gkresult"]}
    else:
        return{"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="product",request_param="type=save", renderer="json")
def saveproduct(request):
    header={"gktoken":request.headers["gktoken"]}
    prdspecs = {}
    proddetails={}
    productdetails={}
    godownflag=False
    taxes =0
    godowns={}
    goid=0
    goopeningstock=0.00
    savedproductcode = ""
    
    for prd in request.params:
        if prd=="type":
            continue
        elif prd=="godownflag":
            if int(request.params[prd]) > 0:
                godownflag=True
            else:
                godownflag=False
        elif prd =="catselect":
            if request.params[prd] !="":
                proddetails["categorycode"] = request.params[prd]
        elif prd == "addproddesc":
            proddetails["productdesc"] = request.params[prd]
        elif prd == "prodmrp":
            proddetails["prodmrp"] = request.params[prd]
        elif prd == "prodsp":
            proddetails["prodsp"] = request.params[prd]
        elif prd == "uom":
            proddetails["uomid"] = request.params[prd]
        elif prd == "openingstock":
            proddetails["openingstock"] = request.params[prd]
        elif prd == "amountdiscount":
            proddetails["amountdiscount"] = request.params[prd]
        elif prd == "percentdiscount":
            proddetails["percentdiscount"] = request.params[prd]
        elif prd == "taxes":
            taxes = json.loads(request.params["taxes"])
        elif prd == "godowns":
            godowns = json.loads(request.params["godowns"])
        elif prd == "newuom":
            continue
        else:
            proddetails["specs"]= json.loads(request.params["specs"])

    if "gscode" in request.params:
        proddetails["gscode"]=request.params["gscode"]
    if "gsflag" in request.params:
        proddetails["gsflag"]=request.params["gsflag"]
    else:
        proddetails["gsflag"] = 7


    productdetails = {"productdetails":proddetails, "godetails":godowns, "godownflag":godownflag}
    if godownflag == True:
            godnames = ""
            j = 1;
            for i in list(godowns.keys()):
                resultgodown = requests.get("http://127.0.0.1:6543/godown?qty=single&goid=%d"%(int(i)), headers=header)
                godnames += resultgodown.json()["gkresult"]["goname"] + "(" + resultgodown.json()["gkresult"]["goaddr"] + ")"
                if j != len(godowns):
                    godnames += ", "
                j += 1

    result = requests.post("http://127.0.0.1:6543/products",data=json.dumps(productdetails),headers=header)

    if result.json()["gkstatus"] == 0:
        savedproductcode = result.json()["gkresult"]
        if godownflag == True:
            gkdata = {"activity":proddetails["productdesc"] + " product created in " + godnames + " godowns"}
        else:
            gkdata = {"activity":proddetails["productdesc"] + " product created"}
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)

    if result.json()["gkstatus"] == 0:
        if len(taxes)>0:
            for tax in taxes:
                taxdata= {"taxname":tax["taxname"],"taxrate":float(tax["taxrate"]),"productcode":result.json()["gkresult"]}
                if tax["state"]!='':
                    taxdata["state"]=tax["state"]

                taxresult = requests.post("http://127.0.0.1:6543/tax",data=json.dumps(taxdata) ,headers=header)
    return {"gkstatus": result.json()["gkstatus"], "gkresult":savedproductcode}


@view_config(route_name="product",request_param="type=edit", renderer="json")
def editproduct(request):
    header={"gktoken":request.headers["gktoken"]}
    proddetails={}
    productdetails={}
    taxes =0
    taxdata = {}
    godownflag=False
    godowns={}
    goid=0
    goopeningstock=0.00
    for prd in request.params:
        if prd=="type":
            continue
        elif prd=="godownflag":
            if int(request.params[prd]) > 0:
                godownflag=True
            else:
                godownflag=False
        elif prd =="productcode":
            proddetails["productcode"] = request.params[prd]
        elif prd =="catselect":
            if request.params[prd] !="":
                proddetails["categorycode"] = request.params[prd]
        elif prd == "editproddesc":
            proddetails["productdesc"] = request.params[prd];
        elif prd == "uom":
            proddetails["uomid"] = request.params[prd]
        elif prd == "prodmrp":
            proddetails["prodmrp"] = request.params[prd]
        elif prd == "prodsp":
            proddetails["prodsp"] = request.params[prd]
        elif prd == "amountdiscount":
            proddetails["amountdiscount"] = request.params[prd]
        elif prd == "percentdiscount":
            proddetails["percentdiscount"] = request.params[prd]
        elif prd == "openingstock":
            proddetails["openingstock"] = request.params[prd]
        elif prd == "taxes":
            taxes = json.loads(request.params["taxes"])
        elif prd == "godowns":
            godowns = json.loads(request.params["godowns"])
        else:
            proddetails["specs"]= json.loads(request.params["specs"])
    if "gscode" in request.params:
        proddetails["gscode"]=request.params["gscode"]
    if "gsflag" in request.params:
        proddetails["gsflag"]=request.params["gsflag"]
    else:
         proddetails["gsflag"] = 7

    productdetails = {"productdetails":proddetails, "godetails":godowns, "godownflag":godownflag}
    result = requests.put("http://127.0.0.1:6543/products", data=json.dumps(productdetails),headers=header)
    if result.json()["gkstatus"] == 0:
        taxid = requests.get("http://127.0.0.1:6543/tax?pscflag=p&productcode=%d"%(int(request.params["productcode"])), headers=header)
        taxids = taxid.json()["gkresult"]
        for tax in taxids:
            taxdata["taxid"] = tax["taxid"]
            taxresult = requests.delete("http://127.0.0.1:6543/tax",data=json.dumps(taxdata) ,headers=header)

        for tax in taxes:
            if len(tax)!=0:
                taxdata= {"taxname":tax["taxname"],"taxrate":float(tax["taxrate"]),"productcode":proddetails["productcode"]}
                if tax["state"]!='':
                    taxdata["state"]=tax["state"]
                taxresult = requests.post("http://127.0.0.1:6543/tax",data=json.dumps(taxdata) ,headers=header)
    return {"gkstatus": result.json()["gkstatus"]}

@view_config(route_name="product",request_param="type=delete", renderer="json")
def deleteproduct(request):
    header={"gktoken":request.headers["gktoken"]}
    resultprod = requests.get("http://127.0.0.1:6543/products?qty=single&productcode=%d"%(int(request.params['productcode'])),headers=header)
    resultgoprod = requests.get("http://127.0.0.1:6543/products?by=godown&productcode=%d"%(int(request.params["productcode"])), headers=header)
    goproddata = resultgoprod.json()["gkresult"]
    result = requests.delete("http://127.0.0.1:6543/products", data=json.dumps({"productcode":request.params["productcode"]}),headers=header)
    if result.json()["gkstatus"] == 0:
        goddetdata = ""
        if len(goproddata) > 0:
            goddetdata = " from godown "
            j = 1
            for goprod in goproddata:
                result = requests.get("http://127.0.0.1:6543/godown?qty=single&goid=%d"%(int(goprod["goid"])), headers=header)
                goddetdata = goddetdata + result.json()["gkresult"]["goname"] + "(" + result.json()["gkresult"]["goaddr"] + ")"
                if j != len(goproddata):
                    goddetdata += ", "
                j += 1
        gkdata = {"activity":resultprod.json()["gkresult"]["productdesc"] + " product deleted" + goddetdata + "."}
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
    return{"gkstatus":result.json()["gkstatus"]}


@view_config(route_name="product",request_param="type=edittab", renderer="gkwebapp:templates/editproduct.jinja2")
def editproducttab(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/products",headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    userrole = requests.get("http://127.0.0.1:6543/user?type=role", headers=header)
    return{"gkresult":result.json()["gkresult"],"gkstatus":result.json()["gkstatus"],"vatorgstflag":resultgstvat.json()["gkresult"], "userrole":userrole.json()["gkresult"]}

@view_config(route_name="product",request_param="type=edittabvat", renderer="gkwebapp:templates/editproductvat.jinja2")
def editProductTabVat(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/products",headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    userrole = requests.get("http://127.0.0.1:6543/user?type=role", headers=header)
    return{"gkresult":result.json()["gkresult"],"gkstatus":result.json()["gkstatus"],"vatorgstflag":resultgstvat.json()["gkresult"], "userrole": userrole.json()["gkresult"]}


@view_config(route_name="product",request_param="type=details", renderer="gkwebapp:templates/editproductspecs.jinja2")
def productdetails(request):
    header={"gktoken":request.headers["gktoken"]}
    prodspecs={}
    result = requests.get("http://127.0.0.1:6543/products?qty=single&productcode=%d"%(int(request.params['productcode'])),headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    userrole = requests.get("http://127.0.0.1:6543/user?type=role", headers=header)
    states = requests.get("http://127.0.0.1:6543/state", headers=header)

    if result.json()["gkresult"]["gsflag"]==7:

        if result.json()["gkresult"]["categorycode"]!=None:
            result1 = requests.get("http://127.0.0.1:6543/categoryspecs?categorycode=%d"%(int(result.json()["gkresult"]["categorycode"])), headers=header)
            prodspecs = result1.json()["gkresult"]
        result2 = requests.get("http://127.0.0.1:6543/unitofmeasurement?qty=all", headers=header)

        result3 = requests.get("http://127.0.0.1:6543/categories", headers=header)
        result4 = requests.get("http://127.0.0.1:6543/godown", headers=header)
        numberofgodowns = int(result.json()["numberofgodowns"])
        return{"proddesc":result.json()["gkresult"],"prodspecs":prodspecs,"uom":result2.json()["gkresult"],"category":result3.json()["gkresult"],"godown":result4.json()["gkresult"],"numberofgodowns":numberofgodowns,"gkstatus":result.json()["gkstatus"],"vatorgstflag":resultgstvat.json()["gkresult"], "states": states.json()["gkresult"], "userrole": userrole.json()["gkresult"]}
    else:
        return{"proddesc":result.json()["gkresult"],"vatorgstflag":resultgstvat.json()["gkresult"], "states": states.json()["gkresult"],"userrole": userrole.json()["gkresult"]}




@view_config(route_name="product",request_param="type=detailsvat", renderer="gkwebapp:templates/editproductspecsvat.jinja2")
def productDetailsVat(request):
    header={"gktoken":request.headers["gktoken"]}
    prodspecs={}
    result = requests.get("http://127.0.0.1:6543/products?qty=single&productcode=%d"%(int(request.params['productcode'])),headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    userrole = requests.get("http://127.0.0.1:6543/user?type=role", headers=header)
    states = requests.get("http://127.0.0.1:6543/state", headers=header)

    if result.json()["gkresult"]["gsflag"]==7:

        if result.json()["gkresult"]["categorycode"]!=None:
            result1 = requests.get("http://127.0.0.1:6543/categoryspecs?categorycode=%d"%(int(result.json()["gkresult"]["categorycode"])), headers=header)
            prodspecs = result1.json()["gkresult"]
        result2 = requests.get("http://127.0.0.1:6543/unitofmeasurement?qty=all", headers=header)

        result3 = requests.get("http://127.0.0.1:6543/categories", headers=header)
        result4 = requests.get("http://127.0.0.1:6543/godown", headers=header)
        numberofgodowns = int(result.json()["numberofgodowns"])
        return{"proddesc":result.json()["gkresult"],"prodspecs":prodspecs,"uom":result2.json()["gkresult"],"category":result3.json()["gkresult"],"godown":result4.json()["gkresult"],"numberofgodowns":numberofgodowns,"gkstatus":result.json()["gkstatus"],"vatorgstflag":resultgstvat.json()["gkresult"], "states": states.json()["gkresult"], "userrole": userrole.json()["gkresult"]}
    else:
        return{"proddesc":result.json()["gkresult"],"vatorgstflag":resultgstvat.json()["gkresult"], "states": states.json()["gkresult"]}


@view_config(route_name="product",request_param="type=list", renderer="gkwebapp:templates/listofstockitems.jinja2")
def listofstockitems(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/products",headers=header)
    return{"gkresult":result.json()["gkresult"],"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="product",request_param="type=printable", renderer="gkwebapp:templates/printlistofstockitems.jinja2")
def printlistofstockitems(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/products",headers=header)

    return{"gkresult":result.json()["gkresult"],"gkstatus":result.json()["gkstatus"]}

'''
This function returns a spreadsheet form of List of Products Report.
The spreadsheet in XLSX format is generated by the backend and sent in base64 encoded format.
It is decoded and returned along with mime information.
'''
@view_config(route_name="product",request_param="type=spreadsheet", renderer="")
def listofstockitemsspreadsheet(request):
    try:
        header={"gktoken":request.headers["gktoken"]}
        result = requests.get("http://127.0.0.1:6543/products", headers=header)
        resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
        resultgstvat = resultgstvat.json()["gkresult"]
        result = result.json()["gkresult"]
        fystart = str(request.params["fystart"]);
        fyend = str(request.params["fyend"]);
        orgname = str(request.params["orgname"])
        # A workbook is opened.
        productwb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = productwb.active
        # Title of the sheet and width of columns are set.
        sheet.title = "List of Products"
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 24
        sheet.column_dimensions['C'].width = 18
        sheet.column_dimensions['D'].width = 24
        sheet.column_dimensions['E'].width = 16
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells('A1:E2')
        # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        # Organisation name and financial year are displayed.
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet.merge_cells('A3:E3')
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A3'] = 'List of Products'
        sheet.merge_cells('A3:E3')
        sheet['A4'] = 'Sr.No.'
        if resultgstvat == "22":
            sheet['B4'] = 'Product'
            sheet['C4'] = 'Category'
            sheet['D4'] = 'UOM'
        else:
            sheet['B4'] = 'Product/service'
            sheet['C4'] = 'Type'
            sheet['D4'] = 'Category'
            sheet['E4'] = 'Uom'
        titlerow = sheet.row_dimensions[4]
        titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
        srno=1
        if resultgstvat == "22":
            row = 5
            for stock in result:
                sheet['A'+str(row)] = srno
                sheet['A'+str(row)].alignment = Alignment(horizontal='left')
                sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['B'+str(row)] = stock['productdesc']
                sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['C'+str(row)] = stock["categoryname"]
                sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['D'+str(row)] = stock["unitname"]
                sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                row +=1
                srno +=1
        else:
            row = 5
            for stock in result:
                sheet['A'+str(row)] = srno
                sheet['A'+str(row)].alignment = Alignment(horizontal='left')
                sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['B'+str(row)] = stock['productdesc']
                sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                if stock["gsflag"] == 7:
                          sheet['C'+str(row)] = 'Product'
                          sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                else:    
                         sheet['C'+str(row)] = 'Service'
                         sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['D'+str(row)] = stock["categoryname"]
                sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['E'+str(row)] = stock["unitname"]
                sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                row +=1
                srno +=1
        output = io.StringIO()
        productwb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','X-Content-Type-Options':'nosniff', 'Set-Cookie':'fileDownload=true; path=/ [;HttpOnly]'}
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true; path=/'}
        return Response(contents, headerlist=list(headerList.items()))
    except:
        return {"gkstatus":3}
    
@view_config(route_name="product",request_param="type=viewstockreport", renderer="gkwebapp:templates/viewstockreport.jinja2")
def viewstockreport(request):

    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/products?invdc=4",headers=header)
    result1 = requests.get("http://127.0.0.1:6543/godown",headers=header)
    result2 = requests.get("http://127.0.0.1:6543/login", headers=header)
    userrole = result2.json()["gkresult"]["userrole"]
    return{"gkresult":result.json()["gkresult"], "godown":result1.json()["gkresult"], "gkstatus":result.json()["gkstatus"], "userrole":userrole}

@view_config(route_name="product",request_param="type=showstockreport")
def showstockreport(request):
    header={"gktoken":request.headers["gktoken"]}
    godownflag = int(request.params["godownflag"])
    goid = int(request.params["goid"])
    goname = request.params["goname"]
    if godownflag==1:
        goaddr = request.params["goaddr"]
    productcode = int(request.params["productcode"])
    calculatefrom = request.params["calculatefrom"]
    calculateto = request.params["calculateto"]
    scalculatefrom = request.params["calculatefrom"]
    scalculateto = request.params["calculateto"]
    productdesc = request.params["productdesc"]
    if int(request.params["backflag"]) > 0:
        scalculatefrom = datetime.strptime(calculatefrom, '%d-%m-%Y').strftime('%Y-%m-%d')
        scalculateto = datetime.strptime(calculateto, '%d-%m-%Y').strftime('%Y-%m-%d')
        stockrefresh = {"productcode":productcode,"calculatefrom":calculatefrom,"calculateto":calculateto,"productdesc":productdesc,"godownflag":godownflag,"goid":goid }
    else:
        stockrefresh = {"productcode":productcode,"calculatefrom":datetime.strptime(calculatefrom, '%Y-%m-%d').strftime('%d-%m-%Y'),"calculateto":datetime.strptime(calculateto, '%Y-%m-%d').strftime('%d-%m-%Y'),"productdesc":productdesc,"godownflag":godownflag,"goid":goid}
    if godownflag>0:
        result = requests.get("http://127.0.0.1:6543/report?type=godownstockreport&goid=%d&productcode=%d&startdate=%s&enddate=%s"%(goid, productcode, scalculatefrom, scalculateto),headers=header)
    else:
        result = requests.get("http://127.0.0.1:6543/report?type=stockreport&productcode=%d&startdate=%s&enddate=%s"%(productcode, scalculatefrom, scalculateto),headers=header)
    if godownflag==1:
        return render_to_response("gkwebapp:templates/showstockreport.jinja2",{"gkresult":result.json()["gkresult"],"stockrefresh":stockrefresh,"godown":goname, "godownadd":goaddr},request=request)
    return render_to_response("gkwebapp:templates/showstockreport.jinja2",{"gkresult":result.json()["gkresult"],"stockrefresh":stockrefresh,"godown":goname},request=request)

@view_config(route_name="product",request_param="type=printablestockreport")
def printablestockreport(request):
    header={"gktoken":request.headers["gktoken"]}
    godownflag = int(request.params["godownflag"])
    goid = int(request.params["goid"])
    goname = request.params["goname"]
    if godownflag==1:
        goaddr = request.params["goaddr"]
    productcode = int(request.params["productcode"])
    scalculatefrom = request.params["calculatefrom"]
    scalculateto = request.params["calculateto"]
    calculatefrom = datetime.strptime(scalculatefrom, '%d-%m-%Y').strftime('%Y-%m-%d')
    calculateto = datetime.strptime(scalculateto, '%d-%m-%Y').strftime('%Y-%m-%d')
    productdesc = request.params["productdesc"]
    stockrefresh = {"productcode":productcode,"calculatefrom":scalculatefrom,"calculateto":scalculateto,"productdesc":productdesc,"godownflag":godownflag,"goid":goid}
    if godownflag > 0:
        result = requests.get("http://127.0.0.1:6543/report?type=godownstockreport&productcode=%d&startdate=%s&enddate=%s&goid=%d&godownflag=%d"%(productcode, calculatefrom, calculateto, goid, godownflag),headers=header)
    else:
        result = requests.get("http://127.0.0.1:6543/report?type=stockreport&productcode=%d&startdate=%s&enddate=%s"%(productcode, calculatefrom, calculateto),headers=header)
    if godownflag==1:
        return render_to_response("gkwebapp:templates/printstockreport.jinja2",{"gkresult":result.json()["gkresult"],"stockrefresh":stockrefresh,"godown":goname, "godownadd":goaddr},request=request)
    return render_to_response("gkwebapp:templates/printstockreport.jinja2",{"gkresult":result.json()["gkresult"],"stockrefresh":stockrefresh,"godown":goname},request=request)

'''
This function returns a spreadsheet form of Product Report.
The spreadsheet in XLSX format is generated by the backend and sent in base64 encoded format.
It is decoded and returned along with mime information.
'''
@view_config(route_name="product",request_param="type=stockreportspreadsheet", renderer="")
def stockreportspreadsheet(request):
    try:
        header={"gktoken":request.headers["gktoken"]}
        godownflag = int(request.params["godownflag"])
        goid = int(request.params["goid"])
        goname = request.params["goname"]
        if godownflag==1:
            goaddr = request.params["goaddr"]
        productcode = int(request.params["productcode"])
        calculatefrom = request.params["calculatefrom"]
        calculateto = request.params["calculateto"]
        scalculatefrom = datetime.strptime(calculatefrom, '%d-%m-%Y').strftime('%Y-%m-%d')
        scalculateto = datetime.strptime(calculateto, '%d-%m-%Y').strftime('%Y-%m-%d')
        productdesc = request.params["productdesc"]
        if godownflag > 0:
            result = requests.get("http://127.0.0.1:6543/report?type=godownstockreport&productcode=%d&startdate=%s&enddate=%s&goid=%d&godownflag=%d"%(productcode, scalculatefrom, scalculateto, goid, godownflag),headers=header)
        else:
            result = requests.get("http://127.0.0.1:6543/report?type=stockreport&productcode=%d&startdate=%s&enddate=%s"%(productcode, scalculatefrom, scalculateto),headers=header)
        result = result.json()["gkresult"]
        fystart = datetime.strptime(request.params["fystart"],'%Y-%m-%d').strftime('%d-%m-%Y')
        fyend = str(request.params["fyend"]);
        orgname = str(request.params["orgname"])
        # A workbook is opened.
        prowb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = prowb.active
        # Title of the sheet and width of columns are set.
        sheet.title = "Product Report"
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 18
        sheet.column_dimensions['C'].width = 22
        sheet.column_dimensions['D'].width = 18
        sheet.column_dimensions['E'].width = 14
        sheet.column_dimensions['F'].width = 14
        sheet.column_dimensions['G'].width = 14
        sheet.column_dimensions['H'].width = 14
        sheet.column_dimensions['I'].width = 14
        sheet.column_dimensions['J'].width = 14
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells('A1:I2')
        # Name and Financial Year of organisation is fetched to be displayed on the first row.
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        # Organisation name and financial year are displayed.
        sheet.merge_cells('A3:J3')
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet.merge_cells('A4:J4')
        sheet['A4'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A4'].alignment = Alignment(horizontal = 'center', vertical='center')
        trn=""
        if godownflag > 0:
            sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
            sheet['A3'] = "Godown Wise Product Report  (Period : "+calculatefrom+" to "+calculateto+")"
            sheet['A4'] = "Name of the Product: "+productdesc
            sheet.merge_cells('A5:J5')
            sheet['A5'].font = Font(name='Liberation Serif',size='14',bold=True)
            sheet['A5'] = "Name of the Godown : "+goname+", Godown Address: "+goaddr
            sheet['A6'] = "Date"
            sheet['B6'] = "Particulars"
            sheet['C6'] = "Document Type"
            sheet['D6'] = "Deli Note No."
            sheet['E6'] = "INV/DR/CR No."
            sheet['F6'] = "RN No."
            sheet['G6'] = "TN No."
            sheet['H6'] = "Inward"
            sheet['I6'] = "Outward"
            sheet['J6'] = "Balance"
            titlerow = sheet.row_dimensions[6]
            titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
            titlerow.alignment = Alignment(horizontal = 'center', vertical='center')
            sheet['H6'].alignment = Alignment(horizontal='right')
            sheet['H6'].font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['I6'].alignment = Alignment(horizontal='right')
            sheet['I6'].font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['J6'].alignment = Alignment(horizontal='right')
            sheet['J6'].font = Font(name='Liberation Serif',size=12,bold=True)
            row = 7

            for stock in result:
                if(stock["trntype"] == "delchal"):
                    trn = "Delivery Note"
                if(stock["trntype"] == "invoice"):
                    trn = "Invoice "
                if(stock["trntype"] == "delchal&invoice"):
                    trn = "Delivery Note & Invoice"
                if(stock["trntype"] == "transfer note"):
                    trn = "Transfer Note "
                if(stock["trntype"] == "Rejection Note"):
                    trn = "Rejection Note"
                if(stock["trntype"] == "Debit Note"):
                    trn = "Debit Note"
                if(stock["trntype"] == "Credit Note"):
                    trn = "Credit Note"

                if stock["particulars"]=="opening stock" and stock["dcno"]=="" and stock["invno"]=="" and stock["date"]=="":
                     sheet['A'+str(row)] =""
                     sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['A'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['B'+str(row)] = stock["particulars"].title()
                     sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['C'+str(row)] =""
                     sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['C'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['D'+str(row)] =""
                     sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['D'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['E'+str(row)] =""
                     sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['E'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['F'+str(row)] =""
                     sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['F'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['G'+str(row)] =""
                     sheet['G'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['G'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['H'+str(row)]=float("%.2f"%float(stock["inward"]))
                     sheet['H'+str(row)].number_format = '0.00'
                     sheet['H'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['H'+str(row)].alignment = Alignment(horizontal='right')
                     sheet['I'+str(row)] =""
                     sheet['I'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['I'+str(row)].alignment = Alignment(horizontal='right')
                     sheet['J'+str(row)] =""
                     sheet['J'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['J'+str(row)].alignment = Alignment(horizontal='right')
                if stock["particulars"]!="Total" and (stock["dcno"]!="" or stock["invno"]!="" or stock["tnno"]!="" or stock["rnno"] != "") and stock["date"]!="":

                     sheet['A'+str(row)] = stock["date"]
                     sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['A'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['B'+str(row)] = stock["particulars"]
                     sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['C'+str(row)] = trn
                     sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['C'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['D'+str(row)] = stock["dcno"]
                     sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['D'+str(row)].alignment = Alignment(horizontal='center')
                     if stock["invno"] != "":

                         sheet['E'+str(row)] = stock["invno"]
                     elif "drcrno" in stock and stock["drcrno"] != "":

                         sheet['E'+str(row)] = stock["drcrno"]
                     else:

                         sheet['E'+str(row)] = ""
                     sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['E'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['F'+str(row)] = stock["rnno"]
                     sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['F'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['G'+str(row)] = stock["tnno"]
                     sheet['G'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['G'+str(row)].alignment = Alignment(horizontal='center')
                     if stock["inwardqty"]!= "":

                         sheet['H'+str(row)]=float("%.2f"%float(stock["inwardqty"]))
                         sheet['H'+str(row)].number_format = '0.00'
                         sheet['H'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                         sheet['H'+str(row)].alignment = Alignment(horizontal='right')
                     if stock["outwardqty"]!= "":

                         sheet['I'+str(row)]=float("%.2f"%float(stock["outwardqty"]))
                         sheet['I'+str(row)].number_format = '0.00'
                         sheet['I'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                         sheet['I'+str(row)].alignment = Alignment(horizontal='right')
                     sheet['J'+str(row)]=float("%.2f"%float(stock["balance"]))
                     sheet['J'+str(row)].number_format = '0.00'
                     sheet['J'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['J'+str(row)].alignment = Alignment(horizontal='right')
                if stock["particulars"]=="Total" and stock["dcno"]=="" and stock["invno"]=="" and stock["date"]=="":

                     sheet['A'+str(row)] =""
                     sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['A'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['B'+str(row)] = stock["particulars"].title()
                     sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['C'+str(row)] =""
                     sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['C'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['D'+str(row)] =""
                     sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['D'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['E'+str(row)] =""
                     sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['E'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['F'+str(row)] =""
                     sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['F'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['G'+str(row)] =""
                     sheet['G'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['G'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['H'+str(row)]=float("%.2f"%float(stock["totalinwardqty"]))
                     sheet['H'+str(row)].number_format = '0.00'
                     sheet['H'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['H'+str(row)].alignment = Alignment(horizontal='right')
                     sheet['I'+str(row)]=float("%.2f"%float(stock["totaloutwardqty"]))
                     sheet['I'+str(row)].number_format = '0.00'
                     sheet['I'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['I'+str(row)].alignment = Alignment(horizontal='right')
                     sheet['J'+str(row)] =""
                     sheet['J'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['J'+str(row)].alignment = Alignment(horizontal='right')
                row  += 1
        else:
            sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
            sheet['A3'] = "Product Report (Period : "+calculatefrom+" to "+calculateto+")"
            sheet['A4'] = "Name of the Product: "+productdesc
            sheet['A5'] = "Date"
            sheet['B5'] = "Particulars"
            sheet['C5'] = "Document Type"
            sheet['D5'] = "Deli Note No."
            sheet['E5'] = "INV/DR/CR No."
            sheet['F5'] = "RN No."
            sheet['G5'] = "Inward"
            sheet['H5'] = "Outward"
            sheet['I5'] = "Balance"
            titlerow = sheet.row_dimensions[5]
            titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
            titlerow.alignment = Alignment(horizontal = 'center', vertical='center')
            sheet['G5'].alignment = Alignment(horizontal='right')
            sheet['G5'].font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['H5'].alignment = Alignment(horizontal='right')
            sheet['H5'].font = Font(name='Liberation Serif',size=12,bold=True)
            sheet['I5'].alignment = Alignment(horizontal='right')
            sheet['I5'].font = Font(name='Liberation Serif',size=12,bold=True)
            row = 6
            
            for stock in result:
                if(stock["trntype"] == "delchal"):
                    trn = "Delivery Note"
                if(stock["trntype"] == "invoice"):
                    trn = "Invoice "
                if(stock["trntype"] == "delchal&invoice"):
                    trn = "Delivery Note & Invoice"
                if(stock["trntype"] == "transfer note"):
                    trn = "Transfer Note "
                if(stock["trntype"] == "Rejection Note"):
                    trn = "Rejection Note"
                if(stock["trntype"] == "Debit Note"):
                    trn = "Debit Note"
                if(stock["trntype"] == "Credit Note"):
                    trn = "Credit Note"

                if stock["particulars"]=="opening stock" and stock["dcno"]=="" and stock["invno"]=="" and stock["date"]=="":
                     sheet['A'+str(row)] =""
                     sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['A'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['B'+str(row)] = stock["particulars"].title()
                     sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['C'+str(row)] =""
                     sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['C'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['D'+str(row)] =""
                     sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['D'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['E'+str(row)] =""
                     sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['E'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['F'+str(row)] =""
                     sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['F'+str(row)].alignment = Alignment(horizontal='right')
                     sheet['G'+str(row)]=float("%.2f"%float(stock["inward"]))
                     sheet['G'+str(row)].number_format = '0.00'
                     sheet['G'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['G'+str(row)].alignment = Alignment(horizontal='right')
                     sheet['H'+str(row)] =""
                     sheet['H'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['H'+str(row)].alignment = Alignment(horizontal='right')
                     sheet['I'+str(row)] =""
                     sheet['I'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['I'+str(row)].alignment = Alignment(horizontal='right')
                if stock["particulars"]!="Total" and (stock["dcno"]!="" or stock["invno"]!="" or stock["rnid"] != "" or stock["drcrno"] != "") and stock["date"]!="":
                     sheet['A'+str(row)] = stock["date"]
                     sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['A'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['B'+str(row)] = stock["particulars"]
                     sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)

                     sheet['C'+str(row)] = trn
                     sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)

                     sheet['C'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['D'+str(row)] = stock["dcno"]
                     sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['D'+str(row)].alignment = Alignment(horizontal='center')
                     if stock["invno"] != "":
                         sheet['E'+str(row)] = stock["invno"]
                     elif "drcrno" in stock and stock["drcrno"] != "":
                         sheet['E'+str(row)] = stock["drcrno"]
                     else:
                         sheet['E'+str(row)] = ""
                     sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['E'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['F'+str(row)] = stock["rnno"]
                     sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['F'+str(row)].alignment = Alignment(horizontal='right')
                     if stock["inwardqty"]!= "":
                         sheet['G'+str(row)]=float("%.2f"%float(stock["inwardqty"]))
                         sheet['G'+str(row)].number_format = '0.00'
                         sheet['G'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                         sheet['G'+str(row)].alignment = Alignment(horizontal='right')
                     if stock["outwardqty"]!= "":
                         sheet['H'+str(row)]=float("%.2f"%float(stock["outwardqty"]))
                         sheet['H'+str(row)].number_format = '0.00'
                         sheet['H'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                         sheet['H'+str(row)].alignment = Alignment(horizontal='right')
                     sheet['I'+str(row)]=float("%.2f"%float(stock["balance"]))
                     sheet['I'+str(row)].number_format = '0.00'
                     sheet['I'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['I'+str(row)].alignment = Alignment(horizontal='right')
                if stock["particulars"]=="Total" and stock["dcno"]=="" and stock["invno"]=="" and stock["date"]=="":
                     sheet['A'+str(row)] =""
                     sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['A'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['B'+str(row)] = stock["particulars"].title()
                     sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['C'+str(row)] =""
                     sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['C'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['D'+str(row)] =""
                     sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['D'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['E'+str(row)] =""
                     sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['E'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['F'+str(row)] =""
                     sheet['F'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['F'+str(row)].alignment = Alignment(horizontal='center')
                     sheet['G'+str(row)]=float("%.2f"%float(stock["totalinwardqty"]))
                     sheet['G'+str(row)].number_format = '0.00'
                     sheet['G'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['G'+str(row)].alignment = Alignment(horizontal='right')
                     sheet['H'+str(row)]=float("%.2f"%float(stock["totaloutwardqty"]))
                     sheet['H'+str(row)].number_format = '0.00'
                     sheet['H'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['H'+str(row)].alignment = Alignment(horizontal='right')
                     sheet['I'+str(row)] =""
                     sheet['I'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
                     sheet['I'+str(row)].alignment = Alignment(horizontal='right')
                row += 1
        output = io.BytesIO()
        prowb.save(output)
        contents = output.getvalue()
        output.close()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx', 'X-Content-Type-Options':'nosniff','Set-Cookie':'fileDownload=true; path=/ [;HttpOnly]'}
        # headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(contents),'Content-Disposition': 'attachment; filename=report.xlsx','Set-Cookie':'fileDownload=true; path=/'}

        return Response(contents, headerlist=list(headerList.items()))
    except:
        return {"gkstatus":3}


@view_config(route_name="product",request_param="type=viewstockonhandreport", renderer="gkwebapp:templates/viewstockonhandreport.jinja2")
def viewStockOnHandReport(request):

    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/products?invdc=4",headers=header)
    result1 = requests.get("http://127.0.0.1:6543/godown",headers=header)
    result2 = requests.get("http://127.0.0.1:6543/login", headers=header)
    userrole = result2.json()["gkresult"]["userrole"]
    return{"gkresult":result.json()["gkresult"], "godown":result1.json()["gkresult"], "gkstatus":result.json()["gkstatus"], "userrole":userrole}

@view_config(route_name="product",request_param="type=viewcategorywisestockonhandreport", renderer="gkwebapp:templates/viewcategorywisestockonhandreport.jinja2")
def viewCategorywiseStockOnHandReport(request):

    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/categories",headers=header)
    result1 = requests.get("http://127.0.0.1:6543/godown",headers=header)
    result2 = requests.get("http://127.0.0.1:6543/login", headers=header)
    userrole = result2.json()["gkresult"]["userrole"]
    return{"gkresult":result.json()["gkresult"], "godown":result1.json()["gkresult"], "gkstatus":result.json()["gkstatus"], "userrole":userrole}

@view_config(route_name="product",request_param="type=showstockonhandreport")
def showstockonhandreport(request):
    header={"gktoken":request.headers["gktoken"]}
    godownflag = int(request.params["godownflag"])
    goid = int(request.params["goid"])
    goname = request.params["goname"]

    if godownflag==1:
        goaddr = request.params["goaddr"]
    productcode = int(request.params["productcode"])
    calculateto = request.params["calculateto"]
    scalculateto = request.params["calculateto"]


    productdesc = request.params["productdesc"]
    if int(request.params["backflag"]) == 1 :
        scalculateto = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%Y-%m-%d')
        date = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%d-%m-%Y')
        stockrefresh = {"productcode":productcode,"calculateto":calculateto,"productdesc":"All Products","godownflag":godownflag,"goid":goid,"date":date }
        result = requests.get("http://127.0.0.1:6543/report?stockonhandreport&productcode=all&enddate=%s"%(scalculateto),headers=header)


    if int(request.params["backflag"]) == 0:
        scalculateto = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%Y-%m-%d')
        date = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%d-%m-%Y')
        stockrefresh = {"productcode":productcode,"calculateto":datetime.strptime(calculateto, '%Y-%m-%d').strftime('%Y-%m-%d'),"productdesc":productdesc,"godownflag":godownflag,"goid":goid,"date":date}
        result = requests.get("http://127.0.0.1:6543/report?stockonhandreport&productcode=%d&enddate=%s"%(productcode,scalculateto),headers=header)

    if godownflag == 1 and int(request.params["backflag"]) == 3 :
        scalculateto = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%Y-%m-%d')
        date = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%d-%m-%Y')
        stockrefresh = {"productcode":productcode,"calculateto":calculateto,"productdesc":productdesc,"godownflag":godownflag,"goid":goid,"date":date}
        result = requests.get("http://127.0.0.1:6543/report?godownwisestockonhand&type=pg&goid=%d&productcode=%d&enddate=%s"%(goid, productcode, scalculateto),headers=header)

    if godownflag == 1 and int(request.params["backflag"]) == 2 and goid == 0:
        scalculateto = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%Y-%m-%d')
        date = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%d-%m-%Y')
        stockrefresh = {"productcode":productcode,"calculateto":calculateto,"productdesc":productdesc,"godownflag":godownflag,"goid":goid,"date":date}
        result = requests.get("http://127.0.0.1:6543/report?godownwisestockonhand&type=pag&productcode=%d&enddate=%s"%(productcode, scalculateto),headers=header)

    if godownflag==1:
        return render_to_response("gkwebapp:templates/showstockonhandreport.jinja2",{"gkresult":result.json()["gkresult"],"stockrefresh":stockrefresh,"godown":goname, "goaddr":goaddr},request=request)

    return render_to_response("gkwebapp:templates/showstockonhandreport.jinja2",{"gkresult":result.json()["gkresult"],"stockrefresh":stockrefresh,"godown":goname},request=request)

@view_config(route_name="product",request_param="type=showcategorywisestockonhandreport")
def showcategorywisestockonhandreport(request):
    header={"gktoken":request.headers["gktoken"]}
    godownflag = int(request.params["godownflag"])
    if godownflag == 0:
        goid = "-1"
    else:
        goid = request.params["goid"]
    categorycode = int(request.params["categorycode"])
    calculateto = request.params["calculateto"]
    categoryname = request.params["categoryname"]
    subcategoryname = request.params["subcategoryname"]
    specname = request.params["specname"]
    stockrefresh = {"categorycode":categorycode,"calculateto":datetime.strptime(calculateto, '%Y-%m-%d').strftime('%Y-%m-%d'),"categoryname":categoryname, "subcategoryname": subcategoryname, "specname":specname, "godownflag":godownflag,"goid":goid,"date":datetime.strptime(calculateto, '%Y-%m-%d').strftime('%d-%m-%Y')}
    if int(request.params["backflag"]) == 1 :
        result = requests.get("http://127.0.0.1:6543/report?stockonhandreport&productcode=all&enddate=%s"%(calculateto),headers=header)
        if godownflag==1:
            if request.params["goid"] == "all":
                return render_to_response("gkwebapp:templates/showcategorywisestockonhandreport.jinja2",{"gkresult":result.json()["gkresult"],"stockrefresh":stockrefresh},request=request)
            else:
                return render_to_response("gkwebapp:templates/showcategorywisestockonhandreport.jinja2",{"gkresult":result.json()["gkresult"],"stockrefresh":stockrefresh,"godown":request.params["goname"], "goaddr":request.params["goaddr"]},request=request)
        else:
            return render_to_response("gkwebapp:templates/showcategorywisestockonhandreport.jinja2",{"gkresult":result.json()["gkresult"],"stockrefresh":stockrefresh},request=request)

    if int(request.params["backflag"]) == 0:
        result = requests.get("http://127.0.0.1:6543/report?type=categorywisestockonhand&categorycode=%d&subcategorycode=%s&speccode=%s&goid=%s&enddate=%s"%(categorycode, request.params["subcategorycode"], request.params["speccode"], goid, calculateto),headers=header)
        if godownflag==1:
            if request.params["goid"] == "all":
                return render_to_response("gkwebapp:templates/showcategorywisestockonhandreport.jinja2",{"gkresult":result.json()["gkresult"],"stockrefresh":stockrefresh},request=request)
            else:
                return render_to_response("gkwebapp:templates/showcategorywisestockonhandreport.jinja2",{"gkresult":result.json()["gkresult"],"stockrefresh":stockrefresh,"godown":request.params["goname"], "goaddr":request.params["goaddr"]},request=request)
        else:
            return render_to_response("gkwebapp:templates/showcategorywisestockonhandreport.jinja2",{"gkresult":result.json()["gkresult"],"stockrefresh":stockrefresh},request=request)

    '''if godownflag == 1 and int(request.params["backflag"]) == 3 :
        scalculateto = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%Y-%m-%d')
        date = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%d-%m-%Y')
        result = requests.get("http://127.0.0.1:6543/report?godownwisestockonhand&type=pg&goid=%d&productcode=%d&enddate=%s"%(goid, productcode, scalculateto),headers=header)

    if godownflag == 1 and int(request.params["backflag"]) == 2 and goid == 0:

        scalculateto = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%Y-%m-%d')
        date = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%d-%m-%Y')
        result = requests.get("http://127.0.0.1:6543/report?godownwisestockonhand&type=pag&productcode=%d&enddate=%s"%(productcode, scalculateto),headers=header)
    '''



@view_config(route_name="product",request_param="type=printablestockonhandreport")
def printablestockonhandreport(request):

    header={"gktoken":request.headers["gktoken"]}
    godownflag = int(request.params["godownflag"])
    goid = int(request.params["goid"])
    goname = request.params["goname"]
    productcode = int(request.params["productcode"])
    calculateto = request.params["calculateto"]
    productdesc = request.params["productdesc"]
    scalculateto = request.params["calculateto"]
    if int(request.params["backflag"]) == 1 :

        scalculateto = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%Y-%m-%d')
        stockrefresh = {"productcode":productcode,"calculateto":calculateto,"productdesc":"all","godownflag":godownflag,"goid":goid }

        result = requests.get("http://127.0.0.1:6543/report?stockonhandreport&productcode=all&enddate=%s"%(scalculateto),headers=header)

    if int(request.params["backflag"]) == 0 :
        scalculateto = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%Y-%m-%d')
        stockrefresh = {"productcode":productcode,"calculateto":calculateto,"productdesc":productdesc,"godownflag":godownflag,"goid":goid}
        result = requests.get("http://127.0.0.1:6543/report?stockonhandreport&productcode=%d&enddate=%s"%(productcode,scalculateto),headers=header)

    if godownflag == 1 and int(request.params["backflag"]) == 3 :

        scalculateto = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%Y-%m-%d')
        stockrefresh = {"productcode":productcode,"calculateto":calculateto,"productdesc":productdesc,"godownflag":godownflag,"goid":goid}
        result = requests.get("http://127.0.0.1:6543/report?godownwisestockonhand&type=pg&goid=%d&productcode=%d&enddate=%s"%(goid, productcode, scalculateto),headers=header)

    if godownflag == 1 and int(request.params["backflag"]) == 2 and goid == 0:

        scalculateto = datetime.strptime(calculateto, '%Y-%m-%d').strftime('%Y-%m-%d')
        stockrefresh = {"productcode":productcode,"calculateto":calculateto,"productdesc":productdesc,"godownflag":godownflag,"goid":goid}
        result = requests.get("http://127.0.0.1:6543/report?godownwisestockonhand&type=pag&productcode=%d&enddate=%s"%(productcode, scalculateto),headers=header)

    return render_to_response("gkwebapp:templates/printstockonhandreport.jinja2",{"gkresult":result.json()["gkresult"],"stockrefresh":stockrefresh,"godown":goname},request=request)

@view_config(route_name="product",request_param="type=productimport", renderer="json")
def ProductImport(request):
    try:
        header={"gktoken":request.headers["gktoken"]}
        xlsxfile = request.POST['xlsxfile'].file
        wb= load_workbook(xlsxfile)
        wb._active_sheet_index = 0
        productSheet = wb.active
        productList = tuple(productSheet.rows)
        uom = requests.get("http://127.0.0.1:6543/unitofmeasurement?qty=all", headers=header).json()["gkresult"]
        godown = requests.get("http://127.0.0.1:6543/godown", headers=header).json()["gkresult"]
        states = requests.get("http://127.0.0.1:6543/state", headers=header).json()["gkresult"]
        statesList=[]
        
        for state in states:
            for value in list(state.values()):
                statesList.append(value)
        proddetails={}
        godowns={}
        godownflag=False
        taxes=[]
        tax={}
        gstrate=["5","18","12","28"]
        gkdata=[]
        errorlist=[]
        duplicateFlag=False
        rowcount=0
        newrows=[]
        multipletax = False
        for productrow in productList[1:]:
            # first row of spreadsheet has column headings. Skipping that row.  
            if (productrow[0].value == None and productrow[1].value == None and productrow[2].value == None and productrow[3].value == None and productrow[4].value == None and productrow[5].value == None and productrow[6].value == None and productrow[7].value == None and productrow[8].value == None and productrow[9].value == None and productrow[10].value == None and productrow[11].value == None):
                continue
            #new entry in name column encountered
            if productrow[0].value != None: 
                if len(proddetails)!=0:
                    temp={}
                    productdetails = {"productdetails":proddetails, "godetails":godowns, "godownflag":godownflag}
                    temp["productdetails"]=productdetails
                    temp["taxes"]=taxes
                    gkdata.append(temp)
                    proddetails={}
                    godowns={}
                    godownflag=False
                    taxes=[]
                    tax={}
                    duplicateFlag=False
                    multipletax = False

                if productrow[0].value != "product name":  
                    newrows.append(rowcount)
                    print((productrow[0].value))
                    #Validations Begin
                    #No product name
                    if productrow[0].value==None:
                        errorlist.append(productrow[0].coordinate)
                    #No HSN code
                    if productrow[1].value==None:
                        errorlist.append(productrow[1].coordinate)
                    #Non float value for Discount in Amount.
                    if productrow[5].value != None and not(is_number(str(productrow[5].value))):
                        errorlist.append(productrow[5].coordinate)
                    #Non float value for Discount in Percent
                    if productrow[6].value != None and not(is_number(str(productrow[6].value))):
                        errorlist.append(productrow[6].coordinate)
                    #Improper taxname
                    if productrow[7].value != None and str(productrow[7].value) not in ["GST","VAT", "CVAT"]:
                        errorlist.append(productrow[7].coordinate)
                    #Null tax rate
                    if productrow[7].value !=None and productrow[9].value==None:
                        errorlist.append(productrow[9].coordinate)
                    #Non float tax rate
                    if productrow[9].value != None and not(is_number(str(productrow[9].value))):
                        errorlist.append(productrow[9].coordinate)
                    if productrow[9].value != None and is_number(str(productrow[9].value)) and int(productrow[9].value) == 0:
                        errorlist.append(productrow[9].coordinate)
                    #Non float godown openingstock
                    if productrow[11].value != None and not(is_number(str(productrow[11].value))):
                        errorlist.append(productrow[11].coordinate)
                    #Sate specified for non VAT tax.
                    if productrow[7].value!="VAT" and productrow[8].value!=None: 
                        errorlist.append(productrow[8].coordinate)
                    #No or wrong name of state for VAT
                    if productrow[7].value=="VAT" and (productrow[8].value==None or productrow[8].value not in statesList):
                        errorlist.append(productrow[8].coordinate)
                    #Improper GST Rate
                    if productrow[7].value=="GST" and str(productrow[9].value) not in gstrate:
                        errorlist.append(productrow[9].coordinate)
                    for taxdict in taxes:
                        taxdicttaxname = taxdict["taxname"]
                        if taxdicttaxname == "IGST":
                            taxdicttaxname = "GST"
                        if productrow[7].value!=None and productrow[7].value!="VAT" and productrow[7].value == taxdicttaxname:
                            multipletax = True
                            errorlist.append(productrow[7].coordinate)
                    #Validations End.

                    proddetails={}        
                    godowns={}
                    godownflag=False
                    taxes=[]
                    proddetails["productdesc"] =  productrow[0].value           
                    proddetails["gscode"]=productrow[1].value
                    if productrow[3].value != None:
                        proddetails["prodmrp"] = productrow[3].value
                    else:
                         proddetails["prodmrp"] = 0.00
                    if productrow[4].value != None:
                        proddetails["prodsp"] = productrow[4].value
                    else:
                        proddetails["prodsp"] = 0.00
                    if productrow[5].value != None:
                        proddetails["amountdiscount"] = productrow[5].value
                    else:
                        proddetails["amountdiscount"] = 0.00
                    if productrow[6].value != None:
                        proddetails["percentdiscount"] = productrow[6].value
                    else:
                        proddetails["percentdiscount"] = 0.00

                    proddetails["specs"] = {}
                    if productrow[2].value !=None :
                        proddetails["gsflag"]=7
                        for i in uom:
                            if i["description"]==productrow[2].value:
                                uomid=i["uomid"]
                                break
                        try:
                            proddetails["uomid"] = uomid
                        except:
                            errorlist.append(productrow[2].coordinate)  
                    else:        
                        proddetails["gsflag"] = 19
                    #proddetails["openingstock"]=0.00
                    if productrow[11].value!=None:
                        proddetails["openingstock"] =productrow[11].value
                    else:
                        proddetails["openingstock"] = 0.00

                    if productrow[7].value!=None and productrow[9].value != None and productrow[9].value != 0 and not multipletax:
                        #tax dictionary generation
                        tax={}
                        if productrow[7].value=="GST": #if GST then no state should be mentioned
                            tax["state"]=None
                            tax["taxname"]="IGST"
                        if productrow[7].value=="VAT":                                            
                            tax["taxname"] = "VAT"
                            tax["state"]=productrow[8].value
                        if productrow[7].value == "CVAT":
                            tax["taxname"] = "CVAT"
                            tax["state"]= None
                        tax["taxrate"]= productrow[9].value
                        if len(tax) > 0:
                            taxes.append(tax)
                    #godown dictionary generation
                    if productrow[10].value!=None:
                        print((productrow[10].value))
                        flag=0
                        for g in godown:
                            print (g)
                            if g["goname"]==productrow[10].value:
                                godownflag=True
                                flag=1
                                godowns[g["goid"]]=productrow[11].value
                                break 
                        if flag==0:
                            errorlist.append(productrow[10].coordinate)                  


            elif productrow[7].value!=None or productrow[10].value!=None:
                multipletax = False
                # if multiple tax or godowns exist
                #Validations Begin
                #Improper taxname
                if productrow[7].value != None and str(productrow[7].value) not in ["GST","VAT", "CVAT"]:
                    errorlist.append(productrow[7].coordinate)
                #Null tax rate
                if productrow[7].value !=None and productrow[9].value==None:
                    errorlist.append(productrow[9].coordinate)
                #Non float tax rate
                if productrow[9].value != None and not(is_number(str(productrow[9].value))):
                    errorlist.append(productrow[9].coordinate)
                if productrow[9].value != None and is_number(str(productrow[9].value)) and int(productrow[9].value) == 0:
                    errorlist.append(productrow[9].coordinate)
                #Non float godown openingstock
                if productrow[11].value != None and not(is_number(str(productrow[11].value))):
                    errorlist.append(productrow[11].coordinate)
                #Sate specified for non VAT tax.
                if productrow[7].value!="VAT" and productrow[8].value!=None: 
                    errorlist.append(productrow[8].coordinate)
                #No or wrong name of state for VAT
                if productrow[7].value=="VAT" and (productrow[8].value==None or productrow[8].value not in statesList):
                    errorlist.append(productrow[8].coordinate)
                #Improper GST Rate
                if productrow[7].value=="GST" and str(productrow[9].value) not in gstrate:
                    errorlist.append(productrow[9].coordinate)
                for taxdict in taxes:
                    taxdicttaxname = taxdict["taxname"]
                    if taxdicttaxname == "IGST":
                        taxdicttaxname = "GST"
                    if productrow[7].value!=None and productrow[7].value!="VAT" and productrow[7].value == taxdicttaxname:
                        multipletax = True
                        errorlist.append(productrow[7].coordinate)
                #Validations End.
                if productrow[7].value!=None and productrow[9].value != None and productrow[9].value != 0 and not multipletax:
                    #tax dictionary generation
                    tax={}
                    if productrow[7].value=="GST": #if GST then no state should be mentioned
                        tax["state"]=None
                        tax["taxname"]="IGST"
                    if productrow[7].value=="VAT":
                        tax["taxname"] = "VAT"
                        tax["state"]=productrow[8].value
                    if productrow[7].value == "CVAT":
                        tax["taxname"] = "CVAT"
                        tax["state"]= None
                    tax["taxrate"] = productrow[9].value

                    if len(tax) > 0:
                        taxes.append(tax)     
                            #godown dictionary generation               
                if productrow[10].value!=None:
                    for g in godown:
                        print (g)
                        if g["goname"]==productrow[10].value:
                            godownflag=True
                            flag=1
                            godowns[g["goid"]]=productrow[11].value
                            break 
                    if flag==0:
                        errorlist.append(productrow[10].coordinate)       

        if len(proddetails)!=0 :
            temp={}
            productdetails = {"productdetails":proddetails, "godetails":godowns, "godownflag":godownflag}
            temp["productdetails"]=productdetails
            temp["taxes"]=taxes
            gkdata.append(temp) 
        #if no errors exist then import else send errors 
        if len(errorlist) ==0: 
            count=-1
            for g in gkdata:                
                try:
                    count+=1
                    godowns=g["productdetails"]["godetails"]
                    productdetails=g["productdetails"]
                    taxes=g["taxes"]
                    godownflag=g["productdetails"]["godownflag"]
                    if g["productdetails"]["godownflag"] == True:
                        godnames = ""
                        j = 1;
                        for i in list(godowns.keys()):
                            resultgodown = requests.get("http://127.0.0.1:6543/godown?qty=single&goid=%d"%(int(i)), headers=header)
                            godnames += resultgodown.json()["gkresult"]["goname"] + "(" + resultgodown.json()["gkresult"]["goaddr"] + ")"
                            if j != len(godowns):
                                godnames += ", "
                            j += 1
                    result = requests.post("http://127.0.0.1:6543/products",data=json.dumps(productdetails),headers=header)                    

                    if result.json()["gkstatus"] == 0:
                        savedproductcode = result.json()["gkresult"]
                        if godownflag == True:
                            godata = {"activity":proddetails["productdesc"] + " product created in " + godnames + " godowns"}
                        else:
                            godata = {"activity":proddetails["productdesc"] + " product created"}
                        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(godata),headers=header)

                    for tax in taxes:
                        if len(tax)!=0:
                            taxdata= {"taxname":tax["taxname"],"taxrate":float(tax["taxrate"]),"productcode":result.json()["gkresult"]}
                            if tax["state"]!='':
                                taxdata["state"]=tax["state"]
                            taxresult = requests.post("http://127.0.0.1:6543/tax",data=json.dumps(taxdata) ,headers=header)
                            
                #if duplicate entries exist then an error is generated while entering tax 
                except Exception as e:           
                    if result.json()["gkstatus"]==1:
                        duplicateFlag=True                        
                        errorlist.append("A"+str(newrows[count]))
                        continue
                        
        if len(errorlist)!=0:            
            errorlist1=[]
            errorrows={}
            for i in errorlist:
                temp=[]
                row=openpyxl.utils.cell.coordinate_from_string(i)[1]-1
                errorlist1.append(openpyxl.utils.cell.coordinate_from_string(i))
                if openpyxl.utils.cell.coordinate_from_string(i)[1] not in list(errorrows.keys()):
                    temprow=productList[row]
                    for cell in temprow:        
                        temp.append(cell.value)
                    errorrows[row+1]=(temp)                                  
            return  {"gkstatus":6,"errorlist":errorlist1,"rows":errorrows,"duplicateFlag":duplicateFlag}    
        else:
            return {"gkstatus":0}
    except Exception as e:
        try:
            return {"gkstatus":result.json()["gkstatus"]}
        except:
            return {"gkstatus":3}
