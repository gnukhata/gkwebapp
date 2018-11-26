from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import (Font,
                             Alignment,
                             PatternFill,
                             NamedStyle,
                             Border,
                             Side)


def style_range_of_cells(ws, cell_range, style_name):
    """Style all cells given a worksheet, stylename and range like `A1:D1`"""
    for row in ws[cell_range]:
        for cell in row:
            cell.style = style_name


def set_alignment(ws, start_cells, alignment):
    """
    This function is used to align all cells in a column
    Example:
        To right align following cells:
        C5, c6, C7...C(max_row) and E5, E6, E7...E(max_row)

        set_alignment(wb["b2b"], ["c5", "E5"], "right")
    This will align all cells D5, D6, D7...Dmax_row to given alignment
    """
    for start_cell in start_cells:
        cell_range = start_cell + ":"+start_cell[0]+str(ws.max_row)
        style_range_of_cells(ws, cell_range, alignment)


def set_number_format(ws, start_cells, number_format):
    """
    This function is used to set number format of columns
    The starting columns are given in start_cells
    Example:
        To set 0.00 number format on following cells:
        C5, C6, C7...C(max_row) and E5, E6, E7...E(max_row)

        set_number_format(wb["b2b"], ["C5", "E5"], "0.00")
    """
    for start_cell in start_cells:
        cell_range = start_cell + ":" + start_cell[0] + str(ws.max_row)
        for row in ws[cell_range]:
            for cell in row:
                cell.number_format = number_format


def float_or_none(number):
    try:
        return float(number)
    except ValueError:
        return ""


def to_datetime(date):
    return datetime.strptime(date, "%d-%b-%y")


def create_styles(wb):
    """Creates all styles required to generate gst_r1 workbook"""
    summary = NamedStyle(name="summary")

    summary.font = Font(name="Times New Roman",
                        bold=True,
                        color="FFFFFFFF",
                        size=11.0)

    summary.alignment = Alignment(horizontal="center",
                                  vertical="top")

    summary.fill = PatternFill(patternType="solid",
                               fgColor="FF0070C0",
                               bgColor="FF0563C1")

    summary.border = Border(outline=True,
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"))

    wb.add_named_style(summary)

    summary_right = NamedStyle(name="summary_right")

    summary_right.font = Font(name="Times New Roman",
                              bold=True,
                              color="FFFFFFFF",
                              size=11.0)

    summary_right.alignment = Alignment(horizontal="right",
                                        vertical="bottom")

    summary_right.fill = PatternFill(patternType="solid",
                                     fgColor="FF0070C0",
                                     bgColor="FF0563C1")

    summary_right.border = Border(outline=True,
                                  left=Side(style="thin"),
                                  right=Side(style="thin"),
                                  top=Side(style="thin"))

    wb.add_named_style(summary_right)

    help_style = NamedStyle(name="help_style")
    help_style.font = Font(name="Calibri",
                           color="FF0563C1",
                           size=11.0)

    help_style.alignment = Alignment(horizontal="center",
                                     vertical="top")

    help_style.border = Border(outline=True,
                               left=Side(style="thin"),
                               right=Side(style="thin"),
                               top=Side(style="thin"))

    wb.add_named_style(help_style)

    count = NamedStyle(name="count")
    count.font = Font(name="Times New Roman",
                      color="FF000000",
                      size=11.0)

    count.alignment = Alignment(horizontal="center",
                                vertical="bottom")

    count.border = Border(outline=True,
                          diagonalDown=False,
                          left=Side(style="thin"),
                          right=Side(style="thin"),
                          top=Side(style="thin"),
                          bottom=Side(style="thin"))

    wb.add_named_style(count)

    sum_style = NamedStyle(name="sum_style")
    sum_style.font = Font(name="Times New Roman",
                          color="FF0A0101",
                          size=11.0)

    sum_style.alignment = Alignment(horizontal="general",
                                    vertical="bottom")

    sum_style.border = Border(outline=True,
                              left=Side(style="thin"),
                              right=Side(style="thin"),
                              top=Side(style="thin"),
                              bottom=Side(style="thin"))
    wb.add_named_style(sum_style)

    sum_style_right = NamedStyle(name="sum_style_right")

    sum_style_right.font = Font(name="Times New Roman",
                                color="FF0A0101",
                                size=11.0)

    sum_style_right.alignment = Alignment(horizontal="general",
                                          vertical="bottom")

    sum_style_right.border = Border(outline=True,
                                    left=Side(style="thin"),
                                    right=Side(style="thin"),
                                    top=Side(style="thin"),
                                    bottom=Side(style="thin"))

    wb.add_named_style(sum_style_right)

    column_header = NamedStyle(name="column_header")

    column_header.font = Font(name="Times New Roman",
                              size=11.0)

    column_header.alignment = Alignment(horizontal="center",
                                        vertical="bottom")

    column_header.fill = PatternFill(patternType="solid",
                                     fgColor="FFF8CBAD",
                                     bgColor="FFBDD7EE")

    column_header.border = Border(outline=True,
                                  left=Side(style="thin"),
                                  right=Side(style="thin"),
                                  top=Side(style="thin"),
                                  bottom=Side(style="thin"))

    wb.add_named_style(column_header)

    column_header_blue = NamedStyle(name="column_header_blue")
    column_header_blue.font = Font(name="Times New Roman",
                                   size=11.0)

    column_header_blue.alignment = Alignment(horizontal="center",
                                             vertical="bottom")

    column_header_blue.fill = PatternFill(patternType="solid",
                                          fgColor="FFB4C7E7",
                                          bgColor="FFBDD7EE")

    column_header_blue.border = Border(outline=True,
                                       left=Side(style="thin"),
                                       right=Side(style="thin"),
                                       top=Side(style="thin"),
                                       bottom=Side(style="thin"))

    wb.add_named_style(column_header_blue)

    column_header_bold = NamedStyle(name="column_header_bold")

    column_header_bold.font = Font(name="Times New Roman",
                                   bold=True,
                                   size=11.0)

    column_header_bold.alignment = Alignment(horizontal="center",
                                             vertical="bottom")

    column_header_bold.fill = PatternFill(patternType="solid",
                                          fgColor="FFF8CBAD",
                                          bgColor="FFBDD7EE")

    column_header_bold.border = Border(outline=True,
                                       left=Side(style="thin"),
                                       right=Side(style="thin"),
                                       top=Side(style="thin"),
                                       bottom=Side(style="thin"))

    wb.add_named_style(column_header_bold)

    column_header_bold_blue = NamedStyle(name="column_header_bold_blue")
    column_header_bold_blue.font = Font(name="Times New Roman",
                                        bold=True,
                                        size=11.0)

    column_header_bold_blue.alignment = Alignment(horizontal="center",
                                                  vertical="bottom")

    column_header_bold_blue.fill = PatternFill(patternType="solid",
                                               fgColor="FFB4C7E7",
                                               bgColor="FFBDD7EE")

    column_header_bold_blue.border = Border(outline=True,
                                            left=Side(style="thin"),
                                            right=Side(style="thin"),
                                            top=Side(style="thin"),
                                            bottom=Side(style="thin"))

    wb.add_named_style(column_header_bold_blue)

    master_sheet_font = NamedStyle(name="master_sheet_font")
    master_sheet_font.font = Font(name="Times New Roman",
                                  color="FF000000",
                                  size=11.0)

    wb.add_named_style(master_sheet_font)

    align_right = NamedStyle(name="right")
    align_right.alignment = Alignment(horizontal="right",
                                      vertical="top")
    wb.add_named_style(align_right)

    align_center = NamedStyle(name="center")
    align_center.alignment = Alignment(horizontal="center",
                                       vertical="top")
    wb.add_named_style(align_center)

    return wb


def b2b_sheet(wb):
    wb.create_sheet(title="b2b")

    style_range_of_cells(wb["b2b"], "A1:B1", "summary")
    style_range_of_cells(wb["b2b"], "A2:M2", "summary")

    for col in range(1, 13):
        i = get_column_letter(col)
        wb["b2b"].column_dimensions[i].width = 25

    wb["b2b"].row_dimensions[4].height = 30.0

    wb["b2b"]["E3"].number_format = "0.00"
    wb["b2b"]["L3"].number_format = "0.00"

    wb["b2b"]["A1"] = """Summary For B2B(4)"""
    wb["b2b"]["A2"] = """No. of Recipients"""
    wb["b2b"]["A3"] = """=SUMPRODUCT((A5:A20004<>"")/COUNTIF(A5:A20004,A5:A20004&""))"""
    wb["b2b"]["A3"].style = "count"
    wb["b2b"]["A4"] = """GSTIN/UIN of Recipient"""
    wb["b2b"]["A4"].style = "column_header"
    wb["b2b"]["B4"] = """Receiver Name"""
    wb["b2b"]["B4"].style = "column_header"
    wb["b2b"]["C2"] = """No. of Invoices"""
    wb["b2b"]["C3"] = """=SUMPRODUCT((C5:C20004<>"")/COUNTIF(C5:C20004,C5:C20004&""))"""
    wb["b2b"]["C3"].style = "count"
    wb["b2b"]["C4"] = """Invoice Number"""
    wb["b2b"]["C4"].style = "column_header"
    wb["b2b"]["D4"] = """Invoice date"""
    wb["b2b"]["D4"].style = "column_header"
    wb["b2b"]["E2"] = """Total Invoice Value"""
    wb["b2b"]["E2"].style = "summary_right"
    wb["b2b"]["E3"] = """=SUMPRODUCT(1/COUNTIF(C5:C20004,C5:C20004&""),E5:E20004)"""
    wb["b2b"]["E3"].style = "sum_style_right"
    wb["b2b"]["E4"] = """Invoice Value"""
    wb["b2b"]["E4"].style = "column_header"
    wb["b2b"]["F4"] = """Place Of Supply"""
    wb["b2b"]["F4"].style = "column_header"
    wb["b2b"]["G4"] = """Reverse Charge"""
    wb["b2b"]["G4"].style = "column_header"
    wb["b2b"]["H4"] = """Applicable % of Tax Rate"""
    wb["b2b"]["H4"].style = "column_header"
    wb["b2b"]["I4"] = """Invoice Type"""
    wb["b2b"]["I4"].style = "column_header"
    wb["b2b"]["J4"] = """E-Commerce GSTIN"""
    wb["b2b"]["J4"].style = "column_header"
    wb["b2b"]["K4"] = """Rate"""
    wb["b2b"]["K4"].style = "column_header"
    wb["b2b"]["L2"] = """Total Taxable Value"""
    wb["b2b"]["L2"].style = "summary_right"
    wb["b2b"]["L3"] = """=SUM(L5:L20004)"""
    wb["b2b"]["L3"].style = "sum_style_right"
    wb["b2b"]["L4"] = """Taxable Value"""
    wb["b2b"]["L4"].style = "column_header"
    wb["b2b"]["M1"] = """HELP"""
    wb["b2b"]["M1"].style = "help_style"
    wb["b2b"]["M2"] = """Total Cess"""
    wb["b2b"]["M2"].style = "summary_right"
    wb["b2b"]["M3"] = """=SUMIF($G$5:$G$20004,"N",M5:M20004)+SUMIF($G$5:$G$20004,"",M5:M20004)"""
    wb["b2b"]["M3"].style = "sum_style_right"
    wb["b2b"]["M4"] = """Cess Amount"""
    wb["b2b"]["M4"].style = "column_header"


def b2ba_sheet(wb):
    wb.create_sheet(title="b2ba")

    wb["b2ba"].merge_cells("B1:D1")
    wb["b2ba"].merge_cells("E1:N1")

    style_range_of_cells(wb["b2ba"], "A1:A1", "summary")
    style_range_of_cells(wb["b2ba"], "A2:O2", "summary")

    for col in range(1, 15):
        i = get_column_letter(col)
        wb["b2ba"].column_dimensions[i].width = 25

    wb["b2ba"].row_dimensions[4].height = 30.0
    wb["b2ba"].row_dimensions[17].height = 13.8

    wb["b2ba"].style = "help_style"
    wb["b2ba"]["A1"] = """Summary For B2BA"""
    wb["b2ba"]["A2"] = """No. of Recipients"""
    wb["b2ba"]["A3"] = """=SUMPRODUCT((A5:A20004<>"")/COUNTIF(A5:A20004,A5:A20004&""))"""
    wb["b2ba"]["A3"].style = "count"
    wb["b2ba"]["A4"] = """GSTIN/UIN of Recipient"""
    wb["b2ba"]["A4"].style = "column_header"
    wb["b2ba"]["B1"] = """Original details """
    wb["b2ba"]["B1"].style = "column_header_bold"
    wb["b2ba"]["B4"] = """Receiver Name"""
    wb["b2ba"]["B4"].style = "column_header"
    wb["b2ba"]["C2"] = """No. of Invoices"""
    wb["b2ba"]["C3"] = """=SUMPRODUCT((C5:C20004<>"")/COUNTIF(C5:C20004,C5:C20004&""))"""
    wb["b2ba"]["C3"].style = "count"
    wb["b2ba"]["C4"] = """Original Invoice Number"""
    wb["b2ba"]["C4"].style = "column_header"
    wb["b2ba"]["D4"] = """Original Invoice date"""
    wb["b2ba"]["D4"].style = "column_header"
    wb["b2ba"]["E1"] = """Revised Details """
    wb["b2ba"]["E1"].style = "column_header_bold_blue"
    wb["b2ba"]["E4"] = """Revised Invoice Number"""
    wb["b2ba"]["E4"].style = "column_header_blue"
    wb["b2ba"]["F4"] = """Revised Invoice date"""
    wb["b2ba"]["F4"].style = "column_header_blue"
    wb["b2ba"]["G2"] = """Total Invoice Value"""
    wb["b2ba"]["G2"].style = "summary_right"
    wb["b2ba"]["G3"] = """=SUMPRODUCT(1/COUNTIF(C5:C20004,C5:C20004&""),G5:G20004)"""
    wb["b2ba"]["G3"].style = "sum_style_right"
    wb["b2ba"]["G4"] = """Invoice Value"""
    wb["b2ba"]["G4"].style = "column_header_blue"
    wb["b2ba"]["H4"] = """Place Of Supply"""
    wb["b2ba"]["H4"].style = "column_header_blue"
    wb["b2ba"]["I4"] = """Reverse Charge"""
    wb["b2ba"]["I4"].style = "column_header_blue"
    wb["b2ba"]["J4"] = """Applicable % of Tax Rate"""
    wb["b2ba"]["J4"].style = "column_header_blue"
    wb["b2ba"]["K4"] = """Invoice Type"""
    wb["b2ba"]["K4"].style = "column_header_blue"
    wb["b2ba"]["L4"] = """E-Commerce GSTIN"""
    wb["b2ba"]["L4"].style = "column_header_blue"
    wb["b2ba"]["M3"] = """=SUMIF($I$5:$I$20004,"N",O5:O20004)+SUMIF($I$5:$I$20004,"",O5:O20004)"""
    wb["b2ba"]["M3"].style = "sum_style_right"
    wb["b2ba"]["M4"] = """Rate"""
    wb["b2ba"]["M4"].style = "column_header_blue"
    wb["b2ba"]["N2"] = """Total Taxable Value"""
    wb["b2ba"]["N2"].style = "summary_right"
    wb["b2ba"]["N3"] = """=SUM(N5:N20004)"""
    wb["b2ba"]["N3"].style = "sum_style_right"
    wb["b2ba"]["N4"] = """Taxable Value"""
    wb["b2ba"]["N4"].style = "column_header_blue"
    wb["b2ba"]["O1"] = """HELP"""
    wb["b2ba"]["O2"] = """Total Cess"""
    wb["b2ba"]["O2"].style = "summary_right"
    wb["b2ba"]["O3"] = """=SUMIF($I$5:$I$20004,"N",O5:O20004)+SUMIF($I$5:$I$20004,"",O5:O20004)"""
    wb["b2ba"]["O3"].style = "sum_style_right"
    wb["b2ba"]["O4"] = """Cess Amount"""
    wb["b2ba"]["O4"].style = "column_header_blue"

    return wb


def b2cl_sheet(wb):
    wb.create_sheet(title="b2cl")

    style_range_of_cells(wb["b2cl"], "A1:A1", "summary")
    style_range_of_cells(wb["b2cl"], "A2:I2", "summary")

    for col in range(1, 11):
        i = get_column_letter(col)
        wb["b2cl"].column_dimensions[i].width = 25

    wb["b2cl"].row_dimensions[4].height = 30.0

    wb["b2cl"]["C3"].number_format = "0.00"
    wb["b2cl"]["G3"].number_format = "0.00"
    wb["b2cl"]["H3"].number_format = "0.00"

    wb["b2cl"]["A1"] = """Summary For B2CL(5)"""
    wb["b2cl"]["A2"] = """No. of Invoices"""
    wb["b2cl"]["A3"] = """=SUMPRODUCT((A5:A20004<>"")/COUNTIF(A5:A20004,A5:A20004&""))"""
    wb["b2cl"]["A3"].style = "count"
    wb["b2cl"]["A4"] = """Invoice Number"""
    wb["b2cl"]["A4"].style = "column_header"
    wb["b2cl"]["B4"] = """Invoice date"""
    wb["b2cl"]["B4"].style = "column_header"
    wb["b2cl"]["C2"] = """Total Inv Value"""
    wb["b2cl"]["C3"] = """=SUMPRODUCT(1/COUNTIF(A5:A20004,A5:A20004&""),C5:C20004)"""
    wb["b2cl"]["C4"] = """Invoice Value"""
    wb["b2cl"]["C4"].style = "column_header"
    wb["b2cl"]["D4"] = """Place Of Supply"""
    wb["b2cl"]["D4"].style = "column_header"
    wb["b2cl"]["E4"] = """Applicable % of Tax Rate"""
    wb["b2cl"]["E4"].style = "column_header"
    wb["b2cl"]["F4"] = """Rate"""
    wb["b2cl"]["F4"].style = "column_header"
    wb["b2cl"]["G2"] = """Total Taxable Value"""
    wb["b2cl"]["G3"] = """=SUM(G5:G20004)"""
    wb["b2cl"]["G3"].style = "sum_style_right"
    wb["b2cl"]["G4"] = """Taxable Value"""
    wb["b2cl"]["G4"].style = "column_header"
    wb["b2cl"]["H2"] = """Total Cess"""
    wb["b2cl"]["H2"].style = "summary_right"
    wb["b2cl"]["H3"] = """=SUM(H5:H20004)"""
    wb["b2cl"]["H3"].style = "sum_style_right"
    wb["b2cl"]["H4"] = """Cess Amount"""
    wb["b2cl"]["H4"].style = "column_header"
    wb["b2cl"]["I1"] = """HELP"""
    wb["b2cl"]["I1"].style = "help_style"
    wb["b2cl"]["I4"] = """E-Commerce GSTIN"""
    wb["b2cl"]["I4"].style = "column_header"
    wb["b2cl"]["J4"] = """Sale from Bonded WH"""
    wb["b2cl"]["J4"].style = "column_header"
    wb["b2cl"]["c3"].style = "sum_style_right"

    return wb


def b2cla_sheet(wb):
    wb.create_sheet(title="b2cla")

    style_range_of_cells(wb["b2cla"], "A1:A1", "summary")
    style_range_of_cells(wb["b2cla"], "A2:K2", "summary")

    wb["b2cla"].merge_cells("B1:C1")
    wb["b2cla"].merge_cells("D1:J1")

    for col in range(1, 14):
        i = get_column_letter(col)
        wb["b2cla"].column_dimensions[i].width = 25

    wb["b2cla"].row_dimensions[4].height = 30

    wb["b2cla"]["A1"] = """Summary For B2CLA"""
    wb["b2cla"]["A2"] = """No. of Invoices"""
    wb["b2cla"]["A3"] = """=SUMPRODUCT((A5:A2004<>"")/COUNTIF(A5:A2004,A5:A2004&""))"""
    wb["b2cla"]["A3"].style = "count"
    wb["b2cla"]["A4"] = """Original Invoice Number"""
    wb["b2cla"]["A4"].style = "column_header"
    wb["b2cla"]["B1"] = """Original details """
    wb["b2cla"]["B1"].style = "column_header"
    wb["b2cla"]["B4"] = """Original Invoice date"""
    wb["b2cla"]["B4"].style = "column_header"
    wb["b2cla"]["C4"] = """Original Place Of Supply"""
    wb["b2cla"]["C4"].style = "column_header"
    wb["b2cla"]["D1"] = """Revised Details """
    wb["b2cla"]["D1"].style = "column_header_bold_blue"
    wb["b2cla"]["D4"] = """Revised Invoice Number"""
    wb["b2cla"]["D4"].style = "column_header_blue"
    wb["b2cla"]["E4"] = """Revised Invoice date"""
    wb["b2cla"]["E4"].style = "column_header_blue"
    wb["b2cla"]["F2"] = """Total Inv Value"""
    wb["b2cla"]["F3"] = """=SUMPRODUCT(1/COUNTIF(A5:A2004,A5:A2004&""),F5:F2004)"""
    wb["b2cla"]["F3"].style = "sum_style_right"
    wb["b2cla"]["F4"] = """Invoice Value"""
    wb["b2cla"]["F4"].style = "column_header_blue"
    wb["b2cla"]["G4"] = """Applicable % of Tax Rate"""
    wb["b2cla"]["G4"].style = "column_header_blue"
    wb["b2cla"]["H4"] = """Rate"""
    wb["b2cla"]["H4"].style = "column_header_blue"
    wb["b2cla"]["I2"] = """Total Taxable Value"""
    wb["b2cla"]["I3"] = """=SUM(I5:I2004)"""
    wb["b2cla"]["I3"].style = "sum_style_right"
    wb["b2cla"]["I4"] = """Taxable Value"""
    wb["b2cla"]["I4"].style = "column_header_blue"
    wb["b2cla"]["J2"] = """Total Cess"""
    wb["b2cla"]["J2"].style = "summary_right"
    wb["b2cla"]["J3"] = """=SUM(J5:J2004)"""
    wb["b2cla"]["J3"].style = "sum_style_right"
    wb["b2cla"]["J4"] = """Cess Amount"""
    wb["b2cla"]["J4"].style = "column_header_blue"
    wb["b2cla"]["K1"] = """HELP"""
    wb["b2cla"]["K1"].style = "help_style"
    wb["b2cla"]["K4"] = """E-Commerce GSTIN"""
    wb["b2cla"]["K4"].style = "column_header_blue"
    wb["b2cla"]["L4"] = """Sale from Bonded WH"""
    wb["b2cla"]["L4"].style = "column_header_blue"


def b2cs_sheet(wb):
    wb.create_sheet(title="b2cs")

    style_range_of_cells(wb["b2cs"], "A1:A1", "summary")
    style_range_of_cells(wb["b2cs"], "A2:G2", "summary")

    for col in range(1, 8):
        i = get_column_letter(col)
        wb["b2cs"].column_dimensions[i].width = 25

    wb["b2cs"].row_dimensions[4].height = 30.0

    wb["b2cs"]["E3"].number_format = "0.00"
    wb["b2cs"]["F3"].number_format = "0.00"

    wb["b2cs"]["A1"] = """Summary For B2CS(7)"""
    wb["b2cs"]["A4"] = """Type"""
    wb["b2cs"]["A4"].style = "column_header"
    wb["b2cs"]["B4"] = """Place Of Supply"""
    wb["b2cs"]["B4"].style = "column_header"
    wb["b2cs"]["C4"] = """Applicable % of Tax Rate"""
    wb["b2cs"]["C4"].style = "column_header"
    wb["b2cs"]["D4"] = """Rate"""
    wb["b2cs"]["D4"].style = "column_header"
    wb["b2cs"]["E2"] = """Total Taxable  Value"""
    wb["b2cs"]["E2"].style = "summary_right"
    wb["b2cs"]["E3"] = """=SUM(E5:E400)"""
    wb["b2cs"]["E3"].style = "sum_style_right"
    wb["b2cs"]["E4"] = """Taxable Value"""
    wb["b2cs"]["E4"].style = "column_header"
    wb["b2cs"]["F2"] = """Total Cess"""
    wb["b2cs"]["F2"].style = "summary_right"
    wb["b2cs"]["F3"] = """=SUM(F5:F400)"""
    wb["b2cs"]["F3"].style = "sum_style_right"
    wb["b2cs"]["F4"] = """Cess Amount"""
    wb["b2cs"]["F4"].style = "column_header"
    wb["b2cs"]["G1"] = """HELP"""
    wb["b2cs"]["G1"].style = "help_style"
    wb["b2cs"]["G2"].style = "summary"
    wb["b2cs"]["G4"] = """E-Commerce GSTIN"""
    wb["b2cs"]["G4"].style = "column_header"


def b2csa_sheet(wb):
    wb.create_sheet(title="b2csa")

    style_range_of_cells(wb["b2csa"], "A1:A1", "summary")
    style_range_of_cells(wb["b2csa"], "A2:I2", "summary")

    wb["b2csa"].merge_cells("C1:H1")

    for col in range(1, 10):
        i = get_column_letter(col)
        wb["b2csa"].column_dimensions[i].width = 25

    wb["b2csa"].row_dimensions[4].height = 30

    wb["b2csa"]["A1"] = """Summary For B2CSA"""
    wb["b2csa"]["A4"] = """Financial Year"""
    wb["b2csa"]["A4"].style = "column_header"
    wb["b2csa"]["B1"] = """Original details """
    wb["b2csa"]["B1"].style = "column_header_bold"
    wb["b2csa"]["B4"] = """Original Month"""
    wb["b2csa"]["B4"].style = "column_header"
    wb["b2csa"]["C1"] = """Revised details"""
    wb["b2csa"]["C1"].style = "column_header_bold_blue"
    wb["b2csa"]["C4"] = """Place Of Supply"""
    wb["b2csa"]["C4"].style = "column_header_blue"
    wb["b2csa"]["D4"] = """Type"""
    wb["b2csa"]["D4"].style = "column_header_blue"
    wb["b2csa"]["E4"] = """Applicable % of Tax Rate"""
    wb["b2csa"]["E4"].style = "column_header_blue"
    wb["b2csa"]["F4"] = """Rate"""
    wb["b2csa"]["F4"].style = "column_header_blue"
    wb["b2csa"]["G2"] = """Total Taxable  Value"""
    wb["b2csa"]["G2"].style = "summary_right"
    wb["b2csa"]["G3"] = """=SUM(G5:G404)"""
    wb["b2csa"]["G3"].style = "sum_style_right"
    wb["b2csa"]["G4"] = """Taxable Value"""
    wb["b2csa"]["G4"].style = "column_header_blue"
    wb["b2csa"]["H2"] = """Total Cess"""
    wb["b2csa"]["H2"].style = "summary_right"
    wb["b2csa"]["H3"] = """=SUM(H5:H404)"""
    wb["b2csa"]["H3"].style = "sum_style_right"
    wb["b2csa"]["H4"] = """Cess Amount"""
    wb["b2csa"]["H4"].style = "column_header_blue"
    wb["b2csa"]["I1"] = """HELP"""
    wb["b2csa"]["I1"].style = "help_style"
    wb["b2csa"]["I4"] = """E-Commerce GSTIN"""
    wb["b2csa"]["I4"].style = "column_header_blue"


def cdnr_sheet(wb):
    wb.create_sheet(title="cdnr")

    style_range_of_cells(wb["cdnr"], "A1:B2", "summary")
    style_range_of_cells(wb["cdnr"], "A2:N2", "summary")

    for col in range(1, 15):
        i = get_column_letter(col)
        wb["cdnr"].column_dimensions[i].width = 25

    wb["cdnr"].row_dimensions[4].height = 30

    wb["cdnr"]["I3"].number_format = "0.00"
    wb["cdnr"]["L3"].number_format = "0.00"
    wb["cdnr"]["M3"].number_format = "0.00"

    wb["cdnr"]["A1"] = """Summary For CDNR(9B)"""
    wb["cdnr"]["A2"] = """No. of Recipients"""
    wb["cdnr"]["A3"] = """=SUMPRODUCT((A5:A1001<>"")/COUNTIF(A5:A1001,A5:A1001&""))"""
    wb["cdnr"]["A3"].style = "count"
    wb["cdnr"]["A4"] = """GSTIN/UIN of Recipient"""
    wb["cdnr"]["A4"].style = "column_header"
    wb["cdnr"]["B4"] = """Receiver Name"""
    wb["cdnr"]["B4"].style = "column_header"
    wb["cdnr"]["C2"] = """No. of Invoices"""
    wb["cdnr"]["C3"] = """=SUMPRODUCT((C5:C1001<>"")/COUNTIF(C5:C1001,C5:C1001&""))"""
    wb["cdnr"]["C3"].style = "count"
    wb["cdnr"]["C4"] = """Invoice/Advance Receipt Number"""
    wb["cdnr"]["C4"].style = "column_header"
    wb["cdnr"]["D4"] = """Invoice/Advance Receipt date"""
    wb["cdnr"]["D4"].style = "column_header"
    wb["cdnr"]["E2"] = """No. of Notes/Vouchers"""
    wb["cdnr"]["E3"] = """=SUMPRODUCT((E5:E1001<>"")/COUNTIF(E5:E1001,E5:E1001&""))"""
    wb["cdnr"]["E3"].style = "count"
    wb["cdnr"]["E4"] = """Note/Refund Voucher Number"""
    wb["cdnr"]["E4"].style = "column_header"
    wb["cdnr"]["F4"] = """Note/Refund Voucher date"""
    wb["cdnr"]["F4"].style = "column_header"
    wb["cdnr"]["G4"] = """Document Type"""
    wb["cdnr"]["G4"].style = "column_header"
    wb["cdnr"]["H4"] = """Place Of Supply"""
    wb["cdnr"]["H4"].style = "column_header"
    wb["cdnr"]["I2"] = """Total Note/Refund Voucher Value"""
    wb["cdnr"]["I2"].style = "summary_right"
    wb["cdnr"]["I3"] = """=SUMPRODUCT(1/COUNTIF(E5:E1001,E5:E1001&""),I5:I1001)"""
    wb["cdnr"]["I3"].style = "sum_style_right"
    wb["cdnr"]["I4"] = """Note/Refund Voucher Value"""
    wb["cdnr"]["I4"].style = "column_header"
    wb["cdnr"]["J4"] = """Applicable % of Tax Rate"""
    wb["cdnr"]["J4"].style = "column_header"
    wb["cdnr"]["K4"] = """Rate"""
    wb["cdnr"]["K4"].style = "column_header"
    wb["cdnr"]["L2"] = """Total Taxable Value"""
    wb["cdnr"]["L2"].style = "summary_right"
    wb["cdnr"]["L3"] = """=SUM(L5:L1001)"""
    wb["cdnr"]["L3"].style = "sum_style_right"
    wb["cdnr"]["L4"] = """Taxable Value"""
    wb["cdnr"]["L4"].style = "column_header"
    wb["cdnr"]["M2"] = """Total Cess"""
    wb["cdnr"]["M2"].style = "summary_right"
    wb["cdnr"]["M3"] = """=SUMIF(G5:G1001,"D",M5:M1001)-(SUMIF(G5:G1001,"C",M5:M1001)+SUMIF(G5:G1001,"R",M5:M1001))"""
    wb["cdnr"]["M3"].style = "sum_style_right"
    wb["cdnr"]["M4"] = """Cess Amount"""
    wb["cdnr"]["M4"].style = "column_header"
    wb["cdnr"]["N1"] = """HELP"""
    wb["cdnr"]["N1"].style = "help_style"
    wb["cdnr"]["N4"] = """Pre GST"""
    wb["cdnr"]["N4"].style = "column_header"


def cdnra_sheet(wb):
    wb.create_sheet(title="cdnra")

    style_range_of_cells(wb["cdnra"], "A1:A1", "summary")
    style_range_of_cells(wb["cdnra"], "A2:P2", "summary")

    wb["cdnra"].merge_cells("B1:F1")
    wb["cdnra"].merge_cells("G1:O1")

    for col in range(1, 17):
        i = get_column_letter(col)
        wb["cdnra"].column_dimensions[i].width = 25

    wb["cdnra"].row_dimensions[4].height = 30

    wb["cdnra"]["a1"] = """summary for cdnra"""
    wb["cdnra"]["A2"] = """No. of Recipients"""
    wb["cdnra"]["A3"] = """=SUMPRODUCT((A5:A1001<>"")/COUNTIF(A5:A1001,A5:A1001&""))"""
    wb["cdnra"]["A3"].style = "count"
    wb["cdnra"]["A4"] = """GSTIN/UIN of Recipient"""
    wb["cdnra"]["A4"].style = "column_header"
    wb["cdnra"]["B1"] = """Original details """
    wb["cdnra"]["B1"].style = "column_header_bold"
    wb["cdnra"]["B4"] = """Receiver Name"""
    wb["cdnra"]["B4"].style = "column_header"
    wb["cdnra"]["C2"] = """No. of Notes/Vouchers"""
    wb["cdnra"]["C3"] = """=SUMPRODUCT((C5:C1001<>"")/COUNTIF(C5:C1001,C5:C1001&""))"""
    wb["cdnra"]["C3"].style = "count"
    wb["cdnra"]["C4"] = """Original Note/Refund Voucher Number"""
    wb["cdnra"]["C4"].style = "column_header"
    wb["cdnra"]["D4"] = """Original Note/Refund Voucher date"""
    wb["cdnra"]["D4"].style = "column_header"
    wb["cdnra"]["E2"] = """No. of Invoices"""
    wb["cdnra"]["E3"] = """=SUMPRODUCT((E5:E1001<>"")/COUNTIF(E5:E1001,E5:E1001&""))"""
    wb["cdnra"]["E3"].style = "count"
    wb["cdnra"]["E4"] = """Original Invoice/Advance Receipt Number"""
    wb["cdnra"]["E4"].style = "column_header"
    wb["cdnra"]["F4"] = """Original Invoice/Advance Receipt date"""
    wb["cdnra"]["F4"].style = "column_header"
    wb["cdnra"]["G1"] = """Revised details"""
    wb["cdnra"]["G1"].style = "column_header_bold_blue"
    wb["cdnra"]["G4"] = """Revised Note/Refund Voucher Number"""
    wb["cdnra"]["G4"].style = "column_header_blue"
    wb["cdnra"]["H4"] = """Revised Note/Refund Voucher date"""
    wb["cdnra"]["H4"].style = "column_header_blue"
    wb["cdnra"]["I4"] = """Document Type"""
    wb["cdnra"]["I4"].style = "column_header_blue"
    wb["cdnra"]["J4"] = """Supply Type"""
    wb["cdnra"]["J4"].style = "column_header_blue"
    wb["cdnra"]["K2"] = """Total Note/Refund Voucher Value"""
    wb["cdnra"]["K2"].style = "summary_right"
    wb["cdnra"]["K3"] = """=SUMPRODUCT(1/COUNTIF(C5:C1001,C5:C1001&""),K5:K1001)"""
    wb["cdnra"]["K3"].style = "sum_style_right"
    wb["cdnra"]["K4"] = """Note/Refund Voucher Value"""
    wb["cdnra"]["K4"].style = "column_header_blue"
    wb["cdnra"]["L4"] = """Applicable % of Tax Rate"""
    wb["cdnra"]["L4"].style = "column_header_blue"
    wb["cdnra"]["M4"] = """Rate"""
    wb["cdnra"]["M4"].style = "column_header_blue"
    wb["cdnra"]["N2"] = """Total Taxable Value"""
    wb["cdnra"]["N2"].style = "summary_right"
    wb["cdnra"]["N3"] = """=SUM(N5:N1004)"""
    wb["cdnra"]["N3"].style = "sum_style_right"
    wb["cdnra"]["N4"] = """Taxable Value"""
    wb["cdnra"]["N4"].style = "column_header_blue"
    wb["cdnra"]["O2"] = """Total Cess"""
    wb["cdnra"]["O2"].style = "summary_right"
    wb["cdnra"]["O3"] = """=SUMIF(I5:I1001,"D",O5:O1001)-(SUMIF(I5:I1001,"C",O5:O1001)+SUMIF(I5:I1001,"R",O5:O1001))"""
    wb["cdnra"]["O3"].style = "sum_style_right"
    wb["cdnra"]["O4"] = """Cess Amount"""
    wb["cdnra"]["O4"].style = "column_header_blue"
    wb["cdnra"]["P1"] = """HELP"""
    wb["cdnra"]["P1"].style = "help_style"
    wb["cdnra"]["P4"] = """Pre GST"""
    wb["cdnra"]["P4"].style = "column_header_blue"


def cdnur_sheet(wb):
    wb.create_sheet(title="cdnur")

    style_range_of_cells(wb["cdnur"], "A1:A1", "summary")
    style_range_of_cells(wb["cdnur"], "A2:M2", "summary")

    for col in range(1, 14):
        i = get_column_letter(col)
        wb["cdnur"].column_dimensions[i].width = 25

    wb["cdnur"].row_dimensions[4].height = 30.0

    wb["cdnur"]["H3"].number_format = "0.00"
    wb["cdnur"]["K3"].number_format = "0.00"
    wb["cdnur"]["L3"].number_format = "0.00"

    wb["cdnur"]["A1"] = """Summary For CDNUR(9B)"""
    wb["cdnur"]["A4"] = """UR Type"""
    wb["cdnur"]["A4"].style = "column_header"
    wb["cdnur"]["B2"] = """No. of Notes/Vouchers"""
    wb["cdnur"]["B3"] = """=SUMPRODUCT((B5:B1001<>"")/COUNTIF(B5:B1001,B5:B1001&""))"""
    wb["cdnur"]["B3"].style = "count"
    wb["cdnur"]["B4"] = """Note/Refund Voucher Number"""
    wb["cdnur"]["B4"].style = "column_header"
    wb["cdnur"]["C4"] = """Note/Refund Voucher date"""
    wb["cdnur"]["C4"].style = "column_header"
    wb["cdnur"]["D4"] = """Document Type"""
    wb["cdnur"]["D4"].style = "column_header"
    wb["cdnur"]["E2"] = """No. of Invoices"""
    wb["cdnur"]["E3"] = """=SUMPRODUCT((E5:E1001<>"")/COUNTIF(E5:E1001,E5:E1001&""))"""
    wb["cdnur"]["E3"].style = "count"
    wb["cdnur"]["E4"] = """Invoice/Advance Receipt Number"""
    wb["cdnur"]["E4"].style = "column_header"
    wb["cdnur"]["F4"] = """Invoice/Advance Receipt date"""
    wb["cdnur"]["F4"].style = "column_header"
    wb["cdnur"]["G4"] = """Place Of Supply"""
    wb["cdnur"]["G4"].style = "column_header"
    wb["cdnur"]["H2"] = """Total Note Value"""
    wb["cdnur"]["H2"].style = "summary_right"
    wb["cdnur"]["H3"] = """=SUMPRODUCT(1/COUNTIF(E5:E1001,E5:E1001&""),H5:H1001)"""
    wb["cdnur"]["H3"].style = "sum_style_right"
    wb["cdnur"]["H4"] = """Note/Refund Voucher Value"""
    wb["cdnur"]["H4"].style = "column_header"
    wb["cdnur"]["I4"] = """Applicable % of Tax Rate"""
    wb["cdnur"]["I4"].style = "column_header"
    wb["cdnur"]["J4"] = """Rate"""
    wb["cdnur"]["J4"].style = "column_header"
    wb["cdnur"]["K2"] = """Total Taxable Value"""
    wb["cdnur"]["K2"].style = "summary_right"
    wb["cdnur"]["K3"] = """=SUM(K5:K1001)"""
    wb["cdnur"]["K3"].style = "sum_style_right"
    wb["cdnur"]["K4"] = """Taxable Value"""
    wb["cdnur"]["K4"].style = "column_header"
    wb["cdnur"]["L2"] = """Total Cess"""
    wb["cdnur"]["L2"].style = "summary_right"
    wb["cdnur"]["L3"] = """=SUMIF(D5:D1001,"D",L5:L1001)-(SUMIF(D5:D1001,"C",L5:L1001)+SUMIF(D5:D1001,"R",L5:L1001))"""
    wb["cdnur"]["L3"].style = "sum_style_right"
    wb["cdnur"]["L4"] = """Cess Amount"""
    wb["cdnur"]["L4"].style = "column_header"
    wb["cdnur"]["M1"] = """HELP"""
    wb["cdnur"]["M1"].style = "help_style"
    wb["cdnur"]["M4"] = """Pre GST"""
    wb["cdnur"]["M4"].style = "column_header"
    wb["cdnur"]["N4"] = """Supply Type"""
    wb["cdnur"]["N4"].style = "column_header"


def cdnura_sheet(wb):
    wb.create_sheet(title="cdnura")

    style_range_of_cells(wb["cdnura"], "A1:A1", "summary")
    style_range_of_cells(wb["cdnura"], "A2:O2", "summary")

    wb["cdnura"].merge_cells("B1:E1")
    wb["cdnura"].merge_cells("F1:J1")


    for col in range(1, 16):
        i = get_column_letter(col)
        wb["cdnura"].column_dimensions[i].width = 25

    wb["cdnura"].row_dimensions[4].height = 30.0

    wb["cdnura"]["A1"] = """Summary For CDNURA"""
    wb["cdnura"]["A4"] = """UR Type"""
    wb["cdnura"]["A4"].style = "column_header"
    wb["cdnura"]["B1"] = """Original details """
    wb["cdnura"]["B1"].style = "column_header_bold"
    wb["cdnura"]["B2"] = """No. of Notes/Vouchers"""
    wb["cdnura"]["B3"] = """=SUMPRODUCT((B5:B1001<>"")/COUNTIF(B5:B1001,B5:B1001&""))"""
    wb["cdnura"]["B3"].style = "count"
    wb["cdnura"]["B4"] = """Original Note/Refund Voucher Number"""
    wb["cdnura"]["B4"].style = "column_header"
    wb["cdnura"]["C4"] = """Original Note/Refund Voucher date"""
    wb["cdnura"]["C4"].style = "column_header"
    wb["cdnura"]["D2"] = """No. of Invoices"""
    wb["cdnura"]["D3"] = """=SUMPRODUCT((D5:D1001<>"")/COUNTIF(D5:D1001,D5:D1001&""))"""
    wb["cdnura"]["D3"].style = "count"
    wb["cdnura"]["D4"] = """Original Invoice/Advance Receipt Number"""
    wb["cdnura"]["D4"].style = "column_header"
    wb["cdnura"]["E4"] = """Original Invoice/Advance Receipt date"""
    wb["cdnura"]["E4"].style = "column_header"
    wb["cdnura"]["F1"] = """Revised details"""
    wb["cdnura"]["F1"].style = "column_header_bold_blue"
    wb["cdnura"]["F4"] = """Revised Note/Refund Voucher Number"""
    wb["cdnura"]["F4"].style = "column_header_blue"
    wb["cdnura"]["G4"] = """Revised Note/Refund Voucher date"""
    wb["cdnura"]["G4"].style = "column_header_blue"
    wb["cdnura"]["H4"] = """Document Type"""
    wb["cdnura"]["H4"].style = "column_header_blue"
    wb["cdnura"]["I4"] = """Supply Type"""
    wb["cdnura"]["I4"].style = "column_header_blue"
    wb["cdnura"]["J2"] = """Total Note Value"""
    wb["cdnura"]["J2"].style = "summary_right"
    wb["cdnura"]["J3"] = """=SUMPRODUCT(1/COUNTIF(B5:B1001,B5:B1001&""),J5:J1001)"""
    wb["cdnura"]["J3"].style = "sum_style_right"
    wb["cdnura"]["J4"] = """Note/Refund Voucher Value"""
    wb["cdnura"]["J4"].style = "column_header_blue"
    wb["cdnura"]["K4"] = """Applicable % of Tax Rate"""
    wb["cdnura"]["K4"].style = "column_header_blue"
    wb["cdnura"]["L4"] = """Rate"""
    wb["cdnura"]["L4"].style = "column_header_blue"
    wb["cdnura"]["M2"] = """Total Taxable Value"""
    wb["cdnura"]["M2"].style = "summary_right"
    wb["cdnura"]["M3"] = """=SUM(M5:M1001)"""
    wb["cdnura"]["M3"].style = "sum_style_right"
    wb["cdnura"]["M4"] = """Taxable Value"""
    wb["cdnura"]["M4"].style = "column_header_blue"
    wb["cdnura"]["N2"] = """Total Cess"""
    wb["cdnura"]["N2"].style = "summary_right"
    wb["cdnura"]["N3"] = """=SUMIF(H5:H1001,"D",N5:N1001)-(SUMIF(H5:H1001,"C",N5:N1001)+SUMIF(H5:H1001,"R",N5:N1001))"""
    wb["cdnura"]["N3"].style = "sum_style_right"
    wb["cdnura"]["N4"] = """Cess Amount"""
    wb["cdnura"]["N4"].style = "column_header_blue"
    wb["cdnura"]["O1"] = """HELP"""
    wb["cdnura"]["O1"].style = "help_style"
    wb["cdnura"]["O4"] = """Pre GST"""
    wb["cdnura"]["O4"].style = "column_header_blue"


def exp_sheet(wb):
    wb.create_sheet(title="exp")

    style_range_of_cells(wb["exp"], "A1:A1", "summary")
    style_range_of_cells(wb["exp"], "A2:K2", "summary")

    for col in range(1, 11):
        i = get_column_letter(col)
        wb["exp"].column_dimensions[i].width = 25

    wb["exp"].row_dimensions[4].height = 30.0

    wb["exp"]["A1"] = """Summary For EXP(6)"""
    wb["exp"]["A4"] = """Export Type"""
    wb["exp"]["A4"].style = "column_header"
    wb["exp"]["B2"] = """No. of Invoices"""
    wb["exp"]["B3"] = """=SUMPRODUCT((B5:B20004<>"")/COUNTIF(B5:B20004,B5:B20004&""))"""
    wb["exp"]["B3"].style = "count"
    wb["exp"]["B4"] = """Invoice Number"""
    wb["exp"]["B4"].style = "column_header"
    wb["exp"]["C4"] = """Invoice date"""
    wb["exp"]["C4"].style = "column_header"
    wb["exp"]["D2"] = """Total Invoice Value"""
    wb["exp"]["D2"].style = "summary_right"
    wb["exp"]["D3"] = """=SUMPRODUCT(1/COUNTIF(B5:B20004,B5:B20004&""),D5:D20004)"""
    wb["exp"]["D3"].style = "sum_style"
    wb["exp"]["D4"] = """Invoice Value"""
    wb["exp"]["D4"].style = "column_header"
    wb["exp"]["E4"] = """Port Code"""
    wb["exp"]["E4"].style = "column_header"
    wb["exp"]["F2"] = """No. of Shipping Bill"""
    wb["exp"]["F3"] = """=SUMPRODUCT((F5:F20004<>"")/COUNTIF(F5:F20004,F5:F20004&""))"""
    wb["exp"]["F3"].style = "count"
    wb["exp"]["F4"] = """Shipping Bill Number"""
    wb["exp"]["F4"].style = "column_header"
    wb["exp"]["G4"] = """Shipping Bill Date"""
    wb["exp"]["G4"].style = "column_header"
    wb["exp"]["H4"] = """Applicable % of Tax Rate"""
    wb["exp"]["H4"].style = "column_header"
    wb["exp"]["I4"] = """Rate"""
    wb["exp"]["I4"].style = "column_header"
    wb["exp"]["J4"] = """Taxable Value"""
    wb["exp"]["J4"].style = "column_header"
    wb["exp"]["K1"] = """HELP"""
    wb["exp"]["K1"].style = "help_style"
    wb["exp"]["K2"] = """Total Taxable Value"""
    wb["exp"]["K2"].style = "summary_right"
    wb["exp"]["K3"] = """=SUM(J5:J20004)"""
    wb["exp"]["K3"].style = "sum_style_right"
    wb["exp"]["K4"] = """Cess Amount"""
    wb["exp"]["K4"].style = "column_header"


def expa_sheet(wb):
    wb.create_sheet(title="expa")

    style_range_of_cells(wb["expa"], "A1:A1", "summary")
    style_range_of_cells(wb["expa"], "A2:M2", "summary")

    wb["expa"].merge_cells("B1:C1")
    wb["expa"].merge_cells("D1:L1")

    for col in range(1, 14):
        i = get_column_letter(col)
        wb["expa"].column_dimensions[i].width = 25

    wb["expa"].row_dimensions[4].height = 30.0

    wb["expa"]["A1"] = """Summary For EXPA"""
    wb["expa"]["A4"] = """Export Type"""
    wb["expa"]["A4"].style = "column_header"
    wb["expa"]["B1"] = """Original details """
    wb["expa"]["B1"].style = "column_header_bold"
    wb["expa"]["B2"] = """No. of Invoices"""
    wb["expa"]["B3"] = """=SUMPRODUCT((B5:B2004<>"")/COUNTIF(B5:B2004,B5:B2004&""))"""
    wb["expa"]["B3"].style = "count"
    wb["expa"]["B4"] = """Original Invoice Number"""
    wb["expa"]["B4"].style = "column_header"
    wb["expa"]["C4"] = """Original Invoice date"""
    wb["expa"]["C4"].style = "column_header"
    wb["expa"]["D1"] = """Revised details"""
    wb["expa"]["D1"].style = "column_header_bold_blue"
    wb["expa"]["D4"] = """Revised Invoice Number"""
    wb["expa"]["D4"].style = "column_header_blue"
    wb["expa"]["E4"] = """Revised Invoice date"""
    wb["expa"]["E4"].style = "column_header_blue"
    wb["expa"]["F2"] = """Total Invoice Value"""
    wb["expa"]["F2"].style = "summary_right"
    wb["expa"]["F3"] = """=SUMPRODUCT(1/COUNTIF(B5:B2004,B5:B2004&""),F5:F2004)"""
    wb["expa"]["F3"].style = "sum_style_right"
    wb["expa"]["F4"] = """Invoice Value"""
    wb["expa"]["F4"].style = "column_header_blue"
    wb["expa"]["G4"] = """Port Code"""
    wb["expa"]["G4"].style = "column_header_blue"
    wb["expa"]["H2"] = """No. of Shipping Bill"""
    wb["expa"]["H3"] = """=SUMPRODUCT((H5:H2004<>"")/COUNTIF(H5:H2004,H5:H2004&""))"""
    wb["expa"]["H3"].style = "count"
    wb["expa"]["H4"] = """Shipping Bill Number"""
    wb["expa"]["H4"].style = "column_header_blue"
    wb["expa"]["I4"] = """Shipping Bill Date"""
    wb["expa"]["I4"].style = "column_header_blue"
    wb["expa"]["J4"] = """Applicable % of Tax Rate"""
    wb["expa"]["J4"].style = "column_header_blue"
    wb["expa"]["K4"] = """Rate"""
    wb["expa"]["K4"].style = "column_header_blue"
    wb["expa"]["L4"] = """Taxable Value"""
    wb["expa"]["L4"].style = "column_header_blue"
    wb["expa"]["M1"] = """HELP"""
    wb["expa"]["M1"].style = "help_style"
    wb["expa"]["M2"] = """Total Taxable Value"""
    wb["expa"]["M2"].style = "summary_right"
    wb["expa"]["M3"] = """=SUM(L5:L2004)"""
    wb["expa"]["M3"].style = "sum_style_right"
    wb["expa"]["M4"] = """Cess Amount"""
    wb["expa"]["M4"].style = "column_header_blue"


def at_sheet(wb):
    wb.create_sheet(title="at")

    style_range_of_cells(wb["at"], "A1:A1", "summary")
    style_range_of_cells(wb["at"], "A2:E2", "summary")

    for col in range(1, 7):
        i = get_column_letter(col)
        wb["at"].column_dimensions[i].width = 25

    wb["at"].row_dimensions[1].height = 28.5
    wb["at"].row_dimensions[4].height = 30.0

    wb["at"]["A1"] = """Summary For Advance Received (11B) """
    wb["at"]["A4"] = """Place Of Supply"""
    wb["at"]["A4"].style = "column_header"
    wb["at"]["B4"] = """Applicable % of Tax Rate"""
    wb["at"]["B4"].style = "column_header"
    wb["at"]["C4"] = """Rate"""
    wb["at"]["C4"].style = "column_header"
    wb["at"]["D2"] = """Total Advance Received"""
    wb["at"]["D2"].style = "summary_right"
    wb["at"]["D3"] = """=SUM(D5:D4004)"""
    wb["at"]["D3"].style = "sum_style_right"
    wb["at"]["D4"] = """Gross Advance Received"""
    wb["at"]["D4"].style = "column_header"
    wb["at"]["E1"] = """HELP"""
    wb["at"]["E1"].style = "help_style"
    wb["at"]["E2"] = """Total Cess"""
    wb["at"]["E2"].style = "summary_right"
    wb["at"]["E3"] = """=SUM(E5:E4004)"""
    wb["at"]["E3"].style = "sum_style_right"
    wb["at"]["E4"] = """Cess Amount"""
    wb["at"]["E4"].style = "column_header"


def ata_sheet(wb):
    wb.create_sheet(title="ata")

    style_range_of_cells(wb["ata"], "A1:A1", "summary")
    style_range_of_cells(wb["ata"], "A2:G2", "summary")

    wb["ata"].merge_cells("B1:C1")
    wb["ata"].merge_cells("D1:F1")

    for col in range(1, 8):
        i = get_column_letter(col)
        wb["ata"].column_dimensions[i].width = 25

    wb["ata"].row_dimensions[1].height = 45.0
    wb["ata"].row_dimensions[4].height = 30.0

    wb["ata"]["A1"] = """Summary For Amended Tax Liability(Advance Received) """
    wb["ata"]["A4"] = """Financial Year"""
    wb["ata"]["A4"].style = "column_header"
    wb["ata"]["B1"] = """Original details """
    wb["ata"]["B1"].style = "column_header_bold"
    wb["ata"]["B4"] = """Original Month"""
    wb["ata"]["B4"].style = "column_header"
    wb["ata"]["C4"] = """Original Place Of Supply"""
    wb["ata"]["C4"].style = "column_header"
    wb["ata"]["D1"] = """Revised details"""
    wb["ata"]["D1"].style = "column_header_bold_blue"
    wb["ata"]["D4"] = """Applicable % of Tax Rate"""
    wb["ata"]["D4"].style = "column_header_blue"
    wb["ata"]["E4"] = """Rate"""
    wb["ata"]["E4"].style = "column_header_blue"
    wb["ata"]["F2"] = """Total Advance Received"""
    wb["ata"]["F2"].style = "summary_right"
    wb["ata"]["F3"] = """=SUM(F5:F20004)"""
    wb["ata"]["F3"].style = "sum_style_right"
    wb["ata"]["F4"] = """Gross Advance Received"""
    wb["ata"]["F4"].style = "column_header_blue"
    wb["ata"]["G1"] = """HELP"""
    wb["ata"]["G1"].style = "help_style"
    wb["ata"]["G2"] = """Total Cess"""
    wb["ata"]["G2"].style = "summary_right"
    wb["ata"]["G3"] = """=SUM(G5:G20004)"""
    wb["ata"]["G3"].style = "sum_style_right"
    wb["ata"]["G4"] = """Cess Amount"""
    wb["ata"]["G4"].style = "column_header_blue"


def atadj_sheet(wb):
    wb.create_sheet(title="atadj")

    style_range_of_cells(wb["atadj"], "A1:A1", "summary")
    style_range_of_cells(wb["atadj"], "A2:E2", "summary")

    for col in range(1, 5):
        i = get_column_letter(col)
        wb["atadj"].column_dimensions[i].width = 25

    wb["atadj"].row_dimensions[1].height = 28.5
    wb["atadj"].row_dimensions[4].height = 30.0

    wb["atadj"]["A1"] = """Summary For Advance Adjusted (11B) """
    wb["atadj"]["A4"] = """Place Of Supply"""
    wb["atadj"]["A4"].style = "column_header"
    wb["atadj"]["B4"] = """Applicable % of Tax Rate"""
    wb["atadj"]["B4"].style = "column_header"
    wb["atadj"]["C4"] = """Rate"""
    wb["atadj"]["C4"].style = "column_header"
    wb["atadj"]["D2"] = """Total Advance Adjusted"""
    wb["atadj"]["D2"].style = "summary_right"
    wb["atadj"]["D3"] = """=SUM(D5:D4004)"""
    wb["atadj"]["D3"].style = "sum_style_right"
    wb["atadj"]["D4"] = """Gross Advance Adjusted"""
    wb["atadj"]["D4"].style = "column_header"
    wb["atadj"]["E1"] = """HELP"""
    wb["atadj"]["E1"].style = "help_style"
    wb["atadj"]["E2"] = """Total Cess"""
    wb["atadj"]["E2"].style = "summary_right"
    wb["atadj"]["E3"] = """=SUM(E5:E4004)"""
    wb["atadj"]["E3"].style = "sum_style_right"
    wb["atadj"]["E4"] = """Cess Amount"""
    wb["atadj"]["E4"].style = "column_header"


def atadja_sheet(wb):
    wb.create_sheet(title="atadja")
    style_range_of_cells(wb["atadja"], "A1:A1", "summary")
    style_range_of_cells(wb["atadja"], "A2:G2", "summary")

    wb["atadja"].merge_cells("B1:C1")
    wb["atadja"].merge_cells("D1:F1")

    for col in range(1, 8):
        i = get_column_letter(col)
        wb["atadja"].column_dimensions[i].width = 25

    wb["atadja"].row_dimensions[1].height = 36.75
    wb["atadja"].row_dimensions[4].height = 30.0

    wb["atadja"]["A1"] = """Summary For Amendement Of Adjustment Advances"""
    wb["atadja"]["A4"] = """Financial Year"""
    wb["atadja"]["A4"].style = "column_header"
    wb["atadja"]["B1"] = """Original details """
    wb["atadja"]["B1"].style = "column_header_bold"
    wb["atadja"]["B4"] = """Original Month"""
    wb["atadja"]["B4"].style = "column_header"
    wb["atadja"]["C4"] = """Original Place Of Supply"""
    wb["atadja"]["C4"].style = "column_header"
    wb["atadja"]["D1"] = """Revised details"""
    wb["atadja"]["D1"].style = "column_header_bold_blue"
    wb["atadja"]["D4"] = """Applicable % of Tax Rate"""
    wb["atadja"]["D4"].style = "column_header_blue"
    wb["atadja"]["E4"] = """Rate"""
    wb["atadja"]["E4"].style = "column_header_blue"
    wb["atadja"]["F2"] = """Total Advance Adjusted"""
    wb["atadja"]["F2"].style = "summary_right"
    wb["atadja"]["F3"] = """=SUM(F5:F4004)"""
    wb["atadja"]["F3"].style = "sum_style_right"
    wb["atadja"]["F4"] = """Gross Advance Adjusted"""
    wb["atadja"]["F4"].style = "column_header_blue"
    wb["atadja"]["G1"] = """HELP"""
    wb["atadja"]["G1"].style = "help_style"
    wb["atadja"]["G2"] = """Total Cess"""
    wb["atadja"]["G2"].style = "summary_right"
    wb["atadja"]["G3"] = """=SUM(G5:G4004)"""
    wb["atadja"]["G3"].style = "sum_style_right"
    wb["atadja"]["G4"] = """Cess Amount"""
    wb["atadja"]["G4"].style = "column_header_blue"


def exemp_sheet(wb):
    wb.create_sheet(title="exemp")

    style_range_of_cells(wb["exemp"], "A1:A1", "summary")
    style_range_of_cells(wb["exemp"], "A2:D2", "summary")

    for col in range(1, 5):
        i = get_column_letter(col)
        wb["exemp"].column_dimensions[i].width = 25

    wb["exemp"].row_dimensions[1].height = 28.5
    wb["exemp"].row_dimensions[4].height = 30.0

    wb["exemp"]["A1"] = """Summary For Nil rated, exempted and non GST outward supplies (8)"""
    wb["exemp"]["A4"] = """Description"""
    wb["exemp"]["A4"].style = "column_header"
    wb["exemp"]["B2"] = """Total Nil Rated Supplies"""
    wb["exemp"]["B2"].style = "summary_right"
    wb["exemp"]["B3"] = """=SUM(B5:B8)"""
    wb["exemp"]["B3"].style = "sum_style_right"
    wb["exemp"]["B4"] = """Nil Rated Supplies"""
    wb["exemp"]["B4"].style = "column_header"
    wb["exemp"]["C2"] = """Total Exempted Supplies"""
    wb["exemp"]["C2"].style = "summary_right"
    wb["exemp"]["C3"] = """=SUM(C5:C8)"""
    wb["exemp"]["C3"].style = "sum_style_right"
    wb["exemp"]["C4"] = """Exempted(other than nil rated/non GST supply)"""
    wb["exemp"]["C4"].style = "column_header"
    wb["exemp"]["D1"] = """HELP"""
    wb["exemp"]["D1"].style = "help_style"
    wb["exemp"]["D2"] = """Total Non-GST Supplies"""
    wb["exemp"]["D2"].style = "summary_right"
    wb["exemp"]["D3"] = """=SUM(D5:D8)"""
    wb["exemp"]["D3"].style = "sum_style_right"
    wb["exemp"]["D4"] = """Non-GST Supplies"""
    wb["exemp"]["D4"].style = "column_header"


def hsn_sheet(wb):
    wb.create_sheet(title="hsn")
    sheet=wb.active
    n = sheet.max_row
    print n
    style_range_of_cells(wb["hsn"], "A1:A1", "summary")
    style_range_of_cells(wb["hsn"], "A2:J2", "summary")
    
    for col in range(1, 11):
        i = get_column_letter(col)
        wb["hsn"].column_dimensions[i].width = 25
    
    wb["hsn"].row_dimensions[4].height = 30

    wb["hsn"]["A1"] = """Summary For HSN(12)"""
    wb["hsn"]["A2"] = """No. of HSN"""
    wb["hsn"]["A3"] = """=SUMPRODUCT((A5:A2004<>"")/COUNTIF(A5:A2004,A5:A2004&""))"""
    wb["hsn"]["A3"].style = "count"
    wb["hsn"]["A4"] = """HSN"""
    wb["hsn"]["A4"].style = "column_header"
    wb["hsn"]["B4"] = """Description"""
    wb["hsn"]["B4"].style = "column_header"
    wb["hsn"]["C4"] = """UQC"""
    wb["hsn"]["C4"].style = "column_header"
    wb["hsn"]["D4"] = """Total Quantity"""
    wb["hsn"]["D4"].style = "column_header"
    wb["hsn"]["E2"] = """Total Value"""
    wb["hsn"]["E2"].style = "summary_right"
    wb["hsn"]["E3"] = """=SUM(E5:E2004)"""
    wb["hsn"]["E3"].style = "sum_style_right"
    wb["hsn"]["E4"] = """Total Value"""
    wb["hsn"]["E4"].style = "column_header"
    wb["hsn"]["F2"] = """Total Taxable Value"""
    wb["hsn"]["F2"].style = "summary_right"
    wb["hsn"]["F3"] = """=SUM(F5:F2004)"""
    wb["hsn"]["F3"].style = "sum_style_right"
    wb["hsn"]["F4"] = """Taxable Value"""
    wb["hsn"]["F4"].style = "column_header"
    wb["hsn"]["G2"] = """Total Integrated Tax"""
    wb["hsn"]["G2"].style = "summary_right"
    wb["hsn"]["G3"] = """=SUM(G5:G2004)"""
    wb["hsn"]["G3"].style = "sum_style_right"
    wb["hsn"]["G4"] = """Integrated Tax Amount"""
    wb["hsn"]["G4"].style = "column_header"
    wb["hsn"]["H2"] = """Total Central Tax"""
    wb["hsn"]["H2"].style = "summary_right"
    wb["hsn"]["H3"] = """=SUM(H5:H2004)"""
    wb["hsn"]["H3"].style = "sum_style_right"
    wb["hsn"]["H4"] = """Central Tax Amount"""
    wb["hsn"]["H4"].style = "column_header"
    wb["hsn"]["I2"] = """Total State/UT Tax"""
    wb["hsn"]["I2"].style = "summary_right"
    wb["hsn"]["I3"] = """=SUM(I5:I2004)"""
    wb["hsn"]["I3"].style = "sum_style_right"
    wb["hsn"]["I4"] = """State/UT Tax Amount"""
    wb["hsn"]["I4"].style = "column_header"
    wb["hsn"]["J1"] = """HELP"""
    wb["hsn"]["J1"].style = "help_style"
    wb["hsn"]["J2"] = """Total Cess"""
    wb["hsn"]["J2"].style = "summary_right"
    wb["hsn"]["J3"] = """=SUM(J5:J2004)"""
    wb["hsn"]["J3"].style = "sum_style_right"
    wb["hsn"]["J4"] = """Cess Amount"""
    wb["hsn"]["J4"].style = "column_header"


def docs_sheet(wb):
    wb.create_sheet(title="docs")

    style_range_of_cells(wb["docs"], "A1:A1", "summary")
    style_range_of_cells(wb["docs"], "A2:E2", "summary")

    for col in range(1, 6):
        i = get_column_letter(col)
        wb["docs"].column_dimensions[i].width = 25

    wb["docs"].row_dimensions[4].height = 30

    wb["docs"]["A1"] = """Summary of documents issued during the tax period (13)"""
    wb["docs"]["A4"] = """Nature of Document"""
    wb["docs"]["A4"].style = "column_header"
    wb["docs"]["B4"] = """Sr. No. From"""
    wb["docs"]["B4"].style = "column_header"
    wb["docs"]["C4"] = """Sr. No. To"""
    wb["docs"]["C4"].style = "column_header"
    wb["docs"]["D2"] = """Total Number"""
    wb["docs"]["D2"].style = "summary_right"
    wb["docs"]["D3"] = """=SUM(D5:D1004)"""
    wb["docs"]["D3"].style = "sum_style_right"
    wb["docs"]["D4"] = """Total Number"""
    wb["docs"]["D4"].style = "column_header"
    wb["docs"]["E1"] = """HELP"""
    wb["docs"]["E1"].style = "help_style"
    wb["docs"]["E2"] = """Total Cancelled"""
    wb["docs"]["E2"].style = "summary_right"
    wb["docs"]["E3"] = """=SUM(E5:E1004)"""
    wb["docs"]["E3"].style = "sum_style_right"
    wb["docs"]["E4"] = """Cancelled"""
    wb["docs"]["E4"].style = "column_header"


def fill_data(result, wb):
    for entry in result["gkdata"]["b2b"]:

        row = (
            entry["gstin"], entry["receiver"],
            entry["invoice_number"], to_datetime(entry["invoice_date"]),
            float_or_none(entry["invoice_value"]), entry["place_of_supply"],
            entry["reverse_charge"], entry["applicable_tax_rate"],
            entry["invoice_type"], entry["ecommerce_gstin"],
            float_or_none(entry["rate"]), float_or_none(entry["taxable_value"]),
            float_or_none(entry["cess"])
        )

        wb["b2b"].append(row)

        set_alignment(wb["b2b"], ["E5", "K5", "L5"], "right")
        set_number_format(wb["b2b"], ["E5", "K5", "L5"], "0.00")
        set_number_format(wb["b2b"], ["D5"], "dd-mmm-yy")

    for entry in result["gkdata"]["b2cl"]:

        row = (
            entry["invoice_number"], to_datetime(entry["invoice_date"]),
            float_or_none(entry["invoice_value"]), entry["place_of_supply"],
            float_or_none(entry["applicable_tax_rate"]), float_or_none(entry["rate"]),
            float_or_none(entry["taxable_value"]), float_or_none(entry["cess"]),
            entry["ecommerce_gstin"], entry["sale_from_bonded_wh"]
        )

        wb["b2cl"].append(row)

    set_alignment(wb["b2cl"], ["C5", "F5", "G5", "H5"], "right")
    set_number_format(wb["b2cl"], ["C5", "F5", "G5", "H5", "B5"], "0.00")
    set_number_format(wb["b2cl"], ["B5"], "dd-mmm-yy")

    for entry in result["gkdata"]["b2cs"]:

        row = (
            entry["type"], entry["place_of_supply"],
            float_or_none(entry["applicable_tax_rate"]), float_or_none(entry["rate"]),
            float_or_none(entry["taxable_value"]), float_or_none(entry["cess"]),
            entry["ecommerce_gstin"]
        )

        wb["b2cs"].append(row)

    set_alignment(wb["b2cs"], ["D5", "E5", "F5"], "right")
    set_number_format(wb["b2cs"], ["D5", "E5", "F5"], "0.00")

    for entry in result["gkdata"]["cdnr"]:

        row = (
            entry["gstin"], entry["receiver"],
            entry["invoice_number"], to_datetime(entry["invoice_date"]),
            entry["voucher_number"], to_datetime(entry["voucher_date"]),
            entry["document_type"], entry["place_of_supply"],
            float_or_none(entry["refund_voucher_value"]), float_or_none(entry["applicable_tax_rate"]),
            float_or_none(entry["rate"]), float_or_none(entry["taxable_value"]),
            float_or_none(entry["cess"]), entry["pregst"]
        )

        wb["cdnr"].append(row)

    set_alignment(wb["cdnr"], ["I5", "K5", "L5", "M5"], "right")
    set_number_format(wb["cdnr"], ["I5", "K5", "L5", "M5"], "0.00")
    set_number_format(wb["cdnr"], ["D5", "F5"], "dd-mmm-yy")

    for entry in result["gkdata"]["cdnur"]:

        row = (
            entry["ur_type"], entry["voucher_number"],
            to_datetime(entry["voucher_date"]), entry["document_type"],
            entry["invoice_number"], to_datetime(entry["invoice_date"]),
            entry["place_of_supply"], float_or_none(entry["refund_voucher_value"]),
            float_or_none(entry["applicable_tax_rate"]), float_or_none(entry["rate"]),
            float_or_none(entry["taxable_value"]), float_or_none(entry["cess"]),
            entry["pregst"], entry["supply_type"]
        )

        wb["cdnur"].append(row)

    set_number_format(wb["cdnur"], ["H5", "J5", "K5", "L5"], "0.00")
    set_number_format(wb["cdnur"], ["C5", "F5"], "dd-mmm-yy")

    l = len(result["gkdata"]["hsn1"]) -1
    del result["gkdata"]["hsn1"][l]
    for entry in result["gkdata"]["hsn1"]:
        
        row = (
            entry["hsnsac"],
             entry["prodctname"],
            entry["uqc"],
            float_or_none(entry["qty"]),float_or_none(entry["totalvalue"]),
            float_or_none(entry["taxableamt"]), float_or_none(entry["IGSTamt"]),
            float_or_none(entry["SGSTamt"]),float_or_none(entry["SGSTamt"]), float_or_none(entry["CESSamt"])
        )

        wb["hsn"].append(row)

    set_number_format(wb["hsn"], ["D5","E5","F5","G5","H5", "I5","J5"], "0.00")
    

        
def gst_r1_template(result):
    wb = openpyxl.Workbook()
    wb.active.title = "Help Instruction"
    create_styles(wb)
    b2b_sheet(wb)
    b2ba_sheet(wb)
    b2cl_sheet(wb)
    b2cla_sheet(wb)
    b2cs_sheet(wb)
    b2csa_sheet(wb)
    cdnr_sheet(wb)
    cdnra_sheet(wb)
    cdnur_sheet(wb)
    cdnura_sheet(wb)
    exp_sheet(wb)
    expa_sheet(wb)
    at_sheet(wb)
    ata_sheet(wb)
    atadj_sheet(wb)
    atadja_sheet(wb)
    exemp_sheet(wb)
    hsn_sheet(wb)
    docs_sheet(wb)
    wb.create_sheet(title="master")

    fill_data(result, wb)

    return wb
