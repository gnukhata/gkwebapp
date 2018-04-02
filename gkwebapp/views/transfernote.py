
"""
        Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
        Copyright (C) 2017, 2018 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

          This file is part of GNUKhata:A modular,robust and Free Accounting System.

          GNUKhata is Free Software; you can redistribute it and/or modify
          it under the terms of the GNU Affero General Public License as
          published by the Free Software Foundation; either version 3 of
          the License, or (at your option) any later version.

          GNUKhata is distributed in the hope that it will be useful, but
          WITHOUT ANY WARRANTY; without even the implied warranty of
          MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
          GNU Affero General Public License for more details.

          You should have received a copy of the GNU Affero General Public
          License along with GNUKhata (COPYING); if not, write to the
          Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
          Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


        Contributors:
        "Krishnakant Mane" <kk@dff.org.in>
        "Arun Kelkar" <arunkelkar@dff.org.in>
        "Ishan Masdekar " <imasdekar@dff.org.in>
        "Navin Karkera" <navin@dff.org.in>
        "Sachin Patil" <sachpatil@openmailbox.org>
        'Prajkta Patkar' <prajakta.dff.org.in>
        "Mohd. Talha Pawaty" <mtalha456@gmail.com>
        "Abhijith Balan" <abhijithb21@openmailbox.org>
               
"""
from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from pyramid.response import Response
import os
import openpyxl
from openpyxl.styles import Font, Alignment

@view_config(route_name="transfernotes",renderer="gkwebapp:templates/transfernote.jinja2")
def showtransfernote(request):
                return {"status":True}

@view_config(route_name="transfernotes",request_param="action=showadd",renderer="gkwebapp:templates/createtransfernote.jinja2")
def showcreatetransfernote(request):
                header={"gktoken":request.headers["gktoken"]}
                products = requests.get("http://127.0.0.1:6543/products?invdc=4", headers=header)
                godowns = requests.get("http://127.0.0.1:6543/godown", headers=header)
                fromgodowns = requests.get("http://127.0.0.1:6543/godown?value=1", headers=header)
                return {"products": products.json()["gkresult"],"godowns":godowns.json()["gkresult"],"fromgodowns":fromgodowns.json()["gkresult"]}



@view_config(route_name="transfernotes",request_param="action=showreceived",renderer="gkwebapp:templates/receivedtransfernote.jinja2")
def showreceivedtransfernote(request):
                header={"gktoken":request.headers["gktoken"]}
                transfernote = requests.get("http://127.0.0.1:6543/transfernote?tn=all",headers=header)
                return {"transfernote":transfernote.json()["gkresult"],"nooftransfernotes":len(transfernote.json()["gkresult"])}

@view_config(route_name="transfernotes",request_param="action=viewlist",renderer="gkwebapp:templates/viewlistoftransfernotes.jinja2")
def viewlistoftransfernotes(request):
                header={"gktoken":request.headers["gktoken"]}
                result = requests.get("http://127.0.0.1:6543/godown", headers=header)
                goddata=[]
                for record in result.json()["gkresult"]:
                    gdata= {"godownname":str(record["goname"]),"godownid":str(record["goid"]),"godownaddress": str(record["goaddr"])}
                    goddata.append(gdata)
                transfernotes = requests.get("http://127.0.0.1:6543/transfernote?type=all", headers=header)
                return {"gkresult":goddata, "numberoftransfernotes":len(transfernotes.json()["gkresult"])}

@view_config(route_name="transfernotes",request_param="action=printlist",renderer="gkwebapp:templates/printlistoftransfernotes.jinja2")
def printlistoftransfernotes(request):
                header={"gktoken":request.headers["gktoken"]}
                startDate =str(request.params["startdate"])
                endDate =str(request.params["enddate"])
                godownname = ""
                godownaddress = ""
                goid = 0
                if request.params.has_key("goid"):
                    goid = int(request.params["goid"])
                    transfernotes = requests.get("http://127.0.0.1:6543/transfernote?type=list&startdate=%s&enddate=%s&goid=%d"%(startDate, endDate, goid),headers=header)
                    godown = requests.get("http://127.0.0.1:6543/godown?qty=single&goid=%d"%(int(request.params["goid"])), headers=header)
                    godownname = godown.json()["gkresult"]["goname"]
                    godownaddress = godown.json()["gkresult"]["goaddr"]
                else:
                    transfernotes = requests.get("http://127.0.0.1:6543/transfernote?type=list&startdate=%s&enddate=%s"%(startDate, endDate),headers=header)
                return {"transfernotes":transfernotes.json()["gkresult"], "startdate":startDate, "enddate":endDate, "godownname":godownname, "godownaddress":godownaddress, "goid":goid}


'''
This function returns a spreadsheet form of List of Transfer Notes Report.
The spreadsheet in XLSX format is generated by the frontend.
'''
@view_config(route_name="transfernotes",request_param="action=generatespreadsheet", renderer="")
def listoftransfernotesspreadsheet(request):
    try :
        header={"gktoken":request.headers["gktoken"]}
        fystart = str(request.params["fystart"]);
        fyend = str(request.params["fyend"]);
        orgname = str(request.params["orgname"])
        orgname += " (FY: " + fystart+" to "+fyend +")"
        startDate =str(request.params["startdate"])
        endDate =str(request.params["enddate"])
        godownname = ""
        godownaddress = ""
        goid = 0
        transfernotewb = openpyxl.Workbook()
        sheet = transfernotewb.active
        sheet.title = "List of Transfer Notes"
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 14
        sheet.column_dimensions['D'].width = 24
        sheet.column_dimensions['E'].width = 24
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 16
        sheet.column_dimensions['H'].width = 14
        sheet.merge_cells('A1:H2')
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        # Organisation name and financial year are displayed.
        sheet['A1'] = orgname
        sheet.merge_cells('A3:H3')
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A3'] = 'List of Transfer Notes'
        sheet.merge_cells('A4:H4')
        sheet['A4'] = 'Period: ' + startDate + ' to ' + endDate
        sheet['A4'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A4'].alignment = Alignment(horizontal = 'center', vertical='center')
        titlerow = 5
        #If an id of a godown is received it will give all transfernotes involving that godown with godownname and godownaddress.
        if request.params.has_key("goid"):
            goid = int(request.params["goid"])
            transfernotes = requests.get("http://127.0.0.1:6543/transfernote?type=list&startdate=%s&enddate=%s&goid=%d"%(startDate, endDate, goid),headers=header)
            godown = requests.get("http://127.0.0.1:6543/godown?qty=single&goid=%d"%(int(request.params["goid"])), headers=header)
            godownname = godown.json()["gkresult"]["goname"]
            godownaddress = godown.json()["gkresult"]["goaddr"]
            nameofgodown = "Name of Godown: "+godownname+" Godown Address: "+godownaddress
            sheet.merge_cells('A5:H5')
            sheet['A5'] = nameofgodown
            sheet['A5'].font = Font(name='Liberation Serif',size='14',bold=True)
            sheet['A5'].alignment = Alignment(horizontal = 'center', vertical='center')
            titlerow =6
        else:
            transfernotes = requests.get("http://127.0.0.1:6543/transfernote?type=list&startdate=%s&enddate=%s"%(startDate, endDate),headers=header)
        transfernotes = transfernotes.json()["gkresult"]
        sheet['A'+str(titlerow)] = 'Sr. No.'
        sheet['B'+str(titlerow)] = 'TN No.'
        sheet['C'+str(titlerow)] = 'Date'
        sheet['D'+str(titlerow)] = 'Dispatch From'
        sheet['E'+str(titlerow)] = 'To be Delivered At'
        sheet['F'+str(titlerow)] = 'Products'
        sheet['G'+str(titlerow)] = 'Quantity'
        sheet['H'+str(titlerow)] = 'Status'
        sheet['A'+str(titlerow)].alignment = Alignment(horizontal='center')
        sheet['B'+str(titlerow)].alignment = Alignment(horizontal='center')
        sheet['C'+str(titlerow)].alignment = Alignment(horizontal='center')
        sheet['D'+str(titlerow)].alignment = Alignment(horizontal='center')
        sheet['E'+str(titlerow)].alignment = Alignment(horizontal='center')
        sheet['F'+str(titlerow)].alignment = Alignment(horizontal='center')
        sheet['G'+str(titlerow)].alignment = Alignment(horizontal='right')
        sheet['H'+str(titlerow)].alignment = Alignment(horizontal='center')
        sheet['A'+str(titlerow)].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['B'+str(titlerow)].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['C'+str(titlerow)].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['D'+str(titlerow)].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['E'+str(titlerow)].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['F'+str(titlerow)].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['G'+str(titlerow)].font = Font(name='Liberation Serif',size=12,bold=True)
        sheet['H'+str(titlerow)].font = Font(name='Liberation Serif',size=12,bold=True)
        row = titlerow + 1
        # Looping each dictionaries in list transfernotes to store data in cells and apply styles.
        for transfernote in transfernotes:
            sheet['A'+str(row)] = transfernote["srno"]
            sheet['A'+str(row)].alignment = Alignment(horizontal='center')
            sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['B'+str(row)] = transfernote["transfernoteno"]
            sheet['B'+str(row)].alignment = Alignment(horizontal='center')
            sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['C'+str(row)] = transfernote["transfernotedate"]
            sheet['C'+str(row)].alignment = Alignment(horizontal='center')
            sheet['C'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['D'+str(row)] = transfernote["fromgodown"]
            sheet['D'+str(row)].alignment = Alignment(horizontal='left')
            sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['E'+str(row)] = transfernote["togodown"]
            sheet['E'+str(row)].alignment = Alignment(horizontal='left')
            sheet['E'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            subrow = row
            for productqty in transfernote["productqty"]:
                sheet['F'+str(subrow)] = productqty["productdesc"]
                sheet['F'+str(subrow)].alignment = Alignment(horizontal='left')
                sheet['F'+str(subrow)].font = Font(name='Liberation Serif', size='12', bold=False)
                sheet['G'+str(subrow)] = productqty["quantity"] + " " + productqty["uom"]
                sheet['G'+str(subrow)].alignment = Alignment(horizontal='right')
                sheet['G'+str(subrow)].font = Font(name='Liberation Serif', size='12', bold=False)
                subrow +=1
            if transfernote["receivedflag"]:
                sheet['H'+str(row)] = "Received"
                sheet['H'+str(row)].alignment = Alignment(horizontal='center')
                sheet['H'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            else:
                sheet['H'+str(row)] = "Pending"
                sheet['H'+str(row)].alignment = Alignment(horizontal='center')
                sheet['H'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            if subrow == row:
                row += 1
            else:
                row = subrow

        transfernotewb.save('report.xlsx')
        xlsxfile = open("report.xlsx","r")
        reportxslx = xlsxfile.read()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(reportxslx),'Content-Disposition': 'attachment; filename=report.xlsx', 'Set-Cookie':'fileDownload=true; path=/'}
        xlsxfile.close()
        os.remove("report.xlsx")
        return Response(reportxslx, headerlist=headerList.items())
    except:
        print "File not found"
        return{"gkstatus":3}
    

@view_config(route_name="transfernotes",request_param="action=showlist",renderer="gkwebapp:templates/listoftransfernotes.jinja2")
def showlistoftransfernotes(request):
                header={"gktoken":request.headers["gktoken"]}
                startDate =str(request.params["startdate"])
                endDate =str(request.params["enddate"])
                godownname = ""
                godownaddress = ""
                goid = 0
                if request.params.has_key("goid"):
                    goid = int(request.params["goid"])
                    transfernotes = requests.get("http://127.0.0.1:6543/transfernote?type=list&startdate=%s&enddate=%s&goid=%d"%(startDate, endDate, goid),headers=header)
                    godown = requests.get("http://127.0.0.1:6543/godown?qty=single&goid=%d"%(int(request.params["goid"])), headers=header)
                    godownname = godown.json()["gkresult"]["goname"]
                    godownaddress = godown.json()["gkresult"]["goaddr"]
                else:
                    transfernotes = requests.get("http://127.0.0.1:6543/transfernote?type=list&startdate=%s&enddate=%s"%(startDate, endDate),headers=header)
                return {"transfernotes":transfernotes.json()["gkresult"], "startdate":startDate, "enddate":endDate, "godownname":godownname, "godownaddress":godownaddress, "goid":goid}

@view_config(route_name="transfernotes",request_param="action=get",renderer="json")
def gettransfernote(request):
                header={"gktoken":request.headers["gktoken"]}
                result = requests.get("http://127.0.0.1:6543/transfernote?tn=single&transfernoteid=%d"%(int(request.params["transfernoteid"])), headers=header)
                return {"gkstatus": result.json()["gkstatus"], "gkresult": result.json()["gkresult"]}

@view_config(route_name="transfernotes",request_param="action=showtn",renderer="gkwebapp:templates/viewsingletransfernote.jinja2")
def showsingletransfernote(request):
                header={"gktoken":request.headers["gktoken"]}
                result = requests.get("http://127.0.0.1:6543/transfernote?tn=single&transfernoteid=%d"%(int(request.params["transfernoteid"])), headers=header)
                return {"gkstatus": result.json()["gkstatus"], "transfernote": result.json()["gkresult"]}

@view_config(route_name="transfernotes",request_param="action=save",renderer="json")
def savetransfernote(request):
                header={"gktoken":request.headers["gktoken"]}
                transferdata = {"transfernoteno":request.params["transfernoteno"],"transfernotedate":request.params["transfernotedate"],"togodown":request.params["togodown"],"fromgodown":request.params["fromgodown"],"transportationmode":request.params["transportationmode"],"issuername":request.params["issuername"],"designation":request.params["designation"]}
                if request.params["nopkt"]!='':
                        transferdata["nopkt"]=request.params["nopkt"]
                
                if request.params.has_key("duedate"):
                    transferdata["duedate"]=request.params["duedate"]
                if request.params.has_key("grace"):
                    transferdata["grace"]=request.params["grace"]
                
                products = {}
                for  row in json.loads(request.params["products"]):
                        products[row["productcode"]] = row["qty"]
                stockdata = {"items":products}
                
                tnwholedata = {"transferdata":transferdata,"stockdata":stockdata}
                result=requests.post("http://127.0.0.1:6543/transfernote",data=json.dumps(tnwholedata),headers=header)
                return {"gkstatus":result.json()["gkstatus"]}



@view_config(route_name="transfernotes",request_param="action=received",renderer="json")
def recieved(request):
                header={"gktoken":request.headers["gktoken"]}
                dataset={"transfernoteid":int(request.params["transfernoteid"]),"recieveddate":request.params["receiveddate"]}
                result=requests.put("http://127.0.0.1:6543/transfernote?received=true",data=json.dumps(dataset),headers=header)
                return {"gkstatus":result.json()["gkstatus"]}

@view_config(route_name="transfernotes",request_param="action=print",renderer="gkwebapp:templates/printtransfernote.jinja2")
def tnprint(request):
        header={"gktoken":request.headers["gktoken"]}
        org = requests.get("http://127.0.0.1:6543/organisation", headers=header)
        fromgodown=requests.get("http://127.0.0.1:6543/godown?qty=single&goid=%d"%(int(request.params["fromgodown"])), headers=header)
        togodown=requests.get("http://127.0.0.1:6543/godown?qty=single&goid=%d"%(int(request.params["togodown"])), headers=header)
        tableset = json.loads(request.params["printset"])
        return {"gkstatus":org.json()["gkstatus"],"org":org.json()["gkdata"],
                "tableset":tableset,"transfernoteno":request.params["transfernoteno"],"transfernotedate":request.params["transfernotedate"],"receiveddate":request.params["receiveddate"],"duedate":request.params["duedate"],"grace":request.params["grace"],
        "togodown":togodown.json()["gkresult"],"transportationmode":request.params["transportationmode"],"issuername":request.params["issuername"],
        "designation":request.params["designation"],"nopkt":request.params["nopkt"],"fromgodown":fromgodown.json()["gkresult"]}

@view_config(route_name="transfernotes",request_param="action=getprod",renderer="json")
def getproductsFromGodown(request):
                header={"gktoken":request.headers["gktoken"]}
                result = requests.get("http://127.0.0.1:6543/products?from=godown&godownid=%d"%int(request.params["godownid"]),headers=header)
                proddata = []
                for record in result.json()["gkresult"]:
                        product = requests.get("http://127.0.0.1:6543/products?qty=single&productcode=%d"%(int(record["productcode"])), headers=header)
                        pdata= {"productcode":str(product.json()["gkresult"]["productcode"]),"productdesc":str(product.json()["gkresult"]["productdesc"])}
                        proddata.append(pdata)
                return {"gkstatus": result.json()["gkstatus"], "products": proddata}

@view_config(route_name="transfernotes",request_param="type=stock",renderer="json")
def showstockreport(request):
        header={"gktoken":request.headers["gktoken"]}
        goid = int(request.params["goid"])
        productcode = int(request.params["productcode"])
        scalculateto = request.params["endDate"]
        result = requests.get("http://127.0.0.1:6543/report?godownwisestockonhand&type=pg&goid=%d&productcode=%d&enddate=%s"%(goid, productcode, scalculateto),headers=header)
        return {"gkresult":result.json()["gkresult"][0]["balance"]}


@view_config(route_name="transfernotes",request_param="action=getgodowns", renderer="json")
def listofgodowns(request):
        header={"gktoken":request.headers["gktoken"]}
        result = requests.get("http://127.0.0.1:6543/godown?type=togodown", headers=header)
        goddata=[]
        for record in result.json()["gkresult"]:
                gdata= {"goid": str(record["goid"]), "goname" : str(record["goname"]), "goaddr": str(record["goaddr"])}
                goddata.append(gdata)
        return {"gkstatus": result.json()["gkstatus"], "godowns":goddata}
