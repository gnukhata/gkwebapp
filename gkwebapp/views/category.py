
"""
Copyright (C) 2013, 2014, 2015, 2016 Digital Freedom Foundation
Copyright (C) 2017, 2018 Digital Freedom Foundation & Accion Labs Pvt. Ltd.

  This file is part of GNUKhata:A modular,robust and Free Accounting System.

  GNUKhata is Free Software; you can redistribute it and/or modify
  it under the terms of the GNU Affero General Public License as
  published by the Free Software Foundation; either version 3 of
  the License, or (at your option) any later version.

  GNUKhata is distributed in the hope that it will be useful, but
  WITHOUT ANY WARRANTY; without even the implied warranty of
  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
  GNU Affero General Public License for more details.

  You should have received a copy of the GNU Affero General Public
  License along with GNUKhata (COPYING); if not, write to the
  Free Software Foundation, Inc., 51 Franklin Street, Fifth Floor,
  Boston, MA  02110-1301  USA59 Temple Place, Suite 330,


Contributors:
"Krishnakant Mane" <kk@dff.org.in>
"Arun Kelkar" <arunkelkar@dff.org.in>
"Ishan Masdekar " <imasdekar@dff.org.in>
"Navin Karkera" <navin@dff.org.in>
"Vanita Rajpurohit" <vanita.rajpurohit9819@gmail.com>
"""

from pyramid.view import view_config
import requests, json
from datetime import datetime
from pyramid.renderers import render_to_response
from pyramid.response import Response
import openpyxl
from openpyxl.styles import Font, Alignment
import os

@view_config(route_name="category",renderer="gkwebapp:templates/category.jinja2")
def showcategory(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/categories", headers=header)
    categories = result.json()["gkresult"]
    gkresult = len(categories)
    """ gkresult used to decide whether to show Edit Category button or not """
    return {"status":True, "gkresult": gkresult}

@view_config(route_name="category",request_param="action=showadd",renderer="gkwebapp:templates/addcategory.jinja2")
def showaddcategory(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/categories", headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    states = requests.get("http://127.0.0.1:6543/state", headers=header)
    return {"gkstatus": result.json()["gkstatus"], "gkresult": result.json()["gkresult"],"categorycount":len(result.json()["gkresult"]),"vatorgstflag":resultgstvat.json()["gkresult"], "states": states.json()["gkresult"]}

@view_config(route_name="category",request_param="action=showedit",renderer="gkwebapp:templates/editcategory.jinja2")
def showeditcategory(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/categories?type=editablecat", headers=header)
    resultgstvat = requests.get("http://127.0.0.1:6543/products?tax=vatorgst",headers=header)
    states = requests.get("http://127.0.0.1:6543/state", headers=header)
    return {"gkstatus": result.json()["gkstatus"], "gkresult": result.json()["gkresult"],"vatorgstflag":resultgstvat.json()["gkresult"], "states": states.json()["gkresult"],"category":len(result.json()["gkresult"]),"status":True}

@view_config(route_name="category",request_param="action=getspecs",renderer="json")
def getspecs(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/categoryspecs?categorycode=%d"%(int(request.params["categorycode"])), headers=header)
    return {"gkstatus": result.json()["gkstatus"], "gkresult": result.json()["gkresult"]}

@view_config(route_name="category",request_param="action=gettax",renderer="json")
def gettax(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/tax?pscflag=c&categorycode=%d"%(int(request.params["categorycode"])), headers=header)
    return {"gkstatus": result.json()["gkstatus"], "gkresult": result.json()["gkresult"]}

@view_config(route_name="category",request_param="action=getCategory",renderer="json")
def getCategory(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/categories?type=single&categorycode=%d"%(int(request.params["categorycode"])), headers=header)
    return {"gkstatus": result.json()["gkstatus"], "gkresult": result.json()["gkresult"]}

@view_config(route_name="category",request_param="action=save",renderer="json")
def savespecs(request):
    header={"gktoken":request.headers["gktoken"]}
    if request.params["subcategoryof"]!='':
        categorydata={"categoryname":request.params["categoryname"],"subcategoryof":int(request.params["subcategoryof"])}
    else:
        categorydata={"categoryname":request.params["categoryname"]}
    result = requests.post("http://127.0.0.1:6543/categories",data=json.dumps(categorydata) ,headers=header)
    if result.json()["gkstatus"]==0:
        gkdata = {"activity":request.params["categoryname"] + " category created"}
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
        specs = json.loads(request.params["specs"])
        for spec in specs:
            specdata= {"attrname":spec["attrname"],"attrtype":int(spec["attrtype"]),"categorycode":result.json()["gkresult"]}
            specresult = requests.post("http://127.0.0.1:6543/categoryspecs",data=json.dumps(specdata) ,headers=header)
        taxes = json.loads(request.params["taxes"])
        for tax in taxes:
            taxdata= {"taxname":tax["taxname"],"taxrate":float(tax["taxrate"]),"categorycode":result.json()["gkresult"]}
            if tax["state"]!='':
                taxdata["state"]=tax["state"]
            taxresult = requests.post("http://127.0.0.1:6543/tax",data=json.dumps(taxdata) ,headers=header)
        return {"gkstatus": result.json()["gkstatus"], "gkresult":result.json()["gkresult"]}
    else:
        return {"gkstatus": result.json()["gkstatus"]}

@view_config(route_name="category",request_param="action=edit",renderer="json")
def editspecs(request):
    header={"gktoken":request.headers["gktoken"]}
    if request.params["subcategoryof"]!='':
        categorydata={"categoryname":request.params["categoryname"],"subcategoryof":int(request.params["subcategoryof"]),"categorycode":int(request.params["categorycode"])}
    else:
        categorydata={"categoryname":request.params["categoryname"],"subcategoryof":None,"categorycode":int(request.params["categorycode"])}
    result = requests.put("http://127.0.0.1:6543/categories",data=json.dumps(categorydata) ,headers=header)
    if result.json()["gkstatus"]==0:
        specs = json.loads(request.params["specs"])
        deletedspecs = json.loads(request.params["deletedspecs"])
        for spec in specs:
            if spec["spcode"] != "New":
                specdata= {"attrname":spec["attrname"],"attrtype":int(spec["attrtype"]),"spcode":int(spec["spcode"]),"categorycode":int(request.params["categorycode"])}
                specresult = requests.put("http://127.0.0.1:6543/categoryspecs",data=json.dumps(specdata) ,headers=header)
            else:
                specdata= {"attrname":spec["attrname"],"attrtype":int(spec["attrtype"]),"categorycode":int(request.params["categorycode"])}
                specresult = requests.post("http://127.0.0.1:6543/categoryspecs",data=json.dumps(specdata) ,headers=header)
        if len(deletedspecs)>0:
            for deletedspec in deletedspecs:
                deletedspecdata= {"spcode":int(deletedspec)}
                deletedspecresult = requests.delete("http://127.0.0.1:6543/categoryspecs",data=json.dumps(deletedspecdata) ,headers=header)
        taxs = json.loads(request.params["taxes"])
        deletedtaxs = json.loads(request.params["deletedtaxs"])
        for tax in taxs:
            if tax["taxid"] != "New":
                taxdata= {"taxid":tax["taxid"],"taxname":tax["taxname"],"taxrate":float(tax["taxrate"]),"categorycode":int(request.params["categorycode"])}
                if tax["state"]!='':
                    taxdata["state"]=tax["state"]
                else:
                    taxdata["state"]=None
                taxresult = requests.put("http://127.0.0.1:6543/tax",data=json.dumps(taxdata) ,headers=header)
            else:
                taxdata= {"taxname":tax["taxname"],"taxrate":float(tax["taxrate"]),"categorycode":int(request.params["categorycode"])}
                if tax["state"]!='':
                    taxdata["state"]=tax["state"]
                taxresult = requests.post("http://127.0.0.1:6543/tax",data=json.dumps(taxdata) ,headers=header)
        if len(deletedtaxs)>0:
            for deletedtax in deletedtaxs:
                deletedtaxdata= {"taxid":int(deletedtax)}
                deletedtaxresult = requests.delete("http://127.0.0.1:6543/tax",data=json.dumps(deletedtaxdata) ,headers=header)
    return {"gkstatus": result.json()["gkstatus"]}

@view_config(route_name="category",request_param="action=delete",renderer="json")
def deletecategory(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/categories?type=single&categorycode=%d"%(int(request.params["categorycode"])), headers=header)
    categoryname = result.json()["gkresult"]["categoryname"]
    categorydata = {"categorycode":int(request.params["categorycode"])}
    result = requests.delete("http://127.0.0.1:6543/categories",data=json.dumps(categorydata), headers=header)
    if result.json()["gkstatus"] == 0:
        gkdata = {"activity":categoryname + " category deleted"}
        resultlog = requests.post("http://127.0.0.1:6543/log", data =json.dumps(gkdata),headers=header)
    return {"gkstatus": result.json()["gkstatus"]}

@view_config(route_name="category",request_param="action=list",renderer="gkwebapp:templates/listofcategories.jinja2")
def listofcategories(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/categories", headers=header)
    categories = result.json()["gkresult"]
    categorydata = []
    for category in categories:
        children = requests.get("http://127.0.0.1:6543/categories?type=children&categorycode=%d"%(int(category["categorycode"])), headers=header)
        children = children.json()["gkresult"]
        childcategories = []
        for child in children:
            childcategories.append({"categoryname":child["categoryname"]})
        categorydata.append({"srno":category["srno"], "categorycode":category["categorycode"], "categoryname":category["categoryname"], "categorystatus":category["categorystatus"],"children":childcategories})
    return {"gkstatus": result.json()["gkstatus"], "gkresult": categorydata}

@view_config(route_name="category",request_param="action=printable", renderer="gkwebapp:templates/printlistofcategories.jinja2")
def printlistofgodowns(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/categories", headers=header)
    categories = result.json()["gkresult"]
    categorydata = []
    for category in categories:
        children = requests.get("http://127.0.0.1:6543/categories?type=children&categorycode=%d"%(int(category["categorycode"])), headers=header)
        children = children.json()["gkresult"]
        childcategories = []
        for child in children:
            childcategories.append({"categoryname":child["categoryname"]})
        categorydata.append({"srno":category["srno"], "categorycode":category["categorycode"], "categoryname":category["categoryname"], "categorystatus":category["categorystatus"],"children":childcategories})
    return {"gkstatus": result.json()["gkstatus"], "gkresult": categorydata}

'''
This function returns a spreadsheet form of List of Categories Report.
The spreadsheet in XLSX format is generated by the backend and sent in base64 encoded format.
It is decoded and returned along with mime information.
'''
@view_config(route_name="category",request_param="action=spreadsheet", renderer="")
def listofgodownssspreadsheet(request):
    try:
        header={"gktoken":request.headers["gktoken"]}
        result = requests.get("http://127.0.0.1:6543/categories", headers=header)
        result = result.json()["gkresult"]
        fystart = str(request.params["fystart"]);
        fyend = str(request.params["fyend"]);
        orgname = str(request.params["orgname"])
        # A workbook is opened.
        categorywb = openpyxl.Workbook()
        # The new sheet is the active sheet as no other sheet exists. It is set as value of variable - sheet.
        sheet = categorywb.active
        # Title of the sheet and width of columns are set.
        sheet.title = "List of Categories"
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 24
        sheet.column_dimensions['C'].width = 24
        sheet.column_dimensions['D'].width = 14
        # Cells of first two rows are merged to display organisation details properly.
        sheet.merge_cells('A1:G2')
        # Font and Alignment of cells are set. Each cell can be identified using the cell index - column name and row number.
        sheet['A1'].font = Font(name='Liberation Serif',size='16',bold=True)
        sheet['A1'].alignment = Alignment(horizontal = 'center', vertical='center')
        # Organisation name and financial year are displayed.
        sheet['A1'] = orgname + ' (FY: ' + fystart + ' to ' + fyend +')'
        sheet.merge_cells('A3:G3')
        sheet['A3'].font = Font(name='Liberation Serif',size='14',bold=True)
        sheet['A3'].alignment = Alignment(horizontal = 'center', vertical='center')
        sheet['A3'] = 'List of Categories'
        sheet.merge_cells('A3:G3')
        sheet['A4'] = 'Sr. No.'
        sheet['B4'] = 'Category'
        sheet['C4'] = 'Sub-Category'
        sheet['D4'] = 'Status'
        titlerow = sheet.row_dimensions[4]
        titlerow.font = Font(name='Liberation Serif',size=12,bold=True)
        row=5
        srno = 1
        for category in result:
            sheet['A'+str(row)] = srno
            sheet['A'+str(row)].alignment = Alignment(horizontal='left')
            sheet['A'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            sheet['B'+str(row)] = category['categoryname']
            sheet['B'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)
            children = requests.get("http://127.0.0.1:6543/categories?type=children&categorycode=%d"%(int(category["categorycode"])), headers=header)
            children = children.json()["gkresult"]
            subrow = row
            for child in children:
                sheet['C'+str(subrow)] = child["categoryname"]
                sheet['C'+str(subrow)].font = Font(name='Liberation Serif', size='12', bold=False) 
                subrow +=1
            sheet['D'+str(row)] = category["categorystatus"]
            sheet['D'+str(row)].font = Font(name='Liberation Serif', size='12', bold=False)            
            if subrow == row:
                row += 1
            else:
                row = subrow
            srno += 1  
        categorywb.save('report.xlsx')
        xlsxfile = open("report.xlsx","r")
        reportxslx = xlsxfile.read()
        headerList = {'Content-Type':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' ,'Content-Length': len(reportxslx),'Content-Disposition': 'attachment; filename=report.xlsx', 'Set-Cookie':'fileDownload=true; path=/'}
        xlsxfile.close()
        os.remove("report.xlsx")
        return Response(reportxslx, headerlist=headerList.items())       
    except:
        return {"gkstatus":3}
    
@view_config(route_name="category",request_param="action=tree",renderer="gkwebapp:templates/treeviewofcategories.jinja2")
def treeviewofcategories(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/categories?type=topcategories", headers=header)
    return {"gkstatus": result.json()["gkstatus"], "gkresult": result.json()["gkresult"], "numberofcategories": len(result.json()["gkresult"])}

@view_config(route_name="category",request_param="action=treechildren",renderer="json")
def childrenofcategories(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/categories?type=children&categorycode=%d"%(int(request.params["categorycode"])), headers=header)
    return {"gkstatus": result.json()["gkstatus"], "gkresult": result.json()["gkresult"]}

@view_config(route_name="category",request_param="type=countcategory",renderer="json")
def countcategory(request):
    header={"gktoken":request.headers["gktoken"]}
    result = requests.get("http://127.0.0.1:6543/categories", headers=header)
    return {"gkstatus": result.json()["gkstatus"], "categorycount": len(result.json()["gkresult"])}
